from decimal import Decimal

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from contracts.models import PriceName, PriceCoast
from directory.models import Researches


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл с картами пациентов + диагноз Д-учета
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        Испорт цен услуг
        Если услуга(id) существует записать в новый ф-л уже существующие, иначе создать новую запись
        :param args:
        :param kwargs:
        :return:
        """
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        identify = 0
        price_code = 0
        coast = 0
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "id" in cells and "код_прайс" in cells and "цена" in cells:
                    starts = True
                    identify = cells.index("id")
                    price_code = cells.index("код_прайс")
                    coast = cells.index("цена")
            else:
                price_obj = PriceName.objects.filter(pk=int(cells[price_code])).first()
                research_obj = Researches.objects.filter(pk=int(cells[identify])).first()
                if cells[coast]:
                    coast_value = Decimal(cells[coast])
                    if price_obj and research_obj:
                        PriceCoast.objects.update_or_create(price_name=price_obj, research=research_obj, defaults={'coast': coast_value})
