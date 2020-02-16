import datetime

from django.core.management.base import BaseCommand
from django.db.models import Q
from openpyxl import load_workbook

import clients.models as clients


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл с пациентами
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)

        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "снилс" in cells and "полис" in cells:
                    starts = True
                    num_card = cells.index("карта")
                    distict_num = cells.index("участок")
                    district_gin = cells.index("участок-жк")
                    pasport_serial = cells.index("серия")
                    pasport_num = cells.index("номер")
                    snils = cells.index("снилс")
                    polis = cells.index("полис")
                    base_l2 = clients.CardBase.objects.get(internal_type=True)
            else:
                # если есть индивидуал по документам
                ind = clients.Document.objects.filter(
                    Q(document_type__title__iexact="СНИЛС", number=cells[snils]) |
                    Q(document_type__title__iexact="Полис ОМС", number=cells[polis]) |
                    Q(document_type__title__iexact="Паспорт гражданина РФ", serial=cells[pasport_serial], number=cells[pasport_num])).first()
                if ind:
                    i = ind.individual
                    if clients.Card.objects.filter(individual=i, base=base_l2).exists():
                        district = clients.District.objects.filter(code_poliklinika=cells[distict_num], is_ginekolog=False).first()
                        district_ginikolog = clients.District.objects.filter(code_poliklinika=cells[district_gin], is_ginekolog=True).first()
                        clients.Card.objects.filter(individual=i, base=base_l2).update(number_poliklinika=cells[num_card],
                                                                                       district=district, ginekolog_district=district_ginikolog)
