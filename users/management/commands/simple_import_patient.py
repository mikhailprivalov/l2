import datetime
from django.core.management.base import BaseCommand
from django.db.models import Q
from openpyxl import load_workbook
import clients.models as clients


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл с пациентами и участками
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)

        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        born_date = "xxx"
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "снилс" in cells and "полис" in cells:
                    starts = True
                    num_card = cells.index("карта")
                    distict_num = cells.index("участок")
                    lastname = cells.index("фамилия")
                    name = cells.index("имя")
                    patronymic = cells.index("отчество")
                    sex = cells.index("пол")
                    snils = cells.index("снилс")
                    polis = cells.index("полис")
                    born_date = cells.index("дата рождения")
                    phone_number = cells.index("телефон")
                    base_l2 = clients.CardBase.objects.filter(internal_type=True)[0]
            else:
                district_obj = clients.District.objects.filter(title=cells[distict_num]).first()
                if district_obj is None:
                    district_obj = clients.District.objects.create(title=cells[distict_num])
                ind = clients.Document.objects.filter(Q(document_type__title__iexact="СНИЛС", number=cells[snils]) | Q(document_type__title__iexact="Полис ОМС", number=cells[polis])).first()
                if ind:
                    i = ind.individual
                    if clients.Card.objects.filter(individual=i, base=base_l2).exists():
                        for c in clients.Card.objects.filter(individual=i, base=base_l2):
                            c.number_poliklinika = cells[num_card]
                            c.district = district_obj
                            c.phone = cells[phone_number]
                            c.save(update_fields=['number_poliklinika', 'district', 'phone'])
                    else:
                        c = clients.Card.objects.create(
                            number=clients.Card.next_l2_n(),
                            base=base_l2,
                            individual=i,
                            number_poliklinika=cells[num_card],
                            district=district_obj,
                            phone=cells[phone_number],
                        )
                else:
                    ind = clients.Individual.objects.create(
                        family=cells[lastname].lower().title(),
                        name=cells[name].lower().title(),
                        patronymic=cells[patronymic].lower().title(),
                        birthday=datetime.datetime.strptime(cells[born_date], "%Y-%m-%d %H:%M:%S").date(),
                        sex=cells[sex],
                    )

                    if cells[snils]:
                        snils_object = clients.DocumentType.objects.get(title__iexact='СНИЛС')
                        clients.Document.objects.create(document_type=snils_object, number=cells[snils], individual=ind)

                    polis_object = clients.DocumentType.objects.get(title__iexact='Полис ОМС')
                    document_polis = clients.Document.objects.create(document_type=polis_object, number=cells[polis], individual=ind) if cells[polis] else None
                    c = clients.Card.objects.create(
                        individual=ind,
                        number=clients.Card.next_l2_n(),
                        base=base_l2,
                        number_poliklinika=cells[num_card],
                        polis=document_polis,
                        district=district_obj,
                        phone=cells[phone_number],
                    )
