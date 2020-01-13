from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from users.models import DoctorProfile


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл с кодами МКБ10.2019 + расшифровка
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "код" in cells and "rmis_login" and "rmis_password" in cells:
                    starts = True
                    code = cells.index("код")
                    rmis_login = cells.index("rmis_login")
                    rmis_password = cells.index("rmis_password")
            else:
                doctor = DoctorProfile.objects.filter(pk=cells[code]).first()
                if doctor:
                    doctor.rmis_login = cells[rmis_login]
                    doctor.rmis_password = cells[rmis_password]
                    doctor.save(update_fields=['rmis_login', 'rmis_password'])
                    print("Обновлен", doctor)
