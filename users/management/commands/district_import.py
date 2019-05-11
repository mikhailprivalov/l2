from django.core.management.base import BaseCommand
from openpyxl import load_workbook

class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)


    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        u_dict = {}
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "код" in cells and "название" in cells:
                    print('зашел')
                    starts = True
                    code = cells.index("код")
                    district_name = cells.index("название")
                    continue
            else:
                print(cells[code])
                print(cells[district_name])

                u_dict[cells[code]] = cells[district_name]

        print(u_dict)




