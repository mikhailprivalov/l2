from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from directory.models import Researches, ResearchSite


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл с картами пациентов + диагноз Д-учета
        """
        parser.add_argument('path', type=str)


    def handle(self, *args, **kwargs):
        """
        Испорт услуг консультацй, Лечения, Стоматологии, Стационар
        Если услуга(id) существует обновиться внутренний код, иначе создать новую
        :param args:
        :param kwargs:
        :return:
        """
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        ws1 = wb[wb.sheetnames[1]]
        starts = False
        x = 0

        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "код_внутренний" in cells and "услуга" in cells and "тип" in cells and "место" in cells and "id" in cells:
                    starts = True
                    identify = cells.index("id")
                    int_code = cells.index("код_внутренний")
                    research = cells.index("услуга")
                    type_research = cells.index("тип")
                    place_research = cells.index("место")
                    podr = cells.index("подразделение")
                    pay = cells.index("платно")

                    def insert_data(ins):
                        c1 = ws1.cell(row=x, column=1)
                        c2 = ws1.cell(row=x, column=2)
                        c3 = ws1.cell(row=x, column=3)
                        c4 = ws1.cell(row=x, column=4)
                        c5 = ws1.cell(row=x, column=5)
                        c6 = ws1.cell(row=x, column=6)
                        c6 = ws1.cell(row=x, column=7)

                        c1.value = ins
                        c2.value = cells[int_code]
                        c3.value = cells[research]
                        c4.value = cells[type_research]
                        c5.value = cells[place_research]
                        c6.value = cells[podr]
                        c6.value = cells[pay]

            else:
                if cells[identify] == '-1':
                    if Researches.objects.filter(internal_code=cells[int_code]).exists():
                        r_o = Researches.objects.values_list('pk').get(internal_code=cells[int_code])
                        insert_data(int(r_o[0]))
                        continue
                    else:
                        treatment = True if cells[type_research] == 'is_treatment' else False
                        doc_refferal = True if cells[type_research] == 'is_doc_refferal' else False
                        stom = True if cells[type_research] == 'is_stom' else False
                        hospital = True if cells[type_research] == 'is_hospital' else False
                        s_t = ResearchSite.objects.get(pk=int(cells[place_research]))
                        c = Researches.objects.create(title=cells[research], site_type=s_t,
                                                      internal_code=cells[int_code],
                                                      is_treatment=treatment, is_doc_refferal=doc_refferal,
                                                      is_hospital=hospital, is_stom=stom)
                        insert_data(int(c.pk))
                        print('добавлен услуга:', c.title, c.pk, c.internal_code)
                else:
                    pk_research = int(cells[identify])
                    res = Researches.objects.get(pk=pk_research)
                    if res:
                        Researches.objects.filter(pk=pk_research).update(internal_code=cells[int_code])
                        print('обновлена услуга (title, pk, internal_code):', res.title, res.pk, cells[int_code])
                        insert_data(int(res.pk))

        wb.save(fp + 'import')
