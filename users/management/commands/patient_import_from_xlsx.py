import datetime

from django.core.management.base import BaseCommand
from django.db.models import Q
from openpyxl import load_workbook

import clients.models as clients
from contracts.models import Company


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл с пациентами
               path_distr - файл с участками
        """
        parser.add_argument('path', type=str)
        parser.add_argument('path_distr', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        fd = kwargs["path_distr"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fd)
        ws = wb[wb.sheetnames[0]]
        distr = {}
        starts = False
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "код" in cells and "название" in cells:
                    starts = True
                    code = cells.index("код")
                    district_name = cells.index("название")
                    continue
            else:
                distr[cells[code]] = cells[district_name]
        print('загружены коды:участки' + '\n', distr)  # noqa: T001

        def get_district(uch='', gin_uch=''):
            """
            парам1 - код для обычного участка
            парам2 - код для гинекологического участка
            """
            uch.strip()
            gin_uch.strip()
            obj_uch = None
            obj_gin = None

            if uch:
                title_uch = distr.get(uch)
                if title_uch is not None:
                    obj_uch, created = clients.District.objects.get_or_create(code_poliklinika=uch, defaults={'title': title_uch, 'is_ginekolog': False, 'sort_weight': '0'})
                    if created:
                        print('Добавлен участок' + '\n', obj_uch)  # noqa: T001

            if gin_uch:
                title_gin = distr.get(gin_uch)
                if title_gin is not None:
                    obj_gin, gin_created = clients.District.objects.get_or_create(code_poliklinika=gin_uch, defaults={'title': title_gin, 'is_ginekolog': True, 'sort_weight': '0'})
                    if gin_created:
                        print('Добавлен гинекологический участок' + '\n', obj_uch)  # noqa: T001

            return [obj_uch, obj_gin]

        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "снилс" in cells and "полис" in cells:
                    starts = True
                    num_card = cells.index("карта")
                    distict_num = cells.index("участок")
                    lastname = cells.index("фамилия")
                    name = cells.index("имя")
                    patronymic = cells.index("отчество")
                    sex = cells.index("пол")
                    house = cells.index("дом")
                    room = cells.index("квартриа")
                    district_gin = cells.index("участок-жк")
                    company = cells.index("работа")
                    city = cells.index("город")
                    street = cells.index("улица")
                    pasport_serial = cells.index("серия")
                    pasport_num = cells.index("номер")
                    snils = cells.index("снилс")
                    polis = cells.index("полис")
                    born_date = cells.index("дата рождения")
                    base_l2 = clients.CardBase.objects.filter(internal_type=True)[0]
            else:
                # если есть индивидуал по документам
                ind = clients.Document.objects.filter(
                    Q(document_type__title__iexact="СНИЛС", number=cells[snils])
                    | Q(document_type__title__iexact="Полис ОМС", number=cells[polis])
                    | Q(document_type__title__iexact="Паспорт гражданина РФ", serial=cells[pasport_serial], number=cells[pasport_num])
                ).first()
                current_company = None
                if cells[company]:
                    companies = Company.objects.filter(cells[company]).first()
                    if companies:
                        current_company = companies.title
                    else:
                        try:
                            cp = Company(title=cells[company], short_title=cells[company])
                            cp.save()
                            current_company = cp
                        except:
                            current_company = None

                if ind:
                    add_dist = get_district(cells[distict_num], cells[district_gin])
                    i = ind.individual
                    if clients.Card.objects.filter(individual=i, base=base_l2).exists():
                        for c in clients.Card.objects.filter(individual=i, base=base_l2):
                            # .update(number_poliklinika=cells[num_card], district=add_dist[0], ginekolog_district=add_dist[1])
                            c.number_poliklinika = cells[num_card]
                            c.district = add_dist[0]
                            c.ginekolog_district = add_dist[1]
                            c.work_place_db = current_company
                            c.save(update_fields=['number_poliklinika', 'district', 'ginekolog_district', 'work_place_db'])
                    else:
                        # создать карту L2
                        m_address = ' '.join('{}, {}, д.{}, кв.{}'.format(cells[city], cells[street], cells[house], cells[room]).strip().split())
                        c = clients.Card.objects.create(
                            number=clients.Card.next_l2_n(),
                            base=base_l2,
                            individual=i,
                            number_poliklinika=cells[num_card],
                            district=add_dist[0],
                            ginekolog_district=add_dist[1],
                            main_address=m_address,
                            fact_address=m_address,
                            work_place_db=current_company,
                        )
                        print('Добавлена карта: \n', c)  # noqa: T001
                else:
                    # создать индивидуал, документы, карты в l2.
                    ind = clients.Individual.objects.create(
                        family=cells[lastname],
                        name=cells[name],
                        patronymic=cells[patronymic],
                        birthday=datetime.datetime.strptime(cells[born_date], "%Y-%m-%d %H:%M:%S").date(),
                        sex=cells[sex],
                    )

                    if cells[snils]:
                        snils_object = clients.DocumentType.objects.get(title__iexact='СНИЛС')
                        clients.Document.objects.create(document_type=snils_object, number=cells[snils], individual=ind)

                    if cells[pasport_serial] and cells[pasport_num]:
                        passport_object = clients.DocumentType.objects.get(title__iexact='Паспорт гражданина РФ')
                        clients.Document.objects.create(document_type=passport_object, number=cells[pasport_num], serial=cells[pasport_serial], individual=ind)

                    polis_object = clients.DocumentType.objects.get(title__iexact='Полис ОМС')
                    document_polis = clients.Document.objects.create(document_type=polis_object, number=cells[polis], individual=ind) if cells[polis] else None

                    add_dist = get_district(cells[distict_num], cells[district_gin])
                    m_address = str(cells[city]).strip() + ', ' + str(cells[street]).strip() + ', д.' + str(cells[house]).strip() + ', кв.' + str(cells[room]).strip()
                    c = clients.Card.objects.create(
                        individual=ind,
                        number=clients.Card.next_l2_n(),
                        base=base_l2,
                        number_poliklinika=cells[num_card],
                        polis=document_polis,
                        district=add_dist[0],
                        ginekolog_district=add_dist[1],
                        main_address=m_address,
                        fact_address=m_address,
                        work_place_db=current_company,
                    )
                    print('Добавлена карта: \n', c)  # noqa: T001
