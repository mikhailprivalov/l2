from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from directory.models import DispensaryRouteSheet
from directory.models import Researches


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        row_count = ws.max_row
        column_count = ws.max_column
        sex = ws.cell(row=2, column=1).value
        for r in range(2, row_count + 1):
            value_pk = ws.cell(row=r, column=4).value
            if value_pk > 0:
                research_object = Researches.objects.filter(pk=value_pk).first()
                if research_object:
                    for c in range(6, column_count + 1):
                        col_value = int(ws.cell(row=r, column=c).value)
                        if col_value > 0:
                            age = ws.cell(row=1, column=c).value
                            if DispensaryRouteSheet.objects.filter(age_client=age, sex_client=sex, research=research_object).first():
                                print(f'запись для {sex}-{age}-{research_object} существует')
                                continue
                            dr = DispensaryRouteSheet(age_client=age, sex_client=sex, research=research_object)
                            dr.save()
                            print(dr)
