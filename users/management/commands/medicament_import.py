from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from directions.models import KeyValue


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл с кодами МКБ10.2019 + расшифровка
        """
        parser.add_argument('path', type=str)


    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "мнн" in cells:
                    starts = True
                    mnn = cells.index("мнн")
                    KeyValue.objects.filter(key='mnn').delete()
            else:
                KeyValue.objects.create(key='mnn', value=cells[mnn].strip())
                print(f'добавлен МНН:{cells[mnn]}')