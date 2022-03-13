from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from users.models import Position


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл с Должностями со столбцами:
        Уникальный идентификатор
        Полное наименование
        """

        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        code_nsi, title, = (
            '',
            '',
        )
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Код" in cells:
                    code_nsi = cells.index("Код")
                    title = cells.index("Наименование")
                    starts = True
            else:

                r = Position.objects.filter(n3_id=cells[code_nsi])
                if not r.exists():
                    Position(n3_id=cells[code_nsi], title=cells[title]).save()
                    print('сохранено', cells[code_nsi], cells[title])  # noqa: T001
                elif r.exists():
                    r = r[0]
                    updated = []
                    if r.title != cells[title]:
                        r.title = cells[title]
                        updated.append('title')
                    if updated:
                        r.save(update_fields=updated)
                        print('обновлено', cells[code_nsi])  # noqa: T001
                    else:
                        print('не обновлено', cells[code_nsi])  # noqa: T001
