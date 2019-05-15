from django.core.management.base import BaseCommand
from openpyxl import load_workbook
import clients.models as clients
import datetime

class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл с картами пациентов + диагноз Д-учета
        """
        parser.add_argument('path', type=str)


    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "карта" in cells and "диагноз" in cells and "дата1" in cells:
                    starts = True
                    num_card = cells.index("карта")
                    diag = cells.index("диагноз")
                    date_start = cells.index("дата1")
            else:
                if clients.Card.objects.filter(number_poliklinika=cells[num_card]).first():
                    card = clients.Card.objects.filter(number_poliklinika=cells[num_card]).first()
                    day_start = datetime.datetime.strptime(cells[date_start], "%Y-%m-%d %H:%M:%S").date()
                    clients.DispensaryReg.objects.update_or_create(card=card, diagnos=cells[diag],
                            defaults={'date_start':day_start})
                    print('добавлен/обновлен Д-учет: \n', card, )
                    print('Диагноз:дата постановки: \n', cells[diag], day_start)


