from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from pharmacotherapy.models import Drugs


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл с кодами МКБ10.2019 + расшифровка
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        print(ws)
        starts = False
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "мнн" in cells and "торг" in cells:
                    starts = True
                    torg = cells.index("торг")
                    mnn = cells.index("мнн")
            else:
                print(cells[mnn])
                if cells[mnn] == "~":
                    continue
                Drugs.objects.create(mnn=cells[mnn][:255], trade_name=cells[torg][:255])
                print(f'добавлен МНН:{cells[mnn]}:{cells[torg]}')
