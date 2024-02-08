from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from external_system.models import ProfessionsWorkersPositionsRefbook


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл с ФСЛИ со столбцами:
        ID,
        NAME
        """
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False

        id_position, name_position = '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "ID" in cells:
                    id_position = cells.index("ID")
                    name_position = cells.index("NAME")
                    starts = True
            else:
                r = ProfessionsWorkersPositionsRefbook.objects.filter(code=cells[id_position])
                if not r.exists():
                    ProfessionsWorkersPositionsRefbook(code=cells[id_position], title=cells[name_position]).save()
                    self.stdout.write('сохранено', cells[name_position])
                elif r.exists():
                    r = r[0]
                    updated = []
                    if r.title != cells[name_position]:
                        r.title = cells[name_position]
                        updated.append('title')
                    if updated:
                        r.save(update_fields=updated)
                        self.stdout.write('обновлено', cells[name_position])
