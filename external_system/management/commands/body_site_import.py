from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from external_system.models import BodySiteRefbook


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл:
        code
        title
        """

        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        code, title = '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Код" in cells:
                    code = cells.index("Код")
                    title = cells.index("Наименование")
                    starts = True
            else:
                r = BodySiteRefbook.objects.filter(code=cells[code])
                if not r.exists():
                    BodySiteRefbook(
                        code=cells[code],
                        title=cells[title],
                    ).save()
                    print('сохранено', cells[code])  # noqa: T001
                elif r.exists():
                    r = r[0]
                    updated = []
                    if r.title != cells[title]:
                        r.title = cells[title]
                        updated.append('title')
                    if updated:
                        r.save(update_fields=updated)
                        print('обновлено', cells[code])  # noqa: T001
                    else:
                        print('не обновлено', cells[code])  # noqa: T001
