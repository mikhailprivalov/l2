from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from external_system.models import FsliRefbookTest


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл с ФСЛИ со столбцами:
        Уникальный идентификатор
        Код LOINC
        Полное наименование
        Английское наименование
        Краткое наименование
        Синонимы
        Аналит
        Характеристика аналита
        Размерность
        Единица измерения
        Образец
        Временная характеристика образца
        Тип метода
        Тип шкалы измерения
        Статус
        Группа тестов
        Код НМУ
        Порядок сортировки"""

        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        code_fsli, code_loinc, title, english_title, short_title, synonym, analit, analit_props, dimension = '', '', '', '', '', '', '', '', ''
        unit, sample, time_characteristic_sample, method_type, scale_type, actual, test_group, code_nmu, sort_num, starts = '', '', '', '', '', '', '', '', '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Уникальный идентификатор" in cells:
                    code_fsli = cells.index("Уникальный идентификатор")
                    code_loinc = cells.index("Код LOINC")
                    title = cells.index("Полное наименование")
                    english_title = cells.index("Английское наименование")
                    short_title = cells.index("Краткое наименование")
                    synonym = cells.index("Синонимы")
                    analit = cells.index("Аналит")
                    analit_props = cells.index("Характеристика аналита")
                    dimension = cells.index("Размерность")
                    unit = cells.index("Единица измерения")
                    sample = cells.index("Образец")
                    time_characteristic_sample = cells.index("Временная характеристика образца")
                    method_type = cells.index("Тип метода")
                    scale_type = cells.index("Тип шкалы измерения")
                    actual = cells.index("Статус")
                    test_group = cells.index("Группа тестов")
                    code_nmu = cells.index("Код НМУ")
                    sort_num = cells.index("Порядок сортировки")
                    starts = True
            else:
                active = True
                if cells[actual].lower() == "устаревший":
                    active = False
                r = FsliRefbookTest.objects.filter(code_fsli=cells[code_fsli])
                if not r.exists():
                    FsliRefbookTest(
                        code_fsli=cells[code_fsli],
                        code_loinc=cells[code_loinc],
                        title=cells[title],
                        english_title=cells[english_title],
                        short_title=cells[short_title],
                        synonym=cells[synonym],
                        analit=cells[analit],
                        analit_props=cells[analit_props],
                        dimension=cells[dimension],
                        unit=cells[unit],
                        sample=cells[sample],
                        time_characteristic_sample=cells[time_characteristic_sample],
                        method_type=cells[method_type],
                        scale_type=cells[scale_type],
                        actual=cells[actual],
                        test_group=cells[test_group],
                        code_nmu=cells[code_nmu],
                        ordering=cells[sort_num] or -1,
                        active=active,
                    ).save()
                    print('сохранено', cells[code_fsli])  # noqa: T001
                elif r.exists():
                    r = r[0]
                    updated = []
                    if r.code_loinc != cells[code_loinc]:
                        r.code_loinc = cells[code_loinc]
                        updated.append('code_loinc')
                    if r.title != cells[title]:
                        r.title = cells[title]
                        updated.append('title')
                    if r.english_title != cells[english_title]:
                        r.english_title = cells[english_title]
                        updated.append('english_title')
                    if r.actual != cells[actual]:
                        r.actual = cells[actual]
                        updated.append('actual')
                    if r.active != active:
                        r.active = active
                        updated.append('active')
                    ordering = cells[sort_num] or None
                    ordering = None if ordering == 'None' else ordering
                    if r.ordering != ordering:
                        r.ordering = ordering
                        updated.append('ordering')
                    if updated:
                        r.save(update_fields=updated)
                        print('обновлено', cells[code_fsli])  # noqa: T001
                    else:
                        print('не обновлено', cells[code_fsli])  # noqa: T001
