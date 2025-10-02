from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from directory.models import Researches
from external_system.models import InstrumentalResearchRefbook
from podrazdeleniya.models import Podrazdeleniya


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл с ФСЛИ со столбцами:
        code_nsi
        title
        method
        area
        localization
        active
        code_nmu
        """

        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        code, title, method, area, localization, actual, code_nmu, short_title = '', '', '', '', '', '', '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Код" in cells:
                    code = cells.index("Код")
                    title = cells.index("Наименование")
                    method = cells.index("Метод")
                    area = cells.index("Область")
                    localization = cells.index("Локализация")
                    code_nmu = cells.index("НМУ")
                    actual = cells.index("Статус")
                    short_title = cells.index("Короткое")
                    starts = True
            else:
                if cells[actual].lower() != "актуальный":
                    continue
                r = InstrumentalResearchRefbook.objects.filter(code_nsi=cells[code])
                if not r.exists():
                    InstrumentalResearchRefbook(
                        code_nsi=cells[code],
                        title=cells[title],
                        short_title=cells[short_title],
                        method=cells[method],
                        area=cells[area],
                        code_nmu=cells[code_nmu],
                        localization=cells[localization],
                    ).save()
                    print('сохранено', cells[code])  # noqa: T001
                elif r.exists():
                    r = r[0]
                    updated = []
                    if r.title != cells[title]:
                        r.title = cells[title]
                        updated.append('title')
                    if r.localization != cells[localization]:
                        r.localization = cells[localization]
                        updated.append('localization')
                    if r.code_nmu != cells[code_nmu]:
                        r.code_nmu = cells[code_nmu]
                        updated.append('code_nmu')
                    if r.area != cells[area]:
                        r.area = cells[area]
                        updated.append('area')
                    if r.method != cells[method]:
                        r.method = cells[method]
                        updated.append('method')
                    if updated:
                        r.save(update_fields=updated)
                        print('обновлено', cells[code])  # noqa: T001
                    else:
                        print('не обновлено', cells[code])  # noqa: T001

                if Researches.objects.filter(nsi_id=cells[code]).exists():
                    r = Researches.objects.filter(nsi_id=cells[code]).first()
                    r.short_title=cells[short_title]
                    r.title=cells[title]
                    r.is_paraclinic=True
                    r.code=cells[code_nmu]
                    r.save()
                else:
                    podrazdeleniye = Podrazdeleniya.objects.filter(title=cells[method]).first()
                    if not podrazdeleniye:
                        podrazdeleniye = Podrazdeleniya(title=cells[method], p_type=3).save()
                    Researches(
                        title=cells[title],
                        short_title=cells[short_title],
                        nsi_id=cells[code],
                        podrazdeleniye=podrazdeleniye,
                        is_paraclinic=True,
                        code=cells[code_nmu]
                    ).save()
                    print('добавлен услуга:', cells[title], cells[code])  # noqa: T001
