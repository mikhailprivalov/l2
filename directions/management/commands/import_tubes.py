import re

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from researches.models import Tubes


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл со столбцами:
        Контейнер
        """

        colors = {
            "красн": "#FF0000",
            "оранж": "#FF8C00",
            "желт": "#FFFF00",
            "жёлт": "#FFFF00",
            "зелен": "#008000",
            "голуб": "#00FFFF",
            "синий": "#0000FF",
            "синего": "#0000FF",
            "фиолет": "#9400D3",
            "сирен": "#c8a2c8",
            "розов": "#FF1493",
            "белый": "#FFFFFF",
            "белая": "#FFFFFF",
            "белого": "#FFFFFF",
            "белой": "#FFFFFF",
            "серый": "#808080",
            "серая": "#808080",
            "серого": "#808080",
            "серой": "#808080",
            "корич": "#8B4513",
        }
        colors_string = "|".join(list(colors.keys()))
        fp = kwargs['path']
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        tubes_title = ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if 'Контейнер' in cells:
                    tubes_title = cells.index('Контейнер')
                    starts = True
            elif starts and cells[tubes_title] != 'None':
                normalized_title = re.split('([/;])', cells[tubes_title])[0].strip()
                color = colors.get('серый')
                color_in_string = re.search(f'({colors_string})', normalized_title, flags=re.IGNORECASE)
                if color_in_string:
                    color = colors.get(color_in_string.group(0))
                current_tubes = Tubes.objects.filter(title=normalized_title).first()
                if not current_tubes:
                    new_tubes = Tubes(title=normalized_title, color=color)
                    new_tubes.save()
                    self.stdout.write(f'Пробирка добавлена - {new_tubes.title}')
