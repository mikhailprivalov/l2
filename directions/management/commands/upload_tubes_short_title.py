from django.core.management.base import BaseCommand
from django.db.models import Q
from openpyxl.workbook import Workbook

from appconf.manager import SettingManager
from directory.models import Researches, Fractions, Unit
from openpyxl import load_workbook

from external_system.models import FsliRefbookTest
from researches.models import Tubes


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        title_idx, short_title_idx, = '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Полное название" in cells:
                    title_idx = cells.index("Полное название")
                    short_title_idx = cells.index("Краткое")
                    starts = True
            else:
                title = cells[title_idx].strip()
                short_title = cells[short_title_idx].strip()

                if short_title == "None":
                    continue
                if len(short_title) > 6:
                    self.stdout.write(f'{title} - Не больше 6 символов')
                    continue
                tube: Tubes = Tubes.objects.filter(title__iexact=title).first()
                tube.short_title = short_title
                tube.save()
                self.stdout.write(f'Пробирка {tube.title} обновлена')
