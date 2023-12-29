
from django.core.management.base import BaseCommand
from django.db.models import Q
from openpyxl import load_workbook

from directory.models import Researches

class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл со столбцами:
        Название
        Код
        Контейнер
        """

        fp = kwargs['path']
        self.stdout.write('Path: ' + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        research_title_idx, research_code_idx, research_tubes_idx = -1, -1, -1,
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts and 'Код' in cells:
                research_title_idx = cells.index('Название')
                research_code_idx = cells.index('Код')
                research_tubes_idx = cells.index('Контейнер')
                starts = True
            elif starts and cells[research_title_idx] != 'None':
                current_research = Researches.objects.filter(Q(internal_code=cells[research_code_idx]) | Q(title=cells[research_title_idx])).first()
                print(current_research)
                if not current_research:
                    new_research = Researches(title=cells[research_title_idx], internal_code=cells[research_code_idx], )
                    print(new_research)
