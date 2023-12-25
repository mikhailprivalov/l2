from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from clients.models import District, Document, CardDocUsage
import logging

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл с участками и снилсами
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        r = 0
        for row in ws.rows:
            r += 1
            cells = [str(x.value) for x in row]
            if not starts:
                if "участок" in cells:
                    starts = True
                    distict_num = cells.index("участок")
                    snils = cells.index("снилс")
            else:
                snils_number = cells[snils]
                district_number = cells[distict_num].replace("'", '').strip()
                distict_obj = District.objects.filter(title=district_number, is_ginekolog=False).first()
                if not distict_obj:
                    distict_obj = District(title=district_number, is_ginekolog=False)
                    distict_obj.save()
                document = Document.objects.filter(number=snils_number, document_type__title="СНИЛС").first()
                if document:
                    cards_document = CardDocUsage.objects.filter(document=document)
                    if cards_document:
                        for card_doc in cards_document:
                            card = card_doc.card
                            card.district = distict_obj
                            card.save()
                            logger.error(f'{snils_number};загружен')  # noqa: T001
                else:
                    logger.error(f'{snils_number};Не загружен')  # noqa: T001
