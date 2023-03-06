from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from ecp_integration.integration import search_patient_ecp_by_fio, attach_patient_ecp
from utils.dates import normalize_dots_date
import logging

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл с пациентами - участок ecp
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        r = 0
        for row in ws.rows:
            r += 1
            cells = [str(x.value) for x in row]
            if not starts:
                if "полис" in cells:
                    starts = True
                    distict_num = cells.index("участок")
                    lastname = cells.index("Фамилия")
                    name = cells.index("Имя")
                    patronymic = cells.index("Отчество")
                    birthday = cells.index("дата рождения")
            else:
                # поиск person_id пациента в ЕЦП
                patient = {'family': '', 'name': '', 'patronymic': '', 'birthday': ''}
                patient['family'] = cells[lastname]
                patient['name'] = cells[name]
                patient['patronymic'] = cells[patronymic]
                patient['birthday'] = normalize_dots_date(cells[birthday]).split(' ')[0]
                patient['snils'] = ''

                person_id = search_patient_ecp_by_fio(patient)
                if not person_id:
                    logger.error(f"{r};Нет в ЕЦП; -; -; {cells[lastname]};{cells[name]};{cells[patronymic]};{cells[birthday]}")
                    continue
                district = cells[distict_num].replace("'", '')

                if len(district) < 15:
                    logger.error(f"{r};Нет учатска в ЕЦП; -; -; {cells[lastname]};{cells[name]};{cells[patronymic]};{cells[birthday]};{district}")
                    continue
                else:
                    attach_data = attach_patient_ecp(person_id, district, "2022-02-22", "10379", "2")
                attach_data = "None" if attach_data is None else attach_data
                logger.error(f'{r};Добавлен участок; {person_id};{attach_data};{cells[lastname]};{cells[name]};{cells[patronymic]};{cells[birthday]}')  # noqa: T001
