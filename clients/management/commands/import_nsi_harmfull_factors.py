from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from clients.models import HarmfulFactor


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл со столбцами:
        Идентификатор
        Код
        Наименование
        """

        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        harmfull_title, description = '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "ИД" in cells:
                    nsi_id = cells.index("ИД")
                    description = cells.index("Описание")
                    harmfull_title = cells.index("Номер")
                    starts = True
            else:
                harmful_factor = HarmfulFactor.objects.filter(title=cells[harmfull_title]).first()
                if harmful_factor:
                    harmful_factor.nsi_id = cells[nsi_id]
                    harmful_factor.save()
                    self.stdout.write(f'Фактор {harmful_factor.title} обновлён, nsi_id={harmful_factor.nsi_id}')
                else:
                    new_harmful_factor = HarmfulFactor(title=cells[harmfull_title], description=cells[description], nsi_id=cells[nsi_id])
                    new_harmful_factor.save()
                    self.stdout.write(f'Фактор {cells[harmfull_title]} добавлен')
