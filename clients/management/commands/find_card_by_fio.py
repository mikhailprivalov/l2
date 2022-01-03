import re
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from api.directions.sql_func import get_confirm_direction_patient_year, get_lab_podr
from api.patients.views import full_patient_search_data
from clients.models import Individual, Card
import datetime


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл с ФСЛИ со столбцами:
        ФИО
        Дата рождения
        """
        data_arg = kwargs["path"].split(";")
        fp = data_arg[0]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        fio, born = '', ''
        p = re.compile(r'^[а-яё]{3}[0-9]{8}$', re.IGNORECASE)
        start_date = data_arg[1]
        end_date = data_arg[2]
        is_lab, is_paraclinic, is_doc_refferal = True, False, False
        if len(data_arg) > 3 and data_arg[3] == "is_paraclinic":
            is_lab, is_paraclinic, is_doc_refferal = False, True, False
        elif len(data_arg) > 3 and data_arg[3] == "is_doc_refferal":
            is_lab, is_paraclinic, is_doc_refferal = False, False, True
        lab_podr = get_lab_podr()
        lab_podr = [i[0] for i in lab_podr]
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Ф.И.О." in cells:
                    fio = cells.index("Ф.И.О.")
                    born = cells.index("Дата рождения")
                    starts = True
            else:
                query = f"{cells[fio]} {cells[born]}"
                if not cells[fio]:
                    break
                f, n, p, rmis_req, split = full_patient_search_data(p, query)
                if len(split) > 3 or (len(split) == 3 and split[-1].isdigit()):
                    objects = Individual.objects.filter(family__istartswith=f, name__istartswith=n, card__base__internal_type=True, birthday=datetime.datetime.strptime(cells[born], "%d.%m.%Y").date())
                    if len(split) > 3:
                        objects.filter(patronymic__istartswith=p)
                    objects = objects[:10]
                else:
                    objects = Individual.objects.filter(family__istartswith=f, name__istartswith=n, patronymic__istartswith=p, card__base__internal_type=True)[:10]
                cards = Card.objects.filter(base__internal_type=True, individual__in=objects)
                if cards:
                    for c in cards:
                        confirmed_directions = get_confirm_direction_patient_year(start_date, end_date, lab_podr, c.pk, is_lab, is_paraclinic, is_doc_refferal)
                        if not confirmed_directions:
                            continue
                        directions = []
                        for d in confirmed_directions:
                            if d.direction not in directions:
                                directions.append(d.direction)
                        print(f"{c.pk};{c.number};{c.individual};{c.individual.birthday};{directions}") # noqa: T001
