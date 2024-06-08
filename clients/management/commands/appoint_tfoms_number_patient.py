from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from clients.models import Document, CardDocUsage
import logging

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл с участками и снилсами
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        r = 0
        count = 0
        enp, snils, number_card = None, None, None
        for row in ws.rows:
            r += 1
            cells = [str(x.value) for x in row]
            if not starts:
                if "номер" in cells:
                    starts = True
                    number_card = cells.index("номер")
                    enp = cells.index("полис")
            else:
                enp_number = cells[enp]
                number_poliklinika = cells[number_card].replace("'", '').strip()
                document = Document.objects.filter(number=enp_number, document_type__title="Полис ОМС").first()
                if document:
                    cards_documents = CardDocUsage.objects.filter(document=document)
                    if cards_documents:
                        for card_doc in cards_documents:
                            card = card_doc.card
                            card.number_poliklinika = number_poliklinika.strip()
                            card.save()
                            count += 1
                            logger.error(f'{number_poliklinika};загружен')  # noqa: T001

                else:
                    logger.error(f'{number_poliklinika}')  # noqa: T001
            logger.error(f'{count}; кол-во')  # noqa: T001
