from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from clients.models import HarmfulFactor


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл со столбцами:
        Идентификатор
        Код
        Наименование
        """

        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        cpp_key, title, description = '', '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Идентификатор" in cells:
                    cpp_key = cells.index("Идентификатор")
                    title = cells.index("Код")
                    description = cells.index("Наименование")
                    starts = True
            else:
                harmful_factor = HarmfulFactor.objects.filter(title=cells[title]).first()
                if harmful_factor:
                    harmful_factor.cpp_key = cells[cpp_key]
                    harmful_factor.save()
                    self.stdout.write(f'Фактор {harmful_factor.title} обновлён, UUID={harmful_factor.cpp_key}')
                else:
                    new_harmful_factor = HarmfulFactor(title=cells[title], description=cells[description], cpp_key=cells[cpp_key])
                    new_harmful_factor.save()
                    self.stdout.write(f'Фактор {new_harmful_factor.title} создан, UUID={new_harmful_factor.cpp_key}')
