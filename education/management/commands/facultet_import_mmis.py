from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from education.models import Faculties


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл со столбцами:
        Код
        Факультет"""

        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        code, title, short_title = '', '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Код" in cells:
                    code = cells.index("Код")
                    title = cells.index("Факультет")
                    short_title = cells.index("Сокращение")
                    starts = True
            else:
                r = Faculties.objects.filter(mmis_id=cells[code])
                if not r.exists():
                    Faculties(mmis_id=cells[code], title=cells[title], short_title=cells[short_title]).save()
                elif r.exists():
                    r = r[0]
                    updated = []
                    if r.title != cells[title]:
                        r.title = cells[title]
                        updated.append('title')
                    if r.short_title != cells[short_title]:
                        r.short_title = cells[short_title]
                        updated.append('short_title')
                    if updated:
                        r.save(update_fields=updated)
