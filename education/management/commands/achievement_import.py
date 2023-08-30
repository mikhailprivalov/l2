from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from education.models import AchievementType


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл со столбцами:
        КодДостижения
        Название
        Сокращение
        Балл
        Год
        """

        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        mmis_id, title, short_title, grade, year = '', '', '', '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "КодДостижения" in cells:
                    mmis_id = cells.index("КодДостижения")
                    title = cells.index("Название")
                    short_title = cells.index("Сокращение")
                    grade = cells.index("Балл")
                    year = cells.index("Год")
                    starts = True
            else:
                r = AchievementType.objects.filter(mmis_id=cells[mmis_id])
                if not r.exists():
                    AchievementType(
                        mmis_id=cells[mmis_id],
                        title=cells[title],
                        short_title=cells[short_title],
                        grade=cells[grade],
                        year=cells[year],
                    ).save()
                elif r.exists():
                    r = r[0]
                    updated = []
                    if r.title != cells[title]:
                        r.title = cells[title]
                        updated.append('title')
                    if r.short_title != cells[short_title]:
                        r.short_title = cells[short_title]
                        updated.append('short_title')
                    if r.year != cells[year]:
                        r.year = cells[year]
                        updated.append('year')
                    if r.grade != cells[grade]:
                        r.grade = cells[grade]
                        updated.append('grade')
                    if updated:
                        r.save(update_fields=updated)
