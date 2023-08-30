from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from education.models import EducationSpeciality


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл со столбцами:
        Код
        ОКСО
        Код_Факультета
        Квалификация
        Название_Спец
        Срок_Обучения
        Шифр
        ОО
        ЦН
        СН
        Всего
        ГодНабора
        """

        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        mmis_id, title, okso, cipher, faculties_mmis_id, qualification_title, period_study, year_start_study = "", "", "", "", "", "", "", "",
        oo_count, cn_count, sn_count, total_count = "", "", "", ""

        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Код" in cells:
                    mmis_id = cells.index("Код")
                    title = cells.index("Название_Спец")
                    okso = cells.index("ОКСО")
                    cipher = cells.index("Шифр")
                    faculties_mmis_id = cells.index("Код_Факультета")
                    qualification_title = cells.index("Квалификация")
                    period_study = cells.index("Срок_Обучения")
                    year_start_study = cells.index("ГодНабора")
                    oo_count = cells.index("ОО")
                    cn_count = cells.index("ЦН")
                    sn_count = cells.index("СН")
                    total_count = cells.index("Всего")
                    starts = True
            else:
                r = EducationSpeciality.objects.filter(mmis_id=cells[mmis_id])
                data_sn_count = cells[sn_count]
                data_cn_count = cells[cn_count]
                data_oo_count = cells[oo_count]
                data_total_count = cells[total_count]
                if data_sn_count == "None":
                    data_sn_count = 0
                if data_cn_count == "None":
                    data_cn_count = 0
                if data_oo_count == "None":
                    data_oo_count = 0
                if data_total_count == "None":
                    data_total_count = 0
                if not r.exists():
                    EducationSpeciality(
                        mmis_id=cells[mmis_id],
                        title=cells[title],
                        okso=cells[okso],
                        cipher=cells[cipher],
                        faculties_mmis_id=cells[faculties_mmis_id],
                        qualification_title=cells[qualification_title],
                        period_study=cells[period_study],
                        year_start_study=cells[year_start_study],
                        oo_count=data_oo_count,
                        cn_count=data_cn_count,
                        sn_count=data_sn_count,
                        total_count=data_total_count
                    ).save()
                    print('сохранено', cells[title])  # noqa: T001
                elif r.exists():
                    r = r[0]
                    updated = []
                    if r.title != cells[title]:
                        r.title = cells[title]
                        updated.append('title')
                    if r.okso != cells[okso]:
                        r.okso = cells[okso]
                        updated.append('okso')
                    if r.cipher != cells[cipher]:
                        r.cipher = cells[cipher]
                        updated.append('cipher')
                    if r.faculties_mmis_id != cells[faculties_mmis_id]:
                        r.faculties_mmis_id = cells[faculties_mmis_id]
                        updated.append('faculties_mmis_id')
                    if r.qualification_title != cells[qualification_title]:
                        r.qualification_title = cells[qualification_title]
                        updated.append('qualification_title')
                    if r.oo_count != cells[oo_count]:
                        if cells[oo_count] == "None":
                            r.oo_count = 0
                        else:
                            r.oo_count = cells[oo_count]
                        updated.append('oo_count')
                    if r.cn_count != cells[cn_count]:
                        if cells[cn_count] == "None":
                            r.cn_count = 0
                        else:
                            r.cn_count = cells[cn_count]
                        updated.append('cn_count')
                    if r.sn_count != cells[sn_count]:
                        if cells[sn_count] == "None":
                            r.sn_count = 0
                        else:
                            r.sn_count = cells[sn_count]
                        updated.append('sn_count')
                    if r.total_count != cells[total_count]:
                        if cells[total_count] == "None":
                            r.total_count = 0
                        else:
                            r.total_count = cells[total_count]
                        updated.append('total_count')
                    if updated:
                        r.save(update_fields=updated)
                        print('обновлено', cells[title])  # noqa: T001
                    else:
                        print('не обновлено', cells[title])  # noqa: T0
