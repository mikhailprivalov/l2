from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from education.models import Subjects


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл со столбцами:
        Дисциплина
        Код
        ДисСокращ
        """

        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        mmis_id, title, short_title = '', '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Код" in cells:
                    mmis_id = cells.index("Код")
                    title = cells.index("Дисциплина")
                    short_title = cells.index("ДисСокращ")
                    starts = True
            else:
                r = Subjects.objects.filter(mmis_id=cells[mmis_id])
                if not r.exists():
                    Subjects(mmis_id=cells[mmis_id], title=cells[title],short_title=cells[short_title]).save()
                    print('сохранено', cells[short_title])  # noqa: T001
                elif r.exists():
                    r = r[0]
                    updated = []
                    if r.title != cells[title]:
                        r.title = cells[title]
                        updated.append('title')
                    if r.short_title != cells[short_title]:
                        r.short_title = cells[short_title]
                        updated.append('short_title')
                    if updated:
                        r.save(update_fields=updated)
                        print('обновлено', cells[short_title])  # noqa: T001
                    else:
                        print('не обновлено', cells[short_title])  # noqa: T0
