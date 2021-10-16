from django.core.management.base import BaseCommand
import simplejson as json
from openpyxl import load_workbook
from directory.models import ParaclinicInputField


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('field_id', type=int)
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]

        field_id = kwargs["field_id"]
        starts = False
        start_val = []
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "значение" in cells:
                    starts = True
                    value = cells.index("значение")
                    field = ParaclinicInputField.objects.get(pk=field_id)
                    field.input_templates = "[]"
                    field.field_type = 10
            else:
                start_val.append(cells[value])

        field = ParaclinicInputField.objects.get(pk=field_id)
        print(field)
        print(start_val)
        field.input_templates = json.dumps(start_val)
        field.save()
