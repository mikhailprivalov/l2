from django.core.management.base import BaseCommand
from django.db.models import Q
from openpyxl.workbook import Workbook

from appconf.manager import SettingManager
from directory.models import Researches, Fractions, Unit
from openpyxl import load_workbook

from external_system.models import FsliRefbookTest


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        result_wb = Workbook()
        result_ws = result_wb[result_wb.sheetnames[0]]
        result_ws.append(['Название', 'Список услуг', 'Код ФСЛИ', 'статус'])

        starts = False
        title, unit, research, fsli, code = '', '', '', '', ''
        count = 0
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Код ФСЛИ" in cells:
                    title = cells.index("Название")
                    unit = cells.index("Ед.Изм.")
                    research = cells.index("Список услуг")
                    fsli = cells.index("Код ФСЛИ")
                    code = cells.index("Код")
                    starts = True
            else:
                research_nmy_code = cells[research].split(' - ')[0]
                if research_nmy_code != "None":
                    fraction_fsli_code = cells[fsli].split(' ')[0]
                    current_research = Researches.objects.filter(code=research_nmy_code).first()
                    if current_research:
                        current_fractions = Fractions.objects.filter(research_id=current_research.pk).order_by('sort_weight')
                        need_add_fractions = True
                        relation_id = None
                        sort_weight = None
                        for fraction in current_fractions:
                            if fraction.fsli == fraction_fsli_code:
                                need_add_fractions = False
                            relation_id = fraction.relation_id
                            sort_weight = fraction.sort_weight
                        if need_add_fractions:
                            unit_string = cells[unit].strip() if cells[unit] != "None" else ""
                            unit_id = None
                            unit_db = None
                            if unit_string:
                                unit_db = Unit.objects.filter(Q(short_title__iexact=unit_string) | Q(short_title__icontains=unit_string) | Q(title__icontains=unit_string)).first()
                            else:
                                fsli_test = FsliRefbookTest.objects.filter(code_fsli=fraction_fsli_code).first()
                                if fsli_test and fsli_test.unit:
                                    unit_db = Unit.objects.filter(Q(short_title__iexact=fsli_test.unit) | Q(short_title__icontains=fsli_test.unit) | Q(title__icontains=fsli_test.unit)).first()
                            if unit_db:
                                unit_id = unit_db.pk
                            if relation_id is not None:
                                new_fraction = Fractions(
                                    research_id=current_research.pk,
                                    relation_id=relation_id,
                                    title=cells[title],
                                    unit_id=unit_id,
                                    units=unit_string,
                                    fsli=fraction_fsli_code,
                                    sort_weight=sort_weight + 1,
                                    external_code=cells[code],
                                )
                                new_fraction.save()
                                self.stdout.write(f'Услуге: {current_research.title} добавлена фракция: {new_fraction.title}')
                                result_ws.append([cells[title], cells[research], cells[fsli], '+'])
                        else:
                            result_ws.append([cells[title], cells[research], cells[fsli], 'уже есть'])
                    else:
                        result_ws.append([cells[title], cells[research], cells[fsli], 'Услуга не найдена'])
                else:
                    result_ws.append([cells[title], cells[research], cells[fsli], 'Нет НМУ кода'])

        dir_tmp = SettingManager.get("dir_param")
        result_wb.save(f"{dir_tmp}/result_import_fractions.xlsx")
