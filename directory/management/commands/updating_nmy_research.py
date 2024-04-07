from django.core.management.base import BaseCommand
from openpyxl.workbook import Workbook

from appconf.manager import SettingManager
from directory.models import Researches
from openpyxl import load_workbook


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        result_wb = Workbook()
        result_ws = result_wb[result_wb.sheetnames[0]]
        result_ws.append(['НМУ', 'Название', 'Обновлено'])

        starts = False
        internal_code, nmy_cod, title = '', '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Код по прайсу" in cells:
                    internal_code = cells.index("Код по прайсу")
                    nmy_cod = cells.index("Код ОКМУ")
                    title = cells.index("Наименование услуги")
                    starts = True
            else:
                if cells[internal_code] != "None" and cells[nmy_cod] != "None":
                    current_research = Researches.objects.filter(internal_code=cells[internal_code]).first()
                    if current_research:
                        current_research.code = cells[nmy_cod]
                        current_research.save()
                        self.stdout.write(f"Услуге: {current_research.title} присвоен код НМУ: {current_research.code}")
                        result_ws.append([current_research.code, current_research.title, '+'])
                    else:
                        result_ws.append([cells[nmy_cod], cells[title], '-'])
                else:
                    result_ws.append([cells[nmy_cod], cells[title], '-'])
        dir_tmp = SettingManager.get("dir_param")
        result_ws.save(f"{dir_tmp}/result_update_nmy.xlsx")
