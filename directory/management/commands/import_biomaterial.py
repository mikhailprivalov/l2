from django.core.management.base import BaseCommand
from directory.models import LaboratoryMaterial
from openpyxl import load_workbook


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        title = ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Материал" in cells:
                    title = cells.index("Материал")
                    starts = True
            else:
                material_obj = LaboratoryMaterial.objects.filter(title=cells[title].strip()).first()
                if not material_obj:
                    LaboratoryMaterial(title=cells[title]).save()
                    self.stdout.write(f'Материал добавлен - {cells[title]}')
