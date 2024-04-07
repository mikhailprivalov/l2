from django.core.management.base import BaseCommand
from directory.models import Researches
from openpyxl import load_workbook


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        internal_code, nmy_cod, = '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Код по прайсу" in cells:
                    internal_code = cells.index("Код по прайсу")
                    nmy_cod = cells.index("Код ОКМУ")
                    starts = True
            else:
                if cells[internal_code] != "None" and cells[nmy_cod] != "None":
                    current_research = Researches.objects.filter(internal_code=cells[internal_code]).first()
                    if current_research:
                        current_research.code = cells[nmy_cod]
                        current_research.save()
                        self.stdout.write(f"Услуге: {current_research.title} присвоен код НМУ: {current_research.code}")

