from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from directory.models import Antibiotic, GroupAntibiotic


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл с микроорганизмами со столбцами:
        Название, Группа, LIS(код)
        """

        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        title, group, lis = '', '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Название" in cells:
                    title = cells.index("Название")
                    group = cells.index("Группа")
                    lis = cells.index("LIS")
                    starts = True
            else:
                group_culture = GroupAntibiotic.objects.filter(title=cells[group])
                if not group_culture.exists():
                    GroupAntibiotic(title=cells[group]).save()
                    print('группа создана', cells[group])  # noqa: T001
                antibiotic = Antibiotic.objects.filter(title=cells[title])
                if not antibiotic.exists():
                    group_antibiotic = GroupAntibiotic.objects.filter(title=cells[group]).first()
                    Antibiotic(title=cells[title], group_antibiotic=group_antibiotic, lis=cells[lis]).save()
                    print('антибиотик сохранен', cells[group])  # noqa: T001
