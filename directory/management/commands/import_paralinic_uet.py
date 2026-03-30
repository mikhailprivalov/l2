from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from directory.models import Researches


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        code_price, uet = '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Код по прайсу" in cells:
                    code_price = cells.index("Код по прайсу")
                    uet = cells.index("УЕТ")
                    starts = True
            else:
                research = Researches.objects.filter(internal_code=cells[code_price]).first()
                if not research:
                    continue
                research.uet_refferal_doc = float(cells[uet])
                research.save()
