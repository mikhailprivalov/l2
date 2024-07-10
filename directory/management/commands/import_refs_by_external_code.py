from django.core.management.base import BaseCommand
from openpyxl.workbook import Workbook

from appconf.manager import SettingManager
from directory.models import Fractions
from openpyxl import load_workbook


def is_float(str_float: str):
    if "." in str_float:
        return True
    return False


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл
        Фаил с  колонками "Код", "Условия", "Ед. изм", Нижняя Гр., Верхняя Гр.
        для импорта референсов в фракции по внешнему коду
        """
        parser.add_argument('path', type=str)
        parser.add_argument('path1', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]

        starts = False
        code_idx, conditions_idx, start_ref_idx, end_ref_idx = None, None, None, None
        current_code = None
        ref_m, ref_f = {}, {}
        result_first_file = {}
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Условия" in cells:
                    code_idx = cells.index("Код")
                    conditions_idx = cells.index("Условия")
                    start_ref_idx = cells.index("Нижняя Гр.")
                    end_ref_idx = cells.index("Верхняя Гр.")
                    starts = True
            else:
                code = cells[code_idx].strip()
                conditions = cells[conditions_idx].strip()
                start = cells[start_ref_idx].strip()
                end = cells[end_ref_idx].strip()
                if start == "None" or end == "None":
                    continue
                tmp_cond_str = conditions.split("Пол: ")
                age = ""
                gender = ""
                tmp_cond_str2 = ""
                if len(tmp_cond_str) < 2:
                    gender = "общий"
                    age = None
                elif len(tmp_cond_str) >= 2:
                    tmp_cond_str2 = tmp_cond_str[1].split("Возр.: ")
                if len(tmp_cond_str2) > 1:
                    gender = tmp_cond_str2[0].strip()
                    age = tmp_cond_str2[1].strip()
                else:
                    gender = tmp_cond_str2[0].strip()
                if age and is_float(age):
                    age_range = age.split("-")
                    try:
                        age_start = float(age_range[0].strip()) * 365
                        age_end = float(age_range[1].strip()) * 365
                        age = f"дней {int(age_start)}-{int(age_end)}"
                    except Exception:
                        self.stdout.write("Не удалось преобразовать в дни")
                        continue
                elif not age:
                    age = "Все"
                if code != "None" and code != current_code:
                    current_code = code
                    
                    if gender.lower() == "общий":
                        if age == "Все":
                            ref_m = {"all": True, "data": {age: f"{start}-{end}"}}
                            ref_f = {"all": True, "data": {age: f"{start}-{end}"}}
                        else:
                            ref_m = {"all": False, "data": {age: f"{start}-{end}"}}
                            ref_f = {"all": False, "data": {age: f"{start}-{end}"}}
                    elif gender.lower() == "мужской":
                        if age == "Все":
                            ref_m = {"all": True, "data": {age: f"{start}-{end}"}}
                        else:
                            ref_m = {"all": False, "data": {age: f"{start}-{end}"}}
                    elif gender.lower() == "женский":
                        if age == "Все":
                            ref_f = {"all": True, "data": {age: f"{start}-{end}"}}
                        else:
                            ref_f = {"all": False, "data": {age: f"{start}-{end}"}}
                    result_first_file[current_code] = {"fsli": "", "ref_m": ref_m, "ref_f": ref_f}
                else:
                    if gender.lower() == "общий":
                        if not result_first_file[current_code]["ref_m"]["all"]:
                            result_first_file[current_code]["ref_m"]["data"][age] = f"{start}-{end}"
                        if not result_first_file[current_code]["ref_f"]["all"]:
                            result_first_file[current_code]["ref_f"]["data"][age] = f"{start}-{end}"
                    elif gender.lower() == "мужской":
                        if not result_first_file[current_code]["ref_m"]["all"]:
                            result_first_file[current_code]["ref_m"]["data"][age] = f"{start}-{end}"
                    elif gender.lower() == "женский":
                        if not result_first_file[current_code]["ref_f"]["all"]:
                            result_first_file[current_code]["ref_f"]["data"][age] = f"{start}-{end}"
        fp1 = kwargs["path1"]
        wb1 = load_workbook(filename=fp1)
        ws1 = wb1[wb1.sheetnames[0]]

        starts = False
        code_idx, fsli_idx = None, None
        for row in ws1.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Код ФСЛИ" in cells:
                    code_idx = cells.index("Код")
                    fsli_idx = cells.index("Код ФСЛИ")
                    starts = True
            else:
                code = cells[code_idx].strip()
                fsli = cells[fsli_idx].strip()
                if code == "None" or fsli == "None" or fsli == "-":
                    continue
                if not result_first_file.get(code):
                    continue
                result_first_file[code]["fsli"] = fsli.split(" ")[0].strip()
        result_wb = Workbook()
        result_ws = result_wb[result_wb.sheetnames[0]]
        result_ws.append(['Код', 'ФСЛИ', 'статус', 'кол-во'])
        for key, value in result_first_file.items():
            tmp_result_count = 0
            fractions = Fractions.objects.filter(fsli=value["fsli"])
            if not fractions.exists():
                continue
            for fraction in fractions:
                if value["ref_m"]["data"]:
                    fraction.ref_m = value["ref_m"]["data"]
                if value["ref_f"]["data"]:
                    fraction.ref_f = value["ref_f"]["data"]
                fraction.save()
                tmp_result_count += 1
            result_ws.append([key, value["fsli"], "+", tmp_result_count])
        dir_tmp = SettingManager.get("dir_param")
        result_wb.save(f"{dir_tmp}/result_import_fractions.xlsx")
