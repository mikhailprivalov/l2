from django.core.management.base import BaseCommand
from directory.models import Fractions, LaboratoryMaterial, Researches, ReleationsFT
from openpyxl import load_workbook

from podrazdeleniya.models import Podrazdeleniya
from researches.models import Tubes


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        internal_code, title, material, container, department, laboratory_duration = '', '', '', '', '', ''
        step = 0
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Код" in cells:
                    internal_code = cells.index("Код")
                    title = cells.index("Название")
                    material = cells.index("Материал")
                    container = cells.index("Контейнер")
                    department = cells.index("Подразделение")
                    laboratory_duration = cells.index("Готовность")
                    starts = True
            else:
                if Researches.objects.filter(title=cells[title].strip()):
                    continue
                material_obj = LaboratoryMaterial.objects.filter(title=cells[material].strip()).first()
                department_obj = Podrazdeleniya.objects.filter(title=cells[department].strip()).first()
                if cells[laboratory_duration] == "None":
                    cells[laboratory_duration] = ''
                research = Researches(
                    title=cells[title].strip(),
                    internal_code=cells[internal_code],
                    laboratory_material=material_obj,
                    podrazdeleniye=department_obj,
                    laboratory_duration=cells[laboratory_duration]
                )
                research.save()

                tube = Tubes.objects.filter(title=cells[container].strip()).first()
                fraction_data, relation_f = None, None
                if ReleationsFT.objects.filter(tube=tube).exists():
                    relation_tmp = None
                    fraction_data = {}
                    for i in ReleationsFT.objects.filter(tube=tube):
                        if i.pk != relation_tmp:
                            fraction_obj = Fractions.objects.filter(relation_id=i.pk).first()
                            fraction_data[i.pk] = fraction_obj.research.laboratory_material
                        relation_tmp = i.pk

                if not ReleationsFT.objects.filter(tube=tube).exists():
                    relation_f = ReleationsFT(tube=tube)
                    relation_f.save()

                if fraction_data:
                    for k, v in fraction_data.items():
                        if v == material_obj:
                            relation_f = ReleationsFT.objects.filter(pk=k).first()
                            break

                if not relation_f:
                    relation_f = ReleationsFT(tube=tube)
                    relation_f.save()

                fraction = Fractions(
                    research=research,
                    title=cells[title],
                    relation=relation_f
                )
                fraction.save()
                step += 1
                self.stdout.write(f'{step}-Услуга добавлена - {research.title}, фракция - {fraction.title}')
