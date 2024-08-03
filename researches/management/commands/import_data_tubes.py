from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from researches.models import Tubes


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        title, color = '', ''
        step = 0
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "название" in cells:
                    title = cells.index("название")
                    color = cells.index("цвет")
                    starts = True
            else:
                if Tubes.objects.filter(title=cells[title].strip()):
                    tube = Tubes.objects.filter(title=cells[title].strip()).first()
                    tube.color = cells[color].strip()
                    tube.save()
                    step += 1
                    self.stdout.write(f'{step}-Контейнер изменен - {tube.title}, фракция - {tube.color}')
                    continue
                else:
                    tube = Tubes(title=cells[title].strip(), color=cells[color].strip())
                    tube.save()
                    step += 1
                    self.stdout.write(f'{step}-Контейнер добавлен - {tube.title}, фракция - {tube.color}')
