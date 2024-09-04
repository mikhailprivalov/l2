import os
from collections import defaultdict
from copy import deepcopy

import pytz_deprecation_shim as pytz
import simplejson as json
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.utils import timezone, dateformat
from django.views.decorators.csrf import csrf_exempt
from openpyxl.reader.excel import load_workbook

import directory.models as directory
import slog.models as slog
from api.directions.sql_func import get_lab_podr
from appconf.manager import SettingManager
from clients.models import CardBase
from contracts.models import PriceName, PriceCoast, Company
from contracts.sql_func import get_research_coast_by_prce
from directions.models import Napravleniya, TubesRegistration, IstochnikiFinansirovaniya, Result, RMISOrgs, ParaclinicResult
from directory.models import Researches
from hospitals.models import Hospitals
from laboratory import settings
from laboratory import utils
from researches.models import Tubes
from results.sql_func import get_expertis_child_iss_by_issledovaniya, get_expertis_results_by_issledovaniya
from users.models import DoctorProfile
from users.models import Podrazdeleniya
from utils.dates import try_parse_range, normalize_date, normalize_dots_date
from utils.parse_sql import death_form_result_parse, get_unique_directions, weapon_form_result_parse
from . import sql_func
from . import structure_sheet
import datetime
import calendar
import openpyxl
from .report import (
    call_patient,
    swab_covid,
    cert_notwork,
    dispanserization,
    dispensary_data,
    custom_research,
    consolidates,
    commercial_offer,
    harmful_factors,
    base_data,
    expertise_report,
    registry_profit,
    appointed_research,
    lab_result,
    partner_coast_data,
)
from .sql_func import (
    attached_female_on_month,
    screening_plan_for_month_all_patient,
    must_dispensarization_from_screening_plan_for_month,
    sql_pass_screening,
    sql_pass_screening_in_dispensarization,
    screening_plan_for_month_all_count,
    sql_pass_pap_analysis_count,
    sql_pass_pap_fraction_result_value,
    sql_card_dublicate_pass_pap_fraction_not_not_enough_adequate_result_value,
    sql_get_result_by_direction,
    sql_get_documents_by_card_id,
    get_all_harmful_factors_templates,
    get_researches_by_templates,
    get_expertise_grade,
    get_confirm_protocol_by_date_extract,
)

from laboratory.settings import (
    PAP_ANALYSIS_ID,
    PAP_ANALYSIS_FRACTION_QUALITY_ID,
    PAP_ANALYSIS_FRACTION_CONTAIN_ID,
    DEATH_RESEARCH_PK,
    COVID_QUESTION_ID,
    RESEARCH_SPECIAL_REPORT,
    DISPANSERIZATION_SERVICE_PK,
    UNLIMIT_PERIOD_STATISTIC_RESEARCH,
    UNLIMIT_PERIOD_STATISTIC_GROUP,
)
from .statistic_func import save_file_disk, initial_work_book


# @ratelimit(key=lambda g, r: r.user.username + "_stats_" + (r.POST.get("type", "") if r.method == "POST" else r.GET.get("type", "")), rate="20/m", block=True)
@csrf_exempt
@login_required
def statistic_xls(request):
    """Генерация XLS"""
    from directions.models import Issledovaniya
    import xlwt
    from collections import OrderedDict

    wb = xlwt.Workbook(encoding='utf-8')
    response = HttpResponse(content_type='application/ms-excel')

    request_data = request.POST if request.method == "POST" else request.GET
    pk = request_data.get("pk", "")
    tp = request_data.get("type", "")
    date_start_o = request_data.get("date-start", "")
    date_end_o = request_data.get("date-end", "")
    users_o = request_data.get("users", "[]")
    user_o = request_data.get("user")
    date_values_o = request_data.get("values", "{}")
    date_type = request_data.get("date_type", "d")
    depart_o = request_data.get("department")

    if tp == 'lab' and pk == '0':
        tp = 'all-labs'

    symbols = (u"абвгдеёжзийклмнопрстуфхцчшщъыьэюяАБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ", u"abvgdeejzijklmnoprstufhzcss_y_euaABVGDEEJZIJKLMNOPRSTUFHZCSS_Y_EUA")  # Словарь для транслитерации
    tr = {ord(a): ord(b) for a, b in zip(*symbols)}  # Перевод словаря для транслита

    borders = xlwt.Borders()
    borders.left = xlwt.Borders.THIN
    borders.right = xlwt.Borders.THIN
    borders.top = xlwt.Borders.THIN
    borders.bottom = xlwt.Borders.THIN

    if "-" in date_start_o:
        date_start_o = normalize_date(date_start_o)
        date_end_o = normalize_date(date_end_o)

    date_start, date_end = try_parse_range(date_start_o, date_end_o)

    if date_start and date_end and tp not in ["lab_sum", "covid_sum", "lab_details", "statistics-consolidate"]:
        for i in UNLIMIT_PERIOD_STATISTIC_GROUP:
            if i not in [str(x) for x in request.user.groups.all()]:
                pk_research = request_data.get("research")
                delta = date_end - date_start
                if abs(delta.days) > 60 and tp == "statistics-research" and int(pk_research) not in UNLIMIT_PERIOD_STATISTIC_RESEARCH:
                    slog.Log(key=tp, type=101, body=json.dumps({"pk": pk_research, "date": {"start": date_start_o, "end": date_end_o}}), user=request.user.doctorprofile).save()
                    return JsonResponse({"error": "period max - 60 days"})

    if date_start_o != "" and date_end_o != "":
        slog.Log(key=tp, type=100, body=json.dumps({"pk": pk, "date": {"start": date_start_o, "end": date_end_o}}), user=request.user.doctorprofile).save()

    # Отчет по динамике анализов
    if tp == "directions_list_dynamic":
        pk = json.loads(pk)
        dn = Napravleniya.objects.filter(pk__in=pk)
        cards = {}

        napr_client = set()
        depart_napr = OrderedDict()
        depart_fraction = OrderedDict()
        one_param = "one_param"

        for d in dn:
            if d.department() is None or d.department().p_type != 2:
                continue
            c = d.client
            napr_client.add(c.pk)
            # Проверить, что все направления относятся к одной карте. И тип "Лаборатория"
            if len(napr_client) > 1:
                response['Content-Disposition'] = str.translate("attachment; filename=\"Назначения.xls\"", tr)
                ws = wb.add_sheet("Вакцинация")
                row_num = 0
                row = [
                    ("Пациент", 7000),
                    ("Карта", 6000),
                    ("Направление", 4000),
                    ("Дата", 4000),
                    ("Назначение", 7000),
                ]
                wb.save(response)
                return response
            # Распределить направления по подразделениям: "depart_napr"
            # {БИО:[напр1, напр2, напр3], КДЛ: [напр11, напр21, напр31], ИММ: [напр41, напр42, напр43]}

            tmp_num_dir = []
            department_title = d.department().id
            department_id = d.department().id

            if department_title in depart_napr.keys():
                tmp_num_dir = depart_napr.get(department_title)
                tmp_num_dir.append(d.pk)
                depart_napr[department_title] = tmp_num_dir
            else:
                tmp_num_dir.append(d.pk)
                depart_napr[department_title] = tmp_num_dir

            # По исследованиям строим структуру "depart_fraction":
            # Будущие заголовки в Excel. Те исследования у, к-рых по одной фракции в общий подсловарь,
            # у к-рых больше одного показателя (фракции) в самостоятельные подсловари. Выборка из справочника, НЕ из "Результатов"
            # пример стр-ра: {биохим: {услуги, имеющие по 1 фракции:[фр1-усл1, фр2-усл2, фр3-усл3],
            #                   усл1:[фр1, фр2, фр3],усл2:[фр1, фр2, фр3],
            #                   усл2:[фр1, фр2, фр3],усл2:[фр1, фр2, фр3]}
            # порядок фракций "По весу".

            one_param_temp = OrderedDict()

            for i in Issledovaniya.objects.filter(napravleniye=d):
                dict_research_fraction = OrderedDict()
                research_iss = i.research_id
                dict_research_fraction = {
                    p: str(t) + ',' + str(u) for p, t, u in directory.Fractions.objects.values_list('pk', 'title', 'units').filter(research=i.research).order_by("sort_weight")
                }

                if depart_fraction.get(department_id) is not None:
                    if len(dict_research_fraction.keys()) == 1:
                        one_param_temp = depart_fraction[department_id][one_param]
                        one_param_temp.update(dict_research_fraction)
                        depart_fraction[department_id].update({one_param: one_param_temp})
                    else:
                        depart_fraction[department_id].update({research_iss: dict_research_fraction})
                else:
                    depart_fraction.update({department_id: {}})
                    if len(dict_research_fraction) == 1:
                        depart_fraction[department_id].update({one_param: dict_research_fraction})
                    else:
                        depart_fraction[department_id].update({research_iss: dict_research_fraction})
                        depart_fraction[department_id].update({one_param: {}})

        # Все возможные анализы в направлениях - стр-ра А
        # направления по лабораториям (тип лаборатории, [номера направлений])
        obj = []
        for type_lab, l_napr in depart_napr.items():
            a = [
                [p, r, n, datetime.datetime.strftime(utils.localtime(t), "%d.%m.%y")]
                for p, r, n, t in Issledovaniya.objects.values_list('pk', 'research_id', 'napravleniye_id', 'time_confirmation').filter(napravleniye_id__in=l_napr)
            ]
            obj.append(a)

        for i in obj:
            for j in i:
                result_k = {fr_id: val for fr_id, val in Result.objects.values_list('fraction', 'value').filter(issledovaniye_id=j[0])}
                j.append(result_k)

        finish_obj = []
        for i in obj:
            for j in i:
                j.pop(0)
                finish_obj.append(j)

        # Строим стр-ру {тип лаборатория: id-анализа:{(направление, дата):{id-фракции:результат,id-фракции:результат}}}
        finish_ord = OrderedDict()
        for t_lab, name_iss in depart_fraction.items():
            finish_ord[t_lab] = {}
            for iss_id, fract_dict in name_iss.items():
                if fract_dict:
                    frac = True
                else:
                    frac = False
                finish_ord[t_lab][iss_id] = {}
                opinion_dict = {
                    (
                        'напр',
                        'дата',
                    ): fract_dict
                }
                val_dict = fract_dict.copy()
                finish_ord[t_lab][iss_id].update(opinion_dict)
                for k, v in fract_dict.items():
                    val_dict[k] = ''

                # Строим стр-ру {id-анализа:{(направление, дата,):{id-фракции:результат,id-фракции:результат}}}
                # one_param - это анализы у которых несколько параметров-фракции (ОАК, ОАМ)
                if iss_id != 'one_param' or iss_id != '' or iss_id is not None:
                    for d in finish_obj:
                        tmp_dict = {}
                        if iss_id == d[0]:
                            for i, j in d[3].items():
                                val_dict[i] = j
                            tmp_dict[
                                (
                                    d[1],
                                    d[2],
                                )
                            ] = deepcopy(val_dict)
                            finish_ord[t_lab][iss_id].update(tmp_dict)

                # Строим стр-ру {one_param:{(направление, дата,):{id-фракции:результат,id-фракции:результат}}}
                # one_param - это анализы у которых только один параметр-фракции (холестерин, глюкоза и др.)
                key_tuple = (
                    (
                        0,
                        0,
                    ),
                )
                if iss_id == 'one_param' and frac:
                    tmp_dict = {}
                    for d in finish_obj:
                        if key_tuple != (
                            d[1],
                            d[2],
                        ):
                            for k, v in fract_dict.items():
                                val_dict[k] = ''
                        for u, s in val_dict.items():
                            if d[3].get(u):
                                val_dict[u] = d[3].get(u)
                                tmp_dict[
                                    (
                                        d[1],
                                        d[2],
                                    )
                                ] = deepcopy(val_dict)
                                key_tuple = (
                                    d[1],
                                    d[2],
                                )

                    finish_ord[t_lab][iss_id].update(tmp_dict)

        response['Content-Disposition'] = str.translate("attachment; filename=\"Назначения.xls\"", tr)
        font_style = xlwt.XFStyle()
        font_style.alignment.wrap = 1
        font_style.borders = borders
        font_style_b = xlwt.XFStyle()
        font_style_b.alignment.wrap = 1
        font_style_b.font.bold = True
        font_style_b.borders = borders
        ws = wb.add_sheet("Динамика")
        row_num = 0

        for k, v in finish_ord.items():
            col_num = 0
            ws.write(row_num, 0, label=Podrazdeleniya.objects.values_list('title').get(pk=k))
            row_num += 1
            col_num = 0
            for name_iss, fr_id in v.items():
                if name_iss != 'one_param':
                    ws.write(row_num, 0, label=Researches.objects.values_list('title').get(pk=name_iss))
                else:
                    ws.write(row_num, 0, label=name_iss)
                row_num += 1
                a, b = '', ''
                for i, j in fr_id.items():
                    col_num = 0
                    a, b = i
                    ws.write(row_num, col_num, label=a)
                    col_num += 1
                    ws.write(row_num, col_num, label=b)
                    ss = ''
                    for g, h in j.items():
                        col_num += 1
                        ss = str(h)
                        ws.write(row_num, col_num, label=ss)
                    row_num += 1
                    col_num += 1
                row_num += 1
            row_num += 1

    if tp == "directions_list":
        pk = json.loads(pk)
        dn = Napravleniya.objects.filter(pk__in=pk)
        cards = {}
        for d in dn:
            c = d.client
            if c.pk not in cards:
                cards[c.pk] = {
                    "card": c.number_with_type(),
                    "fio": c.individual.fio(),
                    "bd": c.individual.bd(),
                    "hn": d.history_num,
                    "d": {},
                }
            cards[c.pk]["d"][d.pk] = {"r": [], "dn": str(dateformat.format(d.data_sozdaniya.date(), settings.DATE_FORMAT)), "tubes": []}

            tubes = TubesRegistration.objects.filter(issledovaniya__napravleniye=d).distinct()
            direction_tube = [tube.number for tube in tubes]
            cards[c.pk]["d"][d.pk]["tubes"] = ','.join(map(str, [t for t in direction_tube]))

            for i in Issledovaniya.objects.filter(napravleniye=d):
                cards[c.pk]["d"][d.pk]["r"].append(
                    {
                        "title": i.research.title,
                    }
                )

        response['Content-Disposition'] = str.translate("attachment; filename=\"Назначения.xls\"", tr)
        font_style = xlwt.XFStyle()
        font_style.alignment.wrap = 1
        font_style.borders = borders

        font_style_b = xlwt.XFStyle()
        font_style_b.alignment.wrap = 1
        font_style_b.font.bold = True
        font_style_b.borders = borders

        ws = wb.add_sheet("Вакцинация")
        row_num = 0
        row = [
            ("Пациент", 7000),
            ("Карта", 6000),
            ("Направление", 4000),
            ("Дата", 4000),
            ("Номер Емкости", 7000),
            ("Назначение", 7000),
        ]

        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num][0], font_style_b)
            ws.col(col_num).width = row[col_num][1]
        row_num += 1

        for ck in cards.keys():
            c = cards[ck]
            started = False
            for dk in c["d"].keys():
                if not started:
                    row = [
                        "{} {}".format(c["fio"], c["bd"]),
                        c["card"],
                    ]
                    started = True
                else:
                    row = ["", ""]

                s2 = False

                for r in c["d"][dk]["r"]:
                    if not s2:
                        s2 = True
                        row.append(str(dk))
                        row.append(c["d"][dk]["dn"])
                        row.append(c["d"][dk]["tubes"])
                    else:
                        row.append("")
                        row.append("")
                        row.append("")
                        row.append("")
                    row.append(r["title"])

                    for col_num in range(len(row)):
                        ws.write(row_num, col_num, row[col_num], font_style)
                    row_num += 1
                    row = []

    if tp == "statistics-visits":
        date_start, date_end = try_parse_range(date_start_o, date_end_o)
        t = request.GET.get("t", "sum")
        fio = request.user.doctorprofile.get_full_fio()
        dep = request.user.doctorprofile.podrazdeleniye.get_title()
        dirs = Napravleniya.objects.filter(
            visit_date__range=(
                date_start,
                date_end,
            ),
            visit_who_mark=request.user.doctorprofile,
        ).order_by("visit_date")

        if t == "sum":
            response['Content-Disposition'] = str.translate("attachment; filename=\"Суммарный отчёт по посещениям.xls\"", tr)
            font_style = xlwt.XFStyle()
            font_style.alignment.wrap = 1
            font_style.borders = borders

            font_style_b = xlwt.XFStyle()
            font_style_b.alignment.wrap = 1
            font_style_b.font.bold = True
            font_style_b.borders = borders

            ws = wb.add_sheet("Посещения")
            row_num = 0
            row = [
                (fio, 7000),
                (dep, 7000),
                ("", 3000),
            ]
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num][0], font_style)
                ws.col(col_num).width = row[col_num][1]
            row_num += 1
            row = [
                date_start_o + " - " + date_end_o,
                "",
                "",
            ]
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)
            row_num += 1
            row = [
                "",
                "",
                "",
            ]
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style if col_num > 0 else font_style_b)
            row_num += 1
            row = [
                "Услуга",
                "Источник финансирования",
                "Количество",
            ]
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style_b)
            row_num += 1
            iss = {}
            for d in dirs:
                for i in Issledovaniya.objects.filter(napravleniye=d).order_by("research__title").order_by("napravleniye__istochnik_f"):
                    rt = i.research.title
                    istf = i.napravleniye.istochnik_f.base.title + " - " + i.napravleniye.fin_title
                    if rt not in iss:
                        iss[rt] = {}

                    if istf not in iss[rt]:
                        iss[rt][istf] = 0

                    iss[rt][istf] += 1
            for k in iss:
                for istf in iss[k]:
                    row = [
                        k,
                        istf,
                        iss[k][istf],
                    ]
                    for col_num in range(len(row)):
                        ws.write(row_num, col_num, row[col_num], font_style)
                    row_num += 1
    elif tp == "vac":
        date_start, date_end = try_parse_range(date_start_o, date_end_o)
        response['Content-Disposition'] = str.translate("attachment; filename=\"Вакцинация.xls\"", tr)
        font_style = xlwt.XFStyle()
        font_style.alignment.wrap = 1
        font_style.borders = borders

        font_style_b = xlwt.XFStyle()
        font_style_b.alignment.wrap = 1
        font_style_b.font.bold = True
        font_style_b.borders = borders

        ts = ["Название", "Доза", "Серия", "Срок годности", "Способ введения", "Дата постановки вакцины"]

        ws = wb.add_sheet("Вакцинация")
        row_num = 0
        row = [("Исполнитель", 6000), ("Подтверждено", 5000), ("RMIS UID", 5000), ("Вакцина", 5000), ("Код", 4000)]

        for t in ts:
            row.append((t, 4000))
        row.append(("Этап", 2500))

        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num][0], font_style_b)
            ws.col(col_num).width = row[col_num][1]
        row_num += 1

        for i in Issledovaniya.objects.filter(
            research__podrazdeleniye__vaccine=True,
            time_confirmation__range=(
                date_start,
                date_end,
            ),
        ).order_by("time_confirmation"):
            if i.napravleniye:
                row = [
                    i.doc_confirmation_fio,
                    i.time_confirmation.astimezone(pytz.timezone(settings.TIME_ZONE)).strftime("%d.%m.%Y %X"),
                    i.napravleniye.client.individual.get_rmis_uid_fast(),
                    i.research.title,
                    i.research.code,
                ]
            else:
                continue
            v = {}
            for p in ParaclinicResult.objects.filter(issledovaniye=i):
                field_type = p.get_field_type()
                if p.field.get_title(force_type=field_type) in ts:
                    if field_type == 1:
                        v_date = p.value.replace("-", ".")
                        v[p.field.get_title(force_type=field_type)] = v_date
                    else:
                        v[p.field.get_title(force_type=field_type)] = p.value
            for t in ts:
                row.append(v.get(t, ""))
            row.append("V")
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)
            row_num += 1

    elif tp == "statistics-tickets-print":
        data_date = request_data.get("date_values")
        data_date = json.loads(data_date)

        if request_data.get("date_type") == 'd':
            d1 = datetime.datetime.strptime(data_date['date'], '%d.%m.%Y')
            d2 = datetime.datetime.strptime(data_date['date'], '%d.%m.%Y')
            month_obj = ''
        else:
            month_obj = int(data_date['month']) + 1
            _, num_days = calendar.monthrange(int(data_date['year']), month_obj)
            d1 = datetime.date(int(data_date['year']), month_obj, 1)
            d2 = datetime.date(int(data_date['year']), month_obj, num_days)

        type_fin = request_data.get("fin")
        title_fin = IstochnikiFinansirovaniya.objects.filter(pk=type_fin).first()
        if title_fin.title == 'ОМС' and title_fin.base.internal_type:
            can_null = 1
        else:
            can_null = 0
        users_o = json.loads(user_o)
        us_o = None
        if users_o != -1:
            us = int(users_o)
            us_o = [DoctorProfile.objects.get(pk=us)]
        elif depart_o != -1:
            depart = Podrazdeleniya.objects.get(pk=depart_o)
            us_o = DoctorProfile.objects.filter(podrazdeleniye=depart)

        wb = openpyxl.Workbook()
        wb.remove(wb.get_sheet_by_name('Sheet'))
        styles_obj = structure_sheet.style_sheet()
        wb.add_named_style(styles_obj[0])

        start_date = datetime.datetime.combine(d1, datetime.time.min)
        end_date = datetime.datetime.combine(d2, datetime.time.max)
        # Проверить, что роль у объекта Врач-Лаборант, или Лаборант, или Врач параклиники, или Лечащий врач
        if us_o:
            for i in us_o:
                if i.is_member(["Лечащий врач", "Врач-лаборант", "Врач параклиники", "Лаборант", "Врач консультаций"]):
                    res_oq = sql_func.direct_job_sql(i.pk, start_date, end_date, type_fin, can_null)
                    res_job = sql_func.indirect_job_sql(i.pk, start_date, end_date)
                    if res_job:
                        ws = wb.create_sheet(f'{i.get_fio()}-Косвенные')
                        ws = structure_sheet.inderect_job_base(ws, i, d1, d2)
                        dict_job = {}
                        for r_j in res_job:
                            key_type_job = r_j[1]
                            key_date = utils.strfdatetime(r_j[0], "%d.%m.%Y")
                            value_total = r_j[2]
                            temp_dict = dict_job.get(key_date, {})
                            temp_dict.update({key_type_job: value_total})
                            dict_job[key_date] = temp_dict
                        structure_sheet.inderect_job_data(ws, dict_job)

                    ws = wb.create_sheet(i.get_fio())
                    ws = structure_sheet.statistics_tickets_base(ws, i, type_fin, d1, d2, styles_obj[0], styles_obj[1])
                    ws = structure_sheet.statistics_tickets_data(ws, res_oq, i, styles_obj[2])

                    if month_obj:
                        # issledovaniye_id(0), research_id(1), date_confirm(2), doc_confirmation_id(3), def_uet(4),
                        # co_executor_id(5), co_executor_uet(6), co_executor2_id(7), co_executor2_uet(8), research_id(9),
                        # research_title(10), research - co_executor_2_title(11)
                        # строим стр-ру {дата:{наименование анализа:УЕТ за дату, СО2:УЕТ за дату}}
                        total_report_dict = OrderedDict()
                        r_sql = sql_func.total_report_sql(i.pk, start_date, end_date, type_fin)
                        titles_set = OrderedDict()
                        for n in r_sql:
                            titles_set[n[10]] = ''
                            titles_set[n[11]] = ''
                            temp_uet, temp_uet2 = 0, 0
                            if i.pk == n[3]:
                                temp_uet = n[4] if n[4] else 0
                            if i.pk == n[5] and n[5] != n[3]:
                                temp_uet = n[6] if n[6] else 0
                            if i.pk == n[7]:
                                temp_uet2 = n[8] if n[8] else 0
                            # попытка получить значения за дату
                            if total_report_dict.get(n[2]):
                                temp_d = total_report_dict.get(n[2])
                                # попытка получить такие же анализы
                                current_uet = temp_d.get(n[10], 0)
                                current_uet2 = temp_d.get(n[11], 0)
                                current_uet = current_uet + temp_uet
                                current_uet2 = current_uet2 + temp_uet2
                                temp_dict = {n[10]: current_uet, n[11]: current_uet2}
                                total_report_dict[int(n[2])].update(temp_dict)
                            else:
                                total_report_dict[int(n[2])] = {n[10]: temp_uet, n[11]: temp_uet2}

                        titles_list = list(titles_set.keys())
                        ws = wb.create_sheet(i.get_fio() + ' - Итог')
                        ws = structure_sheet.job_total_base(ws, month_obj, type_fin)

                        ws, cell_research = structure_sheet.jot_total_titles(ws, titles_list)
                        ws = structure_sheet.job_total_data(ws, cell_research, total_report_dict)

        response['Content-Disposition'] = str.translate("attachment; filename=\"Статталоны.xlsx\"", tr)
        wb.save(response)
        return response

    elif tp == "statistics-passed":
        d_s = request_data.get("date-start")
        d_e = request_data.get("date-end")
        d1 = datetime.datetime.strptime(d_s, '%d.%m.%Y')
        d2 = datetime.datetime.strptime(d_e, '%d.%m.%Y')
        start_date = datetime.datetime.combine(d1, datetime.time.min)
        end_date = datetime.datetime.combine(d2, datetime.time.max)
        passed_oq = sql_func.passed_research(start_date, end_date)
        wb = openpyxl.Workbook()
        wb.remove(wb.get_sheet_by_name('Sheet'))
        ws = wb.create_sheet(f'{d_s}-{d_e}')
        ws = structure_sheet.passed_research_base(ws, d_s)
        ws = structure_sheet.passed_research_data(ws, passed_oq)

        response['Content-Disposition'] = str.translate("attachment; filename=\"Движения.xlsx\"", tr)
        wb.save(response)
        return response

    elif tp == "call-patient":
        return call_patient.call_patient(request_data, response, tr, COVID_QUESTION_ID)
    elif tp == "swab-covidt":
        return swab_covid.swab_covid(request_data, response, tr, COVID_QUESTION_ID)
    elif tp == "cert-not-workt":
        return cert_notwork.cert_notwork(request_data, response, tr, COVID_QUESTION_ID)

    elif tp == "statistics-onco":
        d_s = request_data.get("date-start")
        d_e = request_data.get("date-end")
        d1 = datetime.datetime.strptime(d_s, '%d.%m.%Y')
        d2 = datetime.datetime.strptime(d_e, '%d.%m.%Y')
        start_date = datetime.datetime.combine(d1, datetime.time.min)
        end_date = datetime.datetime.combine(d2, datetime.time.max)
        onco_query = sql_func.disp_diagnos('U999', start_date, end_date)
        wb = openpyxl.Workbook()
        wb.remove(wb.get_sheet_by_name('Sheet'))
        ws = wb.create_sheet(f'{d_s}-{d_e}')
        ws = structure_sheet.onco_base(ws, d_s, d_e)
        ws = structure_sheet.passed_onco_data(ws, onco_query)
        response['Content-Disposition'] = str.translate("attachment; filename=\"Онкоподозрения.xlsx\"", tr)
        wb.save(response)
        return response

    elif tp == "statistics-research":
        response['Content-Disposition'] = str.translate("attachment; filename=\"Услуги.xlsx\"", tr)
        pk = request_data.get("research")
        user_groups = request.user.groups.values_list('name', flat=True)
        research_id = int(pk)
        data_date = request_data.get("date_values")
        data_date = json.loads(data_date)
        purposes = request_data.get("purposes", "")
        special_fields = request_data.get("special-fields", "false")
        is_lab_result = request_data.get("is-lab-result", "false")
        medical_exam = request_data.get("medical-exam", "false")
        by_create_directions = request_data.get("by-create-directions", "false")
        is_purpose = 0
        if purposes != "-1":
            purposes = tuple(purposes.split(","))
            is_purpose = 1
        if request_data.get("date_type") == 'd':
            d1 = datetime.datetime.strptime(data_date['date'], '%d.%m.%Y')
            d2 = datetime.datetime.strptime(data_date['date'], '%d.%m.%Y')
            month_obj = ''
        elif request_data.get("date_type") == 'm':
            month_obj = int(data_date['month']) + 1
            _, num_days = calendar.monthrange(int(data_date['year']), month_obj)
            d1 = datetime.date(int(data_date['year']), month_obj, 1)
            d2 = datetime.date(int(data_date['year']), month_obj, num_days)
        else:
            d_s = request_data.get("date-start")
            d_e = request_data.get("date-end")
            d1 = datetime.datetime.strptime(d_s, '%d.%m.%Y')
            d2 = datetime.datetime.strptime(d_e, '%d.%m.%Y')

        wb = openpyxl.Workbook()
        wb.remove(wb.get_sheet_by_name('Sheet'))
        ws = wb.create_sheet("Отчет")
        research = Researches.objects.get(pk=research_id)
        research_title = research.title
        start_date = datetime.datetime.combine(d1, datetime.time.min)
        end_date = datetime.datetime.combine(d2, datetime.time.max)
        hospital_id = request.user.doctorprofile.hospital_id
        if 'Статистика-все МО' in user_groups:
            hospital_id = -1
        if research_id == DEATH_RESEARCH_PK:
            if 'Свидетельство о смерти-доступ' not in user_groups:
                return JsonResponse({"error": "Нет доступа к данному отчету"})
            if 'Статистика свидетельство о смерти-все МО' in user_groups:
                hospital_id = -1
            researches_sql = sql_func.statistics_death_research(research_id, start_date, end_date, hospital_id)
            unique_issledovaniya = get_unique_directions(researches_sql)
            child_iss = get_expertis_child_iss_by_issledovaniya(unique_issledovaniya) if unique_issledovaniya else None
            expertise_final_data = {}
            if child_iss:
                data = {i.child_iss: i.parent_id for i in child_iss}
                child_iss_tuple = tuple(set([i.child_iss for i in child_iss]))
                result_expertise = get_expertis_results_by_issledovaniya(child_iss_tuple)
                result_val = {}
                for i in result_expertise:
                    if not result_val.get(i.issledovaniye_id, ""):
                        result_val[i.issledovaniye_id] = "Экспертиза;"
                    if i.value.lower() == "да":
                        result_val[i.issledovaniye_id] = f"{result_val[i.issledovaniye_id]} {i.title};"

                for k, v in result_val.items():
                    if not expertise_final_data.get(data.get(k, "")):
                        expertise_final_data[data.get(k)] = ""
                    expertise_final_data[data.get(k)] = f"{expertise_final_data[data.get(k)]} {v}"

            data_death = death_form_result_parse(researches_sql, reserved=False)
            wb.remove(wb.get_sheet_by_name('Отчет'))
            ws = wb.create_sheet("По документам")
            ws = structure_sheet.statistic_research_death_base(ws, d1, d2, research_title[0])
            ws = structure_sheet.statistic_research_death_data(ws, data_death, expertise_final_data)

            reserved_researches_sql = sql_func.statistics_reserved_number_death_research(research_id, start_date, end_date, hospital_id)
            data_death_reserved = death_form_result_parse(reserved_researches_sql, reserved=True)
            ws2 = wb.create_sheet("Номера в резерве")
            ws2 = structure_sheet.statistic_reserved_research_death_base(ws2, d1, d2, research_title[0])
            ws2 = structure_sheet.statistic_reserved_research_death_data(ws2, data_death_reserved)

            card_has_death_date = sql_func.card_has_death_date(research_id, start_date, end_date)
            card_tuple = tuple(set([i.id for i in card_has_death_date]))
            if card_tuple:
                temp_data = sql_func.statistics_death_research_by_card(research_id, card_tuple, hospital_id)
                prev_card = None
                prev_direction = None
                final_data = []
                count = 0
                for k in temp_data:
                    if k.client_id == prev_card and prev_direction != k.napravleniye_id and count != 0:
                        continue
                    else:
                        final_data.append(k)
                    prev_card = k.client_id
                    prev_direction = k.napravleniye_id
                    count += 1

                data_death_card = death_form_result_parse(final_data, reserved=False)
                ws3 = wb.create_sheet("По людям")
                ws3 = structure_sheet.statistic_research_death_base_card(ws3, d1, d2, research_title[0])
                ws3 = structure_sheet.statistic_research_death_data_card(ws3, data_death_card)
        elif research_id in [RESEARCH_SPECIAL_REPORT.get("weapon_research_pk", -1), RESEARCH_SPECIAL_REPORT.get("driver_research", -1)]:
            researches_sql = sql_func.statistics_death_research(research_id, start_date, end_date, hospital_id)
            data = weapon_form_result_parse(researches_sql, reserved=False)
            ws = structure_sheet.statistic_research_wepon_base(ws, d1, d2, research_title[0])
            ws = structure_sheet.statistic_research_weapon_data(ws, data)
        elif is_purpose == 1:
            ws = structure_sheet.statistic_research_base(ws, d1, d2, research_title[0])
            researches_sql = sql_func.statistics_research(research_id, start_date, end_date, hospital_id, is_purpose, purposes)
            ws = structure_sheet.statistic_research_data(ws, researches_sql)
        elif research.podrazdeleniye and research.podrazdeleniye.p_type == 2 and is_lab_result == "true":
            researches_sql = sql_func.lab_result_statistics_research(research_id, start_date, end_date, hospital_id)
            result = lab_result.custom_lab_research_field_fractions(researches_sql)
            ws = lab_result.lab_result_research_base(ws, d1, d2, result, research_title[0])
            ws = custom_research.custom_research_fill_data(ws, result)
        elif special_fields == "true":
            researches_sql = sql_func.custom_statistics_research(research_id, start_date, end_date, hospital_id, medical_exam)
            if Researches.objects.filter(pk=research_id).first().is_monitoring:
                result = custom_research.custom_monitoring_research_data(researches_sql)
                ws = custom_research.custom_monitorimg_research_base(ws, d1, d2, result, research_title[0])
            else:
                result = custom_research.custom_research_data(researches_sql)
                ws = custom_research.custom_research_base(ws, d1, d2, result, research_title[0])
            ws = custom_research.custom_research_fill_data(ws, result)
        else:
            ws = structure_sheet.statistic_research_base(ws, d1, d2, research_title[0])
            if by_create_directions != "true":
                researches_sql = sql_func.statistics_research(research_id, start_date, end_date, hospital_id)
            else:
                researches_sql = sql_func.statistics_research_create_directions(research_id, start_date, end_date, hospital_id)
            ws = structure_sheet.statistic_research_data(ws, researches_sql)

    elif tp == "statistics-create-research":
        response['Content-Disposition'] = str.translate("attachment; filename=\"Услуги.xlsx\"", tr)
        research = request_data.get("research")
        users_docprofile_id = request_data.get("users")
        d_s = request_data.get("date-start")
        d_e = request_data.get("date-end")
        d1 = datetime.datetime.strptime(d_s, '%d.%m.%Y')
        d2 = datetime.datetime.strptime(d_e, '%d.%m.%Y')
        wb = openpyxl.Workbook()
        wb.remove(wb.get_sheet_by_name('Sheet'))
        ws = wb.create_sheet("Отчет")
        start_date = datetime.datetime.combine(d1, datetime.time.min)
        end_date = datetime.datetime.combine(d2, datetime.time.max)
        researches = research.split(",")
        researches = tuple([int(i) for i in researches])
        if users_docprofile_id == "undefined":
            users_docprofile = sorted(list(DoctorProfile.objects.all().values_list('pk', flat=True)))
        else:
            users_docprofile = sorted([int(json.loads(users_docprofile_id))])
        researches = tuple([int(i) for i in researches])
        research_data = {r.pk: {"title": r.title, "count": 0} for r in Researches.objects.filter(pk__in=list(researches))}
        research_data[-999] = {"title": ""}
        users_final_data = {k: deepcopy(research_data) for k in users_docprofile}
        researches_sql = sql_func.statistics_research_create_directions(researches, start_date, end_date, tuple(users_docprofile))
        ws = appointed_research.appointed_base(ws, d_s, d_e, research_data)
        doctors_researches_count = appointed_research.parse_data(researches_sql, users_final_data, research_data)
        ws = appointed_research.fill_appointed_research_by_doctors(ws, doctors_researches_count)

    elif tp == "journal-get-material":
        access_to_all = 'Просмотр статистики' in request.user.groups.values_list('name', flat=True) or request.user.is_superuser
        users = [x for x in json.loads(users_o) if (access_to_all or (x.isdigit() and int(x) == request.user.doctorprofile.pk)) and DoctorProfile.objects.filter(pk=x).exists()]
        date_values = json.loads(date_values_o)
        monthes = {
            "0": "Январь",
            "1": "Февраль",
            "2": "Март",
            "3": "Апрель",
            "4": "Май",
            "5": "Июнь",
            "6": "Июль",
            "7": "Август",
            "8": "Сентябрь",
            "9": "Октябрь",
            "10": "Ноябрь",
            "11": "Декабрь",
        }
        date_values["month_title"] = monthes[date_values["month"]]
        response['Content-Disposition'] = str.translate("attachment; filename=\"Статистика_Забор_биоматериала.xls\"", tr)
        font_style = xlwt.XFStyle()
        font_style.alignment.wrap = 1
        font_style.borders = borders

        font_style_b = xlwt.XFStyle()
        font_style_b.alignment.wrap = 1
        font_style_b.font.bold = True
        font_style_b.borders = borders

        for user_pk in users:
            user_row = DoctorProfile.objects.get(pk=user_pk)
            ws = wb.add_sheet("{} {}".format(user_row.get_fio(dots=False), user_pk))
            row_num = 0
            row = [("Исполнитель: ", 4000), (user_row.get_full_fio(), 7600)]
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num][0], font_style)
                ws.col(col_num).width = row[col_num][1]

            row_num += 1
            row = ["Подразделение: ", user_row.podrazdeleniye.title]
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)

            row_num += 1
            row = ["Дата: ", date_values["date"] if date_type == "d" else "{month_title} {year}".format(**date_values)]
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)
            daterow = row_num
            row_num += 3
            row = [
                ("№", 4000),
                ("ФИО", 7600),
                ("Возраст", 3000),
                ("Карта", 6000),
                ("Число направлений", 5000),
                ("Номера направлений", 6000),
                ("Наименования исследований", 20000),
            ]
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num][0], font_style_b)
                ws.col(col_num).width = row[col_num][1]

            row_num += 1

            if date_type == "d":
                day = date_values.get("date", "01.01.2015")
                day1 = datetime.date(int(day.split(".")[2]), int(day.split(".")[1]), int(day.split(".")[0]))
                day2 = day1 + datetime.timedelta(days=1)
            elif date_type == "m":
                month = int(date_values.get("month", "0")) + 1
                next_m = month + 1 if month < 12 else 1
                year = int(date_values.get("year", "2015"))
                next_y = year + 1 if next_m == 1 else year
                day1 = datetime.date(year, month, 1)
                day2 = datetime.date(next_y, next_m, 1)
            else:
                day1 = day2 = timezone.now()

            iss_list = (
                Issledovaniya.objects.filter(tubes__doc_get=user_row, tubes__time_get__isnull=False, tubes__time_get__range=(day1, day2))
                .order_by("napravleniye__client__individual__patronymic", "napravleniye__client__individual__name", "napravleniye__client__individual__family")
                .distinct()
            )
            patients = {}
            for iss in iss_list:
                k = iss.napravleniye.client.individual_id
                if k not in patients:
                    client = iss.napravleniye.client.individual
                    patients[k] = {"fio": client.fio(short=True, dots=True), "age": client.age_s(direction=iss.napravleniye), "directions": [], "researches": [], "cards": []}
                if iss.napravleniye_id not in patients[k]["directions"]:
                    patients[k]["directions"].append(iss.napravleniye_id)
                kn = iss.napravleniye.client.number_with_type()
                if kn not in patients[k]["cards"]:
                    patients[k]["cards"].append(kn)
                patients[k]["researches"].append(iss.research.title)

            n = 0
            for p_pk in patients:
                n += 1
                row = [
                    str(n),
                    patients[p_pk]["fio"],
                    patients[p_pk]["age"],
                    ", ".join(patients[p_pk]["cards"]),
                    len(patients[p_pk]["directions"]),
                    ", ".join([str(x) for x in patients[p_pk]["directions"]]),
                    ", ".join(patients[p_pk]["researches"]),
                ]
                for col_num in range(len(row)):
                    ws.write(row_num, col_num, row[col_num], font_style)

                row_num += 1

            row = ["Число пациентов: ", str(len(patients))]
            for col_num in range(len(row)):
                ws.write(daterow + 1, col_num, row[col_num], font_style)

    elif tp == "lab":

        lab = Podrazdeleniya.objects.get(pk=int(pk))
        response['Content-Disposition'] = str.translate("attachment; filename=\"Статистика_Лаборатория_{}_{}-{}.xls\"".format(lab.title.replace(" ", "_"), date_start_o, date_end_o), tr)

        import directions.models as d
        from operator import itemgetter

        date_start, date_end = try_parse_range(date_start_o, date_end_o)

        for card_base in list(CardBase.objects.filter(hide=False)) + [None]:
            cb_title = "Все базы" if not card_base else card_base.short_title
            for finsource in list(IstochnikiFinansirovaniya.objects.filter(base=card_base)) + [False]:
                finsource_title = "Все источники"

                if isinstance(finsource, IstochnikiFinansirovaniya):
                    finsource_title = finsource.title

                ws = wb.add_sheet(cb_title + " " + finsource_title + " выполн.")

                font_style = xlwt.XFStyle()
                font_style.borders = borders
                row_num = 0
                row = ["Период: ", "{0} - {1}".format(date_start_o, date_end_o)]

                for col_num in range(len(row)):
                    if col_num == 0:
                        ws.write(row_num, col_num, row[col_num], font_style)
                    else:
                        ws.write_merge(row_num, row_num, col_num, col_num + 2, row[col_num], style=font_style)

                row_num += 1

                font_style = xlwt.XFStyle()
                font_style.borders = borders

                row = [(lab.title, 16000)]

                for col_num in range(len(row)):
                    ws.write(row_num, col_num, row[col_num][0], font_style)
                    ws.col(col_num).width = row[col_num][1]
                    ws.write(row_num, col_num + 1, "", font_style)

                row_num = 2
                row = ["Выполнено исследований", cb_title + " " + finsource_title]

                for col_num in range(len(row)):
                    if col_num == 0:
                        ws.write(row_num, col_num, row[col_num], font_style)
                    else:
                        ws.write_merge(row_num, row_num, col_num, col_num + 1, row[col_num], style=font_style)

                font_style = xlwt.XFStyle()
                font_style.alignment.wrap = 1
                font_style.borders = borders
                pki = int(pk)
                otds = {pki: defaultdict(lambda: 0)}
                otds_pat = {pki: defaultdict(lambda: 0)}

                ns = 0
                for obj in directory.Researches.objects.filter(podrazdeleniye__pk=lab.pk):
                    if finsource is not False:
                        iss_list = Issledovaniya.objects.filter(
                            research__pk=obj.pk, time_confirmation__isnull=False, time_confirmation__range=(date_start, date_end), napravleniye__istochnik_f=finsource
                        )
                    elif card_base:
                        iss_list = Issledovaniya.objects.filter(
                            research__pk=obj.pk, time_confirmation__isnull=False, time_confirmation__range=(date_start, date_end), napravleniye__istochnik_f__base=card_base
                        )
                    else:
                        iss_list = Issledovaniya.objects.filter(research__pk=obj.pk, time_confirmation__isnull=False, time_confirmation__range=(date_start, date_end))

                    iss_list = iss_list.filter(napravleniye__isnull=False)

                    for researches in iss_list:
                        n = False
                        for x in d.Result.objects.filter(issledovaniye=researches):
                            x = x.value.lower().strip()
                            n = any([y in x for y in ["забор", "тест", "неправ", "ошибк", "ошибочный", "кров", "брак", "мало", "недостаточно", "реактив"]]) or x == "-"
                        if n:
                            continue
                        if researches.napravleniye:
                            otd_pk = "external-" + str(researches.napravleniye.imported_org_id) if not researches.napravleniye.doc else researches.napravleniye.doc.podrazdeleniye_id
                        else:
                            otd_pk = "empty"
                        if otd_pk not in otds:
                            otds[otd_pk] = defaultdict(lambda: 0)
                        otds[otd_pk][obj.pk] += 1
                        otds[pki][obj.pk] += 1
                        if any([x.get_is_norm()[0] == "normal" for x in researches.result_set.all()]):
                            continue
                        if otd_pk not in otds_pat:
                            otds_pat[otd_pk] = defaultdict(lambda: 0)
                        otds_pat[otd_pk][obj.pk] += 1
                        otds_pat[pki][obj.pk] += 1

                style = xlwt.XFStyle()
                style.borders = borders
                font = xlwt.Font()
                font.bold = True
                style.font = font
                otd_local_keys = [x for x in otds.keys() if isinstance(x, int)]
                otd_external_keys = [int(x.replace("external-", "")) for x in otds.keys() if isinstance(x, str) and "external-" in x and x != "external-None"]
                for otdd in (
                    list(Podrazdeleniya.objects.filter(pk=pki))
                    + list(Podrazdeleniya.objects.filter(pk__in=[x for x in otd_local_keys if x != pki]))
                    + list(RMISOrgs.objects.filter(pk__in=otd_external_keys))
                ):
                    row_num += 2
                    row = [
                        otdd.title if otdd.pk != pki else "Сумма по всем отделениям",
                        "" if otdd.pk != pki else "Итого",
                    ]
                    for col_num in range(len(row)):
                        ws.write(row_num, col_num, row[col_num], style=style)
                    rows = []
                    ok = otds.get(otdd.pk, otds.get("external-{}".format(otdd.pk), {}))
                    for obj in directory.Researches.objects.filter(pk__in=[x for x in ok.keys()]):
                        row = [
                            obj.title,
                            ok[obj.pk],
                        ]
                        rows.append(row)
                        ns += 1
                    for row in sorted(rows, key=itemgetter(0)):
                        row_num += 1
                        for col_num in range(len(row)):
                            ws.write(row_num, col_num, row[col_num], font_style)

                ws_pat = wb.add_sheet(cb_title + " " + finsource_title + " паталог.")

                row_num = 0
                row = ["Период: ", "{0} - {1}".format(date_start_o, date_end_o)]

                for col_num in range(len(row)):
                    if col_num == 0:
                        ws_pat.write(row_num, col_num, row[col_num], font_style)
                    else:
                        ws_pat.write_merge(row_num, row_num, col_num, col_num + 2, row[col_num], style=font_style)

                row_num = 1
                row = [
                    (lab.title, 16000),
                ]

                for col_num in range(len(row)):
                    ws_pat.write(row_num, col_num, row[col_num][0], font_style)
                    ws_pat.col(col_num).width = row[col_num][1]
                    ws_pat.write(row_num, col_num + 1, "", font_style)

                font_style = xlwt.XFStyle()
                font_style.borders = borders

                row_num = 2
                row = ["Паталогии", cb_title + " " + finsource_title]

                for col_num in range(len(row)):
                    if col_num == 0:
                        ws_pat.write(row_num, col_num, row[col_num], font_style)
                    else:
                        ws_pat.write_merge(row_num, row_num, col_num, col_num + 1, row[col_num], style=font_style)

                otd_local_keys = [x for x in otds_pat.keys() if isinstance(x, int)]
                otd_external_keys = [int(x.replace("external-", "")) for x in otds_pat.keys() if isinstance(x, str) and "external-" in x]

                for otdd in (
                    list(Podrazdeleniya.objects.filter(pk=pki))
                    + list(Podrazdeleniya.objects.filter(pk__in=[x for x in otd_local_keys if x != pki]))
                    + list(RMISOrgs.objects.filter(pk__in=otd_external_keys))
                ):
                    row_num += 2
                    row = [
                        otdd.title,
                        "" if otdd.pk != pki else "Итого",
                    ]
                    for col_num in range(len(row)):
                        ws_pat.write(row_num, col_num, row[col_num], style=style)
                    rows = []
                    ok = otds_pat.get(otdd.pk, otds_pat.get("external-{}".format(otdd.pk), {}))
                    for obj in directory.Researches.objects.filter(pk__in=[x for x in otds_pat.get(otdd.pk, ok.keys())]):
                        row = [
                            obj.title,
                            ok[obj.pk],
                        ]
                        rows.append(row)
                    for row in sorted(rows, key=itemgetter(0)):
                        row_num += 1
                        for col_num in range(len(row)):
                            ws_pat.write(row_num, col_num, row[col_num], font_style)
                if ns == 0:
                    ws.sheet_visible = False
                    ws_pat.sheet_visible = False

    elif tp == "lab_sum":
        response['Content-Disposition'] = str.translate("attachment; filename=\"Статистика_Лаборатория_Колво_{}-{}.xls\"".format(date_start_o, date_end_o), tr)
        wb = openpyxl.Workbook()
        wb.remove(wb.get_sheet_by_name('Sheet'))
        ws = wb.create_sheet("Кол-во по лаборатории")

        d1 = datetime.datetime.strptime(date_start_o, '%d.%m.%Y')
        d2 = datetime.datetime.strptime(date_end_o, '%d.%m.%Y')
        start_date = datetime.datetime.combine(d1, datetime.time.min)
        end_date = datetime.datetime.combine(d2, datetime.time.max)
        lab_podr = get_lab_podr()
        lab_podr = tuple([i[0] for i in lab_podr])
        researches_by_sum = sql_func.statistics_sum_research_by_lab(lab_podr, start_date, end_date)
        ws = structure_sheet.statistic_research_by_sum_lab_base(ws, d1, d2, "Кол-во по лабораториям")
        ws = structure_sheet.statistic_research_by_sum_lab_data(ws, researches_by_sum)
    elif tp == "lab_details":
        response['Content-Disposition'] = str.translate("attachment; filename=\"Статистика_Лаборатория_детали_{}-{}.xls\"".format(date_start_o, date_end_o), tr)
        wb = openpyxl.Workbook()
        wb.remove(wb.get_sheet_by_name('Sheet'))
        ws = wb.create_sheet("Детали по лаборатории")

        d1 = datetime.datetime.strptime(date_start_o, '%d.%m.%Y')
        d2 = datetime.datetime.strptime(date_end_o, '%d.%m.%Y')
        start_date = datetime.datetime.combine(d1, datetime.time.min)
        end_date = datetime.datetime.combine(d2, datetime.time.max)
        lab_podr = get_lab_podr()
        lab_podr = tuple([i[0] for i in lab_podr])
        researches_deatails = sql_func.statistics_details_research_by_lab(lab_podr, start_date, end_date)
        ws = structure_sheet.statistic_research_by_details_lab_base(ws, d1, d2, "Детали по лаборатории")
        ws = structure_sheet.statistic_research_by_details_lab_data(ws, researches_deatails)
    elif tp == "statistics-hosp-expertise":
        response['Content-Disposition'] = str.translate("attachment; filename=\"Экспертиза_{}-{}.xls\"".format(date_start_o, date_end_o), tr)
        wb = openpyxl.Workbook()
        wb.remove(wb.get_sheet_by_name('Sheet'))
        ws = wb.create_sheet("Экспертиза")
        d1 = normalize_dots_date(date_start_o)
        d2 = normalize_dots_date(date_end_o)
        extract_researches_id = list(directory.HospitalService.objects.values_list("slave_research_id", flat=True).filter(site_type=7))
        field_id_for_extract_date = list(directory.ParaclinicInputField.objects.values_list("pk", flat=True).filter(group__research__in=extract_researches_id, title="Дата выписки"))
        result_extract = get_confirm_protocol_by_date_extract(tuple(field_id_for_extract_date), d1, d2)  # Найти выписки с датой выписки в периоде
        result_expertise_data = {i.iss_protocol_extract: {"title_research": i.main_extract_research, "direction_main_extract_dir": i.direction_main_extract_dir} for i in result_extract}
        iss_protocol_extract = list(result_expertise_data.keys())
        result_expertise_grade = get_expertise_grade(tuple(iss_protocol_extract))  # Результаты экспертизы
        for i in result_expertise_grade:
            if i.level_value and i.level_value.lower() == "третий":
                result_expertise_data[i.parent_id]['третий'] = i.grade_value
            elif i.level_value and i.level_value.lower() == "второй":
                result_expertise_data[i.parent_id]['второй'] = i.grade_value
            else:
                result_expertise_data[i.parent_id]['без уровня'] = i.grade_value
        final_result = {}
        for i in result_expertise_data.values():
            if not final_result.get(i["title_research"]):
                final_result[i["title_research"]] = [i]
            else:
                final_result[i["title_research"]].append(i)
        ws = expertise_report.expertise_data(ws, final_result)
    elif tp == "statistics-dispanserization":
        response['Content-Disposition'] = str.translate("attachment; filename=\"Статистика_Диспансеризация_{}-{}.xls\"".format(date_start_o, date_end_o), tr)
        wb = openpyxl.Workbook()
        wb.remove(wb.get_sheet_by_name('Sheet'))
        ws = wb.create_sheet("Диспансеризация")
        d1 = datetime.datetime.strptime(date_start_o, '%d.%m.%Y')
        d2 = datetime.datetime.strptime(date_end_o, '%d.%m.%Y')
        start_date = datetime.datetime.combine(d1, datetime.time.min)
        end_date = datetime.datetime.combine(d2, datetime.time.max)
        services_start = DISPANSERIZATION_SERVICE_PK.get("pkServiceStart", [])
        service_end = DISPANSERIZATION_SERVICE_PK.get("pkServiceEnd", [])
        services = services_start.copy()
        services.extend(service_end)
        query = sql_func.statistics_dispanserization(tuple(services), start_date, end_date)
        doctors = dispanserization.get_doctors_dispanserization(query)
        doctors_count_pass_patient = sql_func.doctors_pass_count_patient_by_date(tuple(doctors), start_date, end_date)
        result_dates = dispanserization.dispanserization_data(query, services_start, service_end, doctors_count_pass_patient)

        ws = dispanserization.dispanserization_base(ws, d1, d2, result_dates)
        ws = dispanserization.dispanserization_fill_data(ws, result_dates)
    elif tp == "disp-plan":
        response['Content-Disposition'] = str.translate("attachment; filename=\"План Д-учет_{}-{}.xls\"".format(date_start_o, date_end_o), tr)
        wb = openpyxl.Workbook()
        wb.remove(wb.get_sheet_by_name('Sheet'))
        ws = wb.create_sheet("Дисп-учет")
        month = request_data.get("month", "")
        year = request_data.get("year", "")
        month_obj = int(month)
        monthes = {
            "1": "Январь",
            "2": "Февраль",
            "3": "Март",
            "4": "Апрель",
            "5": "Май",
            "6": "Июнь",
            "7": "Июль",
            "8": "Август",
            "9": "Сентябрь",
            "10": "Октябрь",
            "11": "Ноябрь",
            "12": "Декабрь",
        }
        _, num_days = calendar.monthrange(int(year), month_obj)
        d1 = datetime.date(int(year), month_obj, 1)
        d2 = datetime.date(int(year), month_obj, num_days)
        start_date = f"{d1.strftime('%Y-%m-%d')} 00:00:00"
        end_date = f"{d2.strftime('%Y-%m-%d')} 23:59:59"
        query = sql_func.dispansery_plan(start_date, end_date)
        cards_pk = [i.card_id for i in query]
        ws = dispensary_data.dispansery_plan_base(ws, monthes[month], year)
        if len(cards_pk) > 0:
            query_diagnoses_pk = sql_func.dispansery_card_diagnos(tuple(cards_pk))
            result = dispensary_data.handle_query(query, query_diagnoses_pk)
            ws = dispensary_data.dispansery_plan_fill_data(ws, result)
    elif tp == "disp-registered":
        response['Content-Disposition'] = str.translate("attachment; filename=\"План Д-учет_{}-{}.xls\"".format(date_start_o, date_end_o), tr)
        wb = openpyxl.Workbook()
        wb.remove(wb.get_sheet_by_name('Sheet'))
        ws = wb.create_sheet("Дисп-учет зарегистрировано")
        data_date = request_data.get("date_values")
        data_date = json.loads(data_date)

        if request_data.get("date_type") == 'd':
            d1 = datetime.datetime.strptime(data_date['date'], '%d.%m.%Y')
        else:
            month_obj = int(data_date['month']) + 1
            _, num_days = calendar.monthrange(int(data_date['year']), month_obj)
            d1 = datetime.date(int(data_date['year']), month_obj, num_days)
        child = sql_func.dispansery_registered_by_year_age(18, d1, 1)
        adult = sql_func.dispansery_registered_by_year_age(18, d1, 0)
        result = [{"adult": len(adult), "child": len(child)}]
        ws = dispensary_data.dispansery_reg_count_base(ws, d1)
        ws = dispensary_data.dispansery_reg_count_fill_data(ws, result)
    elif tp == "covid_sum":
        response['Content-Disposition'] = str.translate("attachment; filename=\"Статистика_Лаборатория_Колво_{}-{}.xls\"".format(date_start_o, date_end_o), tr)
        wb = openpyxl.Workbook()
        wb.remove(wb.get_sheet_by_name('Sheet'))
        ws = wb.create_sheet("Кол-во по Ковид")

        pk = request_data.get("research")
        d1 = datetime.datetime.strptime(date_start_o, '%d.%m.%Y')
        d2 = datetime.datetime.strptime(date_end_o, '%d.%m.%Y')
        start_date = datetime.datetime.combine(d1, datetime.time.min)
        end_date = datetime.datetime.combine(d2, datetime.time.max)
        result_patient = sql_get_result_by_direction(pk, start_date, end_date)
        cards = tuple(set([i.client_id for i in result_patient]))
        document_card = sql_get_documents_by_card_id(cards)
        patient_docs = {}
        document_type = {4: "снилс", 5: "рождение", 1: "паспорт", 3: "полис"}
        for doc in document_card:
            data = None
            if doc.document_type_id in [4, 3]:
                data = {document_type.get(doc.document_type_id): doc.number}
            elif doc.document_type_id in [1, 5]:
                data = {document_type.get(doc.document_type_id): f"{doc.serial}@{doc.number}"}
            if patient_docs.get(doc.card_id, None):
                temp_docs = patient_docs.get(doc.card_id)
                temp_docs.append(data)
                patient_docs[doc.card_id] = temp_docs
            else:
                if data:
                    patient_docs[doc.card_id] = [data]

        ws = structure_sheet.statistic_research_by_covid_base(ws, d1, d2, "Кол-во по ковид")
        ws = structure_sheet.statistic_research_by_covid_data(ws, result_patient, patient_docs)
    elif tp == "lab-staff":
        lab = Podrazdeleniya.objects.get(pk=int(pk))
        researches = list(directory.Researches.objects.filter(podrazdeleniye=lab, hide=False).order_by('title').order_by("sort_weight").order_by("direction_id"))
        pods = list(Podrazdeleniya.objects.filter(p_type=Podrazdeleniya.DEPARTMENT).order_by("title"))
        response['Content-Disposition'] = str.translate(
            "attachment; filename=\"Статистика_Исполнители_Лаборатория_{0}_{1}-{2}.xls\"".format(lab.title.replace(" ", "_"), date_start_o, date_end_o), tr
        )

        import directions.models as d
        from operator import itemgetter

        date_start, date_end = try_parse_range(date_start_o, date_end_o)
        iss = Issledovaniya.objects.filter(research__podrazdeleniye=lab, time_confirmation__isnull=False, time_confirmation__range=(date_start, date_end))

        font_style_wrap = xlwt.XFStyle()
        font_style_wrap.alignment.wrap = 1
        font_style_wrap.borders = borders
        font_style_vertical = xlwt.easyxf('align: rotation 90')
        font_style_vertical.borders = borders

        def val(v):
            return "" if v == 0 else v

        def nl(v):
            return v + ("" if len(v) > 19 else "\n")

        for executor in DoctorProfile.objects.filter(user__groups__name__in=("Врач-лаборант", "Лаборант"), podrazdeleniye__p_type=Podrazdeleniya.LABORATORY).order_by("fio").distinct():

            cnt_itogo = {}
            ws = wb.add_sheet(executor.get_fio(dots=False) + " " + str(executor.pk))

            row_num = 0
            row = [("Исполнитель", 5500), ("Отделение", 5000)]

            from django.utils.text import Truncator

            for research in researches:
                row.append(
                    (
                        Truncator(research.title).chars(30),
                        1300,
                    )
                )

            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num][0], font_style_wrap if col_num < 2 else font_style_vertical)
                ws.col(col_num).width = row[col_num][1]

            row_num += 1
            itogo_row = [executor.get_fio(dots=True), nl("Итого")]
            empty_row = ["", ""]
            cnt_local_itogo = {}
            for pod in pods:
                row = [executor.get_fio(dots=True), nl(pod.title)]
                cnt = {}
                for research in researches:
                    if research.title not in cnt.keys():
                        cnt[research.title] = 0
                    if research.title not in cnt_local_itogo.keys():
                        cnt_local_itogo[research.title] = 0
                    if research.title not in cnt_itogo.keys():
                        cnt_itogo[research.title] = 0

                    for i in iss.filter(doc_confirmation=executor, napravleniye__doc__podrazdeleniye=pod, research=research):
                        isadd = False
                        allempty = True
                        for r in Result.objects.filter(issledovaniye=i):
                            value = r.value.lower().strip()
                            if value != "":
                                allempty = False
                                n = any([y in value for y in ["забор", "тест", "неправ", "ошибк", "ошибочный", "кров", "брак", "мало", "недостаточно", "реактив"]])
                                if not n:
                                    isadd = True

                        if not isadd or allempty:
                            continue

                        cnt[research.title] += 1
                        cnt_itogo[research.title] += 1
                        cnt_local_itogo[research.title] += 1
                for research in researches:
                    row.append(val(cnt[research.title]))
                    # data["otds"][pod.title] += 1
                    # data["all"][pod.title] += 1
                    # cnt_all[pod.title] += 1
                for col_num in range(len(row)):
                    ws.write(row_num, col_num, row[col_num], font_style_wrap)
                row_num += 1

            for research in researches:
                itogo_row.append(val(cnt_local_itogo[research.title]))
                empty_row.append("")
            for col_num in range(len(itogo_row)):
                ws.write(row_num, col_num, itogo_row[col_num], font_style_wrap)
            row_num += 1
    elif tp == "otd":
        otd = Podrazdeleniya.objects.get(pk=int(pk))
        response['Content-Disposition'] = str.translate("attachment; filename=\"Статистика_Отделение_{0}_{1}-{2}.xls\"".format(otd.title.replace(" ", "_"), date_start_o, date_end_o), tr)

        ws = wb.add_sheet("Выписано направлений")

        font_style = xlwt.XFStyle()
        row_num = 0
        row = ["За период: ", "{0} - {1}".format(date_start_o, date_end_o)]

        date_start_o, date_end_o = try_parse_range(date_start_o, date_end_o)

        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)
        row_num += 1

        font_style = xlwt.XFStyle()

        row = [otd.title]

        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

        font_style = xlwt.XFStyle()

        row_num += 1
        row = [
            (u"Всего выписано", 6000),
            (str(Napravleniya.objects.filter(doc__podrazdeleniye=otd, data_sozdaniya__range=(date_start_o, date_end_o)).count()), 3000),
        ]

        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num][0], font_style)
            ws.col(col_num).width = row[col_num][1]

        row_num += 1
        researches = Issledovaniya.objects.filter(napravleniye__doc__podrazdeleniye=otd, napravleniye__data_sozdaniya__range=(date_start_o, date_end_o), time_confirmation__isnull=False)
        naprs = len(set([v.napravleniye_id for v in researches]))
        row = [u"Завершенных", str(naprs)]

        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    elif tp == "list-users":
        response['Content-Disposition'] = str.translate("attachment; filename=\"Список_пользователей.xls\"", tr)
        ws = wb.add_sheet("Пользователи")
        row_num = 0
        font_style = xlwt.XFStyle()
        for p in Podrazdeleniya.objects.filter(hide=False).order_by("title"):
            has = False
            for u in DoctorProfile.objects.filter(podrazdeleniye=p).exclude(user__username="admin").order_by("fio"):
                has = True
                row = [("ID отделения %s" % p.pk, 9000), (p.title, 9000), ("ID пользователя %s" % u.pk, 9000), (u.user.username, 5000), (u.fio, 10000)]
                for col_num in range(len(row)):
                    ws.write(row_num, col_num, row[col_num][0], font_style)
                    ws.col(col_num).width = row[col_num][1]
                row_num += 1
            if has:
                row_num += 1
    elif tp == "lab-receive":
        lab = Podrazdeleniya.objects.get(pk=int(pk))
        response['Content-Disposition'] = str.translate(
            "attachment; filename=\"Статистика_Принято_емкостей_{0}_{1}-{2}.xls\"".format(lab.title.replace(" ", "_"), date_start_o, date_end_o), tr
        )

        import directions.models as d
        from operator import itemgetter

        date_start, date_end = try_parse_range(date_start_o, date_end_o)
        ws = wb.add_sheet(lab.title)

        font_style_wrap = xlwt.XFStyle()
        font_style_wrap.alignment.wrap = 1
        font_style_wrap.borders = borders
        font_style = xlwt.XFStyle()
        font_style.borders = borders

        row_num = 0
        row = [
            (lab.title + ", принято емкостей за {0}-{1}".format(date_start_o, date_end_o), 16000),
        ]

        replace = [{"from": "-", "to": " "}, {"from": ".", "to": " "}, {"from": " и ", "to": " "}]
        n = len(row) - 1
        pods = Podrazdeleniya.objects.filter(p_type=Podrazdeleniya.DEPARTMENT).order_by("title")
        for pod in pods:
            n += 1
            title = pod.title
            for rep in replace:
                title = title.replace(rep["from"], rep["to"])

            tmp = title.split()
            title = []
            nx = 0
            for x in tmp:
                x = x.strip()
                if len(x) == 0:
                    continue

                title.append(x if x.isupper() else x[0].upper() + ("" if nx > 0 else x[1:7]))
                nx += 1

            row.append(
                (
                    "".join(title),
                    3700,
                )
            )
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num][0], font_style)
            ws.col(col_num).width = row[col_num][1]
        row_num += 1

        for tube in directory.Tubes.objects.filter(releationsft__fractions__research__podrazdeleniye=lab).distinct().order_by("title"):
            row = [tube.title]
            for pod in pods:
                gets = (
                    d.TubesRegistration.objects.filter(issledovaniya__research__podrazdeleniye=lab, type__tube=tube, time_recive__range=(date_start, date_end), doc_get__podrazdeleniye=pod)
                    .filter(Q(notice="") | Q(notice__isnull=True))
                    .distinct()
                )
                row.append("" if not gets.exists() else str(gets.count()))

            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)
            row_num += 1

    elif tp == "all-labs":
        labs = Podrazdeleniya.objects.filter(p_type=Podrazdeleniya.LABORATORY).exclude(title="Внешние организации")
        response['Content-Disposition'] = str.translate("attachment; filename=\"Статистика_Все_Лаборатории_{0}-{1}.xls\"".format(date_start_o, date_end_o), tr)
        ws = wb.add_sheet("Выполненных анализов")

        font_style = xlwt.XFStyle()
        row_num = 0
        row = ["За период: ", "{0} - {1}".format(date_start_o, date_end_o)]

        date_start_o, date_end_o = try_parse_range(date_start_o, date_end_o)

        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)
        row_num += 1

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = [
            (u"Лаборатория", 9000),
            (u"Выполнено анализов", 8000),
        ]

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num][0], font_style)
            ws.col(col_num).width = columns[col_num][1]

        font_style = xlwt.XFStyle()
        font_style.alignment.wrap = 1
        all = 0
        for lab in labs:
            row_num += 1
            c = Issledovaniya.objects.filter(research__podrazdeleniye=lab, time_confirmation__isnull=False, time_confirmation__range=(date_start_o, date_end_o)).count()
            row = [lab.title, c]
            all += c
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)
        row_num += 1
        row = [
            "",
            "Всего: " + str(all),
        ]
        font_style = xlwt.XFStyle()
        font_style.alignment.wrap = 3
        font_style.alignment.horz = 3
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)
    elif tp == "tubes-using":
        response['Content-Disposition'] = str.translate("attachment; filename=\"Статистика_Использование_Емкостей_{0}-{1}.xls\"".format(date_start_o, date_end_o), tr)

        per = "{0} - {1}".format(date_start_o, date_end_o)

        ws = wb.add_sheet("Общее использование емкостей")
        font_style = xlwt.XFStyle()
        row_num = 0
        row = ["За период: ", per]

        date_start_o, date_end_o = try_parse_range(date_start_o, date_end_o)

        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)
        row_num += 1

        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        columns = [
            (u"Тип емкости", 9000),
            (u"Материал взят в процедурном каб", 9000),
            (u"Принято лабораторией", 8000),
            (u"Не принято лабораторией", 8000),
            (u"Потеряны", 4000),
        ]

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num][0], font_style)
            ws.col(col_num).width = columns[col_num][1]

        font_style = xlwt.XFStyle()
        font_style.alignment.wrap = 1
        all_get = 0
        all_rec = 0
        all_nrec = 0
        all_lost = 0
        for tube in Tubes.objects.all():
            row_num += 1
            c_get = TubesRegistration.objects.filter(type__tube=tube, time_get__isnull=False, time_get__range=(date_start_o, date_end_o)).count()
            c_rec = TubesRegistration.objects.filter(type__tube=tube, time_recive__isnull=False, notice="", time_get__range=(date_start_o, date_end_o)).count()
            c_nrec = TubesRegistration.objects.filter(type__tube=tube, time_get__isnull=False, time_get__range=(date_start_o, date_end_o)).exclude(notice="").count()
            str1 = ""
            str2 = ""
            if c_nrec > 0:
                str1 = str(c_nrec)
            if c_get - c_rec - all_nrec > 0:
                str2 = str(c_get - c_rec - all_nrec)
                all_lost += c_get - c_rec - all_nrec

            row = [tube.title, c_get, c_rec, str1, str2]
            all_get += c_get
            all_rec += c_rec
            all_nrec += c_nrec
            for col_num in range(len(row)):
                font_style.alignment.wrap = 1
                font_style.alignment.horz = 1
                if col_num > 0:
                    font_style.alignment.wrap = 3
                    font_style.alignment.horz = 3
                ws.write(row_num, col_num, row[col_num], font_style)

        labs = Podrazdeleniya.objects.filter(p_type=Podrazdeleniya.LABORATORY).exclude(title="Внешние организации")
        for lab in labs:
            ws = wb.add_sheet(lab.title)
            font_style = xlwt.XFStyle()
            row_num = 0
            row = ["За период: ", per]
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)
            row_num += 1

            font_style = xlwt.XFStyle()
            font_style.font.bold = True
            columns = [
                (u"Тип емкости", 9000),
                (u"Материал взят в процедурном каб", 9000),
                (u"Принято лабораторией", 8000),
                (u"Не принято лабораторией", 8000),
                (u"Потеряны", 4000),
            ]

            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num][0], font_style)
                ws.col(col_num).width = columns[col_num][1]

            font_style = xlwt.XFStyle()
            font_style.alignment.wrap = 1
            all_get = 0
            all_rec = 0
            all_nrec = 0
            all_lost = 0
            for tube in Tubes.objects.all():
                row_num += 1
                c_get = TubesRegistration.objects.filter(
                    issledovaniya__research__podrazdeleniye=lab, type__tube=tube, time_get__isnull=False, time_get__range=(date_start_o, date_end_o)
                ).count()
                c_rec = TubesRegistration.objects.filter(
                    issledovaniya__research__podrazdeleniye=lab, type__tube=tube, time_recive__isnull=False, notice="", time_get__range=(date_start_o, date_end_o)
                ).count()
                c_nrec = (
                    TubesRegistration.objects.filter(issledovaniya__research__podrazdeleniye=lab, type__tube=tube, time_get__isnull=False, time_get__range=(date_start_o, date_end_o))
                    .exclude(notice="")
                    .count()
                )
                str1 = ""
                str2 = ""
                if c_nrec > 0:
                    str1 = str(c_nrec)
                if c_get - c_rec - all_nrec > 0:
                    str2 = str(c_get - c_rec - all_nrec)
                    all_lost += c_get - c_rec - all_nrec

                row = [tube.title, c_get, c_rec, str1, str2]
                all_get += c_get
                all_rec += c_rec
                all_nrec += c_nrec
                for col_num in range(len(row)):
                    font_style.alignment.wrap = 1
                    font_style.alignment.horz = 1
                    if col_num > 0:
                        font_style.alignment.wrap = 3
                        font_style.alignment.horz = 3
                    ws.write(row_num, col_num, row[col_num], font_style)

    elif tp == "uets":
        usrs = DoctorProfile.objects.filter(podrazdeleniye__p_type=Podrazdeleniya.LABORATORY).order_by("podrazdeleniye__title")
        response['Content-Disposition'] = str.translate("attachment; filename=\"Статистика_УЕТс_{0}-{1}.xls\"".format(date_start_o, date_end_o), tr)

        ws = wb.add_sheet("УЕТы")

        font_style = xlwt.XFStyle()
        row_num = 0
        row = ["За период: ", "{0} - {1}".format(date_start_o, date_end_o)]

        date_start_o, date_end_o = try_parse_range(date_start_o, date_end_o)

        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        row_num += 1
        row = [
            (u"Лаборатория", 8000),
            (u"ФИО", 8000),
            (u"УЕТы", 2500),
        ]

        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num][0], font_style)
            ws.col(col_num).width = row[col_num][1]

        font_style = xlwt.XFStyle()
        for usr in usrs:
            researches_uets = {}
            researches = Issledovaniya.objects.filter(doc_save=usr, time_save__isnull=False, time_save__range=(date_start_o, date_end_o))
            for issledovaniye in researches:
                if usr.labtype == 1:
                    uet_tmp = sum([v.uet_doc for v in directory.Fractions.objects.filter(research=issledovaniye.research)])
                else:
                    uet_tmp = sum([v.uet_lab for v in directory.Fractions.objects.filter(research=issledovaniye.research)])
                researches_uets[issledovaniye.pk] = {"uet": uet_tmp}
            researches = Issledovaniya.objects.filter(doc_confirmation=usr, time_confirmation__isnull=False, time_confirmation__range=(date_start_o, date_end_o))
            for issledovaniye in researches:
                if usr.labtype == 1:
                    uet_tmp = sum([v.uet_doc for v in directory.Fractions.objects.filter(research=issledovaniye.research)])
                else:
                    uet_tmp = sum([v.uet_lab for v in directory.Fractions.objects.filter(research=issledovaniye.research)])
                researches_uets[issledovaniye.pk] = {"uet": uet_tmp}
            uets = sum([researches_uets[v]["uet"] for v in researches_uets.keys()])
            row_num += 1
            row = [
                usr.podrazdeleniye.title,
                usr.get_full_fio(),
                uets,
            ]
            for col_num in range(len(row)):
                font_style.alignment.wrap = 1
                font_style.alignment.horz = 1
                if col_num > 2:
                    font_style.alignment.wrap = 3
                    font_style.alignment.horz = 3
                ws.write(row_num, col_num, row[col_num], font_style)
    elif tp == "message-ticket":
        filters = {'pk': int(request_data.get("hospital"))}
        any_hospital = request.user.doctorprofile.all_hospitals_users_control
        if not any_hospital:
            filters['pk'] = request.user.doctorprofile.get_hospital_id()

        response['Content-Disposition'] = str.translate(f"attachment; filename=\"Обращения {date_start_o.replace('.', '')} {date_end_o.replace('.', '')} {filters['pk']}.xlsx\"", tr)

        wb = openpyxl.Workbook()
        wb.remove(wb.get_sheet_by_name('Sheet'))
        ws = wb.create_sheet("Обращения")
        styles_obj = structure_sheet.style_sheet()
        wb.add_named_style(styles_obj[0])
        if int(filters['pk']) == -1 and any_hospital:
            filters = {}
        rows_hosp = list(Hospitals.objects.values_list('pk', flat=True).filter(hide=False, **filters))
        d1 = datetime.datetime.strptime(date_start_o, '%d.%m.%Y')
        d2 = datetime.datetime.strptime(date_end_o, '%d.%m.%Y')
        ws = structure_sheet.statistic_message_ticket_base(ws, date_start_o, date_end_o, styles_obj[3])
        start_date = datetime.datetime.combine(d1, datetime.time.min)
        end_date = datetime.datetime.combine(d2, datetime.time.max)
        message_ticket_sql = sql_func.message_ticket(rows_hosp, start_date, end_date)
        ws = structure_sheet.statistic_message_ticket_data(ws, message_ticket_sql, styles_obj[3])
        ws = wb.create_sheet("Итоги-Обращения")
        message_total_purpose_sql = sql_func.message_ticket_purpose_total(rows_hosp, start_date, end_date)
        ws = structure_sheet.statistic_message_purpose_total_data(ws, message_total_purpose_sql, date_start_o, date_end_o, styles_obj[3])

    elif tp == "statistics-consolidate":
        response['Content-Disposition'] = str.translate("attachment; filename=\"Свод пациенты-услуги_{}-{}.xls\"".format(date_start_o, date_end_o), tr)
        wb = openpyxl.Workbook()
        wb.remove(wb.get_sheet_by_name('Sheet'))
        ws = wb.create_sheet("Сводный")
        d1 = datetime.datetime.strptime(date_start_o, '%d.%m.%Y')
        d2 = datetime.datetime.strptime(date_end_o, '%d.%m.%Y')
        start_date = datetime.datetime.combine(d1, datetime.time.min)
        end_date = datetime.datetime.combine(d2, datetime.time.max)

        type_fin = request_data.get("fin")
        title_fin = IstochnikiFinansirovaniya.objects.filter(pk=type_fin).first()
        set_research = int(request_data.get("research-set", -1))
        company_id = int(request_data.get("company", -1))
        query = None
        is_research_set = -1
        head_data = {}
        if set_research > 0:
            set_research_d = directory.SetOrderResearch.objects.filter(set_research_id=set_research).order_by("order")
            head_data = {i.research_id: i.research.title for i in set_research_d}
            is_research_set = 1
        if company_id > 0:
            def_value_data = {k: 0 for k in head_data.keys()}
            price = get_price_company(company_id, start_date, end_date)
            if price:
                research_coast = PriceCoast.get_coast_by_researches(price, list(def_value_data.keys()))
            else:
                price = title_fin.contracts.price
                research_coast = PriceCoast.get_coast_by_researches(price, list(def_value_data.keys()))
            query = sql_func.statistics_by_research_sets_company(start_date, end_date, type_fin, tuple(def_value_data.keys()), company_id)
            head_data_coast = {k: research_coast.get(k, "") for k, v in head_data.items()}
            if company_id > 0:
                company = Company.objects.get(pk=company_id)
                company_title = company.title
            else:
                company_title = ""
            ws, start_research_column = consolidates.consolidate_research_sets_base(ws, d1, d2, title_fin.title, head_data, company_title, head_data_coast)
            ws = consolidates.consolidate_research_sets_fill_data(ws, query, def_value_data, start_research_column)
        else:
            def_value_data = {k: 0 for k in head_data.keys()}
            researche_ids = (-1,)
            if is_research_set == 1:
                researche_ids = tuple(def_value_data.keys())
            query = sql_func.statistics_consolidate_research(start_date, end_date, type_fin, is_research_set, researche_ids)
            ws = consolidates.consolidate_base(ws, d1, d2, title_fin.title)
            ws = consolidates.consolidate_fill_data(ws, query)
    elif tp == "consolidate-type-department":
        response['Content-Disposition'] = str.translate("attachment; filename=\"Свод пациенты-услуги_{}-{}.xls\"".format(date_start_o, date_end_o), tr)
        wb = openpyxl.Workbook()
        wb.remove(wb.get_sheet_by_name('Sheet'))
        ws = wb.create_sheet("Врачи-сводный по подразделению")
        d1 = datetime.datetime.strptime(date_start_o, '%d.%m.%Y')
        d2 = datetime.datetime.strptime(date_end_o, '%d.%m.%Y')
        start_date = datetime.datetime.combine(d1, datetime.time.min)
        end_date = datetime.datetime.combine(d2, datetime.time.max)

        type_fin = int(request_data.get("fin", -1))
        detail_patient = int(request_data.get("detail-patient", -1))
        if type_fin == -100:
            type_fin = tuple(IstochnikiFinansirovaniya.objects.values_list('id', flat=True).filter(base__internal_type=True))
        else:
            type_fin = (type_fin,)
        type_department = int(request_data.get("type-department", -1))
        doctors = tuple(DoctorProfile.objects.values_list('id', flat=True).filter(podrazdeleniye__p_type=type_department, position__title__icontains="врач"))
        fin_source_data_doctors = IstochnikiFinansirovaniya.objects.values_list('id', 'title').filter(id__in=type_fin).order_by('order_weight')
        fin_source_data = {}
        for i in fin_source_data_doctors:
            fin_source_data[i[0]] = i[1]
        query_doctors = sql_func.consolidate_doctors_by_type_department(start_date, end_date, type_fin, doctors)
        ws_and_finish_order = consolidates.consolidate_base_doctors_by_type_department(ws, d1, d2, fin_source_data)
        ws = ws_and_finish_order[0]
        if detail_patient == 1:
            ws = consolidates.consolidate_fill_data_doctors_by_type_department_detail_patient(ws, query_doctors, ws_and_finish_order[1])
        else:
            ws = consolidates.consolidate_fill_data_doctors_by_type_department(ws, query_doctors, ws_and_finish_order[1])
    elif tp == "statistics-registry-profit":
        response['Content-Disposition'] = str.translate("attachment; filename=\"Реестр_{}-{}.xls\"".format(date_start_o, date_end_o), tr)
        wb = openpyxl.Workbook()
        wb.remove(wb.get_sheet_by_name('Sheet'))
        ws = wb.create_sheet("Реестр")
        d1 = datetime.datetime.strptime(date_start_o, '%d.%m.%Y')
        d2 = datetime.datetime.strptime(date_end_o, '%d.%m.%Y')
        start_date = datetime.datetime.combine(d1, datetime.time.min)
        end_date = datetime.datetime.combine(d2, datetime.time.max)
        type_fin = int(request_data.get("fin", -1))
        data = sql_func.statistics_registry_profit(start_date, end_date, type_fin)
        companies_id = set([int(i.company_id) for i in data if i.company_id is not None])
        companies_price = {}
        price_companies = {}
        for company_id in list(companies_id):
            price = get_price_company(company_id, start_date, end_date)
            if price:
                companies_price[company_id] = price.id
                price_companies[price.id] = {"company_id": company_id}
        coast_research_price = get_research_coast_by_prce(tuple(price_companies.keys()))

        for coast in coast_research_price:
            price_companies[coast.price_name_id][coast.research_id] = float(coast.coast)

        result = {}
        for d in data:
            coast = None
            coast_price_research = None
            if d.company_id:
                id_price = companies_price.get(d.company_id)
                if id_price:
                    coast_price_research = price_companies.get(id_price)
                if coast_price_research:
                    coast = coast_price_research.get(d.research_id)
            if not coast:
                coast = 0
            if not result.get(d.doc_confirmation_id):
                result[d.doc_confirmation_id] = {
                    "fio": f"{d.doc_family} {d.doc_name} {d.doc_patronymic}",
                    "position": d.position_title,
                    "researches": {
                        d.research_id: {
                            "companies": {
                                d.company_id: {
                                    "coasts": {coast: 1},
                                    "company_title": d.company_title,
                                },
                            },
                            "research_title": d.research_title,
                        }
                    },
                }
            else:
                tmp_doctor_researches = result[d.doc_confirmation_id]["researches"]
                if not tmp_doctor_researches.get(d.research_id):
                    tmp_doctor_researches[d.research_id] = {
                        "companies": {
                            d.company_id: {
                                "coasts": {coast: 1},
                                "company_title": d.company_title,
                            },
                        },
                        "research_title": d.research_title,
                    }
                    result[d.doc_confirmation_id]["researches"] = tmp_doctor_researches.copy()
                else:
                    tmp_research = tmp_doctor_researches.get(d.research_id)
                    if not tmp_research["companies"].get(d.company_id):
                        tmp_research["companies"][d.company_id] = {
                            "coasts": {coast: 1},
                            "company_title": d.company_title,
                        }
                        tmp_doctor_researches[d.research_id] = tmp_research["companies"].copy()
                    else:
                        coasts = tmp_research["companies"][d.company_id]["coasts"]
                        if not coasts.get(coast):
                            coasts[coast] = 1
                            tmp_research["companies"][d.company_id]["coasts"] = coasts.copy()
                        else:
                            coast_data = coasts.get(coast)
                            coast_data += 1
                            coasts[coast] = coast_data
                            tmp_research["companies"][d.company_id]["coasts"] = coasts.copy()
                    tmp_doctor_researches[d.research_id] = tmp_research.copy()
                    result[d.doc_confirmation_id]["researches"] = tmp_doctor_researches.copy()
        final_result = []
        for k, v in result.items():
            for id_research, data_research in v['researches'].items():
                for id_company, data_company in data_research['companies'].items():
                    if id_company != "research_title":
                        for coats, count in data_company['coasts'].items():
                            final_result.append(
                                {
                                    "fio": v.get("fio", "-"),
                                    "position": v.get("position", "-"),
                                    "research": data_research.get("research_title", "-"),
                                    "company": data_company.get("company_title"),
                                    "coast": coats,
                                    "count": count,
                                }
                            )

        ws = registry_profit.profit_base(ws, date_start_o, date_end_o)
        ws = registry_profit.profit_data(ws, final_result)

    elif tp == "statistics-corp-create":
        response['Content-Disposition'] = str.translate("attachment; filename=\"Услуги.xlsx\"", tr)
        user_groups = request.user.groups.values_list('name', flat=True)
        data_date = request_data.get("date_values")
        data_date = json.loads(data_date)
        if request_data.get("date_type") == 'd':
            d1 = datetime.datetime.strptime(data_date['date'], '%d.%m.%Y')
            d2 = datetime.datetime.strptime(data_date['date'], '%d.%m.%Y')
            month_obj = ''
        elif request_data.get("date_type") == 'm':
            month_obj = int(data_date['month']) + 1
            _, num_days = calendar.monthrange(int(data_date['year']), month_obj)
            d1 = datetime.date(int(data_date['year']), month_obj, 1)
            d2 = datetime.date(int(data_date['year']), month_obj, num_days)
        else:
            d_s = request_data.get("date-start")
            d_e = request_data.get("date-end")
            d1 = datetime.datetime.strptime(d_s, '%d.%m.%Y')
            d2 = datetime.datetime.strptime(d_e, '%d.%m.%Y')

        wb = openpyxl.Workbook()
        wb.remove(wb.get_sheet_by_name('Sheet'))
        ws = wb.create_sheet("Отчет")
        start_date = datetime.datetime.combine(d1, datetime.time.min)
        end_date = datetime.datetime.combine(d2, datetime.time.max)
        hospital_id = request.user.doctorprofile.hospital_id
        by_create_direction = False
        if request_data.get("by-create-direction", "false") == "true":
            by_create_direction = True
        if 'Статистика-все МО' in user_groups:
            hospital_id = -1

        if by_create_direction:
            researches_sql = sql_func.statistics_corp_by_create_direction(start_date, end_date)
        else:
            researches_sql = sql_func.statistics_corp_by_confirm_direction(start_date, end_date)

        result = []
        step = 0
        old_patient_summ = 0
        old_hosp_title = None
        old_patient_card = None
        old_patient_family = None
        old_patient_name = None
        old_patient_patronymic = None
        old_patient_birthday = None

        for i in researches_sql:
            if step != 0 and (i.patient_card != old_patient_card):
                old_patient_summ = 0
                result.append(
                    {
                        "direction_num": "-",
                        "target_date": "-",
                        "research_title": "-",
                        "hospital_title": old_hosp_title,
                        "patient_family": old_patient_family,
                        "patient_name": old_patient_name,
                        "patient_patronymic": old_patient_patronymic,
                        "patient_birthday": old_patient_birthday,
                        "patient_card": "Итого по пациенту",
                        "research_coast": old_patient_summ,
                        "tube_number": "-",
                    }
                )

            old_hosp_title = i.hospital_title
            old_patient_card = i.patient_card
            old_patient_family = i.patient_family
            old_patient_name = i.patient_name
            old_patient_patronymic = i.patient_patronymic
            old_patient_birthday = i.patient_birthday
            old_patient_summ += i.research_coast if i.research_coast else 0
            result.append(
                {
                    "direction_num": i.direction_num,
                    "target_date": i.target_date,
                    "research_title": i.research_title,
                    "hospital_title": i.hospital_title,
                    "patient_family": i.patient_family,
                    "patient_name": i.patient_name,
                    "patient_patronymic": i.patient_patronymic,
                    "patient_birthday": i.patient_birthday,
                    "patient_card": i.patient_card,
                    "research_coast": i.research_coast if i.research_coast else 0,
                    "tube_number": i.tube_number if i.tube_number else "-",
                }
            )
            step += 1

        result.append(
            {
                "direction_num": "-",
                "target_date": "-",
                "research_title": "-",
                "hospital_title": old_hosp_title,
                "patient_family": old_patient_family,
                "patient_name": old_patient_name,
                "patient_patronymic": old_patient_patronymic,
                "patient_birthday": old_patient_birthday,
                "patient_card": "Итого по пациенту",
                "research_coast": old_patient_summ,
                "tube_number": "-",
            }
        )

        ws = partner_coast_data.partner_coast_base(ws)
        ws = partner_coast_data.partner_coast_fill_data(ws, result)

    wb.save(response)
    return response


@csrf_exempt
@login_required
def sreening_xls(request):
    request_data = request.POST if request.method == "POST" else request.GET
    month = request_data.get("month", "")
    year = request_data.get("year", "")
    month_obj = int(month)
    _, num_days = calendar.monthrange(int(year), month_obj)
    d1 = datetime.date(int(year), month_obj, 1)
    d2 = datetime.date(int(year), month_obj, num_days)
    datetime_start = f"{d1.strftime('%Y-%m-%d')} 00:00:00"
    datetime_end = f"{d2.strftime('%Y-%m-%d')} 23:59:59"
    last_day_month = f"{d2.strftime('%Y-%m-%d')}"

    # кол-во прикрепленных по возрасту всего
    min_age = 18
    max_age = 69
    screening_data = {}

    attached_count_age_for_month = attached_female_on_month(last_day_month, min_age, max_age)
    screening_data['attached_count_age_for_month'] = attached_count_age_for_month[0].count

    # кол-во в плане по скринингу в текущем месяце
    count_regplan_for_month = screening_plan_for_month_all_count(year, month)

    screening_data['count_regplan_for_month'] = count_regplan_for_month[0].count

    sreening_plan_individuals = screening_plan_for_month_all_patient(year, month)
    sreening_people_cards = tuple([i.card_id for i in sreening_plan_individuals])

    # из них подлежащих при диспансеризации (кол-во)
    # получить карты и "research(уникальные)" "возраста на конец года" из screening_regplan_for_month -> проверить возраст
    # далее првоерить в DispensaryRouteSheet пары
    count_dispensarization_from_screening = must_dispensarization_from_screening_plan_for_month(year, month, f'{year}-12-31')
    screening_data['count_dispensarization_from_screening'] = count_dispensarization_from_screening[0].count

    # Число женщин 30-65 лет, прошедших скрининг
    pass_screening = sql_pass_screening(year, month, datetime_start, datetime_end, sreening_people_cards)
    screening_data['pass_screening'] = pass_screening[0].count

    # Число женщин 30-65 лет, прошедших скрининг из них при диспансеризации
    pass_screening_in_dispensarization = sql_pass_screening_in_dispensarization(year, month, datetime_start, datetime_end, f'{year}-12-31')
    screening_data['pass_screening_in_dispensarization'] = pass_screening_in_dispensarization[0].count

    # кто прошел тест папаниколау
    pass_pap_analysis = sql_pass_pap_analysis_count(datetime_start, datetime_end, sreening_people_cards, tuple(PAP_ANALYSIS_ID))
    screening_data['pass_pap_analysis'] = pass_pap_analysis[0].count

    # адекватных
    pass_pap_adequate_result_value = sql_pass_pap_fraction_result_value(
        datetime_start, datetime_end, sreening_people_cards, tuple(PAP_ANALYSIS_ID), tuple(PAP_ANALYSIS_FRACTION_QUALITY_ID), "^адекватный"
    )
    screening_data['pass_pap_adequate_result_value'] = pass_pap_adequate_result_value[0].count

    # недостаточно адекватный
    pass_pap_not_enough_adequate_result_value = sql_pass_pap_fraction_result_value(
        datetime_start, datetime_end, sreening_people_cards, tuple(PAP_ANALYSIS_ID), tuple(PAP_ANALYSIS_FRACTION_QUALITY_ID), "^недостаточно"
    )
    screening_data['pass_pap_not_enough_adequate_result_value'] = pass_pap_not_enough_adequate_result_value[0].count

    # неадекватный
    pass_pap_not_adequate_result_value = sql_pass_pap_fraction_result_value(
        datetime_start, datetime_end, sreening_people_cards, tuple(PAP_ANALYSIS_ID), tuple(PAP_ANALYSIS_FRACTION_QUALITY_ID), "неадекватный"
    )
    screening_data['pass_pap_not_adequate_result_value'] = pass_pap_not_adequate_result_value[0].count

    # карты с недостаточно адекватный и неадекватным результатом к-рым дважды взяли мазок
    pass_pap_not_not_enough_adequate_result_value = sql_card_dublicate_pass_pap_fraction_not_not_enough_adequate_result_value(
        datetime_start, datetime_end, sreening_people_cards, tuple(PAP_ANALYSIS_ID), tuple(PAP_ANALYSIS_FRACTION_QUALITY_ID), "неадекватный", "недостаточно адекватный", count_param=2
    )
    # people_cards_not_not_enough_adequate = tuple([i.client_id for i in pass_pap_not_not_enough_adequate_result_value])
    count_people_dublicate = len(pass_pap_not_not_enough_adequate_result_value)
    screening_data['count_people_dublicate'] = count_people_dublicate

    # АSCUS
    pass_pap_ascus_result_value = sql_pass_pap_fraction_result_value(
        datetime_start, datetime_end, sreening_people_cards, tuple(PAP_ANALYSIS_ID), tuple(PAP_ANALYSIS_FRACTION_CONTAIN_ID), "ASCUS"
    )
    screening_data['pass_pap_ascus_result_value'] = pass_pap_ascus_result_value[0].count

    # CIN-I
    pass_pap_cin_i_result_value = sql_pass_pap_fraction_result_value(
        datetime_start, datetime_end, sreening_people_cards, tuple(PAP_ANALYSIS_ID), tuple(PAP_ANALYSIS_FRACTION_CONTAIN_ID), "дисплазии CIN-I$", "", count_param=1
    )
    screening_data['pass_pap_cin_i_result_value'] = pass_pap_cin_i_result_value[0].count

    # CIN I-II, II
    pass_pap_cin_i_ii_result_value = sql_pass_pap_fraction_result_value(
        datetime_start, datetime_end, sreening_people_cards, tuple(PAP_ANALYSIS_ID), tuple(PAP_ANALYSIS_FRACTION_CONTAIN_ID), "CIN-I-II$", "дисплазии CIN-II$", count_param=2
    )
    screening_data['pass_pap_cin_i_ii_result_value'] = pass_pap_cin_i_ii_result_value[0].count

    # CIN-II-III, III
    pass_pap_cin_ii_iii_result_value = sql_pass_pap_fraction_result_value(
        datetime_start, datetime_end, sreening_people_cards, tuple(PAP_ANALYSIS_ID), tuple(PAP_ANALYSIS_FRACTION_CONTAIN_ID), "CIN-II-III$", "CIN-III$", count_param=2
    )
    screening_data['pass_pap_cin_ii_iii_result_value'] = pass_pap_cin_ii_iii_result_value[0].count

    # cr in situ
    pass_pap_cr_in_situ_result_value = sql_pass_pap_fraction_result_value(
        datetime_start, datetime_end, sreening_people_cards, tuple(PAP_ANALYSIS_ID), tuple(PAP_ANALYSIS_FRACTION_CONTAIN_ID), "cr in situ", "", count_param=1
    )
    screening_data['pass_pap_cr_in_situ_result_value'] = pass_pap_cr_in_situ_result_value[0].count

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = "attachment; filename=\"Screening.xlsx\""
    wb = openpyxl.Workbook()
    wb.remove(wb.get_sheet_by_name('Sheet'))
    ws = wb.create_sheet("Обращения")
    styles_obj = structure_sheet.style_sheet()
    wb.add_named_style(styles_obj[0])
    hospital_id = request.user.doctorprofile.hospital_id
    researches_sql = sql_func.statistics_research(PAP_ANALYSIS_ID[0], datetime_start, datetime_end, hospital_id)
    screening_data['count_pap_analysys'] = len(researches_sql)
    ws = structure_sheet.statistic_screening_month_data(ws, screening_data, month, year, styles_obj[3])
    wb.save(response)
    return response


def commercial_offer_xls_save_file(data_offer, patients, research_price):
    wb = openpyxl.Workbook()
    wb.remove(wb.get_sheet_by_name('Sheet'))
    ws = wb.create_sheet("Спецификация")
    ws = commercial_offer.offer_base(ws)
    ws = commercial_offer.offer_fill_data(ws, data_offer)

    ws = wb.create_sheet("Реестр")
    ws = commercial_offer.register_base(ws)
    ws = commercial_offer.register_data(ws, patients, research_price)

    dir_param = SettingManager.get("dir_param", default='/tmp', default_type='s')
    today = datetime.datetime.now()
    date_now1 = datetime.datetime.strftime(today, "%y%m%d%H%M%S%f")[:-3]
    date_now_str = "offer" + str(date_now1)
    file_dir = os.path.join(dir_param, date_now_str + '.xlsx')
    wb.save(filename=file_dir)
    return file_dir


def data_xls_save_file(data, sheet_name):
    wb, ws = initial_work_book(sheet_name)
    ws = base_data.fill_base(ws, data)
    ws = base_data.fill_data(ws, data)
    file_dir = save_file_disk(wb)
    return file_dir


def data_xls_save_headers_file(meta_patients, head_data, sheet_name, name_func):
    wb, ws = initial_work_book(sheet_name)
    ws = base_data.fill_default_base(ws, head_data)
    if name_func == "fill_xls_check_research_exam_data":
        ws = base_data.fill_xls_check_research_exam_data(ws, meta_patients)
    file_dir = save_file_disk(wb)
    return file_dir


@csrf_exempt
@login_required
def get_harmful_factors(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = "attachment; filename=\"Specification.xlsx\""
    wb = openpyxl.Workbook()
    wb.remove(wb.get_sheet_by_name('Sheet'))
    ws = wb.create_sheet("Спецификация")

    ws = harmful_factors.harmful_factors_base(ws)
    data_template = get_all_harmful_factors_templates()
    data_template_ids = tuple([i.template_id for i in data_template])
    date_researches = get_researches_by_templates(data_template_ids)
    data_template_meta = {
        i.template_id: {"harmfulfactor_title": i.harmfulfactor_title, "description": i.description, "template_title": i.template_title, "research_title": ""} for i in data_template
    }
    for k in date_researches:
        data_template_meta[k.template_id]['research_title'] = f"{data_template_meta[k.template_id]['research_title']};  {k.title}"
    ws = harmful_factors.harmful_factors_fill_data(ws, data_template_meta)
    wb.save(response)
    return response


@csrf_exempt
@login_required
def open_xls(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = "attachment; filename=\"Register.xlsx\""
    request_data = request.POST if request.method == "POST" else request.GET
    file_name = request_data.get("file").replace('"', "")
    dir_param = SettingManager.get("dir_param", default='/tmp', default_type='s')
    file_dir = os.path.join(dir_param, file_name)
    wb = load_workbook(file_dir)
    wb.save(response)
    os.remove(file_dir)
    return response


def get_price_company(company_id, start_date, end_date):
    return PriceName.get_company_price_by_date(company_id, start_date.date(), end_date.date())


def get_price_hospital(hospital_id, start_date, end_date):
    return PriceName.get_hospital_price_by_date(hospital_id, start_date, end_date, is_subcontract=True)
