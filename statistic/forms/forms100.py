from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter

from clients.models import HarmfulFactor
from clients.sql_func import researches_by_harmfull_factor_id, harmfull_factor_data
from laboratory.settings import CONTROL_AGE_MEDEXAM, CONTROL_AGE_MEDEXAM_MALE, CONTROL_AGE_MEDEXAM_FEMALE
from laboratory.utils import current_year
from statistic.forms.forms100_sql_func import (
    closed_company_cases_by_date,
    directions_by_parent_cases_issledovaniye,
    search_value_where_done_custom_research,
    search_value_type_medical_inspection,
)


def form_01(ws1, data):
    style_border = NamedStyle(name="style_border_ca")
    bd = Side(style="thin", color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=10)
    style_border.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

    style_border2 = NamedStyle(name="style_border2")
    bd = Side(style="thin", color="000000")
    style_border2.font = Font(bold=False, size=12)
    style_border2.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    custom_research = data["custom_research"]
    custom_researches_id = {i: 0 for i in custom_research.keys()}
    custom_researches_title = list(custom_research.values())

    # получить ЗАКРЫТЫЕ случаи за дату по компании
    last_date_year = f"{current_year()}-12-31"
    closed_id = closed_company_cases_by_date(data['start_date'], data['end_date'], data['company_id'], last_date_year)

    male = CONTROL_AGE_MEDEXAM.get("м")
    female = CONTROL_AGE_MEDEXAM.get("ж")
    adds_harmfull_title = set([CONTROL_AGE_MEDEXAM_MALE[i.age_year] if i.sex == "м" else CONTROL_AGE_MEDEXAM_FEMALE[i.age_year] for i in closed_id])
    adds_harmfull = {i.title: i.id for i in HarmfulFactor.objects.filter(title__in={*adds_harmfull_title})}
    factors_id = {*set([i.factor_id for i in closed_id]), *adds_harmfull.values()}
    researches_harmfull_factors = researches_by_harmfull_factor_id(tuple(factors_id))
    # структура уникальных услуг для всех пациентов по все факторам
    harmfull_factors_research_id_title = {i.research_id: {"title": i.research_title, "code": i.code, "internal_code": i.internal_code} for i in researches_harmfull_factors}
    research_id_is_doc_refferal = set([i.research_id for i in researches_harmfull_factors if i.is_doc_refferal])

    # структура факторы - услуги {factor_id: [research_id, research_id]}
    researches_harmfull_data = {}
    for i in researches_harmfull_factors:
        if not researches_harmfull_data.get(i.harmfull_factor_id):
            researches_harmfull_data[i.harmfull_factor_id] = [i.research_id]
        else:
            researches_harmfull_data[i.harmfull_factor_id].append(i.research_id)

    # стр-ра по закрытм случаям
    # {"id_case_iss": {"fio": "", "sex": "", "birthday": "", "factors": "",
    #                  "custom_researches": {"id_research": "Да|Нет", "id_research": "Да|Нет"},
    #                  "result_researches": {
    #                                         "id_research": {price: "", "factor": ""}, "id_research": {price: "", "factor": ""}
    #                                        }
    #                  }
    #  }
    closed_case_structure_data = {}
    factors_ids = set()
    cases_issledovaniye_ids = {}
    custom_research_ids = data["custom_research"].keys()
    harmfull_factors = harmfull_factor_data()
    harmfull_factors_id_title = {i.id: i.title for i in harmfull_factors}
    for i in closed_id:
        if not closed_case_structure_data.get(i.case_issledovaniye_id):
            closed_case_structure_data[i.case_issledovaniye_id] = {
                "fio": f"{i.patient_family} {i.patient_name} {i.patient_patronymic}",
                "sex": i.sex,
                "birthday": i.patient_birthday,
                "age_year": i.age_year,
                "date_end": i.date_end,
                "factors": [i.factor_id],
                "factors_title": [harmfull_factors_id_title.get(i.factor_id)],
                "custom_researches": custom_researches_id.copy(),
                "type_inspection": "-",
                "result_researches": {
                    research_id: {
                        "price": 0,
                        "date_confirm": "",
                        "iss_id": "",
                        "where_done": 0,
                        "research_title": harmfull_factors_research_id_title.get(research_id)["title"],
                        "code": harmfull_factors_research_id_title.get(research_id)["code"],
                        "internal_code": harmfull_factors_research_id_title.get(research_id)["internal_code"],
                    }
                    for research_id in researches_harmfull_data.get(i.factor_id)
                },
            }
        else:
            closed_case_structure_data[i.case_issledovaniye_id]["factors"].append(i.factor_id)
            closed_case_structure_data[i.case_issledovaniye_id]["factors_title"].append(harmfull_factors_id_title.get(i.factor_id))

            closed_case_structure_data[i.case_issledovaniye_id]["result_researches"].update(
                {
                    research_id: {
                        "price": 0,
                        "date_confirm": "",
                        "iss_id": "",
                        "where_done": 0,
                        "research_title": harmfull_factors_research_id_title.get(research_id)["title"],
                        "code": harmfull_factors_research_id_title.get(research_id)["code"],
                        "internal_code": harmfull_factors_research_id_title.get(research_id)["internal_code"],
                    }
                    for research_id in researches_harmfull_data.get(i.factor_id)
                }
            )

        factors_ids.add(i.factor_id)
        cases_issledovaniye_ids[i.case_issledovaniye_id] = i.direction_num

    for v in closed_case_structure_data.values():
        if v["sex"] == "м":
            for k in sorted(male.keys()):
                if v["age_year"] < k and researches_harmfull_data.get(adds_harmfull.get(male[k])):
                    v["factors_title"].append(male[k])
                    v["result_researches"].update(
                        {
                            research_id: {
                                "price": 0,
                                "date_confirm": "",
                                "iss_id": "",
                                "where_done": 0,
                                "research_title": harmfull_factors_research_id_title.get(research_id)["title"],
                                "code": harmfull_factors_research_id_title.get(research_id)["code"],
                                "internal_code": harmfull_factors_research_id_title.get(research_id)["internal_code"],
                            }
                            for research_id in researches_harmfull_data.get(adds_harmfull.get(male[k]))
                        }
                    )
                    break
        if v["sex"] == "ж":
            for k in sorted(female.keys()):
                if v["age_year"] < k and researches_harmfull_data.get(adds_harmfull.get(female[k])):
                    v["factors_title"].append(female[k])
                    v["result_researches"].update(
                        {
                            research_id: {
                                "price": 0,
                                "date_confirm": "",
                                "iss_id": "",
                                "where_done": 0,
                                "research_title": harmfull_factors_research_id_title.get(research_id)["title"],
                                "code": harmfull_factors_research_id_title.get(research_id)["code"],
                                "internal_code": harmfull_factors_research_id_title.get(research_id)["internal_code"],
                            }
                            for research_id in researches_harmfull_data.get(adds_harmfull.get(female[k]))
                        }
                    )
                    break

    cases_iss = tuple(cases_issledovaniye_ids.keys())
    # получить все исследования, у к-рых в направлении родитель ссылка на случай
    result_iss_id = directions_by_parent_cases_issledovaniye(cases_iss)

    result_iss_id_structure = {i.iss_id: {"parent": i.parent_case_iss_id, "research_id": i.research_id, "date_confirm": i.date_confirm} for i in result_iss_id}
    result_iss_id_structure_by_parent = {}
    for k, v in result_iss_id_structure.items():
        if not result_iss_id_structure_by_parent.get(v["parent"]):
            result_iss_id_structure_by_parent[v["parent"]] = [{"iss_id": k, "research_id": v["research_id"], "date_confirm": v["date_confirm"]}]
        else:
            result_iss_id_structure_by_parent[v["parent"]].append({"iss_id": k, "research_id": v["research_id"], "date_confirm": v["date_confirm"]})

    # выполненные исследование все для всех пациентов
    research_issledovaniye_ids = [i.iss_id for i in result_iss_id]

    # поиск результатов для кастомных услуг среди результатов, в каком учреждении оказана услуга
    result_where_done_custom_research_sql = search_value_where_done_custom_research(tuple(research_issledovaniye_ids), tuple(custom_research.keys()))

    result_type_medical_inspection_sql = search_value_type_medical_inspection(tuple(cases_iss))
    result_type_medical_inspection = {i.issledovaniye_id: i.result_value for i in result_type_medical_inspection_sql}

    # взять значения, в каком учреждении пройдено обследование по исследованию
    result_where_done_custom_research = {i.issledovaniye_id: {"research_id": i.research_id, "value": i.result_value} for i in result_where_done_custom_research_sql}
    total_sum_by_specialist = {}
    total_sum_by_instrumental_and_lab = {}

    for k, v in closed_case_structure_data.items():
        closed_case_structure_data[k]["type_inspection"] = result_type_medical_inspection.get(k)
        for i in result_iss_id_structure_by_parent[k]:
            result_where_done = None
            current_research_id = i.get("research_id")
            if closed_case_structure_data[k]["result_researches"].get(current_research_id):
                closed_case_structure_data[k]["result_researches"][current_research_id]["date_confirm"] = i.get("date_confirm")
                closed_case_structure_data[k]["result_researches"][current_research_id]["iss_id"] = i.get("iss_id")
                if i.get("date_confirm"):
                    result_where_done = 1
                    closed_case_structure_data[k]["result_researches"][current_research_id]["where_done"] = result_where_done
                else:
                    result_where_done = 0
                    closed_case_structure_data[k]["result_researches"][current_research_id]["where_done"] = result_where_done
            if data["research_coast"].get(current_research_id):
                try:
                    closed_case_structure_data[k]["result_researches"][current_research_id]["price"] = data["research_coast"].get(current_research_id)
                except:
                    pass

            if result_where_done_custom_research.get(i.get("iss_id")):
                if result_where_done_custom_research.get(i.get("iss_id"))["research_id"] in custom_research_ids:
                    target_research_id = result_where_done_custom_research[i.get("iss_id")]["research_id"]
                    where_done = result_where_done_custom_research[i.get("iss_id")]["value"]
                    if where_done.lower() == "да":
                        result_where_done = 0
                    else:
                        result_where_done = 1
                    closed_case_structure_data[k]["custom_researches"][target_research_id] = result_where_done
                    if closed_case_structure_data[k]["result_researches"].get(target_research_id):
                        closed_case_structure_data[k]["result_researches"][target_research_id]["where_done"] = result_where_done

            if result_where_done == 1:
                if current_research_id in research_id_is_doc_refferal:
                    if not total_sum_by_specialist.get(current_research_id):
                        total_sum_by_specialist[current_research_id] = 1
                    else:
                        total_sum_by_specialist[current_research_id] += 1
                else:
                    if not total_sum_by_instrumental_and_lab.get(current_research_id):
                        total_sum_by_instrumental_and_lab[current_research_id] = 1
                    else:
                        total_sum_by_instrumental_and_lab[current_research_id] += 1

    ws1.merge_cells("A8:Q8")
    megre_cell = ws1["A8"]
    megre_cell.value = f"Спецификация {data['executor']}"
    megre_cell.style = style_border2

    ws1.merge_cells("A9:Q9")
    megre_cell = ws1["A9"]
    megre_cell.value = f"Заказчик: {data['customer_title']}"
    megre_cell.style = style_border2

    ws1.merge_cells("A10:Q10")
    megre_cell = ws1["A10"]
    megre_cell.value = f'Реестр оказанных медицинских услуг за период с {data["start_date"]} по {data["end_date"]}'
    megre_cell.style = style_border2

    columns = [
        ("№ п/п", 5),
        ("№ Договора", 20),
        ("ФИО", 40),
        ("Дата рождения", 10),
        ("Возраст", 10),
        ("Дата окончания", 10),
        ("Пол", 8),
        ("Вид медосмотра (периодическмй / предварительный и т.д.)", 11),
    ]

    columns2 = [(i, 7) for i in custom_researches_title]
    columns.extend(columns2)

    start_column_for_custom_filed = 7
    list_researches = list(custom_researches_id.keys())
    # номер столбца для custom researches
    custom_researches_id_number_columns = {i: start_column_for_custom_filed + list_researches.index(i) for i in list_researches}
    end_column_for_custom_filed = start_column_for_custom_filed + len(list_researches)

    columns3 = [
        ("Дата оказания услуги", 10),
        ("Диагноз (код по МКБ)", 10),
        ("Код услуги", 12),
        ("Муж", 6),
        ("Жен", 6),
        ("Номер позиции вредности веществ по Приказу №29н", 11),
        ("Специалисты, обследования", 15),
        ("Кол-во, чел", 6),
        ("Тариф, руб.", 12),
        ("Сумма, руб.", 12),
    ]
    columns.extend(columns3)

    start_row = 13
    row = start_row
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border

    row += 1

    step = 1
    sum_research_col = 1
    total_sum_rows_value = "="
    total_sum_sex_male = "="
    total_sum_sex_female = "="
    col_sex_male = ""
    col_sex_female = ""
    for i in closed_case_structure_data.values():
        start_row_for_sum = row
        ws1.cell(row=row, column=1).value = step
        ws1.cell(row=row, column=2).value = data.get("contract_number")
        ws1.cell(row=row, column=3).value = i.get("fio")
        ws1.cell(row=row, column=4).value = i.get("birthday")
        sex = i.get("sex")
        ws1.cell(row=row, column=5).value = "-"
        ws1.cell(row=row, column=6).value = i.get("date_end")
        ws1.cell(row=row, column=7).value = sex
        ws1.cell(row=row, column=8).value = i.get("type_inspection")

        for k, v in i["custom_researches"].items():
            col_custom = custom_researches_id_number_columns.get(k)
            ws1.cell(row=row, column=col_custom).value = v

        col_sex_male = end_column_for_custom_filed + 3
        ws1.cell(row=row, column=col_sex_male).value = 1 if sex == "м" else 0
        total_sum_sex_male = f"{total_sum_sex_male}{get_column_letter(col_sex_male)}{row}+"

        col_sex_female = end_column_for_custom_filed + 4
        ws1.cell(row=row, column=col_sex_female).value = 1 if sex == "ж" else 0
        total_sum_sex_female = f"{total_sum_sex_female}{get_column_letter(col_sex_female)}{row}+"

        col_factors = end_column_for_custom_filed + 5
        factors_title = i.get("factors_title")
        current_row = row
        for k in factors_title:
            ws1.cell(row=current_row, column=col_factors).value = k
            current_row += 1

        for k, v in i["result_researches"].items():
            col_date = end_column_for_custom_filed
            ws1.cell(row=row, column=col_date).value = v.get("date_confirm")
            col_dig = end_column_for_custom_filed + 1
            ws1.cell(row=row, column=col_dig).value = "Z10.01"

            col_research_code = end_column_for_custom_filed + 2
            ws1.cell(row=row, column=col_research_code).value = v.get("code")

            col_title = end_column_for_custom_filed + 6
            ws1.cell(row=row, column=col_title).value = v.get("research_title")
            where_done_col = col_title + 1
            ws1.cell(row=row, column=where_done_col).value = v.get("where_done")
            price_col = col_title + 2
            ws1.cell(row=row, column=price_col).value = v.get("price")
            sum_research_col = col_title + 3
            ws1.cell(row=row, column=sum_research_col).value = f'={get_column_letter(where_done_col)}{row}*{get_column_letter(price_col)}{row}'
            row += 1
        ws1.cell(row=row, column=1).value = "Итого"
        ws1.cell(row=row, column=sum_research_col).value = f'=SUM({get_column_letter(sum_research_col)}{start_row_for_sum}:{get_column_letter(sum_research_col)}{row - 1})'
        total_sum_rows_value = f"{total_sum_rows_value}{get_column_letter(sum_research_col)}{row}+"

        row += 1
        step += 1
    ws1.cell(row=row, column=1).value = "Итого по всем"
    total_sum_rows_value = total_sum_rows_value.rstrip("+")
    ws1.cell(row=row, column=sum_research_col).value = total_sum_rows_value

    total_sum_sex_male = total_sum_sex_male.rstrip("+")
    total_sum_sex_female = total_sum_sex_female.rstrip("+")

    ws1.cell(row=row, column=col_sex_male).value = total_sum_sex_male
    ws1.cell(row=row, column=col_sex_female).value = total_sum_sex_female

    columns = [
        ("Специалисты", 30),
        ("Кол-во человек", 12),
        ("Цена", 15),
        ("Стоимость", 20),
    ]
    row = start_row
    col = sum_research_col + 2
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx + col).value = column[0]
        ws1.column_dimensions[get_column_letter(idx + col)].width = column[1]
        ws1.cell(row=row, column=idx + col).style = style_border

    for k, v in total_sum_by_specialist.items():
        row += 1
        ws1.cell(row=row, column=1 + col).value = harmfull_factors_research_id_title.get(k)["title"]
        ws1.cell(row=row, column=2 + col).value = v
        ws1.cell(row=row, column=3 + col).value = data["research_coast"].get(k)
        ws1.cell(row=row, column=4 + col).value = f'={get_column_letter(2 + col)}{row}*{get_column_letter(3 + col)}{row}'
    row += 1
    ws1.cell(row=row, column=3 + col).value = "Итого"
    ws1.cell(row=row, column=4 + col).value = f'=SUM({get_column_letter(4 + col)}{start_row + 1}:{get_column_letter(4 + col)}{row - 1})'

    row += 1
    second_total_sum_start_row = row
    for k, v in total_sum_by_instrumental_and_lab.items():
        row += 1
        ws1.cell(row=row, column=1 + col).value = harmfull_factors_research_id_title.get(k)["title"]
        ws1.cell(row=row, column=2 + col).value = v
        ws1.cell(row=row, column=3 + col).value = data["research_coast"].get(k)
        ws1.cell(row=row, column=4 + col).value = f'={get_column_letter(2 + col)}{row}*{get_column_letter(3 + col)}{row}'
        for i in range(col + 1, col + 5):
            ws1.cell(row=row, column=i).style = style_border

    row += 1
    ws1.cell(row=row, column=3 + col).value = "Итого"
    ws1.cell(row=row, column=4 + col).value = f'=SUM({get_column_letter(4 + col)}{second_total_sum_start_row + 1}:{get_column_letter(4 + col)}{row - 1})'

    return ws1
