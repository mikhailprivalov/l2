from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter


def handle_query(query, query_diagnoses_pk):
    card_diagnoses = {}
    for i in query_diagnoses_pk:
        temp_diagnos = card_diagnoses.get(i.card_id, "")
        temp_diagnos = f"{temp_diagnos}, {i.diagnos}"
        card_diagnoses[i.card_id] = temp_diagnos

    return [{"fio": f"{i.family} {i.name} {i.patronymic}", "born": i.born, "card": i.number, "diagnoses": card_diagnoses[i.card_id], "month": i.month} for i in query]


def dispansery_plan_base(ws1, d1, d2):
    style_border = NamedStyle(name="style_border_ca")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=14)
    style_border.alignment = Alignment(wrap_text=True, horizontal='justify', vertical='center')

    ws1.cell(row=1, column=1).value = 'Д-учет план:'
    ws1.cell(row=2, column=1).value = f'{d1} {d2}'

    columns = [
        ('№ карты', 13),
        ('ФИО', 55),
        ('Дата рождения', 15),
        ('Диагноз', 65),
        ('Месяц план', 20),
    ]
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=4, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=4, column=idx).style = style_border

    return ws1


def dispansery_plan_fill_data(ws1, result_query, row=4):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

    r = row
    for i in result_query:
        r += 1
        ws1.cell(row=r, column=1).value = i["card"]
        ws1.cell(row=r, column=2).value = i["fio"]
        ws1.cell(row=r, column=3).value = i["born"]
        ws1.cell(row=r, column=4).value = i["diagnoses"]
        ws1.cell(row=r, column=5).value = i["month"]

        for z in range(1, 6):
            ws1.cell(row=r, column=z).style = style_border1

    return ws1


def dispansery_reg_count_base(ws1, d1):
    style_border = NamedStyle(name="style_border_ca")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=12)
    style_border.alignment = Alignment(wrap_text=True, horizontal='justify', vertical='center')

    ws1.cell(row=1, column=1).value = 'Пациентов на Д-учете'
    ws1.cell(row=2, column=1).value = f'На дату {d1}'

    columns = [
        ('Взрослые 18 и старше', 25),
        ('Дети до 0-17 лет', 25),
    ]
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=4, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=4, column=idx).style = style_border

    return ws1


def dispansery_reg_count_fill_data(ws1, result_query, row=4):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

    r = row
    for i in result_query:
        r += 1
        ws1.cell(row=r, column=1).value = str(i["adult"])
        ws1.cell(row=r, column=2).value = str(i["child"])

        for z in range(1, 3):
            ws1.cell(row=r, column=z).style = style_border1

    return ws1
