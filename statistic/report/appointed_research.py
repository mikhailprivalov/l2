from copy import deepcopy

import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter


def appointed_base(ws1, d1, d2, researches):
    style_border = NamedStyle(name="style_border_ca5")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'

    researches[-999]= {"title": "Врач"}
    researches = dict(sorted(researches.items()))
    columns = [(v.get("title"), 35) for v in researches.values()]

    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border

    return ws1


def parse_data(researches_sql, users_final_data, research_data):
    for data in researches_sql:
        current_id_doctor = None
        current_fio = ""
        if users_final_data.get(data.doctor_id):
            current_id_doctor = data.doctor_id
            current_fio = f"{data.doctor_family} {data.doctor_name} {data.doctor_patronymic}"
        elif users_final_data.get(data.d_from_doc_id):
            current_id_doctor = data.d_from_doc_id
            current_fio = f"{data.d_from_doc_family} {data.d_from_doc_name} {data.d_from_doc_patronymic}"
        tmp_doctor = users_final_data.get(current_id_doctor)
        tmp_doctor[-999]["count"] = current_fio
        count_research = tmp_doctor.get(data.research_id)["count"]
        count_research += 1
        tmp_doctor.get(data.research_id)["count"] = count_research
        users_final_data[current_id_doctor] = dict(sorted(tmp_doctor.items()))
    return users_final_data


def fill_appointed_research_by_doctors(ws1, data, row=6):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    r = row
    for value in data.values():
        column = 1
        for key, research_data in value.items():
            ws1.cell(row=r, column=column).value = research_data.get("count", "-")
            ws1.cell(row=r, column=column).style = style_border1
            column += 1
        r += 1
    return  ws1
