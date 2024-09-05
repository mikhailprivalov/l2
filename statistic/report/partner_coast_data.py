import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter


def partner_coast_base(ws1):
    style_border = NamedStyle(name="style_border_ca5")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    columns = [
        ('Контрагент', 60),
        ('Пациент', 40),
        ('Карта', 20),
        ('Направление', 20),
        ('Дата', 15),
        ('Номер контейнера', 50),
        ('Назначения', 30),
        ('Стоимость', 20),
    ]

    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border

    return ws1


def partner_coast_fill_data(ws1, result_query):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    pink_fill = openpyxl.styles.fills.PatternFill(patternType='solid', start_color='FCD5B4', end_color='FCD5B4')
    green_fill = openpyxl.styles.fills.PatternFill(patternType='solid', start_color='b6d7a8', end_color='b6d7a8')

    r = 5
    fill_rows = []
    total_hosp_fill_rows = []
    for v in result_query:
        r += 1
        ws1.cell(row=r, column=1).value = v.get("hospital_title", "-")
        ws1.cell(row=r, column=2).value = f'{v.get("patient_family", "-")} {v.get("patient_name", "-")} {v.get("patient_patronymic", "-")} {v.get("patient_birthday")}'
        ws1.cell(row=r, column=3).value = v.get("patient_card", "-")
        ws1.cell(row=r, column=4).value = v.get("direction_num", "-")
        ws1.cell(row=r, column=5).value = v.get("target_date", "-")
        ws1.cell(row=r, column=6).value = v.get("tube_number", "-")
        ws1.cell(row=r, column=7).value = v.get("research_title", "-")
        ws1.cell(row=r, column=8).value = v.get("research_coast", "-")

        for c in range(1, 9):
            ws1.cell(row=r, column=c).style = style_border1
        if v.get("patient_card") == "Итого по пациенту":
            fill_rows.append(r)
        if v.get("patient_card") == "Итого по контрагенту":
            total_hosp_fill_rows.append(r)

    for k in fill_rows:
        rows = ws1[f'A{k}:H{k}']
        for row in rows:
            for cell in row:
                cell.fill = pink_fill

    for k in total_hosp_fill_rows:
        rows = ws1[f'A{k}:H{k}']
        for row in rows:
            for cell in row:
                cell.fill = green_fill

    return ws1
