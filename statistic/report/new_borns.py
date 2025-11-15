from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter

from utils.dates import normalize_date
from utils.db import namedtuplefetchall
from laboratory.settings import TIME_ZONE
from django.db import connection
import simplejson as json


def form_01(ws1, data):
    # Для журнала новорожденных
    style_border = NamedStyle(name="style_border_ca")
    bd = Side(style="thin", color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=12)
    style_border.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    style_border2 = NamedStyle(name="style_border2")
    style_border2.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border2.font = Font(bold=False, size=11)
    style_border2.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    columns = [
        ('№ п/п.', 5),
        ('ФИО пациента', 30),
        ('Дата рождения', 30),
        ('Пол', 10),
        ('Родильный дом', 10),
        ('Дата передачи', 30),
        ('Дата выписки', 30),
        ('Участок', 15),
        ('Адрес', 25),
        ('Телефон', 25),
        ('Вес при рождении (г.)', 30),
        ('Вес при выписке (г.)', 30),
        ('БЦЖ-статус', 20),
        ('БЦЖ-дата', 20),
        ('Гепатит-статус', 20),
        ('Гепатит-дата', 20),
        ('ФКУ-статус', 20),
        ('ФКУ-дата', 20),
        ('Дата патронажа (перв)', 20),
    ]
    sql_data_main_research = sql_01(data['research_id'], data['d_s'], data['d_e'])
    card_pks = set([i.client_id for i in sql_data_main_research])
    main_result_by_card_id = {}
    for i in sql_data_main_research:
        if not main_result_by_card_id.get(i.client_id):
            main_result_by_card_id[i.client_id] = {"fio": f"{i.family} {i.name} {i.patronymic}", "sex": i.sex, i.field_title: i.field_value}
        else:
            main_result_by_card_id[i.client_id].update({i.field_title: i.field_value})

    sql_data_child_research = sql_02(data['child_research_id'], card_pks)
    child_result_by_card_id = {}
    max_count_date_for_client_id = {}
    for i in sql_data_child_research:
        if not child_result_by_card_id.get(i.client_id):
            child_result_by_card_id[i.client_id] = {i.char_medical_examination: {i.field_title: i.field_value}}
            max_count_date_for_client_id[i.client_id] = 1
        elif not child_result_by_card_id[i.client_id].get(i.char_medical_examination):
            child_result_by_card_id[i.client_id][i.char_medical_examination] = {i.field_title: i.field_value}
            max_count_date_for_client_id[i.client_id] += 1
        else:
            child_result_by_card_id[i.client_id][i.char_medical_examination].update({i.field_title: i.field_value})
    child_count_patronage = [('Дата патронажа', 20) for i in  range(max(max_count_date_for_client_id.values()))]

    columns.extend(child_count_patronage)
    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border



    previous_date = None
    step = 0
    for card_id, values in main_result_by_card_id.items():
        row += 1
        step += 1
        ws1.cell(row=row, column=1).value = step
        ws1.cell(row=row, column=2).value = values.get("fio")
        ws1.cell(row=row, column=3).value = values.get("Дата рождения")
        ws1.cell(row=row, column=4).value = values.get("пол")
        ws1.cell(row=row, column=5).value = values.get("Родильный дом")
        ws1.cell(row=row, column=6).value = values.get("Дата передачи")
        ws1.cell(row=row, column=7).value = values.get("Дата выписки")
        ws1.cell(row=row, column=8).value = values.get("Участок")
        ws1.cell(row=row, column=9).value = values.get("Адрес")
        ws1.cell(row=row, column=10).value = values.get("Телефон")
        ws1.cell(row=row, column=11).value = values.get("Вес при рождении (г.)")
        ws1.cell(row=row, column=12).value = values.get("Вес при выписке (г.)")

        ws1.cell(row=row, column=13).value = values.get("БЦЖ-статус")
        ws1.cell(row=row, column=14).value = values.get("БЦЖ-дата")
        ws1.cell(row=row, column=15).value = values.get("Гепатит-статус")
        ws1.cell(row=row, column=16).value = values.get("Гепатит-дата")
        ws1.cell(row=row, column=17).value = values.get("ФКУ-статус")
        ws1.cell(row=row, column=18).value = values.get("ФКУ-дата")
        ws1.cell(row=row, column=19).value = values.get("Дата патронажа (перв)")
        column = 19
        for date_child_patronage, value_child_patronage in child_result_by_card_id.get(card_id):
            column += 1
            ws1.cell(row=row, column=column).value = values.get("Дата патронажа")

        for c in range(11):
            ws1.cell(row=row, column=c + 1).style = style_border2

        previous_date = i.get("medical_examination")

    return ws1


def sql_01(research_id, d_s, d_e):
    """
    Для журнала новородок первичные
    :return:
    """
    with connection.cursor() as cursor:
        cursor.execute(
            """
                SELECT
                    directions_paraclinicresult.issledovaniye_id,
                    directions_paraclinicresult.field_id,
                    directions_paraclinicresult.value as field_value,
                    directory_paraclinicinputfield.title as field_title,
                    directions_issledovaniya.medical_examination,
                    directions_napravleniya.client_id,
                    ci.family,
                    ci.name,
                    ci.patronymic,
                    ci.sex
                    FROM directions_paraclinicresult
                    LEFT JOIN directions_issledovaniya ON directions_issledovaniya.id = directions_paraclinicresult.issledovaniye_id
                    LEFT JOIN directory_paraclinicinputfield ON directory_paraclinicinputfield.id = directions_paraclinicresult.field_id
                    LEFT JOIN directions_napravleniya ON directions_napravleniya.id = directions_issledovaniya.napravleniye_id
                    LEFT JOIN clients_card cc ON cc.id = directions_napravleniya.client_id
                    LEFT JOIN clients_individual ci ON ci.id = cc.individual_id
                    WHERE 
                      directions_issledovaniya.research_id=%(research_id)s
                      AND directory_paraclinicinputfield.for_talon = true
                      AND directions_issledovaniya.time_confirmation IS NOT NULL
                      AND directions_issledovaniya.medical_examination AT TIME ZONE %(tz)s BETWEEN %(d_start)s AND %(d_end)s      
                    order by directions_issledovaniya.medical_examination, directions_issledovaniya.napravleniye_id
                """,
            params={'research_id': research_id, 'd_start': d_s, 'd_end': d_e, 'tz': TIME_ZONE},
        )

        rows = namedtuplefetchall(cursor)
    return rows


def sql_02(research_id, card_pks):
    """
    Для журнала новородок повторные приемы
    :return:
    """
    with connection.cursor() as cursor:
        cursor.execute(
            """
                SELECT
                    directions_napravleniya.client_id,
                    directions_paraclinicresult.issledovaniye_id,
                    directions_issledovaniya.medical_examination,
                    to_char(directions_issledovaniya.medical_examination, 'DD.MM.YYYY') as char_medical_examination,
                    directions_paraclinicresult.value as field_value,
                    directory_paraclinicinputfield.title as field_title,
                    directions_issledovaniya.napravleniye_id as direction_number
                    FROM directions_paraclinicresult
                    LEFT JOIN directions_issledovaniya ON directions_issledovaniya.id = directions_paraclinicresult.issledovaniye_id
                    LEFT JOIN directory_paraclinicinputfield ON directory_paraclinicinputfield.id = directions_paraclinicresult.field_id
                    LEFT JOIN directions_napravleniya ON directions_napravleniya.id = directions_issledovaniya.napravleniye_id
                    LEFT JOIN directions_issledovaniya di ON di.id = directions_napravleniya.parent_id
                    WHERE 
                      directions_issledovaniya.research_id=%(research_id)s
                      AND directions_napravleniya.client_id in %(card_pks)s
                      AND directory_paraclinicinputfield.for_talon = true
                      AND directions_issledovaniya.time_confirmation IS NOT NULL      
                    order by directions_napravleniya.client_id, directions_issledovaniya.medical_examination
                """,
            params={'research_id': research_id, 'card_pks': card_pks},
        )

        rows = namedtuplefetchall(cursor)
    return rows
