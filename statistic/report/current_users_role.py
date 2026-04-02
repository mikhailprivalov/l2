from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter
from utils.db import namedtuplefetchall
from laboratory.settings import USE_COMBO_ROLE
from django.db import connection


def form_01(ws1, data):
    # Для журнала ВК - ДЛО
    style_border = NamedStyle(name="style_border_ca")
    bd = Side(style="thin", color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=12)
    style_border.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    style_border2 = NamedStyle(name="style_border2")
    style_border2.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border2.font = Font(bold=False, size=11)
    style_border2.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    columns = [
        ('№ п.п.', 10),
        ('Организация', 30),
        ('Статус МО', 30),
        ('Пользователь', 30),
        ('Статус пользователя', 30),
        ('Роль', 30),
        ('Детали-роли', 50),
        ('Закрепленные клиники', 50),
    ]
    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border

    sql_data = sql_01()
    result = [
        {
            "hospital_title": i.hospital_title,
            "hospital_status": "Скрыт" if i.hospital_hide else "Доступен",
            "doctor": f"{i.doc_family} {i.doc_name} {i.doc_patronymic}",
            "doctor_active": "Не активен" if i.dismissed or i.hospital_hide else "Активен",
            "doctor_detail_role": i.group_name,
            "assigned_hospital": i.hospitals_title,
        }
        for i in sql_data
    ]
    step = 0
    combo_role = list(USE_COMBO_ROLE.keys())
    for i in result:
        row += 1
        step += 1
        ws1.cell(row=row, column=1).value = step
        ws1.cell(row=row, column=2).value = i.get("hospital_title")
        ws1.cell(row=row, column=3).value = i.get("hospital_status")
        ws1.cell(row=row, column=4).value = i.get("doctor")
        ws1.cell(row=row, column=5).value = i.get("doctor_active")
        combo_user = [r for r in i.get('doctor_detail_role') if r in combo_role]
        detail_combo_role = []
        for cu in combo_user:
            detail_combo_role.extend(USE_COMBO_ROLE.get(cu))
        ws1.cell(row=row, column=6).value = ", ".join(combo_user)
        detail_role_str = ", ".join(i.get("doctor_detail_role"))
        if len(detail_combo_role) > 0:
            detail_combo_str = ", ".join(detail_combo_role)
        else:
            detail_combo_str = ""
        ws1.cell(row=row, column=7).value = f"{detail_role_str}, {detail_combo_str}"
        assigned_hospital = ", ".join(i.get("assigned_hospital"))
        ws1.cell(row=row, column=8).value = assigned_hospital
        for k in range(8):
            ws1.cell(row=row, column=k + 1).style = style_border2
    return ws1


def sql_01():
    """
    Для журнала ВК-ДЛО
    :return:
    """
    with connection.cursor() as cursor:
        cursor.execute(
            """
                SELECT
                    hh.title as hospital_title,
                    hh.hide as hospital_hide,
                    users_doctorprofile.family as doc_family,
                    users_doctorprofile.name as doc_name,
                    users_doctorprofile.patronymic as doc_patronymic,
                    users_doctorprofile.dismissed,
                    COALESCE(
                    (
                        SELECT JSON_AGG(auth_group.name)
                        FROM auth_group
                        LEFT JOIN auth_user_groups aug ON auth_group.id=aug.group_id
                        WHERE aug.user_id = users_doctorprofile.user_id
                    ), '[]'::json
                    ) AS group_name,
                    COALESCE(
                    (
                        SELECT JSON_AGG(hh.title)
                        FROM users_permissionhospitalprotocoldoctorprofile
                        LEFT JOIN hospitals_hospitals hh ON hh.id=users_permissionhospitalprotocoldoctorprofile.hospital_id
                        WHERE users_permissionhospitalprotocoldoctorprofile.doctor_profile_id = users_doctorprofile.id
                    ), '[]'::json
                    ) AS hospitals_title
                    FROM users_doctorprofile
                    LEFT JOIN hospitals_hospitals hh ON users_doctorprofile.hospital_id = hh.id    
                    order by hh.id, users_doctorprofile.family
                """,
            params={},
        )
        rows = namedtuplefetchall(cursor)
    return rows
