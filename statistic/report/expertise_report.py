import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter


def expertise_data(ws1, final_result):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=15)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    second_fill = openpyxl.styles.fills.PatternFill(patternType='solid', start_color='c6c9cc', end_color='c6c9cc')
    third_fill = openpyxl.styles.fills.PatternFill(patternType='solid', start_color='b1cbe6', end_color='b1cbe6')
    current_column = 0
    for k, v in final_result.items():
        step = 2
        column = current_column
        ws1.cell(row=step - 1, column=column + 1).value = k
        ws1.cell(row=step - 1, column=column + 1).style = style_border1
        ws1.cell(row=step, column=column + 1).value = "№ истории"
        ws1.column_dimensions[get_column_letter(column + 1)].width = 23
        ws1.cell(row=step, column=column + 1).style = style_border1

        ws1.cell(row=step, column=column + 2).value = "второй"
        ws1.column_dimensions[get_column_letter(column + 2)].width = 15
        ws1.cell(row=step, column=column + 2).style = style_border1
        ws1.cell(row=step, column=column + 2).fill = second_fill

        ws1.cell(row=step, column=column + 3).value = "третий"
        ws1.column_dimensions[get_column_letter(column + 3)].width = 15
        ws1.cell(row=step, column=column + 3).style = style_border1
        ws1.cell(row=step, column=column + 3).fill = third_fill
        for i in v:
            step += 1
            column += 1
            ws1.cell(row=step, column=column).value = i.get("direction_main_extract_dir", "-")
            ws1.cell(row=step, column=column).style = style_border1
            column += 1
            second_result = float(i.get("второй", "0"))
            if second_result == 0:
                second_result = float(i.get("без уровня", "0"))
            ws1.cell(row=step, column=column).value = second_result
            ws1.cell(row=step, column=column).style = style_border1
            ws1.cell(row=step, column=column).fill = second_fill
            column += 1
            ws1.cell(row=step, column=column).value = float(i.get("третий", "0"))
            ws1.cell(row=step, column=column).style = style_border1
            ws1.cell(row=step, column=column).fill = third_fill
            column = current_column
        current_column = column + 3
    return ws1
