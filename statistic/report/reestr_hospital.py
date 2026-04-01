from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter


def reestr_hospital_base(ws1, d1, d2, title):
    style_border = NamedStyle(name="style_border_ca5")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=1, column=1).value = title
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'

    columns = [
        ('Клиника', 25),
        ('ID заявки/номер заявки', 10),
        ('Дата создания (МСК)', 15),
        ('Время создания (МСК)', 15),
        ('Дата подтверждения (МСК)', 15),
        ('Время подтверждения (МСК)', 15),
        ('№ карты', 15),
        ('ФИО пациента', 30),
        ('Код ОКМУ', 20),
        ('Область исследования/ Услуга', 25),
        ('Модальность', 10),
        ('ФИО Врача', 30),
        ('Тариф услуги', 8),
        ('Тариф контраста', 8),
        ('Тариф Срочные', 8),
        ('Тариф Динамика', 8),
        ('Тариф Расширенные', 8),
        ('Итого сумма', 16),
    ]

    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border

    return ws1


def reestr_hospital_fill_data(ws1, result_query):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    style_border2 = NamedStyle(name="style_border1")
    style_border2.font = Font(bold=True, size=11)
    style_border2.alignment = Alignment(wrap_text=True, horizontal='right', vertical='center')

    r = 5
    for v in result_query:
        r += 1
        ws1.cell(row=r, column=1).value = v.get("hospital", "-")
        ws1.cell(row=r, column=2).value = v.get("direction_number", "-")
        ws1.cell(row=r, column=3).value = v.get("date_create", "-")
        ws1.cell(row=r, column=4).value = v.get("time_create", "-")
        ws1.cell(row=r, column=5).value = v.get("date_confirm", "-")
        ws1.cell(row=r, column=6).value = v.get("time_confirm", "-")
        ws1.cell(row=r, column=7).value = v.get("card_number", "-")
        ws1.cell(row=r, column=8).value = f'{v.get("patient_family", "-")} {v.get("patient_name", "-")} {v.get("patient_patronymic", "-")}'
        ws1.cell(row=r, column=9).value = v.get("service_code", "-")
        ws1.cell(row=r, column=10).value = v.get("service", "-")
        ws1.cell(row=r, column=11).value = v.get("department", "-")
        ws1.cell(row=r, column=12).value = f'{v.get("doctor_family", "-")} {v.get("doctor_name", "-")} {v.get("doctor_patronymic", "-")}'
        ws1.cell(row=r, column=13).value = v.get("tarif_coast", "-")
        ws1.cell(row=r, column=14).value = v.get("tarif_contrast", "-")
        ws1.cell(row=r, column=15).value = v.get("tarif_dynamic", "-")
        ws1.cell(row=r, column=16).value = v.get("tarif_extension", "-")
        ws1.cell(row=r, column=17).value = v.get("tarif_night", "-")
        ws1.cell(row=r, column=18).value = f'=SUM({get_column_letter(13)}{r}:{get_column_letter(17)}{r})'
    r_sum = r
    r += 1
    ws1.merge_cells(f"{get_column_letter(14)}{r}:{get_column_letter(15)}{r}")
    megre_cell = ws1[f"{get_column_letter(14)}{r}"]
    megre_cell.value = 'ВСЕГО сумма, руб'
    megre_cell.style = style_border2
    ws1.cell(row=r, column=18).value = f'=SUM({get_column_letter(18)}{6}:{get_column_letter(18)}{r_sum})'
    ws1.cell(row=r, column=18).style = style_border2
    return ws1
