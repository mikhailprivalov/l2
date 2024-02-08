import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter


def harmful_factors_base(ws1):
    style_border = NamedStyle(name="style_border_ca5")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    columns = [
        ('Код вредности', 20),
        ('Описание', 40),
        ('Шаблон', 25),
        ('Услуги', 50),
    ]

    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border

    return ws1


def harmful_factors_fill_data(ws1, result_query):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    pink_fill = openpyxl.styles.fills.PatternFill(patternType='solid', start_color='FCD5B4', end_color='FCD5B4')

    r = 5
    fill_rows = []
    for v in result_query.values():
        r += 1
        ws1.cell(row=r, column=1).value = v.get("harmfulfactor_title", "")
        ws1.cell(row=r, column=2).value = v.get("description", "")
        ws1.cell(row=r, column=3).value = v.get("template_title", "")
        fill_rows.append(r)

        for i in v.get("research_title", "").split(";"):
            if i:
                ws1.cell(row=r, column=4).value = i
                r += 1
            for j in range(1, 5):
                ws1.cell(row=r, column=j).style = style_border1
        r -= 1
    for k in fill_rows:
        rows = ws1[f'A{k}:D{k}']
        for row in rows:
            for cell in row:
                cell.fill = pink_fill
    return ws1
