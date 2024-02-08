from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter

from directory.models import Researches
from utils.dates import normalize_date


def offer_base(ws1):
    style_border = NamedStyle(name="style_border_ca5")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    columns = [
        ('Услуга', 50),
        ('Количество', 18),
        ('Цена', 18),
        ('Сумма', 18),
    ]

    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]

    return ws1


def offer_fill_data(ws1, result_query):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    r = 5
    start_row = r + 1
    for i in result_query:
        r += 1
        ws1.cell(row=r, column=1).value = i.get("code", "")
        ws1.cell(row=r, column=2).value = i.get("title", "")
        ws1.cell(row=r, column=3).value = i.get("count", 0)
        ws1.cell(row=r, column=4).value = float(i.get("coast", 0))
        ws1.cell(row=r, column=5).value = f'= {get_column_letter(4)}{r} * {get_column_letter(3)}{r}'

        for j in range(1, 5):
            ws1.cell(row=r, column=j).style = style_border1
    r += 1
    ws1.cell(row=r, column=1).value = 'Итого'
    ws1.cell(row=r, column=5).value = f'=SUM({get_column_letter(5)}{start_row}:{get_column_letter(5)}{r - 1})'
    for j in range(1, 5):
        ws1.cell(row=r, column=j).style = style_border1

    return ws1


def register_base(ws):
    style_border = NamedStyle(name="style_border_ca5")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    columns = [
        ('№', 10),
        ('Фамилия Имя Отчество', 50),
        ('Должность', 20),
        ('код вредности', 20),
        ('медицинские услуги', 20),
        ('Цена', 20),
    ]

    row = 5
    for idx, column in enumerate(columns, 1):
        ws.cell(row=row, column=idx).value = column[0]
        ws.column_dimensions[get_column_letter(idx)].width = column[1]
        ws.cell(row=row, column=idx).style = style_border

    return ws


def register_data(ws1, patients, research_price):
    r = 5
    start_row = r + 1
    current_patient = ""
    current_patient_step = 0
    sum_rows = []
    count_sum = 0
    for i in patients:
        r += 1
        if current_patient != i.get("fio", "-"):
            current_patient_step += 1
            ws1.cell(row=r, column=1).value = current_patient_step
            born_date = normalize_date(i.get("born"))
            ws1.cell(row=r, column=2).value = f'{i.get("fio", "-")} ({born_date}-{i.get("age", "age")})'
            ws1.cell(row=r, column=3).value = i.get("position", "-")
            ws1.cell(row=r, column=4).value = i.get("harmful_factor", "-")
            start_row = r
            r -= 1
        is_researches = False
        for k in i.get("researches"):
            r += 1
            research_data = research_price.get(k, "-@0").split("@")
            if research_data[0] == "-":
                research_title = Researches.objects.filter(pk=k).first().title
            else:
                research_title = research_data[0]
            ws1.cell(row=r, column=5).value = research_title
            ws1.cell(row=r, column=6).value = float(research_data[1])
            count_sum += float(research_data[1])
            is_researches = True
        if not is_researches:
            r += 1
        r += 1
        ws1.cell(row=r, column=2).value = 'Итого по пациенту'
        ws1.cell(row=r, column=6).value = f'=SUM({get_column_letter(6)}{start_row}:{get_column_letter(6)}{r - 1})'
        sum_rows.append(r)
    r += 1
    ws1.cell(row=r, column=2).value = 'Итого всего'
    formula = "=SUM("
    for i in sum_rows:
        formula = f"{formula} + {get_column_letter(6)}{i}"
    ws1.cell(row=r, column=6).value = f"{formula})"
    ws1.cell(row=r, column=7).value = count_sum
    return ws1
