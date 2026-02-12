from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter
from utils.db import namedtuplefetchall
from laboratory.settings import TIME_ZONE, USE_RMIS_NUMBER_IN_REVISE_REPORT_ECP_SEND
from django.db import connection


def form_01(ws1, data):
    # Для журнала ВК - ДЛО
    style_border = NamedStyle(name="style_border_ca")
    bd = Side(style="thin", color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=12)
    style_border.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    style_border2 = NamedStyle(name="style_border2")
    style_border2.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border2.font = Font(bold=False, size=11)
    style_border2.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    columns = [
        ('№ п.п.', 10),
        ('Врач', 30),
        ('Успех', 10),
        ('Дата подтверждения', 15),
        ('Дата создания', 15),
        ('Номер L2', 25),
        ('Служебный ИД', 25),
        ('Организация', 30),
        ('Пациент', 30),
        ('Дата рождения', 15),
        ('Услуга', 15),
        ('Код услуги', 15),
        ('Сообщение', 15),
        ('Случай', 15),
        ('Посещение', 15),
    ]
    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border

    sql_data = sql_01(tuple(data['research_id']), data['d_s'], data['d_e'], USE_RMIS_NUMBER_IN_REVISE_REPORT_ECP_SEND)
    result = [
        {
            "doctor": f"{i.doc_family} {i.doc_name} {i.doc_patronymic}",
            "status": "Да" if i.result_rmis_send else "Нет",
            "date_confirm": i.date_confirm,
            "date_create": i.rmis_direction_date,
            "direction_id": i.direction_number,
            "rmis_id": i.rmis_number,
            "hospital": i.hospital_title,
            "patient": f"{i.patient_family} {i.patient_name} {i.patient_patronymic}",
            "patient_birthday": i.patient_birthday,
            "service_title": i.research_title,
            "message": i.amd_message,
            "service_code": i.doctor_research_code if i.doctor_research_code else i.research_code,
            "case_num": i.rmis_case_number,
            "visit_num": i.rmis_visit_number,
            "rmis_login": i.rmis_login,
        }
        for i in sql_data
    ]
    step = 0
    for i in result:
        row += 1
        step += 1
        ws1.cell(row=row, column=1).value = step
        ws1.cell(row=row, column=2).value = i.get("doctor")
        ws1.cell(row=row, column=3).value = i.get("status")
        ws1.cell(row=row, column=4).value = i.get("date_confirm")
        ws1.cell(row=row, column=5).value = i.get("date_create")
        ws1.cell(row=row, column=6).value = i.get("direction_id")
        ws1.cell(row=row, column=7).value = i.get("rmis_id")
        ws1.cell(row=row, column=8).value = i.get("hospital")
        ws1.cell(row=row, column=9).value = i.get("patient")
        ws1.cell(row=row, column=10).value = i.get("patient_birthday")
        ws1.cell(row=row, column=11).value = i.get("service_title")
        ws1.cell(row=row, column=12).value = i.get("service_code")
        ws1.cell(row=row, column=13).value = i.get("message")
        ws1.cell(row=row, column=14).value = i.get("case_num")
        ws1.cell(row=row, column=15).value = i.get("visit_num")
        ws1.cell(row=row, column=16).value = i.get("rmis_login")
        for k in range(16):
            ws1.cell(row=row, column=k + 1).style = style_border
    return ws1


def sql_01(research_id, d_s, d_e, use_rmis_number):
    """
    Для журнала ВК-ДЛО
    :return:
    """
    with connection.cursor() as cursor:
        cursor.execute(
            """
                SELECT
                    ud.id as doc_id,
                    ud.family as doc_family,
                    ud.name as doc_name,
                    ud.patronymic as doc_patronymic,
                    dn.result_rmis_send,
                    hh.title as hospital_title,
                    to_char(dn.rmis_direction_date AT TIME ZONE %(tz)s, 'DD.MM.YYYY') as rmis_direction_date,
                    to_char(directions_issledovaniya.time_confirmation AT TIME ZONE %(tz)s, 'DD.MM.YYYY') as date_confirm,
                    directions_issledovaniya.napravleniye_id as direction_number,
                    dn.rmis_number,
                    ci.family as patient_family,
                    ci.name as patient_name,
                    ci.patronymic as patient_patronymic,
                    to_char(ci.birthday, 'DD.MM.YYYY') as patient_birthday,
                    dn.amd_message,
                    dn.rmis_case_number,
                    dn.rmis_visit_number,
                    dr.title as research_title,
                    dr.code as research_code,
                    ud.service_code_ambulatory as doctor_research_code,
                    ud.rmis_login
   
                    FROM directions_issledovaniya
                    LEFT JOIN directions_napravleniya dn ON dn.id = directions_issledovaniya.napravleniye_id
                    LEFT JOIN users_doctorprofile ud ON directions_issledovaniya.doc_confirmation_id=ud.id
                    LEFT JOIN clients_card cc ON cc.id=dn.client_id
                    LEFT JOIN clients_individual ci ON cc.individual_id=ci.id
                    LEFT JOIN hospitals_hospitals hh ON dn.hospital_id=hh.id
                    LEFT JOIN directory_researches dr ON directions_issledovaniya.research_id=dr.id
                    WHERE 
                      directions_issledovaniya.research_id in %(research_id)s
                      AND directions_issledovaniya.time_confirmation AT TIME ZONE %(tz)s BETWEEN %(d_start)s AND %(d_end)s
                      AND dn.parent_id IS NULL
                      AND 
                      CASE when %(use_rmis_number)s = 1 THEN
                         dn.rmis_number IS NOT NULL 
                      when %(use_rmis_number)s = 0 THEN
                         directions_issledovaniya.napravleniye_id IS NOT NULL
                      END                                               
                    order by ud.id, directions_issledovaniya.time_confirmation
                """,
            params={'research_id': research_id, 'd_start': d_s, 'd_end': d_e, 'tz': TIME_ZONE, 'use_rmis_number': use_rmis_number},
        )

        rows = namedtuplefetchall(cursor)
    return rows
