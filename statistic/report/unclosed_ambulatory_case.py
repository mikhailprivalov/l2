from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter
from utils.db import namedtuplefetchall
from laboratory.settings import TIME_ZONE, RESEARCH_ID_CLOSE_CASE
from django.db import connection


def form_01(ws1, data):
    # Для журнал не закрытых случаев
    style_border = NamedStyle(name="style_border_ca")
    bd = Side(style="thin", color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=12)
    style_border.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    style_border2 = NamedStyle(name="style_border2")
    style_border2.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border2.font = Font(bold=False, size=11)
    style_border2.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    columns = [
        ('№ п.п.', 10),
        ('№ случая', 10),
        ('Компания', 50),
        ('ФИО', 40),
        ('Дата рождения', 10),
        ('Пол', 10),
    ]
    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border

    sql_data = sql_01(data['d_s'], data['d_e'])
    result = [
        {
            "direction_id": i.direction_num,
            "patient": f"{i.patient_family} {i.patient_name} {i.patient_patronymic}",
            "patient_birthday": i.patient_birthday,
            "company_title": i.company_title,
            "sex": i.sex,
        }
        for i in sql_data
    ]
    step = 0
    for i in result:
        row += 1
        step += 1
        ws1.cell(row=row, column=1).value = step
        ws1.cell(row=row, column=2).value = i.get("direction_id")
        ws1.cell(row=row, column=3).value = i.get("company_title")
        ws1.cell(row=row, column=4).value = i.get("patient")
        ws1.cell(row=row, column=5).value = i.get("patient_birthday")
        ws1.cell(row=row, column=6).value = i.get("sex")

    return ws1


def sql_01(d_start, d_end):
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT
            dn.id as direction_num,
            dn.work_place_db_id,
            contract_comp.title as company_title,
            directions_issledovaniya.id as case_issledovaniye_id,
            ci.family as patient_family,
            ci.name as patient_name,
            ci.patronymic as patient_patronymic,
            ci.sex,
            to_char(ci.birthday AT TIME ZONE %(tz)s, 'DD.MM.YYYY') as patient_birthday
            FROM directions_issledovaniya
            LEFT JOIN directions_napravleniya dn on directions_issledovaniya.napravleniye_id = dn.id
            LEFT JOIN clients_card cc on cc.id=dn.client_id
            LEFT JOIN clients_individual ci on cc.individual_id = ci.id
            LEFT JOIN contracts_company contract_comp on dn.work_place_db_id = contract_comp.id 
            WHERE
            directions_issledovaniya.time_confirmation is Null
            AND directions_issledovaniya.research_id in %(research_id_case)s
            AND dn.cancel = false
            ORDER BY dn.data_sozdaniya
            """,
            params={'d_start': d_start, 'd_end': d_end, 'tz': TIME_ZONE, 'research_id_case': RESEARCH_ID_CLOSE_CASE},
        )

        rows = namedtuplefetchall(cursor)
    return rows
