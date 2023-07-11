from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter


def profit_base(ws1, d1, d2):
    style_border = NamedStyle(name="style_border_ca")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=1, column=1).value = 'Услуга:'
    ws1.cell(row=1, column=2).value = 'Реестр начислений'
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'

    columns = [
        ('Компания', 45),
        ('ФИО врача', 35),
        ('Должность', 45),
        ('Вид услуги', 45),
        ('Кол-во услуг', 10),
        ('Цена', 15),
        ('Сумма', 15),
    ]
    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border

    return ws1


def profit_data(ws1, final_result):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=15)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    row = 5
    for v in final_result:
        row += 1
        ws1.cell(row=row, column=1).value = v["company"]
        ws1.cell(row=row, column=2).value = v["fio"]
        ws1.cell(row=row, column=3).value = v["position"]
        ws1.cell(row=row, column=4).value = v["research"]
        ws1.cell(row=row, column=5).value = v["coast"]
        ws1.cell(row=row, column=6).value = v["count"]
        ws1.cell(row=row, column=7).value = f'={get_column_letter(5)}{row}*{get_column_letter(6)}{row}'

        for c in range(1, 8):
            ws1.cell(row=row, column=c).style = style_border1

    return ws1
