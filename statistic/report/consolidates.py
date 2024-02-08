from copy import deepcopy

import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter


def consolidate_base(ws1, d1, d2, fin_source):
    style_border = NamedStyle(name="style_border_ca5")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=1, column=1).value = f'Сводный: {fin_source}'
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'

    columns = [
        ('Номер карты', 10),
        ('Возраст', 8),
        ('Фамилия', 16),
        ('Имя', 16),
        ('Отчество', 16),
        ('Комментарий_в_карте', 20),
        ('Наимен. услуги', 30),
        ('№ направления', 15),
        ('Дата подтверждения', 15),
        ('Врач', 25),
        ('Место работы', 55),
        ('Профиль_мед.помощи', 15),
        ('Спец_врача', 25),
        ('Основное', 15),
        ('Подчинение', 15),
        ('Цель', 15),
        ('Категория', 25),
        ('Леч врач', 25),
        ('Кто создал', 25),
    ]

    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border

    return ws1


def consolidate_fill_data(ws1, result_query):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    r = 5
    tmp_data = {}
    for i in result_query:
        if i.dir_id:
            tmp_data[i.id_iss] = {
                "patient_card_num": i.patient_card_num,
                "patient_age": i.patient_age,
                "patient_family": i.patient_family,
                "patient_name": i.patient_name,
                "patient_patronymic": i.patient_patronymic,
                "dir_harmful_factor": i.dir_harmful_factor,
                "research_title": i.research_title,
                "dir_id": i.dir_id,
                "date_confirm": i.date_confirm,
                "fio": f"{i.doc_f} {i.doc_n} {i.doc_p}",
                "patient_workplace": i.patient_workplace,
                "doc_speciality": i.doc_speciality,
                "purpose": i.purpose,
            }
            r += 1
            ws1.cell(row=r, column=1).value = i.patient_card_num
            ws1.cell(row=r, column=2).value = i.patient_age
            ws1.cell(row=r, column=3).value = i.patient_family
            ws1.cell(row=r, column=4).value = i.patient_name
            ws1.cell(row=r, column=5).value = i.patient_patronymic
            ws1.cell(row=r, column=6).value = i.dir_harmful_factor
            ws1.cell(row=r, column=7).value = i.research_title
            ws1.cell(row=r, column=8).value = i.dir_id
            ws1.cell(row=r, column=9).value = i.date_confirm
            ws1.cell(row=r, column=10).value = f"{i.doc_f} {i.doc_n} {i.doc_p}"
            ws1.cell(row=r, column=11).value = i.patient_workplace
            ws1.cell(row=r, column=12).value = ""
            ws1.cell(row=r, column=13).value = i.doc_speciality
            ws1.cell(row=r, column=14).value = i.id_iss
            ws1.cell(row=r, column=15).value = i.parent_iss
            ws1.cell(row=r, column=16).value = i.purpose
            ws1.cell(row=r, column=17).value = i.category_title
            ws1.cell(row=r, column=18).value = f"{i.user_doc_f} {i.user_doc_n} {i.user_doc_p}"
            ws1.cell(row=r, column=19).value = f"{i.doc_who_create_f} {i.doc_who_create_n} {i.doc_who_create_p}"
            for j in range(1, 20):
                ws1.cell(row=r, column=j).style = style_border1

        if i.parent_iss:
            data = tmp_data.get(i.parent_iss)
            if not data:
                continue
            r += 1
            ws1.cell(row=r, column=1).value = data["patient_card_num"]
            ws1.cell(row=r, column=2).value = data["patient_age"]
            ws1.cell(row=r, column=3).value = data["patient_family"]
            ws1.cell(row=r, column=4).value = data["patient_name"]
            ws1.cell(row=r, column=5).value = data["patient_patronymic"]
            ws1.cell(row=r, column=6).value = data["dir_harmful_factor"]
            ws1.cell(row=r, column=7).value = i.research_title
            ws1.cell(row=r, column=8).value = data["dir_id"]
            ws1.cell(row=r, column=9).value = data["date_confirm"]
            ws1.cell(row=r, column=10).value = data["fio"]
            ws1.cell(row=r, column=11).value = data["patient_workplace"]
            ws1.cell(row=r, column=12).value = ""
            ws1.cell(row=r, column=13).value = data["doc_speciality"]
            ws1.cell(row=r, column=14).value = i.id_iss
            ws1.cell(row=r, column=15).value = i.parent_iss
            ws1.cell(row=r, column=16).value = data["purpose"]
            ws1.cell(row=r, column=17).value = i.category_title
            ws1.cell(row=r, column=18).value = f"{i.user_doc_f} {i.user_doc_n} {i.user_doc_p}"
            ws1.cell(row=r, column=19).value = f"{i.doc_who_create_f} {i.doc_who_create_n} {i.doc_who_create_p}"
            for j in range(1, 20):
                ws1.cell(row=r, column=j).style = style_border1

    return ws1


def consolidate_research_sets_base(ws1, d1, d2, fin_source, head_data, company_title, coast_data):
    style_border = NamedStyle(name="style_border_ca5")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=1, column=1).value = f'Сводный: {fin_source}'
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'
    ws1.cell(row=4, column=1).value = f'Контрагент: {company_title}'

    columns = [
        ('Отдел', 20),
        ('№ п/п.', 20),
        ('Карта', 15),
        ('ФИО', 48),
    ]
    start_column_research = len(columns) + 1

    custom_columns = [(i, 13) for i in head_data.values()]
    coast_data_column = [*[("", 10) for k in range(len(columns))], *[(i, 13) for i in coast_data.values()]]
    columns.extend(custom_columns)

    row = 6
    step = 0
    for k, j in zip(columns, coast_data_column):
        step += 1
        ws1.cell(row=row, column=step).value = k[0]
        ws1.cell(row=row + 1, column=step).value = j[0]
        ws1.column_dimensions[get_column_letter(step)].width = k[1]
        ws1.cell(row=row, column=step).style = style_border
        ws1.cell(row=row + 1, column=step).style = style_border

    ws1.cell(row=row, column=step + 1).value = "Итого по человеку"
    ws1.cell(row=row, column=step + 1).style = style_border
    ws1.cell(row=row + 1, column=step + 1).style = style_border
    return (
        ws1,
        start_column_research,
    )


def consolidate_research_sets_fill_data(ws1, query, def_value_data, start_research_column):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    last_patient = None
    base_step = 0
    current_patient_researh_data = deepcopy(def_value_data)
    current_department_id = None
    current_department_title = ""
    last_patient_fio = ""
    last_patient_card = ""
    row = 7
    price_row = row
    start_row = 0
    current_sum_columns = {}
    total_sum_rows = []
    fill_rows = {}
    count_patient_in_department = 0
    for i in query:
        if last_patient != i.client_id and base_step != 0:
            row += 1
            count_patient_in_department += 1
            ws1.cell(row=row, column=1).value = current_department_title
            ws1.cell(row=row, column=2).value = count_patient_in_department
            ws1.cell(row=row, column=3).value = last_patient_card
            ws1.cell(row=row, column=4).value = last_patient_fio
            column = start_research_column
            row_sum = 0
            if start_row == 0:
                start_row = row
            for k in current_patient_researh_data.values():
                ws1.cell(row=row, column=column).value = k
                current_sum_columns[column] = f'=SUM({get_column_letter(column)}{start_row}:{get_column_letter(column)}{row})'
                row_sum = f'{row_sum} + {get_column_letter(column)}{price_row}*{get_column_letter(column)}{row}'
                column += 1
            ws1.cell(row=row, column=column).value = f"={row_sum}"
            current_patient_researh_data = deepcopy(def_value_data)

        current_patient_researh_data[i.research_id] = 1

        if current_department_id != i.department_id and base_step != 0:
            count_patient_in_department = 0
            start_row = 0
            row += 1
            ws1.cell(row=row, column=1).value = f'Итого кол-во {current_department_title}'
            for col, val in current_sum_columns.items():
                ws1.cell(row=row, column=col).value = val
                ws1.cell(row=row + 1, column=col).value = f'={get_column_letter(col)}{price_row}*{get_column_letter(col)}{row}'
            total_sum_rows.append(row)
            row += 1
            ws1.cell(row=row, column=1).value = f'Итого сумма {current_department_title}'
            fill_rows[row] = col
            current_sum_columns = {}
        current_department_id = i.department_id
        current_department_title = i.department_title
        last_patient = i.client_id
        last_patient_fio = f"{i.patient_family} {i.patient_name} {i.patient_patronymic}"
        last_patient_card = i.patient_card_num
        base_step += 1

    row += 1
    count_patient_in_department += 1
    ws1.cell(row=row, column=1).value = current_department_title
    ws1.cell(row=row, column=2).value = count_patient_in_department
    ws1.cell(row=row, column=3).value = last_patient_card
    ws1.cell(row=row, column=4).value = last_patient_fio
    column = start_research_column
    row_sum = 0
    for k in current_patient_researh_data.values():
        if start_row == 0:
            start_row = row
        ws1.cell(row=row, column=column).value = k
        current_sum_columns[column] = f'=SUM({get_column_letter(column)}{start_row}:{get_column_letter(column)}{row})'
        row_sum = f'{row_sum} + {get_column_letter(column)}{price_row}*{get_column_letter(column)}{row}'
        column += 1

    ws1.cell(row=row, column=column).value = f"={row_sum}"

    row += 1
    ws1.cell(row=row, column=1).value = f'Итого кол-во {current_department_title}'
    for col, val in current_sum_columns.items():
        ws1.cell(row=row, column=col).value = val
        ws1.cell(row=row + 1, column=col).value = f'={get_column_letter(col)}{price_row}*{get_column_letter(col)}{row}'
    total_sum_rows.append(row)
    row += 1
    ws1.cell(row=row, column=1).value = f'Итого сумма {current_department_title}'
    fill_rows[row] = col

    row += 1
    ws1.cell(row=row, column=1).value = 'Итого кол-во всего'
    total_sum_end = ""
    for total_col in range(start_research_column, col + 1):
        step = 0
        for total_research in total_sum_rows:
            if step != 0:
                total_sum_end = f'{total_sum_end} + {get_column_letter(total_col)}{total_research}'
            else:
                total_sum_end = f'{get_column_letter(total_col)}{total_research}'
            step += 1
        ws1.cell(row=row, column=total_col).value = f"={total_sum_end}"
        total_sum_end = ""

    row += 1
    ws1.cell(row=row, column=1).value = 'Итого сумма всего'
    for total_col in range(start_research_column, col + 1):
        ws1.cell(row=row, column=total_col).value = f'={get_column_letter(total_col)}{price_row}*{get_column_letter(total_col)}{row-1}'

    total_fill = openpyxl.styles.fills.PatternFill(patternType='solid', start_color='ffcc66', end_color='ffcc66')
    for k, v in fill_rows.items():
        fill_cells(ws1[f'A{k}:{get_column_letter(v)}{k}'], total_fill)
        fill_cells(ws1[f'A{k - 1}:{get_column_letter(v)}{k - 1}'], total_fill)

    return ws1


def fill_cells(rows_fill, total_fill):
    for row_f in rows_fill:
        for cell in row_f:
            cell.fill = total_fill


def consolidate_base_doctors_by_type_department(ws1, d1, d2, fin_source_data):
    style_border = NamedStyle(name="style_border_ca5")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=1, column=1).value = 'Сводный:'
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'

    columns = [
        ('Подразделение', 20),
        ('Сотрудник', 30),
        ('Услуга', 30),
        ('Данные пациента', 40),
    ]
    finish_order = {}
    for k, v in fin_source_data.items():
        start_position_col = len(columns) + 1
        columns.append((f"{v}, шт.", 12))
        finish_order[k] = start_position_col
        start_position_col += 1
        columns.append((f"{v}, ует.", 12))
    columns.append(("Итого, шт.", 12))
    columns.append(("Итого, ует.", 12))
    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border

    return (ws1, finish_order)


def consolidate_fill_data_doctors_by_type_department_detail_patient(ws1, query, fin_source_order):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    total_fill = openpyxl.styles.fills.PatternFill(patternType='solid', start_color='a9d094', end_color='a9d094')
    row = 5
    start_row = row + 1
    old_doctor, old_department, current_doctor, current_department_title = "", "", "", ""
    step = 0
    min_col_val, max_col_val = min(fin_source_order.values()), max(fin_source_order.values()) + 2
    sum_current_department = []
    for i in query:
        row += 1
        current_department_title = i.department_title
        current_doctor = f"{i.family} {i.name} {i.patronymic}"
        if (old_doctor != current_doctor) and (step != 0):
            ws1.cell(row=row, column=1).value = old_department
            ws1.cell(row=row, column=2).value = f"Итого: {old_doctor}"
            ws1 = doctor_summary(ws1, min_col_val, max_col_val, start_row, row, total_fill)
            sum_current_department.append(row)
            if old_department != current_department_title:
                row += 1
                ws1.cell(row=row, column=1).value = f"Итого: {old_department}"
                ws1 = count_sum_from_data_cells(ws1, min_col_val, max_col_val, sum_current_department, row)
                sum_current_department = []
            ws1 = count_sum_by_custom_cells(ws1, start_row, row + 1, min_col_val, max_col_val)
            ws1 = count_sum_by_custom_cells(ws1, start_row, row + 1, min_col_val + 1, max_col_val + 1)
            row += 1
            start_row = row
        ws1.cell(row=row, column=1).value = current_department_title
        ws1.cell(row=row, column=2).value = current_doctor
        ws1.cell(row=row, column=3).value = i.research_title
        ws1.cell(row=row, column=4).value = f"{i.patient_family} {i.patient_name} {i.patient_patronymic}; карта-{i.patient_card_num}; напр-{i.dir_id}; {i.date_confirm}"
        col = fin_source_order.get(i.istochnik_f_id) if fin_source_order.get(i.istochnik_f_id) else 50
        ws1.cell(row=row, column=col).value = 1
        ws1.cell(row=row, column=col + 1).value = i.uet_refferal_doc
        old_doctor, old_department = current_doctor, current_department_title
        step += 1
    row += 1
    ws1.cell(row=row, column=1).value = current_department_title
    ws1.cell(row=row, column=2).value = f"Итого: {current_doctor}"
    ws1 = doctor_summary(ws1, min_col_val, max_col_val, start_row, row, total_fill)
    sum_current_department.append(row)
    row += 1
    ws1.cell(row=row, column=1).value = f"Итого: {old_department}"
    ws1 = count_sum_from_data_cells(ws1, min_col_val, max_col_val, sum_current_department, row)
    ws1 = count_sum_by_custom_cells(ws1, start_row, row + 1, min_col_val, max_col_val)
    ws1 = count_sum_by_custom_cells(ws1, start_row, row + 1, min_col_val + 1, max_col_val + 1)
    return ws1


def consolidate_fill_data_doctors_by_type_department(ws1, query, fin_source_order):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    total_fill = openpyxl.styles.fills.PatternFill(patternType='solid', start_color='a9d094', end_color='a9d094')
    row = 5
    start_row = row + 1
    old_doctor, old_department, current_doctor, current_department_title, current_research, old_research = "", "", "", "", "", ""
    step = 0
    step_department = 0
    min_col_val, max_col_val = min(fin_source_order.values()), max(fin_source_order.values()) + 2
    sum_current_department = []
    research_finsource_count = {k: 0 for k in fin_source_order.keys()}
    uet_finsource_count = {k: 0 for k in fin_source_order.keys()}
    for i in query:
        current_department_title = i.department_title
        current_doctor = f"{i.family} {i.name} {i.patronymic}"
        current_research = i.research_title
        if (old_doctor != current_doctor) and (step != 0) and (step_department != 0):
            row += 1
            ws1.cell(row=row, column=3).value = old_research
            ws1.cell(row=row, column=1).value = old_department
            ws1.cell(row=row, column=2).value = old_doctor
            ws1 = fill_data_cells_by_research(ws1, research_finsource_count, fin_source_order, uet_finsource_count, row)
            row += 1
            ws1 = doctor_summary(ws1, min_col_val, max_col_val, start_row, row, total_fill)
            sum_current_department.append(row)
            ws1.cell(row=row, column=1).value = old_department
            ws1.cell(row=row, column=2).value = f"Итого: {old_doctor}"
            research_finsource_count = {k: 0 for k in fin_source_order.keys()}
            uet_finsource_count = {k: 0 for k in fin_source_order.keys()}
            if old_department != current_department_title:
                row += 1
                ws1.cell(row=row, column=1).value = f"Итого: {old_department}"
                ws1 = count_sum_from_data_cells(ws1, min_col_val, max_col_val, sum_current_department, row)
                sum_current_department = []
                research_finsource_count = {k: 0 for k in fin_source_order.keys()}
                uet_finsource_count = {k: 0 for k in fin_source_order.keys()}
                step_department = 0
            ws1 = count_sum_by_custom_cells(ws1, start_row, row + 1, min_col_val, max_col_val)
            ws1 = count_sum_by_custom_cells(ws1, start_row, row + 1, min_col_val + 1, max_col_val + 1)
            row += 1
            start_row = row

        if old_research != current_research and step != 0 and step_department != 0:
            row += 1
            ws1.cell(row=row, column=3).value = old_research
            ws1 = fill_data_cells_by_research(ws1, research_finsource_count, fin_source_order, uet_finsource_count, row)
            research_finsource_count = {k: 0 for k in fin_source_order.keys()}
            uet_finsource_count = {k: 0 for k in fin_source_order.keys()}

            ws1.cell(row=row, column=1).value = current_department_title
            ws1.cell(row=row, column=2).value = current_doctor
        step_department += 1
        if research_finsource_count.get(i.istochnik_f_id) or research_finsource_count.get(i.istochnik_f_id) == 0:
            research_finsource_count[i.istochnik_f_id] = research_finsource_count[i.istochnik_f_id] + 1
            uet_finsource_count[i.istochnik_f_id] = uet_finsource_count[i.istochnik_f_id] + i.uet_refferal_doc
        old_doctor, old_department, old_research = current_doctor, current_department_title, current_research
        step += 1
    row += 1
    ws1.cell(row=row, column=3).value = old_research
    ws1 = fill_data_cells_by_research(ws1, research_finsource_count, fin_source_order, uet_finsource_count, row)
    ws1.cell(row=row, column=1).value = current_department_title
    ws1.cell(row=row, column=2).value = current_doctor
    row += 1
    ws1.cell(row=row, column=1).value = current_department_title
    ws1.cell(row=row, column=2).value = f"Итого: {current_doctor}"
    ws1 = doctor_summary(ws1, min_col_val, max_col_val, start_row, row, total_fill)
    sum_current_department.append(row)
    row += 1
    ws1.cell(row=row, column=1).value = f"Итого: {old_department}"
    ws1 = count_sum_from_data_cells(ws1, min_col_val, max_col_val, sum_current_department, row)
    ws1 = count_sum_by_custom_cells(ws1, start_row, row + 1, min_col_val, max_col_val)
    ws1 = count_sum_by_custom_cells(ws1, start_row, row + 1, min_col_val + 1, max_col_val + 1)
    return ws1


def count_sum_by_custom_cells(ws2, start_row, end_row, start_col, end_col):
    for s in range(start_row, end_row):
        step_sum = 1
        sum_column = '=SUM('
        for k in range(start_col, end_col):
            step_sum += 1
            if step_sum % 2 == 0 and step_sum != 2:
                sum_column = f"{sum_column},{get_column_letter(k)}{s}"
            elif step_sum % 2 == 0 and step_sum == 2:
                sum_column = f"{sum_column}{get_column_letter(k)}{s}"
        sum_column = f"{sum_column})"
        ws2.cell(row=s, column=end_col).value = f'{sum_column}'
    return ws2


def count_sum_from_data_cells(ws2, start_col, end_col, data_rows, purpose_row):
    for purpose_col in range(start_col, end_col):
        step_sum = 0
        sum_column = '=SUM('
        for current_row in data_rows:
            if step_sum != 0:
                sum_column = f'{sum_column}, {get_column_letter(purpose_col)}{current_row}'
            else:
                sum_column = f'{sum_column}{get_column_letter(purpose_col)}{current_row}'
            step_sum += 1
        sum_column = f"{sum_column})"
        ws2.cell(row=purpose_row, column=purpose_col).value = f"={sum_column}"
    return ws2


def doctor_summary(ws2, star_col, end_col, start_current_row, purpose_row, fill_param):
    for k in range(star_col, end_col):
        ws2.cell(row=purpose_row, column=k).value = f'=SUM({get_column_letter(k)}{start_current_row}:{get_column_letter(k)}{purpose_row - 1})'
    ws2.row_dimensions.group(start_current_row, purpose_row - 1, hidden=True)
    fill_cells(ws2[f'A{purpose_row}:{get_column_letter(end_col + 1)}{purpose_row}'], fill_param)
    return ws2


def fill_data_cells_by_research(ws2, research_finsource_count, fin_source_order, uet_finsource_count, current_row):
    for k, v in research_finsource_count.items():
        col = fin_source_order.get(k) if fin_source_order.get(k) else 50
        ws2.cell(row=current_row, column=col).value = v
    for k, v in uet_finsource_count.items():
        col = fin_source_order.get(k) if fin_source_order.get(k) else 50
        ws2.cell(row=current_row, column=col + 1).value = v
    return ws2
