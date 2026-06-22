from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


def _fmt_period_date(value):
    if hasattr(value, 'strftime'):
        return value.strftime('%d.%m.%Y')
    return str(value)


def _apply_cell_style(cell, bold=False):
    thin = Side(style='thin', color='000000')
    cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)
    cell.font = Font(bold=bold, size=11)
    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')


def hospital_discharge_epicrisis_base(ws1, d1, d2, research_title):
    ws1.cell(row=1, column=1).value = 'Услуга:'
    ws1.cell(row=1, column=2).value = research_title
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {_fmt_period_date(d1)} по {_fmt_period_date(d2)}'

    columns = [
        ('№ п.п.', 8),
        ('ФИО пациента', 35),
        ('дата рождения пациента', 22),
        ('Названия услуги выписки', 35),
        ('ФИО врача', 30),
        ('№ направления выписки', 22),
        ('Дата выписки из протокола', 24),
        ('время выписки из протокола', 22),
        ('№ родительского направления', 26),
        ('Название услуги родительского направления', 40),
    ]
    header_row = 5
    for idx, (title, width) in enumerate(columns, 1):
        cell = ws1.cell(row=header_row, column=idx)
        cell.value = title
        ws1.column_dimensions[get_column_letter(idx)].width = width
        _apply_cell_style(cell, bold=True)

    return ws1


def hospital_discharge_epicrisis_fill_data(ws1, rows, start_row=6):
    r = start_row - 1
    for step, row in enumerate(rows, 1):
        r += 1
        values = [
            step,
            row.patient_fio or '',
            row.patient_birthday or '',
            row.extract_research_title or '',
            row.doc_fio or '',
            row.extract_direction_id or '',
            row.discharge_date or '',
            row.discharge_time or '',
            row.parent_direction_id or '',
            row.parent_research_title or '',
        ]
        for col, value in enumerate(values, 1):
            cell = ws1.cell(row=r, column=col)
            cell.value = value
            _apply_cell_style(cell)

    return ws1
