from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter

from laboratory.settings import DISPANSERIZATION_STATTALON_FIELDS_RESULTS_PK


def dispanserization_data(query_sql, pk_service_start, pk_service_end, doctors_count_pass_patient):
    doctors_count_pass = {}
    for i in doctors_count_pass_patient:
        if not doctors_count_pass.get(i.doc_confirmation_id, None):
            doctors_count_pass[i.doc_confirmation_id] = {i.confirm_time: i.count}
        else:
            tmp_doctors_count_pass = doctors_count_pass.get(i.doc_confirmation_id, {})
            tmp_doctors_count_pass[i.confirm_time] = i.count
            doctors_count_pass[i.doc_confirmation_id] = tmp_doctors_count_pass.copy()

    dates = [i.confirm_time for i in query_sql]
    dates = sorted(set(dates))
    dates_val = [{"val_start": 0, "val_end": 0, "patient_date": 0} for i in range(len(dates))]
    result = []
    prev_doc = None
    step = 0
    current_tmp_data = {}
    check_result_reception_id = False
    if len(DISPANSERIZATION_STATTALON_FIELDS_RESULTS_PK) > 0:
        check_result_reception_id = True
    for i in query_sql:
        current_index = dates.index(i.confirm_time)
        if i.doc_confirmation_id != prev_doc:
            if step != 0:
                tmp_doctor_data = current_tmp_data[prev_doc]
                tmp_dates_val = tmp_doctor_data.get("dates_val", [])
                tmp_dates = tmp_doctor_data.get("dates", [])
                for k, v in doctors_count_pass[prev_doc].items():
                    if k not in tmp_dates:
                        continue
                    tmp_index = tmp_dates.index(k)
                    tmp_data = tmp_dates_val[tmp_index]
                    tmp_data["patient_date"] = v
                    tmp_dates_val[tmp_index] = tmp_data.copy()
                tmp_doctor_data["dates_val"] = tmp_dates_val.copy()
                current_tmp_data[prev_doc] = tmp_doctor_data.copy()
                result.append(current_tmp_data.copy())
            current_doc = {"fio_doc": i.fio_doctor, "dates_val": dates_val.copy(), "dates": dates.copy()}
            current_tmp_data[i.doc_confirmation_id] = current_doc.copy()
        current_doc = current_tmp_data.get(i.doc_confirmation_id, {})
        tmp_val = current_doc.get("dates_val", [])[current_index].copy()
        if i.research_id in pk_service_start:
            val_start = tmp_val.get("val_start", 0)
            val_start += 1
            tmp_val["val_start"] = val_start
        elif i.research_id in pk_service_end:
            if check_result_reception_id and i.result_reception_id in DISPANSERIZATION_STATTALON_FIELDS_RESULTS_PK:
                val_end = tmp_val.get("val_end", 0)
                val_end += 1
                tmp_val["val_end"] = val_end
            elif not check_result_reception_id:
                val_end = tmp_val.get("val_end", 0)
                val_end += 1
                tmp_val["val_end"] = val_end
        current_doc["dates_val"][current_index] = tmp_val.copy()
        current_tmp_data = {i.doc_confirmation_id: current_doc.copy()}
        prev_doc = i.doc_confirmation_id
        step += 1
    tmp_doctor_data = current_tmp_data[prev_doc]
    tmp_dates_val = tmp_doctor_data.get("dates_val", [])
    tmp_dates = tmp_doctor_data.get("dates", [])
    for k, v in doctors_count_pass[prev_doc].items():
        if k not in tmp_dates:
            continue
        tmp_index = tmp_dates.index(k)
        tmp_data = tmp_dates_val[tmp_index]
        tmp_data["patient_date"] = v
        tmp_dates_val[tmp_index] = tmp_data.copy()
    tmp_doctor_data["dates_val"] = tmp_dates_val.copy()
    current_tmp_data[prev_doc] = tmp_doctor_data.copy()
    result.append(current_tmp_data.copy())

    return {"result": result, "dates": list(dates)}


def get_doctors_dispanserization(query_sql):
    return [i.doc_confirmation_id for i in query_sql]


def dispanserization_base(ws1, d1, d2, result_query, row=5):
    style_border = NamedStyle(name=f"style_border_ca{row}")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=1, column=1).value = 'Диспансеризация:'
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'

    columns = [("Врач", 40)]
    columns_date = [("", 40)]
    for i in result_query["dates"]:
        columns.append(("начато", 6))
        columns.append(("завершено", 10))
        columns.append(("принято пациентов", 10))
        columns_date.append((i, 6))
        columns_date.append(("", 10))
        columns_date.append(("", 10))
    columns.append(("начато", 6))
    columns.append(("завершено", 10))
    columns.append(("принято пациентов", 10))
    columns_date.append(("За период", 6))
    columns_date.append(("", 10))
    columns_date.append(("", 10))

    for idx, column in enumerate(columns_date, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border

    column_count = 2
    for i in range(len(result_query["dates"]) + 2):
        ws1.merge_cells(start_row=row, start_column=column_count, end_row=row, end_column=column_count + 2)
        column_count += 3

    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row + 1, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row + 1, column=idx).style = style_border

    return ws1


def dispanserization_fill_data(ws1, result_query, row=6):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    alignment1 = Alignment(wrap_text=True, horizontal='left', vertical='center')
    len_dates_count = len(result_query["dates"]) * 3

    r = row
    start_row = r + 1
    for i in result_query["result"]:
        r += 1
        val_start_sum = '=SUM('
        val_end_sum = '=SUM('
        patient_date_sum = '=SUM('
        for k in i.values():
            ws1.cell(row=r, column=1).value = k["fio_doc"]
            column = 1
            for j in k["dates_val"]:
                column += 1
                ws1.cell(row=r, column=column).value = j["val_start"]
                val_start_sum = f'{val_start_sum} + {get_column_letter(column)}{r}'
                column += 1
                ws1.cell(row=r, column=column).value = j["val_end"]
                val_end_sum = f'{val_end_sum} + {get_column_letter(column)}{r}'
                column += 1
                ws1.cell(row=r, column=column).value = j["patient_date"]
                patient_date_sum = f'{patient_date_sum} + {get_column_letter(column)}{r}'

            val_start_sum = f'{val_start_sum})'
            val_end_sum = f'{val_end_sum})'
            patient_date_sum = f'{patient_date_sum})'
            ws1.cell(row=r, column=column + 1).value = val_start_sum
            ws1.cell(row=r, column=column + 2).value = val_end_sum
            ws1.cell(row=r, column=column + 3).value = patient_date_sum

        for z in range(1, len_dates_count + 5):
            ws1.cell(row=r, column=z).style = style_border1
        ws1.cell(row=r, column=1).alignment = alignment1

    column = 1
    r += 1

    ws1.cell(row=r, column=column).value = "Итого"
    ws1.cell(row=r, column=column).style = style_border1
    ws1.cell(row=r, column=column).alignment = alignment1

    for i in range(len_dates_count + 3):
        column += 1
        ws1.cell(row=r, column=column).value = f'=SUM({get_column_letter(column)}{start_row}:{get_column_letter(column)}{r - 1})'
        ws1.cell(row=r, column=column).style = style_border1

    return ws1
