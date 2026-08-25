from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter

FIELD_LPU = 'ЛПУ сообщившее о заболевании'
FIELD_REG_DATE = 'Дата регистрации'
FIELD_DIAGNOSIS = 'Первичный диагноз'


def match_master_field(master_field_results, title):
    title_norm = title.strip().lower()
    fallback = ''
    for f in master_field_results:
        field_title = (f.get('master_field_title') or '').strip().lower()
        value = f.get('master_value') or ''
        if field_title == title_norm:
            return value
        if title_norm in field_title and not fallback:
            fallback = value
    return fallback


def build_epid_rows(epid_data, confirm_dates, dir_ids):
    rows = []
    for dir_id in dir_ids:
        item = epid_data.get(dir_id) or {}
        fields = item.get('master_field_results') or []
        rows.append(
            {
                'lpu': match_master_field(fields, FIELD_LPU),
                'epid_number': item.get('epid_value') or '',
                'master_dir': item.get('master_dir') or '',
                'reg_date': match_master_field(fields, FIELD_REG_DATE) or confirm_dates.get(dir_id, ''),
                'diagnosis': match_master_field(fields, FIELD_DIAGNOSIS),
            }
        )
    return rows


def epid_numbers_base(ws1, d1, d2):
    style_border = NamedStyle(name="style_border_epid_h")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=1, column=1).value = 'Эпид. номера'
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'

    columns = [
        (FIELD_LPU, 45),
        ('Эпид номер', 22),
        ('№ направления (извещение)', 22),
        (FIELD_REG_DATE, 20),
        (FIELD_DIAGNOSIS, 50),
    ]

    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border

    return ws1


def epid_numbers_fill_data(ws1, rows, row=6):
    style_border1 = NamedStyle(name="style_border_epid_d")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    r = row
    for i in rows:
        ws1.cell(row=r, column=1).value = i.get('lpu', '')
        ws1.cell(row=r, column=2).value = i.get('epid_number', '')
        ws1.cell(row=r, column=3).value = i.get('master_dir', '')
        ws1.cell(row=r, column=4).value = i.get('reg_date', '')
        ws1.cell(row=r, column=5).value = i.get('diagnosis', '')
        for c in range(1, 6):
            ws1.cell(row=r, column=c).style = style_border1
        r += 1
    return ws1
