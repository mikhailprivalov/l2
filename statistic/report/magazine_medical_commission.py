from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter

from utils.dates import normalize_date
from utils.db import namedtuplefetchall
from laboratory.settings import TIME_ZONE
from django.db import connection
import simplejson as json


def form_01(ws1, data):
    # Для журнала ВК - ДЛО
    style_border = NamedStyle(name="style_border_ca")
    bd = Side(style="thin", color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

    columns = [
        ('№ п/п.', 5),
        ('Дата экспертизы', 10),
        ('Наименование ЛПУ, фамилия врача, направившего пациента на экспертизу', 25),
        ('Фамилия, имя, отчество пациента', 25),
        ('Дата рождения', 10),
        ('Характеристика случая экспертизы', 30),
        ('Диагноз основной (МКБ-10)', 8),
        ('№ ф.003/у (№ случая")', 15),
        ('Обоснование заключения. Заключение экспертов, рекомендации', 30),
        ('Основной состав экспертов', 25),
        ("Подписи экспертов", 15),
    ]
    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border

    result = []
    previous_direction_number = None
    tmp_string = {}
    step = 0
    sql_data = sql_01(data['research_id'], data['d_s'], data['d_e'])
    for i in sql_data:
        if i.direction_number != previous_direction_number and step != 0:
            result.append(tmp_string.copy())
            tmp_string = {}
        tmp_string.update({i.field_title: i.field_value, "history_num": i.parent_direction, "doc_fio": i.doc_fio, "med_exam": i.medical_examination})
        previous_direction_number = i.direction_number
        step += 1
    result.append(tmp_string.copy())
    step = 0
    previous_date = None

    for i in result:
        row += 1
        if step != 0 and previous_date != i.get("medical_examination"):
            step = 0
        step += 1
        ws1.cell(row=row, column=1).value = step
        ws1.cell(row=row, column=2).value = normalize_date(i.get("Дата"))
        ws1.cell(row=row, column=3).value = i.get("doc_fio")
        ws1.cell(row=row, column=4).value = i.get("ФИО пациента")
        ws1.cell(row=row, column=5).value = i.get("Дата рождения пациента")
        subject = json.loads(i.get("Предмет рассмотрения"))
        subject_result = ""
        k = 0
        for s in subject.get("rows"):
            if k != 0:
                subject_result = f"{subject_result} \n {s[0]}"
            else:
                subject_result = s[0]
            k += 1
        ws1.cell(row=row, column=6).value = subject_result
        diag = json.loads(i.get("Диагноз основной по МКБ-10"))

        ws1.cell(row=row, column=7).value = diag.get("code")
        ws1.cell(row=row, column=8).value = i.get("history_num")
        ws1.cell(row=row, column=9).value = "-"
        ws1.cell(row=row, column=10).value = "-"


        previous_date = i.get("medical_examination")
        print(i)

    return ws1


def sql_01(research_id, d_s, d_e):
        """
        Для журнала ВК-ДЛО
        :return:
        """
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    directions_paraclinicresult.issledovaniye_id,
                    directions_paraclinicresult.field_id,
                    directions_paraclinicresult.value as field_value,
                    directory_paraclinicinputfield.title as field_title,
                    directions_issledovaniya.napravleniye_id as direction_number,
                    users_doctorprofile.fio as doc_fio,
                    di.napravleniye_id as parent_direction,
                    directions_issledovaniya.medical_examination
                    FROM public.directions_paraclinicresult
                    LEFT JOIN directions_issledovaniya ON directions_issledovaniya.id = directions_paraclinicresult.issledovaniye_id
                    LEFT JOIN directory_paraclinicinputfield ON directory_paraclinicinputfield.id = directions_paraclinicresult.field_id
                    LEFT JOIN directions_napravleniya ON directions_napravleniya.id = directions_issledovaniya.napravleniye_id
                    LEFT JOIN users_doctorprofile ON directions_issledovaniya.doc_confirmation_id=users_doctorprofile.id
                    LEFT JOIN directions_issledovaniya di ON di.id = directions_napravleniya.parent_id
                    WHERE 
                      directions_issledovaniya.research_id=%(research_id)s
                      AND directory_paraclinicinputfield.for_talon = true
                      AND directions_issledovaniya.time_confirmation IS NOT NULL
                      AND directions_issledovaniya.medical_examination AT TIME ZONE %(tz)s BETWEEN %(d_start)s AND %(d_end)s      
                    order by directions_issledovaniya.medical_examination, directions_issledovaniya.napravleniye_id
                """,
                params={'research_id': research_id, 'd_start': d_s, 'd_end': d_e, 'tz': TIME_ZONE},
            )

            rows = namedtuplefetchall(cursor)
        return rows
