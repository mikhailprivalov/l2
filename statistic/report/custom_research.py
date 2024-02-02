from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter
import json


def custom_research_data(query_sql):
    result = []
    prev_direction = None
    step = 0
    tmp_result = {}
    custom_fields = []
    for i in query_sql:
        if prev_direction != i.direction_number and step != 0:
            result.append(tmp_result.copy())
        if prev_direction != i.direction_number:
            parent = f"#{i.parent}" if i.parent else ""
            tmp_result = {
                "Направление": f"{i.direction_number} - {parent}",
                "Источник": i.fin_source,
                "Пациент": i.patient_fio,
                "Пол": i.patient_sex,
                "Дата рождения": i.patient_birthday,
                "Возраст": i.patient_age,
                "Адрес": i.patient_main_address,
                "Исполнитель": i.doc_fio,
                "Код врача": i.personal_code,
            }
        tmp_result[i.field_title] = i.field_value
        if i.field_title not in custom_fields:
            custom_fields.append(i.field_title)
        step += 1
        prev_direction = i.direction_number
    result.append(tmp_result.copy())
    fields = ["Направление", "Источник", "Пациент", "Пол", "Дата рождения", "Возраст", "Адрес", "Исполнитель", "Код врача"]
    fields.extend(custom_fields)
    return {"result": result, "custom_fields": custom_fields, "fields": fields}


def custom_research_base(ws1, d1, d2, result_query, research_title):
    style_border = NamedStyle(name="style_border_ca")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=1, column=1).value = 'Услуга:'
    ws1.cell(row=1, column=2).value = research_title
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'

    columns = [
        ('Направление', 15),
        ('Пациент', 45),
        ('Пол', 10),
        ('Дата рождения', 26),
        ('Возраст', 10),
        ('Адрес', 40),
        ('Исполнитель', 35),
        ('Код врача', 15),
    ]

    columns2 = [(i, 25) for i in result_query["custom_fields"]]
    columns.extend(columns2)
    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border

    return ws1


def custom_research_fill_data(ws1, result_query, row=6):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    r = row
    len_fields = len(result_query["fields"])
    for i in result_query["result"]:
        column = 0
        table_data = {}
        r += 1
        for k in range(len_fields):
            column += 1
            title = result_query["fields"][k]
            try:
                data = i.get(title, '')
            except:
                continue
            is_dict = False
            if type(data) is str and "columns" in data:
                try:
                    table_data = json.loads(data)
                    is_dict = True
                except:
                    is_dict = False
            if not is_dict:
                ws1.cell(row=r, column=column).value = i.get(title, '')
                ws1.cell(row=r, column=column).style = style_border1
            else:
                for i in table_data.get("rows"):
                    json_col = column
                    for col in i:
                        ws1.cell(row=r, column=json_col).value = col
                        json_col += 1
                    r += 1
    return ws1


def custom_monitoring_research_data(query_sql):
    result = []
    prev_direction = None
    step = 0
    tmp_result = {}
    custom_fields = []
    for i in query_sql:
        if prev_direction != i.direction_number and step != 0:
            result.append(tmp_result.copy())
        if prev_direction != i.direction_number:
            tmp_result = {
                "Номер": f"{i.direction_number}",
                "МО": f"{i.hosp_title}",
            }
        tmp_result[i.field_title] = i.field_value
        if i.field_title not in custom_fields:
            custom_fields.append(i.field_title)
        step += 1
        prev_direction = i.direction_number
    result.append(tmp_result.copy())
    fields = ["Номер", "МО"]
    fields.extend(custom_fields)
    return {"result": result, "custom_fields": custom_fields, "fields": fields}


def custom_monitorimg_research_base(ws1, d1, d2, result_query, research_title):
    style_border = NamedStyle(name="style_border_ca")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=1, column=1).value = 'Услуга:'
    ws1.cell(row=1, column=2).value = research_title
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'

    columns = [
        ('номер', 15),
        ('МО', 50),
    ]

    columns2 = [(i, 25) for i in result_query["custom_fields"]]
    columns.extend(columns2)
    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border

    return ws1