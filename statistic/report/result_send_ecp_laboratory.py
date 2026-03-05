from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter

from directory.models import Researches
from utils.db import namedtuplefetchall
from laboratory.settings import TIME_ZONE, USE_RMIS_NUMBER_IN_REVISE_REPORT_ECP_SEND
from django.db import connection


def form_01(ws1, data):
    # Для журнала ВК - ДЛО
    style_border = NamedStyle(name="style_border_ca")
    bd = Side(style="thin", color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=12)
    style_border.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    style_border2 = NamedStyle(name="style_border2")
    style_border2.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border2.font = Font(bold=False, size=11)
    style_border2.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    columns = [
        ('№ п.п.', 10),
        ('Врач', 30),
        ('Дата подтверждения', 15),
        ('Дата создания', 15),
        ('Пациент', 30),
        ('Дата рождения', 15),
        ('Услуги', 45),
        ('Успех', 10),
        ('Номер L2', 25),
        ('Служебный ИД', 25),
        ('Случай', 15),
        ('Сообщение', 15),
        ('Участок', 25),
        ('Устаревший способ', 15),
    ]
    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border

    researches = list(Researches.objects.filter(podrazdeleniye__p_type=data['type_podr']).values_list("pk", flat=True))

    sql_data = sql_01(tuple(researches), data['d_s'], data['d_e'], USE_RMIS_NUMBER_IN_REVISE_REPORT_ECP_SEND)
    result_sql = {}
    for i in sql_data:
        if not result_sql.get(i.direction_number):
            result_sql[i.direction_number] = {
                "doctor": f"{i.doc_family} {i.doc_name} {i.doc_patronymic}",
                "date_confirm": i.date_confirm,
                "date_create": i.date_create,
                "patient": f"{i.patient_family} {i.patient_name} {i.patient_patronymic}",
                "patient_birthday": i.patient_birthday,
                "status": "Да" if i.result_rmis_send else "Нет",
                "rmis_id": i.rmis_number,
                "case_num": i.rmis_case_number,
                "message": i.amd_message,
                "service_title": [i.research_title],
                "disttict": i.ditrict_title,
                "older_yet_send": i.yet_send_services
            }
        else:
            result_sql[i.direction_number]["service_title"].append(i.research_title)

    step = 0
    for k, i in result_sql.items():
        row += 1
        step += 1
        ws1.cell(row=row, column=1).value = step
        ws1.cell(row=row, column=2).value = i.get("doctor")
        ws1.cell(row=row, column=3).value = i.get("date_confirm")
        ws1.cell(row=row, column=4).value = i.get("date_create")
        ws1.cell(row=row, column=5).value = i.get("patient")
        ws1.cell(row=row, column=6).value = i.get("patient_birthday")
        ws1.cell(row=row, column=7).value = ", ".join(i.get("service_title"))
        ws1.cell(row=row, column=8).value = i.get("status")
        ws1.cell(row=row, column=9).value = k
        ws1.cell(row=row, column=10).value = i.get("rmis_id")
        ws1.cell(row=row, column=11).value = i.get("case_num")
        ws1.cell(row=row, column=12).value = i.get("message")
        ws1.cell(row=row, column=13).value = i.get("district")
        ws1.cell(row=row, column=14).value = i.get("older_yet_send")
        for k in range(12):
            ws1.cell(row=row, column=k + 1).style = style_border
    return ws1


def sql_01(research_id, d_s, d_e, use_rmis_number):
    """
    Для журнала ВК-ДЛО
    :return:
    """
    with connection.cursor() as cursor:
        cursor.execute(
            """
                SELECT
                    ud.family as doc_family,
                    ud.name as doc_name,
                    ud.patronymic as doc_patronymic,
                    ci.family as patient_family,
                    ci.name as patient_name,
                    ci.patronymic as patient_patronymic,
                    to_char(ci.birthday, 'DD.MM.YYYY') as patient_birthday,
                    to_char(dn.last_confirmed_at AT TIME ZONE %(tz)s, 'DD.MM.YYYY') as date_confirm,
                    to_char(dn.data_sozdaniya AT TIME ZONE %(tz)s, 'DD.MM.YYYY') as date_create,
                    
                    directions_issledovaniya.napravleniye_id as direction_number,
                    dn.rmis_number,
                    dn.result_rmis_send,
                    dn.rmis_case_number,
                    dn.amd_message,
                    dn.rmis_resend_services as yet_send_services,
                    dr.title as research_title,
                    cd.title as ditrict_title
                    
                    FROM directions_issledovaniya
                    LEFT JOIN directions_napravleniya dn ON dn.id = directions_issledovaniya.napravleniye_id
                    LEFT JOIN users_doctorprofile ud ON directions_issledovaniya.doc_confirmation_id=ud.id
                    LEFT JOIN clients_card cc ON cc.id=dn.client_id
                    LEFT JOIN clients_individual ci ON cc.individual_id=ci.id
                    LEFT JOIN directory_researches dr ON directions_issledovaniya.research_id=dr.id
                    LEFT JOIN clients_district cd ON cc.district_id = cd.id
                    WHERE 
                      directions_issledovaniya.research_id in %(research_id)s 
                      AND
                      dn.total_confirmed = true
                      AND
                      dn.last_confirmed_at AT TIME ZONE %(tz)s BETWEEN %(d_start)s AND %(d_end)s                                               
                    order by dn.last_confirmed_at, directions_issledovaniya.napravleniye_id
                """,
            params={'research_id': research_id, 'd_start': d_s, 'd_end': d_e, 'tz': TIME_ZONE, 'use_rmis_number': use_rmis_number},
        )

        rows = namedtuplefetchall(cursor)
    return rows
