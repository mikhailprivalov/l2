from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter


def fill_base(ws1, header_data):
    style_border = NamedStyle(name="style_border_ca5")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    columns = [
        ('Услуга', 50),
        ('Количество', 18),
        ('Цена', 18),
        ('Сумма', 18),
    ]
    # columns = [(i, 30) for i in header_data[0].keys()]

    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]

    return ws1


def fill_default_base(ws1, header_data):
    style_border = NamedStyle(name="style_border_ca5")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    columns = [(i, 20) for i in header_data.values()]

    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]

    return ws1


def fill_data(ws1, value_data):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    r = 5
    for i in value_data:
        r += 1
        step_col = 0
        for v in i.values():
            step_col += 1
            ws1.cell(row=r, column=step_col).value = v
    return ws1


def fill_xls_check_research_exam_data(ws1, value_data):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    r = 5
    for row in value_data.values():
        r += 1
        ws1.cell(row=r, column=1).value = row.get('card_num', '-')
        ws1.cell(row=r, column=2).value = row.get('district', '-')
        fio = row.get('fio', '-').split(' ')
        ws1.cell(row=r, column=3).value = fio[0]
        ws1.cell(row=r, column=4).value = fio[1]
        ws1.cell(row=r, column=5).value = fio[2]
        ws1.cell(row=r, column=6).value = "-"
        ws1.cell(row=r, column=7).value = "-"
        ws1.cell(row=r, column=8).value = row.get('snils', '-')
        researches = row.get('researches')
        column = 8
        for v in researches.values():
            column += 1
            ws1.cell(row=r, column=column).value = v
    return ws1
