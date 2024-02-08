import json
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils.cell import get_column_letter

from directions.models import IstochnikiFinansirovaniya
from doctor_call.models import DoctorCall
from hospitals.tfoms_hospital import HOSPITAL_TITLE_BY_CODE_TFOMS
from statistic.sql_func import get_pair_iss_direction
from utils.dates import normalize_dash_date, normalize_date
from dateutil.parser import parse as du_parse
from dateutil.relativedelta import relativedelta

month_dict = {1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель', 5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август', 9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'}


def job_total_base(ws1, month, type_fin):
    """
    Основа(каркас) для итоговых данных
    :return:
    """
    ws1.column_dimensions[get_column_letter(1)].width = 22
    for i in range(1, 32):
        ws1.column_dimensions[get_column_letter(1 + i)].width = 4
        ws1.cell(row=4, column=1 + i).value = str(i)

    ws1.cell(row=1, column=1).value = 'Месяц'
    ws1.cell(row=1, column=2).value = month_dict.get(month)
    ws1.cell(row=4, column=1).value = 'Вид работы'
    fin_obj = IstochnikiFinansirovaniya.objects.get(pk=type_fin)
    ws1.cell(row=2, column=1).value = fin_obj.title

    return ws1


def jot_total_titles(ws1, titles):
    """
    Заговловки видов работ
    :param ws1:
    :param titles:
    :return:
    """
    cel_res = OrderedDict()
    for i in range(len(titles)):
        cell_row = 5 + i
        ws1.cell(row=cell_row, column=1).value = titles[i]
        cel_res[titles[i]] = cell_row

    return ws1, cel_res


def job_total_data(ws1, titles, data):
    for k, v in data.items():
        for res, uet in v.items():
            r = titles.get(res)
            ws1.cell(row=r, column=k + 1).value = str(uet)


def passed_research_base(ws1, data_date):
    """
    :param ws1:
    :return:
    """
    style_border = NamedStyle(name="style_border")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=19)
    ws1.cell(row=1, column=1).value = 'ЖУРНАЛ учета приема и отказов в госпитализации за ' + data_date + 'г.(мед.документация Ф№001/У утв. МИНЗДРАВОМ СССР 04.10.1980г. №1030)'
    ws1.cell(row=1, column=1).style = style_border

    # габариты ячеек
    ws1.row_dimensions[2].height = 115
    columns = [
        ('№ п/п', 5),
        ('Время поступления', 8),
        ('Услуга (дата-время подтверждения)', 14),
        ('Направление', 11),
        ('Фамилия, имя, отчество больного', 20),
        ('Дата рождения', 10),
        ('Постоянное место жительства или адрес родственников, близких и N телефона', 23),
        ('Каким учреждением был направлен или доставлен', 15),
        ('Отделение, в которое помещен больной', 12),
        ('N карты (стационарного) больного', 10),
        ('Диагноз направившего учреждения', 7),
        ('Диагноз при поступлении', 7),
        ('№ ДДУ', 16),
        ('Полис', 21),
        ('Примечания', 10),
        ('Выписан, переведен в другой стационар, умер (вписать и указать дату и название стационара, куда переведен', 20),
        ('Отметка о сообщении родственникам или учреждению', 11),
        ('Если не был госпитализирован указать причину и принятые меры', 11),
        ('отказ в приеме первичный, повторный (вписать)', 11),
    ]

    for idx, column in enumerate(columns, 1):
        ws1.cell(row=2, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=2, column=idx).style = style_border

    return ws1


def passed_research_data(ws1, data):
    r = 2
    n = 0
    empty = ' '

    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

    for i in data:
        current_research_title = i[1]
        current_polis_n = i[2] or empty
        current_polis_who_give = i[3] or empty
        current_napravlen = i[4]
        current_datatime_confirm = i[5]
        current_create_napr = i[6]
        current_diagnoz = i[7] or empty
        current_result = i[8] or empty
        current_napr_time_at = i[19] or empty
        current_num_card = i[10]
        current_family = i[11] or empty
        current_name = i[12] or empty
        current_patronymic = i[13] or empty
        current_birthday = i[14] or empty
        current_main_address = i[15] if i[15] else ''
        current_fact_address = i[16] if i[16] else empty
        current_address = current_main_address if current_main_address else current_fact_address
        current_work_place = i[17] or empty
        current_kem_napravlen = i[18] or empty
        r = r + 1
        n = n + 1
        ws1.cell(row=r, column=1).value = n
        ws1.cell(row=r, column=2).value = current_napr_time_at
        ws1.cell(row=r, column=3).value = f'{current_research_title},\n({current_datatime_confirm})'
        ws1.cell(row=r, column=4).value = f'{current_napravlen},\n({current_create_napr})'
        ws1.cell(row=r, column=5).value = current_family + ' ' + current_name + ' ' + current_patronymic
        ws1.cell(row=r, column=6).value = current_birthday
        ws1.cell(row=r, column=7).value = current_address
        ws1.cell(row=r, column=8).value = current_kem_napravlen
        ws1.cell(row=r, column=9).value = 'Приемное'
        ws1.cell(row=r, column=10).value = current_num_card
        ws1.cell(row=r, column=11).value = ' '
        ws1.cell(row=r, column=12).value = current_diagnoz
        ws1.cell(row=r, column=13).value = current_work_place
        ws1.cell(row=r, column=14).value = current_polis_n + ', ' + current_polis_who_give
        ws1.cell(row=r, column=15).value = ' '
        ws1.cell(row=r, column=16).value = current_result
        ws1.cell(row=r, column=17).value = ' '
        ws1.cell(row=r, column=18).value = ' '
        ws1.cell(row=r, column=19).value = ' '
        for j in range(1, 20):
            ws1.cell(row=r, column=j).style = style_border1

    return ws1


def covid_call_patient_base(ws1):
    """
    :param ws1:
    :return:
    """
    style_border = NamedStyle(name="style_border")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=19)
    ws1.cell(row=1, column=1).value = 'Обзвон'
    ws1.cell(row=1, column=1).style = style_border

    # габариты ячеек
    ws1.row_dimensions[2].height = 15
    columns = [
        ('ФИО', 25),
        ('№ карты', 15),
        ('Телефон', 20),
        ('Оператор', 25),
        ('Дата', 25),
    ]

    for idx, column in enumerate(columns, 1):
        ws1.cell(row=2, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=2, column=idx).style = style_border

    return ws1


def covid_call_patient_data(ws1, data):
    r = 3

    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

    for i in data:
        ws1.cell(row=r, column=1).value = i["fio_patient"]
        ws1.cell(row=r, column=2).value = i["number"]
        ws1.cell(row=r, column=3).value = i["Контактный телефон"]
        ws1.cell(row=r, column=4).value = i["Оператор"]
        ws1.cell(row=r, column=5).value = normalize_dash_date(i["Дата следующего звонка"])

        for j in range(1, 6):
            ws1.cell(row=r, column=j).style = style_border1

    return ws1


def covid_swab_base(ws1):
    """
    :param ws1:
    :return:
    """
    style_border = NamedStyle(name="style_border")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=19)
    ws1.cell(row=1, column=1).value = 'Повторный мазок'
    ws1.cell(row=1, column=1).style = style_border

    # габариты ячеек
    ws1.row_dimensions[2].height = 15
    columns = [
        ('ФИО', 25),
        ('№ карты', 15),
        ('Телефон', 20),
        ('Оператор', 25),
        ('Дата', 25),
        ('Адрес', 55),
    ]

    for idx, column in enumerate(columns, 1):
        ws1.cell(row=2, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=2, column=idx).style = style_border

    return ws1


def covid_swab_data(ws1, data):
    r = 3

    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

    for i in data:
        ws1.cell(row=r, column=1).value = i["fio_patient"]
        ws1.cell(row=r, column=2).value = i["number"]
        ws1.cell(row=r, column=3).value = i["Контактный телефон"]
        ws1.cell(row=r, column=4).value = i["Оператор"]
        ws1.cell(row=r, column=5).value = normalize_dash_date(i["Сдача повторного мазка на COVID"])
        ws1.cell(row=r, column=6).value = i["Адрес"]

        for j in range(1, 6):
            ws1.cell(row=r, column=j).style = style_border1
        r += 1

    return ws1


def covid_bl_base(ws1):
    """
    :param ws1:
    :return:
    """
    style_border = NamedStyle(name="style_border")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=19)
    ws1.cell(row=1, column=1).value = 'Продолжение БЛ'
    ws1.cell(row=1, column=1).style = style_border

    # габариты ячеек
    ws1.row_dimensions[2].height = 15
    columns = [
        ('ФИО', 25),
        ('№ карты', 15),
        ('Телефон', 20),
        ('Оператор', 25),
        ('Дата', 25),
    ]

    for idx, column in enumerate(columns, 1):
        ws1.cell(row=2, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=2, column=idx).style = style_border

    return ws1


def covid_bl_data(ws1, data):
    r = 3

    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

    for i in data:
        ws1.cell(row=r, column=1).value = i["fio_patient"]
        ws1.cell(row=r, column=2).value = i["number"]
        ws1.cell(row=r, column=3).value = i["Контактный телефон"]
        ws1.cell(row=r, column=4).value = i["Оператор"]
        ws1.cell(row=r, column=5).value = normalize_dash_date(i["Продолжение БЛ"])
        for j in range(1, 6):
            ws1.cell(row=r, column=j).style = style_border1

    return ws1


def onco_base(ws1, d_s, d_e):
    """
    :param ws1:
    :return:
    """
    style_border = NamedStyle(name="style_border")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=13)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws1.cell(row=1, column=1).value = f'ЖУРНАЛ учета онкоподозрения c {d_s} по {d_e}'
    ws1.cell(row=1, column=1).style = style_border

    # габариты ячеек
    # ws1.row_dimensions[2].height = 85
    columns = [('№ п/п', 5), ('ФИО пациента', 30), ('Дата рождения', 15), ('N карты', 15), ('Врач поставил', 30), ('Дата постановки', 20), ('Врач снял', 30), ('Дата снятия', 20)]
    ws1.row_dimensions[2].height = 15
    ws1.cell(row=2, column=1).value = ''
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=3, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=3, column=idx).style = style_border

    return ws1


def passed_onco_data(ws1, data):
    r = 3
    n = 0
    empty = ' '

    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=12)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

    for i in data:
        current_patient = i[0] or empty
        current_birhday = i[1] or empty
        current_num_card = i[2] or empty
        current_doc_start = i[3] or empty
        current_date_start = i[4] or empty
        current_doc_end = i[5] or empty
        current_date_end = i[6] or empty
        r = r + 1
        n = n + 1
        ws1.cell(row=r, column=1).value = n
        ws1.cell(row=r, column=2).value = current_patient
        ws1.cell(row=r, column=3).value = current_birhday
        ws1.cell(row=r, column=4).value = current_num_card
        ws1.cell(row=r, column=5).value = current_doc_start
        ws1.cell(row=r, column=6).value = current_date_start
        ws1.cell(row=r, column=7).value = current_doc_end
        ws1.cell(row=r, column=8).value = current_date_end
        for j in range(1, 9):
            ws1.cell(row=r, column=j).style = style_border1

    return ws1


def style_sheet():
    style_border = NamedStyle(name="style_border")
    bd = Side(style='thin', color="000000")
    border = Border(left=bd, top=bd, right=bd, bottom=bd)

    style_border.border = border
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True)

    style_border1 = NamedStyle(name="style_border1")
    style_border1.border = border
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True)

    style_o = NamedStyle(name="style_o")
    style_o.font = Font(bold=True, size=11)

    style_border_res = NamedStyle(name="style_border_res")
    bd = Side(style='thin', color="000000")
    style_border_res.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border_res.font = Font(bold=False, size=11)
    style_border_res.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    return (style_border, style_o, style_border1, style_border_res)


def statistics_tickets_base(ws1, i_obj, type_fin, d1, d2, style_border, style_o):
    """
    Назначить ширину колонок. Вход worksheet выход worksheen с размерами
    Заголовки данных
    """

    columns = [
        ('Дата', 13),
        ('Кол-во', 7),
        ('Услуга', 15),
        ('Соисполнитель', 9),
        ('ФИО пациента,\n№ направления', 31),
        ('Дата рождения', 13),
        ('№ карты', 12),
        ('Данные полиса', 27),
        ('Код услуги', 16),
        ('Услуга \n (ует/мин)', 12),
        ('Время \n подтверждения', 18),
        ('Онкоподозрение', 13),
        ('Первичный прием', 12),
        ('Цель \n посещения\n(код)е', 13),
        ('Диагноз \n МКБ', 13),
        ('Впервые', 13),
        ('Результат \n обращения \n(код)', 13),
        ('Исход(код)', 13),
        ('', 13),
        ('Источник по направлению', 13),
        ('Источник по услуге', 13),
        ('Главное направление', 13),
        ('Категория по направлению', 14),
        ('Категория по исследованию', 14),
    ]
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=7, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=7, column=idx).style = style_border

    # Закголовки столбцов
    ws1.cell(row=1, column=1).value = 'Сотрудник'
    ws1.cell(row=1, column=1).style = style_o
    ws1.cell(row=1, column=2).value = i_obj.fio
    ws1.cell(row=2, column=1).value = 'Должность'
    ws1.cell(row=2, column=1).style = style_o
    ws1.cell(row=2, column=2).value = i_obj.specialities.title if i_obj.specialities else ""
    ws1.cell(row=4, column=1).value = 'Период:'
    ws1.cell(row=4, column=1).style = style_o
    ws1.cell(row=5, column=1).value = d1
    ws1.cell(row=5, column=2).value = 'по'
    ws1.cell(row=5, column=3).value = d2
    ws1.cell(row=1, column=5).value = 'Код врача'
    ws1.cell(row=1, column=5).style = style_o
    ws1.cell(row=1, column=6).value = i_obj.personal_code
    ws1.cell(row=3, column=5).value = 'Источник'
    ws1.cell(row=3, column=5).style = style_o
    fin_obj = IstochnikiFinansirovaniya.objects.get(pk=type_fin)
    ws1.cell(row=3, column=6).value = fin_obj.title

    return ws1


def statistics_tickets_data(ws1, issl_obj, i_obj, style_border1):
    # i_obj - обеъект доктор

    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', start_color='a9d094', end_color='a9d094')
    total_fill = openpyxl.styles.fills.PatternFill(patternType='solid', start_color='ffcc66', end_color='ffcc66')

    r = 7
    r1 = r + 1
    total_sum = []
    # one_days = timedelta(1)
    current_date = ''
    parent_iss_pk = [k[28] for k in issl_obj if k[28]]
    parent_iss = tuple(set(parent_iss_pk))
    result = {}
    if len(parent_iss) > 0:
        for k in get_pair_iss_direction(parent_iss):
            result[k.iss_pk] = k.direction_pk

    for issled in issl_obj:
        # Порядок колонок в issled:
        # title, code, is_first_reception, polis_n, polis_who_give, \
        # first_time, napravleniye_id, doc_confirmation_id, def_uet, co_executor_id, \
        # co_executor_uet, co_executor2_id, co_executor2_uet, datetime_confirm, date_confirm, \
        # time_confirm, maybe_onco, purpose, diagnos, iss_result, \
        # outcome, card_number, client_family, client_name, client_patronymic, \
        # birthday
        empty = ' '
        # current_datetime_confirm = issled[13]
        current_date = issled[14]
        # current_count = 1
        current_research_title = issled[0]
        f = issled[22] or empty
        n = issled[23] or empty
        p = issled[24] or empty
        current_napr = str(issled[6])
        current_patient_napr = f'{f} {n} {p}; {current_napr}'
        current_born = issled[25]
        current_card = issled[21]
        polis_n = issled[3] or ''
        polis_who = issled[4] or ''
        current_polis = f'{polis_n};\n{polis_who}'
        current_code_reserch = issled[1]
        current_doc_conf = issled[7]
        current_def_uet = issled[8] or 0
        current_co_exec1 = issled[9]
        current_uet1 = issled[10] or 0
        current_co_exec2 = issled[11]
        current_uet2 = issled[12] or 0
        current_time_confirm = issled[15]
        current_isfirst = issled[2]
        current_onko = issled[16]
        current_purpose = issled[17]
        current_diagnos = issled[18]
        current_firsttime = issled[5]
        current_result = issled[19]
        current_octome = issled[20]
        direction_fin_source = issled[26]
        iss_fin_source = issled[27]
        parent_direction = " -"
        if issled[28]:
            parent_direction = result.get(issled[28], "-")

        direction_price_category = issled[29]
        iss_price_category = issled[30]

        # current_price = ''

        if r != 7 and r != 8:
            befor_date = ws1.cell(row=r, column=1).value
            if current_date != befor_date and not (ws1.cell(row=r, column=1).value).istitle():
                r = r + 1
                ws1.cell(row=r, column=1).value = 'Итого за ' + befor_date[:2]
                ws1.cell(row=r, column=2).value = f'=SUM(B{r1}:B{r - 1})'
                ws1.cell(row=r, column=10).value = f'=SUM(J{r1}:J{r - 1})'

                total_sum.append(r)
                ws1.row_dimensions.group(r1, r - 1, hidden=True)
                rows = ws1[f'A{r}:V{r}']
                for row in rows:
                    for cell in row:
                        cell.fill = my_fill
                r1 = r + 1

        r = r + 1
        ws1.cell(row=r, column=1).value = current_date
        ws1.cell(row=r, column=2).value = 1
        ws1.cell(row=r, column=3).value = current_research_title
        sum_uet = 0
        co_exec = ''
        if (current_doc_conf == i_obj.pk) and (current_co_exec1 == i_obj.pk):
            sum_uet = sum_uet + current_def_uet
            co_exec = co_exec + 'ОСН'

        if (current_doc_conf == i_obj.pk) and (current_co_exec1 != i_obj.pk):
            sum_uet = sum_uet + current_def_uet
            co_exec = co_exec + 'ОСН'

        if (current_doc_conf != i_obj.pk) and (current_co_exec1 == i_obj.pk):
            sum_uet = sum_uet + current_uet1
            co_exec = co_exec + 'СО-1'

        if current_co_exec2 == i_obj.pk:
            sum_uet = sum_uet + current_uet2
            co_exec = co_exec + ', СО-2'
        ws1.cell(row=r, column=4).value = co_exec
        ws1.cell(row=r, column=5).value = current_patient_napr
        ws1.cell(row=r, column=6).value = current_born
        ws1.cell(row=r, column=7).value = current_card

        ws1.cell(row=r, column=8).value = current_polis
        ws1.cell(row=r, column=9).value = current_code_reserch
        ws1.cell(row=r, column=10).value = str(sum_uet)
        ws1.cell(row=r, column=11).value = current_time_confirm

        ws1.cell(row=r, column=12).value = current_onko
        ws1.cell(row=r, column=13).value = current_isfirst
        ws1.cell(row=r, column=14).value = current_purpose
        ws1.cell(row=r, column=15).value = current_diagnos
        ws1.cell(row=r, column=16).value = current_firsttime
        ws1.cell(row=r, column=17).value = current_result
        ws1.cell(row=r, column=18).value = current_octome
        ws1.cell(row=r, column=19).value = ''
        ws1.cell(row=r, column=20).value = direction_fin_source
        ws1.cell(row=r, column=21).value = iss_fin_source
        ws1.cell(row=r, column=22).value = parent_direction
        ws1.cell(row=r, column=23).value = direction_price_category
        ws1.cell(row=r, column=24).value = iss_price_category

        rows = ws1[f'A{r}:X{r}']
        for row in rows:
            for cell in row:
                cell.style = style_border1

    r = r + 1
    ws1.cell(row=r, column=1).value = 'Итого за ' + current_date[:2]
    ws1.cell(row=r, column=2).value = f'=SUM(B{r1}:B{r - 1})'
    ws1.cell(row=r, column=10).value = f'=SUM(J{r1}:J{r - 1})'
    ws1.row_dimensions.group(r1, r - 1, hidden=True)
    total_sum.append(r)
    rows = ws1[f'A{r}:X{r}']
    for row in rows:
        for cell in row:
            cell.fill = my_fill

    t_s = '=SUM('
    t_s_uet = '=SUM('
    for ts in total_sum:
        t_uet = ts
        t_s = t_s + f'(B{ts})' + ','
        t_s_uet = t_s_uet + f'(J{t_uet})' + ','
    t_s = t_s + ')'
    t_s_uet = t_s_uet + ')'
    r = r + 1
    ws1.cell(row=r, column=1).value = 'Итого Всего'
    ws1.cell(row=r, column=2).value = t_s
    ws1.cell(row=r, column=10).value = t_s_uet
    rows = ws1[f'A{r}:X{r}']
    for row in rows:
        for cell in row:
            cell.fill = total_fill

    return ws1


def inderect_job_base(ws1, doc_obj, d1, d2):
    pink_fill = openpyxl.styles.fills.PatternFill(patternType='solid', start_color='FCD5B4', end_color='FCD5B4')
    rows = ws1[f'A{1}:V{1}']
    for row in rows:
        for cell in row:
            cell.fill = pink_fill

    ws1.column_dimensions[get_column_letter(1)].width = 15
    ws1.column_dimensions[get_column_letter(2)].width = 30
    ws1.column_dimensions[get_column_letter(3)].width = 15

    ws1.cell(row=1, column=1).value = "Косвенные услуги"
    ws1.cell(row=2, column=1).value = "Сотрудник"
    ws1.cell(row=2, column=2).value = doc_obj.fio
    ws1.cell(row=3, column=1).value = f'c {d1}'
    ws1.cell(row=3, column=2).value = f'по {d2}'

    return ws1


def inderect_job_data(ws1, indirect_job):
    r = 4
    for k, v in indirect_job.items():
        for k_job, v_job in v.items():
            r = r + 1
            ws1.cell(row=r, column=1).value = k
            ws1.cell(row=r, column=2).value = k_job
            ws1.cell(row=r, column=3).value = v_job

    return ws1


def statistic_research_base(ws1, d1, d2, research_titile):
    style_border = NamedStyle(name="style_border")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=1, column=1).value = 'Услуга:'
    ws1.cell(row=1, column=2).value = research_titile
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'

    columns = [
        ('Исполнитель', 26),
        ('Направление, за дату', 15),
        ('Дата подтверждения', 16.5),
        ('Время подтверждения', 16.5),
        ('Источник', 10),
        ('Цена', 10),
        ('Кол-во', 7),
        ('Скидка', 7.5),
        ('Сумма', 14),
        ('Физлицо', 26),
        ('Дата рождения', 12),
        ('Возраст', 8),
        ('Карта', 15),
        ('МО', 30),
        ('Цель', 14),
        ('Код(Вич)', 14),
        ('Категория по направлению', 14),
        ('Категория по исследованию', 14),
    ]
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=4, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=4, column=idx).style = style_border

    return ws1


def statistic_research_data(ws1, researches):
    """
    res - результат выборки SQL
    порядок возврата:
    napr, date_confirm, time_confirm, create_date_napr, create_time_napr,
    doc_fio, coast, discount, how_many, ((coast + (coast/100 * discount)) * how_many)::NUMERIC(10,2) AS sum_money,
    ist_f, time_confirmation, num_card, ind_family, ind_name,
    patronymic, birthday, date_born, to_char(EXTRACT(YEAR from age(time_confirmation, date_born)), '999') as ind_age
    :return:
    """
    style_border_res = NamedStyle(name="style_border_res")
    bd = Side(style='thin', color="000000")
    style_border_res.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border_res.font = Font(bold=False, size=11)
    style_border_res.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    r = 4
    for res in researches:
        r += 1
        current_doc = res[5]
        current_napr = res[0]
        current_napr_atcreate = res[3]
        current_date_confirm = res[1]
        current_time_confirm = res[2]
        current_ist_f = res[10]
        current_coast = res[6]
        current_how_many = res[8]
        current_discount = res[7]
        current_price_total = res[9]
        current_ind_fio = f'{res[13]} {res[14]} {res[15]}'
        current_born = res[16]
        current_age = res[18]
        current_num_card = res[12]
        current_purpose = res[20]
        vich_code = res[21]
        direction_category_price = res[22]
        iss_category_price = res[23]

        ws1.cell(row=r, column=1).value = current_doc
        ws1.cell(row=r, column=2).value = f'{current_napr}, {current_napr_atcreate}'
        ws1.cell(row=r, column=3).value = current_date_confirm
        ws1.cell(row=r, column=4).value = current_time_confirm
        ws1.cell(row=r, column=5).value = current_ist_f
        ws1.cell(row=r, column=6).value = current_coast
        ws1.cell(row=r, column=7).value = current_how_many
        ws1.cell(row=r, column=8).value = current_discount
        ws1.cell(row=r, column=9).value = current_price_total
        ws1.cell(row=r, column=10).value = current_ind_fio
        ws1.cell(row=r, column=11).value = current_born
        ws1.cell(row=r, column=12).value = current_age
        ws1.cell(row=r, column=13).value = current_num_card
        ws1.cell(row=r, column=14).value = res[19]
        ws1.cell(row=r, column=15).value = current_purpose
        ws1.cell(row=r, column=16).value = vich_code
        ws1.cell(row=r, column=17).value = direction_category_price
        ws1.cell(row=r, column=18).value = iss_category_price

        rows = ws1[f'A{r}:R{r}']
        for row in rows:
            for cell in row:
                cell.style = style_border_res

    return ws1


def statistic_research_death_base(ws1, d1, d2, research_titile):
    style_border = NamedStyle(name="style_border")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=1, column=1).value = 'Услуга:'
    ws1.cell(row=1, column=2).value = research_titile
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'

    columns = [
        ('Серия', 13),
        ('Номер', 15),
        ('Вид МСС', 17),
        ('Медицинская организация выдавшая свидететельство', 18),
        ('Прикрепление пациента', 18),
        ('Участок', 10),
        ('Дата смерти', 11),
        ('Дата рождения', 11),
        ('ФИО умершего пациента', 25),
        ('Пол (м/ж)', 6),
        ('Возраст на дату смерти', 6),
        ('а) болезнь или состояние, непосредст-венно приведшее к смерти', 17),
        ('а) период', 10),
        ('а) Код по МКБ- 10', 9),
        ('б) патологи-ческое состояние, которое привело к болезни или состоянию, непосредст-венно приведшее к смерти', 17),
        ('б) период', 10),
        ('б) Код по МКБ- 10', 9),
        ('в) перво-начальная причина смерти', 17),
        ('в) период', 10),
        ('в) Код по МКБ- 10', 9),
        ('г) внешняя причина при травмах и отравлениях', 17),
        ('г) период', 10),
        ('г) Код по МКБ- 10', 9),
        ('II.Прочие важные состояния способствовавшие смерти', 15),
        ('класс заболевания первоначальной причины смерти', 15),
        ('Место смерти (1/0)', 15),
        ('Название стационара', 15),
        ('ДТП (1/0)', 12),
        ('Материнская смертность (1/0)', 15),
        ('ФИО выдавшего свидетельства', 20),
        ('Тип места смерти', 25),
        ('ОКПО', 16),
        ('ОКАТО', 16),
        ('Экспертиза', 35),
        ('Основания определения', 35),
    ]
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=4, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=4, column=idx).style = style_border

    return ws1


def statistic_reserved_research_death_base(ws1, d1, d2, research_titile):
    style_border = NamedStyle(name="style_border_rz")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=1, column=1).value = 'Услуга:'
    ws1.cell(row=1, column=2).value = research_titile
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'

    columns = [
        ('Медицинская организация', 40),
        ('Номер в резерве', 20),
        ('Дата создания', 22),
        ('ФИО пациента', 35),
        ('Направление', 20),
    ]
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=4, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=4, column=idx).style = style_border

    return ws1


def statistic_research_death_base_card(ws1, d1, d2, research_titile):
    style_border = NamedStyle(name="style_border_ca")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=1, column=1).value = 'Услуга:'
    ws1.cell(row=1, column=2).value = research_titile
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'

    columns = [
        ('Серия', 13),
        ('Номер', 15),
        ('Вид МСС', 17),
        ('Медицинская организация выдавшая свидететельство', 18),
        ('Прикрепление пациента', 18),
        ('Участок', 10),
        ('Дата смерти', 11),
        ('Дата рождения', 11),
        ('ФИО умершего пациента', 25),
        ('Пол (м/ж)', 6),
        ('Возраст на дату смерти', 6),
        ('а) болезнь или состояние, непосредст-венно приведшее к смерти', 17),
        ('а) период', 10),
        ('а) Код по МКБ- 10', 9),
        ('б) патологи-ческое состояние, которое привело к болезни или состоянию, непосредст-венно приведшее к смерти', 17),
        ('б) период', 10),
        ('б) Код по МКБ- 10', 9),
        ('в) перво-начальная причина смерти', 17),
        ('в) период', 10),
        ('в) Код по МКБ- 10', 9),
        ('г) внешняя причина при травмах и отравлениях', 17),
        ('г) период', 10),
        ('г) Код по МКБ- 10', 9),
        ('II.Прочие важные состояния способствовавшие смерти', 15),
        ('класс заболевания первоначальной причины смерти', 15),
        ('Место смерти (1/0)', 15),
        ('Название стационара', 15),
        ('ДТП (1/0)', 12),
        ('Материнская смертность (1/0)', 15),
        ('ФИО выдавшего свидетельства', 20),
        ('Тип места смерти', 25),
        ('ОКПО', 16),
        ('ОКАТО', 16),
        ('Экспертиза', 35),
        ('Основания определения', 35),
    ]
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=4, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=4, column=idx).style = style_border

    return ws1


def statistic_research_death_data(ws1, researches, expertise_final_data):
    """
    :return:
    """
    style_border_res = NamedStyle(name="style_border_res")
    bd = Side(style='thin', color="000000")
    style_border_res.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border_res.font = Font(bold=False, size=11)
    style_border_res.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    r = 4

    for i in researches:
        if not i:
            return ws1
        try:
            type_doc_death = i["Вид медицинского свидетельства о смерти"]["title"]
        except:
            type_doc_death = i.get("Вид медицинского свидетельства о смерти", "")
        if not type_doc_death:
            continue

        r += 1
        ws1.cell(row=r, column=1).value = i.get("Серия", "")
        ws1.cell(row=r, column=2).value = i.get("Номер", "")

        ws1.cell(row=r, column=3).value = type_doc_death
        ws1.cell(row=r, column=4).value = i["hosp_title"]

        mo_attachment, mo_district = "-", "-"
        if i.get("Прикрепление", None):
            attachment_data = i.get("Прикрепление").split("—")
            mo_attachment = HOSPITAL_TITLE_BY_CODE_TFOMS.get(attachment_data[0].strip(), attachment_data[0].strip())
            mo_district = attachment_data[1]

        ws1.cell(row=r, column=5).value = mo_attachment
        ws1.cell(row=r, column=6).value = mo_district
        ws1.cell(row=r, column=7).value = normalize_dash_date(i["Дата смерти"])
        ws1.cell(row=r, column=8).value = i.get("Дата рождения", "-")
        ws1.cell(row=r, column=9).value = i["fio_patient"]
        ws1.cell(row=r, column=10).value = i["sex"]
        d1 = du_parse(i["Дата смерти"])
        try:
            d2 = du_parse(i.get("Дата рождения", ""))
            delta = relativedelta(d1, d2)
            ws1.cell(row=r, column=11).value = delta.years
        except:
            ws1.cell(row=r, column=11).value = "-"
        # а)
        diag_data = get_table_diagnos(i, "а) Болезнь или состояние, непосредственно приведшее к смерти")
        ws1.cell(row=r, column=12).value = f'{diag_data[1]["code"]} {diag_data[1]["title"]}'
        ws1.cell(row=r, column=13).value = diag_data[0]
        ws1.cell(row=r, column=14).value = diag_data[1]["code"]

        # б)
        diag_data = get_table_diagnos(i, "б) патологическое состояние, которое привело к возникновению вышеуказанной причины:")
        ws1.cell(row=r, column=15).value = f'{diag_data[1]["code"]} {diag_data[1]["title"]}'
        ws1.cell(row=r, column=16).value = diag_data[0]
        ws1.cell(row=r, column=17).value = diag_data[1]["code"]

        # в)
        diag_data = get_table_diagnos(i, "в) первоначальная причина смерти:")
        ws1.cell(row=r, column=18).value = f'{diag_data[1]["code"]} {diag_data[1]["title"]}'
        ws1.cell(row=r, column=19).value = diag_data[0]
        ws1.cell(row=r, column=20).value = diag_data[1]["code"]

        # г)
        diag_data = get_table_diagnos(i, "г) внешняя причина при травмах и отравлениях:")
        ws1.cell(row=r, column=21).value = f'{diag_data[1]["code"]} {diag_data[1]["title"]}'
        ws1.cell(row=r, column=22).value = diag_data[0]
        ws1.cell(row=r, column=23).value = diag_data[1]["code"]

        diag_data = get_table_diagnos(i, "II. Прочие важные состояния, способствовавшие смерти, но не связанные с болезнью или патологическим состоянием, приведшим к ней")
        ws1.cell(row=r, column=24).value = f'{diag_data[1]["code"]} {diag_data[1]["title"]} {diag_data[0]}'
        ws1.cell(row=r, column=25).value = ""

        place_death_details = ""
        try:
            place_death_details = json.loads(i["Место смерти"])
            is_dict = True
        except:
            is_dict = False
        if not is_dict:
            try:
                place_death_details = i["Место смерти"].get("address", None)
                is_dict = True
            except:
                is_dict = False
        if not is_dict:
            place_death_details = "-"
        try:
            title_pace_death = place_death_details.get("address")
        except:
            title_pace_death = place_death_details
        ws1.cell(row=r, column=26).value = title_pace_death
        # Название стационара
        ws1.cell(row=r, column=27).value = i.get("МО", "")
        # ДТП

        ws1.cell(row=r, column=28).value = i["ДТП"]
        ws1.cell(row=r, column=29).value = i.get("Беременность", 0)

        if i.get("Заполнил", None):
            who_write = i.get("Заполнил")
        else:
            who_write = ""
        ws1.cell(row=r, column=30).value = who_write
        ws1.cell(row=r, column=31).value = ""
        ws1.cell(row=r, column=32).value = ""
        ws1.cell(row=r, column=33).value = ""
        experise = ""
        if expertise_final_data.get(i.get('issledovaniye_id', ""), ""):
            experise = expertise_final_data.get(i.get('issledovaniye_id', ""), "")
        ws1.cell(row=r, column=34).value = experise
        base_death = i["Основания для определения причины смерти"]
        type_worker = i["Тип медицинского работника"]
        ws1.cell(row=r, column=35).value = base_death.get("title", "-")
        ws1.cell(row=r, column=36).value = type_worker.get("title", "-")

        rows = ws1[f'A{r}:AH{r}']
        for row in rows:
            for cell in row:
                cell.style = style_border_res

    return ws1


def statistic_research_death_data_card(ws1, researches):
    """
    :return:
    """
    style_border_res = NamedStyle(name="style_border_res_ca")
    bd = Side(style='thin', color="000000")
    style_border_res.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border_res.font = Font(bold=False, size=11)
    style_border_res.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    r = 4

    for i in researches:
        if not i:
            return ws1
        try:
            type_doc_death = i["Вид медицинского свидетельства о смерти"]["title"]
        except:
            type_doc_death = i.get("Вид медицинского свидетельства о смерти", "")
        if not type_doc_death:
            continue
        r += 1
        ws1.cell(row=r, column=1).value = i["Серия"]
        ws1.cell(row=r, column=2).value = i["Номер"]

        ws1.cell(row=r, column=3).value = type_doc_death
        ws1.cell(row=r, column=4).value = i["hosp_title"]

        mo_attachment, mo_district = "-", "-"
        if i.get("Прикрепление", None):
            attachment_data = i.get("Прикрепление").split("—")
            mo_attachment = HOSPITAL_TITLE_BY_CODE_TFOMS.get(attachment_data[0].strip(), attachment_data[0].strip())
            mo_district = attachment_data[1]

        ws1.cell(row=r, column=5).value = mo_attachment
        ws1.cell(row=r, column=6).value = mo_district
        ws1.cell(row=r, column=7).value = normalize_dash_date(i["Дата смерти"])
        ws1.cell(row=r, column=8).value = i["Дата рождения"]
        ws1.cell(row=r, column=9).value = i["fio_patient"]
        ws1.cell(row=r, column=10).value = i["sex"]
        d1 = du_parse(i["Дата смерти"])
        try:
            d2 = du_parse(i["Дата рождения"])
            delta = relativedelta(d1, d2)
            ws1.cell(row=r, column=11).value = delta.years
        except:
            ws1.cell(row=r, column=11).value = "-"
        # а)
        diag_data = get_table_diagnos(i, "а) Болезнь или состояние, непосредственно приведшее к смерти")
        ws1.cell(row=r, column=12).value = f'{diag_data[1]["code"]} {diag_data[1]["title"]}'
        ws1.cell(row=r, column=13).value = diag_data[0]
        ws1.cell(row=r, column=14).value = diag_data[1]["code"]

        # б)
        diag_data = get_table_diagnos(i, "б) патологическое состояние, которое привело к возникновению вышеуказанной причины:")
        ws1.cell(row=r, column=15).value = f'{diag_data[1]["code"]} {diag_data[1]["title"]}'
        ws1.cell(row=r, column=16).value = diag_data[0]
        ws1.cell(row=r, column=17).value = diag_data[1]["code"]

        # в)
        diag_data = get_table_diagnos(i, "в) первоначальная причина смерти:")
        ws1.cell(row=r, column=18).value = f'{diag_data[1]["code"]} {diag_data[1]["title"]}'
        ws1.cell(row=r, column=19).value = diag_data[0]
        ws1.cell(row=r, column=20).value = diag_data[1]["code"]

        # г)
        diag_data = get_table_diagnos(i, "г) внешняя причина при травмах и отравлениях:")
        ws1.cell(row=r, column=21).value = f'{diag_data[1]["code"]} {diag_data[1]["title"]}'
        ws1.cell(row=r, column=22).value = diag_data[0]
        ws1.cell(row=r, column=23).value = diag_data[1]["code"]

        diag_data = get_table_diagnos(i, "II. Прочие важные состояния, способствовавшие смерти, но не связанные с болезнью или патологическим состоянием, приведшим к ней")
        ws1.cell(row=r, column=24).value = f'{diag_data[1]["code"]} {diag_data[1]["title"]} {diag_data[0]}'
        ws1.cell(row=r, column=25).value = ""

        place_death_details = ""
        try:
            place_death_details = json.loads(i["Место смерти"])
            is_dict = True
            if is_dict:
                place_death_details = place_death_details.get("address", "-")
        except:
            is_dict = False
        if not is_dict:
            try:
                place_death_details = i["Место смерти"].get("address", None)
                is_dict = True
            except:
                is_dict = False
        if not is_dict:
            place_death_details = "-"

        ws1.cell(row=r, column=26).value = place_death_details
        # Название стационара
        ws1.cell(row=r, column=27).value = i.get("МО", "")
        # ДТП
        ws1.cell(row=r, column=28).value = i["ДТП"]
        ws1.cell(row=r, column=29).value = i.get("Беременность", 0)

        if i.get("Заполнил", None):
            who_write = i.get("Заполнил")
        else:
            who_write = ""
        ws1.cell(row=r, column=30).value = who_write
        try:
            type_where_death = i["Типы мест наступления смерти"]["title"]
        except:
            type_where_death = "-"
        ws1.cell(row=r, column=31).value = type_where_death
        ws1.cell(row=r, column=32).value = i["hosp_okpo"]
        ws1.cell(row=r, column=33).value = i["hosp_okato"]
        base_death = i["Основания для определения причины смерти"]
        type_worker = i["Тип медицинского работника"]
        ws1.cell(row=r, column=35).value = base_death.get("title", "-")
        ws1.cell(row=r, column=36).value = type_worker.get("title", "-")

        rows = ws1[f'A{r}:AG{r}']
        for row in rows:
            for cell in row:
                cell.style = style_border_res

    return ws1


def statistic_reserved_research_death_data(ws1, researches):
    """
    :return:
    """
    style_border_res = NamedStyle(name="style_border_res_rz")
    bd = Side(style='thin', color="000000")
    style_border_res.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border_res.font = Font(bold=False, size=11)
    style_border_res.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    r = 4

    for i in researches:
        if not i:
            return ws1
        r += 1
        if not i.get("Номер", ""):
            continue
        ws1.cell(row=r, column=1).value = i.get("hosp_title", "")
        ws1.cell(row=r, column=2).value = i.get("Номер", "")
        ws1.cell(row=r, column=3).value = i.get("date_create", "")
        ws1.cell(row=r, column=4).value = i.get("fio_patient", "")
        ws1.cell(row=r, column=5).value = i.get("napravleniye_id", "")
        rows = ws1[f'A{r}:E{r}']
        for row in rows:
            for cell in row:
                cell.style = style_border_res

    return ws1


def statistic_research_by_covid_base(ws1, d1, d2, research_titile):
    style_border = NamedStyle(name="style_border")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    columns = [
        ("№ заказа", 23),
        ("Название организации", 33),
        ("ОГРН организации", 33),
        ("Дата заказа", 23),
        ("Код услуги", 33),
        ("Название услуги", 33),
        ("Дата взятия биоматериала", 13),
        ("Дата готовности результата", 13),
        ("Результат", 13),
        ("Тип исследования", 13),
        ("Значение результата", 33),
        ("Фамилия", 33),
        ("Имя", 33),
        ("Отчество", 33),
        ("Пол", 8),
        ("Дата рождения", 13),
        ("Телефон", 13),
        ("e-mail", 13),
        ("Тип ДУЛ", 13),
        ("Номер документа", 13),
        ("Серия документа", 13),
        ("СНИЛС", 23),
        ("ОМС", 23),
        ("Адрес регистрации регион", 23),
        ("Адрес регистрации район", 23),
        ("Адрес регистрации город", 23),
        ("Адрес регистрации улица", 23),
        ("Адрес регистрации дом", 23),
        ("Адрес регистрации строение", 23),
        ("Адрес регистрации квартира", 23),
        ("Адрес факт  регион", 23),
        ("Адрес факт район", 23),
        ("Адрес факт город", 23),
        ("Адрес факт улица", 23),
        ("Адрес факт дом", 23),
        ("Адрес факт строение", 23),
        ("Адрес факт квартира", 23),
        ("Название лаборатории", 23),
        ("ОГРН лаборатории", 23),
        ("ID тест системы (согласно справочнику)", 23),
        ("Серия (партия) тест-системы", 23),
        ("Дата производства тест-системы", 23),
        ("Тест-система", 13),
    ]
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=1, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=1, column=idx).style = style_border

    return ws1


def statistic_research_by_covid_data(ws1, result_patient, patient_docs):
    """
    :return:
    """
    style_border_res = NamedStyle(name="style_border_res")
    bd = Side(style='thin', color="000000")
    style_border_res.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border_res.font = Font(bold=False, size=11)
    style_border_res.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    r = 1
    if not result_patient:
        return ws1

    for i in result_patient:
        r += 1
        ws1.cell(row=r, column=1).value = i.dir_id
        ws1.cell(row=r, column=2).value = i.hosp_title
        ws1.cell(row=r, column=3).value = i.hosp_ogrn
        ws1.cell(row=r, column=4).value = i.date_create
        ws1.cell(row=r, column=5).value = i.research_code
        ws1.cell(row=r, column=6).value = i.research_title
        ws1.cell(row=r, column=7).value = i.date_reciev
        ws1.cell(row=r, column=8).value = i.date_confirm
        if i.value and i.method_title != "ИФА":
            val_param = 0 if 'отриц' in i.value.lower() else 1
            result_val = ""
        else:
            result_val = i.value
            val_param = ""
        ws1.cell(row=r, column=9).value = val_param
        method_val = 2 if i.method_title == "ИФА" else 1
        ws1.cell(row=r, column=10).value = method_val
        ws1.cell(row=r, column=10).value = result_val
        ws1.cell(row=r, column=12).value = i.family
        ws1.cell(row=r, column=13).value = i.name
        ws1.cell(row=r, column=14).value = i.patronymic
        ws1.cell(row=r, column=15).value = 1 if i.sex.lower() == "м" else 2
        ws1.cell(row=r, column=16).value = i.born
        ws1.cell(row=r, column=17).value = ""
        ws1.cell(row=r, column=18).value = ""

        patient_doc = patient_docs.get(i.client_id, None)
        type, serial, number, snils, polis = "", "", "", "", ""
        if patient_doc:
            for pat_doc in patient_doc:
                for k, v in pat_doc.items():
                    if k == "снилс":
                        snils = v
                    elif k == "полис":
                        polis = v
                    elif "паспорт" in k.lower() or "рождение" in k.lower():
                        k_value = "Паспорт гражданина РФ" if "паспорт" in k.lower() else k
                        type = k_value
                        data = v.split("@")
                        serial = data[0]
                        number = data[1]
        ws1.cell(row=r, column=19).value = type
        ws1.cell(row=r, column=20).value = number
        ws1.cell(row=r, column=21).value = serial

        ws1.cell(row=r, column=22).value = snils
        ws1.cell(row=r, column=23).value = polis
        ws1.cell(row=r, column=24).value = "Иркутская область"

        ws1.cell(row=r, column=38).value = i.hosp_title
        ws1.cell(row=r, column=39).value = i.hosp_ogrn
        ws1.cell(row=r, column=40).value = "553407"
        ws1.cell(row=r, column=41).value = i.hosp_ogrn
        ws1.cell(row=r, column=42).value = i.hosp_ogrn
        ws1.cell(row=r, column=43).value = (
            "Набор реагентов для выявления РНК коронавируса SARS-CoV-2 тяжелого острого респираторного синдрома (COVID-19) методом полимеразной цепной "
            "реакции «АмплиПрайм® SARS-CoV-2 DUO» по ТУ 21.20.23-083-09286667-2020"
        )

        rows = ws1[f'A{r}:C{r}']
        for row in rows:
            for cell in row:
                cell.style = style_border_res

    return ws1


def statistic_research_by_sum_lab_base(ws1, d1, d2, research_titile):
    style_border = NamedStyle(name="style_border")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=1, column=2).value = research_titile
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'
    columns = [
        ('Лаборатория', 33),
        ('Услуга', 55),
        ('Кол-во', 25),
    ]
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=4, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=4, column=idx).style = style_border

    return


def statistic_research_by_sum_lab_data(ws1, researches):
    """
    :return:
    """
    style_border_res = NamedStyle(name="style_border_res")
    bd = Side(style='thin', color="000000")
    style_border_res.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border_res.font = Font(bold=False, size=11)
    style_border_res.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    r = 4
    if not researches:
        return ws1

    for i in researches:
        r += 1
        ws1.cell(row=r, column=1).value = i.lab_title
        ws1.cell(row=r, column=2).value = i.research_title
        ws1.cell(row=r, column=3).value = i.sum_research_id

        rows = ws1[f'A{r}:C{r}']
        for row in rows:
            for cell in row:
                cell.style = style_border_res

    return ws1


def statistic_research_by_details_lab_base(ws1, d1, d2, research_titile):
    style_border = NamedStyle(name="style_border")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=1, column=2).value = research_titile
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'
    columns = [
        ('ID', 23),
        ('лаборатория', 15),
        ('анализ', 35),
        ('дата', 15),
        ('время', 15),
        ('аппарат', 15),
        ('дата взятия', 15),
        ('время взятия', 15),
    ]
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=4, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=4, column=idx).style = style_border

    return ws1


def statistic_research_by_details_lab_data(ws1, researches):
    """
    :return:
    """
    style_border_res = NamedStyle(name="style_border_res")
    bd = Side(style='thin', color="000000")
    style_border_res.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border_res.font = Font(bold=False, size=11)
    style_border_res.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    r = 4
    if not researches:
        return ws1

    for i in researches:
        r += 1
        ws1.cell(row=r, column=1).value = i.napravleniye_id if i.napravleniye_id else ""
        ws1.cell(row=r, column=2).value = i.lab_title if i.lab_title else ""
        ws1.cell(row=r, column=3).value = i.research_title if i.research_title else ""
        ws1.cell(row=r, column=4).value = i.date_confirm if i.date_confirm else ""
        ws1.cell(row=r, column=5).value = i.time_confirm if i.time_confirm else ""
        ws1.cell(row=r, column=6).value = i.name if i.name else ""
        ws1.cell(row=r, column=7).value = i.date_tubes if i.date_tubes else ""
        ws1.cell(row=r, column=8).value = i.time_tubes if i.time_tubes else ""

        rows = ws1[f'A{r}:H{r}']
        for row in rows:
            for cell in row:
                cell.style = style_border_res

    return ws1


def statistic_message_ticket_base(ws1, d1, d2, style_border):
    ws1.cell(row=1, column=1).value = 'Обращения'
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'
    columns = [
        ('МО', 20),
        ('Номер', 20),
        ('Создано', 15),
        ('Физ. лицо', 26),
        ('Телефон', 20),
        ('Адрес', 20),
        ('Цель', 20),
        ('Примечания', 26),
        ('Статус', 16),
        ('Источник', 16),
        ('Создатель', 26),
    ]
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=4, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=4, column=idx).style = style_border

    return ws1


def statistic_message_ticket_data(ws1, message_ticket_sql, style_border_res):
    r = 4
    purposes = dict(DoctorCall.PURPOSES)
    statuses = dict(DoctorCall.STATUS)
    for ticket in message_ticket_sql:
        r += 1
        ws1.cell(row=r, column=1).value = ticket.hospital_short_title or ticket.hospital_title
        ws1.cell(row=r, column=2).value = ticket.external_num or ticket.num
        ws1.cell(row=r, column=3).value = ticket.date_create
        ws1.cell(row=r, column=4).value = f'{ticket.family} {ticket.name} {ticket.patronymic}'
        ws1.cell(row=r, column=5).value = ticket.phone
        ws1.cell(row=r, column=6).value = ticket.address
        ws1.cell(row=r, column=7).value = purposes.get(ticket.purpose, '')
        ws1.cell(row=r, column=8).value = ticket.comment
        ws1.cell(row=r, column=9).value = statuses.get(ticket.status, '')
        ws1.cell(row=r, column=10).value = 'интернет' if statuses.get(ticket.is_external) else 'оператор'
        who_create = ""
        if ticket.fio and ticket.short_title:
            who_create = f"{ticket.fio}-{ticket.short_title}"
        ws1.cell(row=r, column=11).value = who_create

        rows = ws1[f'A{r}:K{r}']
        for row in rows:
            for cell in row:
                cell.style = style_border_res

    return ws1


def statistic_message_purpose_total_data(ws1, message_total, d1, d2, style_border_res):
    ws1.cell(row=1, column=1).value = 'Обращения'
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'

    columns = [
        ('Цель', 20),
        ('Всего', 20),
        ('Выполнено', 20),
    ]
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=5, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=5, column=idx).style = style_border_res

    r = 5
    r1 = r
    purposes = dict(DoctorCall.PURPOSES)
    for p in message_total:
        r += 1
        ws1.cell(row=r, column=1).value = purposes.get(p.total_purpose, '')
        ws1.cell(row=r, column=2).value = p.sum_total_purpose
        ws1.cell(row=r, column=3).value = p.sum_execute_purpose or ''
        rows = ws1[f'A{r}:C{r}']
        for row in rows:
            for cell in row:
                cell.style = style_border_res

    ws1.cell(row=r + 1, column=1).value = 'Итого'
    ws1.cell(row=r + 1, column=1).style = style_border_res
    ws1.cell(row=r + 1, column=2).value = f'=SUM(B{r1 + 1}:B{r})'
    ws1.cell(row=r + 1, column=2).style = style_border_res
    ws1.cell(row=r + 1, column=3).value = f'=SUM(C{r1 + 1}:C{r})'
    ws1.cell(row=r + 1, column=3).style = style_border_res

    return ws1


def statistic_screening_month_data(ws1, data, month, year, style_border_res):
    ws1.cell(row=1, column=1).value = 'Скрининг'
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=2, column=2).value = f'{month_dict[int(month)]}-{year}'

    size = 12
    columns = [
        ('Месяц', 10),
        ('Число женщин 18-69 лет, проживающих на прикрепленной территории', size),
        ('Число женщин 30 -65 лет, подлежащих скринингу (всего)', size),
        ('Число женщин 30 -65 лет, подлежащих скринингу при диспансеризации', size),
        ('Число женщин  30-65 лет, прошедших  скрининг', size),
        ('Число женщин 30 -65 лет, подлежащих скринингу при диспансеризации', size),
        ('Число  женщин, которым выполнен ПАП-тест от общего числа прошедших скрининг', size),
        ('Число женщин,  по препаратам которых получили цитологический  результат', size),
        ('Из них, препараты признаны адекватными', size),
        ('Недостаточно адекватными', size),
        ('Не адекватными', size),
        ('Из числа женщин с недостаточно адекватным, неадекватным результатом, число вызванных женщин,  у которых повторно взят материал на цитологическое исследование', size),
        ('В т.ч. АSCUS', size),
        ('В т.ч. легкое интраэпителиальное  поражение CIN I, признаки ВПЧ', size),
        ('Умеренное  интраэпителиальное поражение CIN I-II, II', size),
        ('Тяжелое интраэпителиальное поражение CIN II-III, III', size),
        ('cr in situ', size),
        ('Подозрение на ЗНО шейки матки', size),
        ('Всего по Папа-Николау', size),
    ]

    for idx, column in enumerate(columns, 1):
        ws1.cell(row=5, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=5, column=idx).style = style_border_res

    ws1.cell(row=6, column=1).value = f'{month_dict[int(month)]}'
    ws1.cell(row=6, column=1).style = style_border_res
    for k, v in data.items():
        if k == "attached_count_age_for_month":
            ws1.cell(row=6, column=2).value = v
            ws1.cell(row=6, column=2).style = style_border_res
        if k == "count_regplan_for_month":
            ws1.cell(row=6, column=3).value = v
            ws1.cell(row=6, column=3).style = style_border_res
        if k == "count_dispensarization_from_screening":
            ws1.cell(row=6, column=4).value = v
            ws1.cell(row=6, column=4).style = style_border_res
        if k == "pass_screening":
            ws1.cell(row=6, column=5).value = v
            ws1.cell(row=6, column=5).style = style_border_res
        if k == "pass_screening_in_dispensarization":
            ws1.cell(row=6, column=6).value = v
            ws1.cell(row=6, column=6).style = style_border_res
        if k == "pass_pap_analysis":
            ws1.cell(row=6, column=7).value = v
            ws1.cell(row=6, column=7).style = style_border_res
        if k == "pass_pap_adequate_result_value":
            ws1.cell(row=6, column=9).value = v
            ws1.cell(row=6, column=9).style = style_border_res
        if k == "pass_pap_not_enough_adequate_result_value":
            ws1.cell(row=6, column=10).value = v
            ws1.cell(row=6, column=10).style = style_border_res
        if k == "pass_pap_not_adequate_result_value":
            ws1.cell(row=6, column=11).value = v
            ws1.cell(row=6, column=11).style = style_border_res
        if k == "count_people_dublicate":
            ws1.cell(row=6, column=12).value = v
            ws1.cell(row=6, column=12).style = style_border_res
        if k == "pass_pap_ascus_result_value":
            ws1.cell(row=6, column=13).value = v
            ws1.cell(row=6, column=13).style = style_border_res
        if k == "pass_pap_cin_i_result_value":
            ws1.cell(row=6, column=14).value = v
            ws1.cell(row=6, column=14).style = style_border_res
        if k == "pass_pap_cin_i_ii_result_value":
            ws1.cell(row=6, column=15).value = v
            ws1.cell(row=6, column=15).style = style_border_res
        if k == "pass_pap_cin_ii_iii_result_value":
            ws1.cell(row=6, column=16).value = v
            ws1.cell(row=6, column=16).style = style_border_res
        if k == "pass_pap_cr_in_situ_result_value":
            ws1.cell(row=6, column=17).value = v
            ws1.cell(row=6, column=17).style = style_border_res
        if k == "count_pap_analysys":
            ws1.cell(row=6, column=19).value = v
            ws1.cell(row=6, column=19).style = style_border_res

    return ws1


def get_table_diagnos(diagnos_data, item):
    diag_details = {}
    period_data = ""

    try:
        diagnos_data[item].keys()
        diag_data = diagnos_data[item]
    except:
        diag_data = json.loads(diagnos_data[item])
    try:
        diag_details = json.loads(diag_data["rows"][0][2])
        period_data = f'{diag_data["rows"][0][0]} {diag_data["rows"][0][1]}'
        is_dict = True
    except:
        is_dict = False
    if not is_dict:
        diag_details["code"] = "-"
        diag_details["title"] = "-"
        period_data = "-"

    return (period_data, diag_details)


def statistic_research_wepon_base(ws1, d1, d2, research_titile):
    style_border = NamedStyle(name="style_border")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=1, column=1).value = 'Услуга:'
    ws1.cell(row=1, column=2).value = research_titile
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'

    columns = [
        ('Серия', 13),
        ('Номер', 15),
        ('Медицинская организация выдавшая документ', 38),
        ('Дата выдачи', 11),
        ('Дата рождения', 11),
        ('ФИО пациента', 25),
        ('Адрес пациента', 41),
        ('ФИО выдавшего свидетельства', 20),
        ('Служебный номер', 15),
    ]
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=4, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=4, column=idx).style = style_border

    return ws1


def statistic_research_weapon_data(ws1, researches):
    """
    :return:
    """
    style_border_res = NamedStyle(name="style_border_res")
    bd = Side(style='thin', color="000000")
    style_border_res.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border_res.font = Font(bold=False, size=11)
    style_border_res.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    r = 4

    for i in researches:
        if not i:
            return ws1

        r += 1
        ws1.cell(row=r, column=1).value = i.get("Серия", "")
        ws1.cell(row=r, column=2).value = i.get("Номер", "")
        ws1.cell(row=r, column=3).value = i.get("hosp_title", "")
        if i.get("Дата выдачи", None):
            ws1.cell(row=r, column=4).value = normalize_date(i["Дата выдачи"])
        ws1.cell(row=r, column=5).value = i.get("Дата рождения пациента", "")
        ws1.cell(row=r, column=6).value = i.get("fio_patient", "")
        ws1.cell(row=r, column=7).value = i.get("Место постоянного жительства (регистрации) пациента", "")
        ws1.cell(row=r, column=8).value = i.get("Врач", "")
        ws1.cell(row=r, column=9).value = i.get("napravleniye_id", "")

        rows = ws1[f'A{r}:I{r}']
        for row in rows:
            for cell in row:
                cell.style = style_border_res

    return ws1
