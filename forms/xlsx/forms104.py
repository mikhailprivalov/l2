import calendar
import datetime
from typing import Dict

import openpyxl
import pytils
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from employees.models import TimeTrackingDocument, WorkDayStatus, FactTimeWork
from hospitals.models import Hospitals


def set_style_for_area(work_sheet, min_row, max_row, min_col, max_col, border_style=None, alignment_style=None, fill_style=None):
    for row in work_sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if border_style:
                cell.border = border_style
            if alignment_style:
                cell.alignment = alignment_style
            if fill_style:
                cell.fill = fill_style


def merge_cells_by_row(work_sheet, start_row, end_row, start_col, end_col):
    """
    Объединяет построчно для каждой колонки в диапозоне
    """
    for col in range(start_col, end_col + 1):
        work_sheet.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)


def _parse_cell_data(work_day_statuses: Dict, cell_value: Dict):
    start_time = cell_value.get("startWorkTime")
    end_time = cell_value.get("endWorkTime")
    type_id = cell_value.get("typeId")
    if type_id:
        result = work_day_statuses.get(int(type_id), "")
    elif start_time and end_time:
        result = f"{start_time} {end_time}"
    else:
        result = ""
    return result

def _parse_tabel_cell_data(work_day_statuses: Dict, cell_value: Dict):
    day_hours = cell_value.get("day_hours", 0)
    night_hours = cell_value.get("night_hours", 0)
    type_id = cell_value.get("type_id")
    if type_id:
        result = work_day_statuses.get(int(type_id), "")
    elif day_hours or night_hours:
        result = f"{day_hours + night_hours}"
    else:
        result = ""
    return result


def chunked(data, size):
    for i in range(0, len(data), size):
        yield data[i : i + size]


def form_01(request_data) -> Workbook:
    """
    Создает xlsx файл по форме графика рабочего времени
    """

    request_body = request_data.get("request_body")
    document_id = request_body.get("documentId")
    employees_work_time = request_body.get("employeesWorkTime")
    time_tracking_document: TimeTrackingDocument = TimeTrackingDocument.get_document_for_print(document_id)
    work_day_statuses = WorkDayStatus.get_statuses_dict()
    current_date = datetime.datetime.now()
    current_month_name = pytils.dt.ru_strftime(u"%B", inflected=True, date=current_date)
    current_day = current_date.day
    current_year = current_date.year
    document_date = time_tracking_document.month
    document_month_name = pytils.dt.ru_strftime(u"%B", date=document_date)
    document_year = document_date.year
    document_month = document_date.month
    document_last_day_month = calendar.monthrange(document_year, document_month)[1]
    department_title = time_tracking_document.department.name
    organization: Hospitals = request_data.get("hospital")
    organization_title = organization.safe_short_title

    thin_line = Side(style='thin')
    thin_bottom_border = Border(bottom=thin_line)
    thin_border = Border(left=thin_line, top=thin_line, right=thin_line, bottom=thin_line)
    alignment_center = Alignment(horizontal='center', vertical='center')
    alignment_center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignment_left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    font_bold = Font(bold=True)
    weekend_fill = PatternFill(start_color="ffe699", end_color="ffe699", fill_type="solid")

    work_book: Workbook = openpyxl.Workbook()
    work_book.remove(work_book.get_sheet_by_name("Sheet"))
    max_row_count_in_list = 15
    for index, chunk_data in enumerate(chunked(employees_work_time, max_row_count_in_list)):
        department_words = department_title.replace("-", " ").split(" ")
        list_name = ''.join([word[0] for word in department_words])
        work_sheet = work_book.create_sheet(f"{list_name}_{index}")

        first_row_table_number = 12
        second_row_table_number = 13
        third_row_table_number = 14
        table_data_end_row_number = third_row_table_number + len(chunk_data) - 1

        item_number_col_number = 1
        fio_col_number = 2
        position_col_number = 3
        type_employment_col_number = 4
        occupied_volume_col_number = 5
        norm_hours_col_number = 6
        working_shift_col_number = 7
        start_date_col_number = 8
        end_date_col_number = start_date_col_number + document_last_day_month - 1
        amount_hours_col_number = end_date_col_number + 1
        employees_signature_col_number = amount_hours_col_number + 1

        item_number_col_width = 4.17
        fio_col_width = 17.5
        position_col_width = 18.83
        type_employment_col_width = 12.7
        occupied_volume_col_width = 8.7
        norm_hours_col_width = 9.5
        working_shift_col_width = 6.5
        date_col_width = 5.33
        amount_hours_col_width = 9.83
        employees_signature_col_width = 11.33

        work_sheet.column_dimensions[get_column_letter(item_number_col_number)].width = item_number_col_width
        work_sheet.column_dimensions[get_column_letter(fio_col_number)].width = fio_col_width
        work_sheet.column_dimensions[get_column_letter(position_col_number)].width = position_col_width
        work_sheet.column_dimensions[get_column_letter(type_employment_col_number)].width = type_employment_col_width
        work_sheet.column_dimensions[get_column_letter(occupied_volume_col_number)].width = occupied_volume_col_width
        work_sheet.column_dimensions[get_column_letter(norm_hours_col_number)].width = norm_hours_col_width
        work_sheet.column_dimensions[get_column_letter(working_shift_col_number)].width = working_shift_col_width
        for col_number in range(start_date_col_number, end_date_col_number + 1):
            work_sheet.column_dimensions[get_column_letter(col_number)].width = date_col_width
        work_sheet.column_dimensions[get_column_letter(amount_hours_col_number)].width = amount_hours_col_width
        work_sheet.column_dimensions[get_column_letter(employees_signature_col_number)].width = employees_signature_col_width

        work_sheet.row_dimensions[first_row_table_number].height = 47.25
        for row_number in range(third_row_table_number, table_data_end_row_number + 1):
            work_sheet.row_dimensions[row_number].height = 45

        work_sheet["AL1"] = "Приложение № 2 к"
        work_sheet["B2"] = "Председатель ППО"
        work_sheet["C2"] = "________________"
        work_sheet["AD2"] = "УТВЕРЖДАЮ"
        work_sheet["AL2"] = "приказу от '20' февраля 2025 г № 37"
        work_sheet.merge_cells('A3:AC3')
        work_sheet["A3"].alignment = alignment_center
        work_sheet["A3"].font = font_bold
        work_sheet["A3"] = "ГРАФИК РАБОЧЕГО ВРЕМЕНИ"
        work_sheet["AD3"].font = font_bold
        work_sheet["AD3"] = "Руководитель учреждения ______________________"
        work_sheet["AK4"] = "подпись"
        work_sheet.merge_cells('A5:AC5')
        work_sheet["A5"].alignment = alignment_center
        work_sheet["A5"].font = font_bold
        work_sheet["A5"] = organization_title
        work_sheet["AF5"].font = font_bold
        work_sheet["AF5"] = f'"{current_day}"{current_month_name} {current_year}г.'
        work_sheet.merge_cells('A7:AC7')
        work_sheet["A7"].alignment = alignment_center
        work_sheet["A7"].font = font_bold
        work_sheet["A7"] = department_title
        cell_a_number = 1
        cell_ag_number = 33
        set_style_for_area(work_sheet, 7, 7, cell_a_number, cell_ag_number, thin_bottom_border)
        work_sheet["AD7"] = f"календарных дней {document_last_day_month}"
        work_sheet.merge_cells('A8:AC8')
        work_sheet["A8"].alignment = alignment_center
        work_sheet["A8"] = "(подразделение)"
        work_sheet["AD8"] = "рабочих дней"
        cell_ad_number = 30
        set_style_for_area(work_sheet, 8, 8, cell_ad_number, cell_ag_number, thin_bottom_border)
        work_sheet.merge_cells('A9:AC9')
        work_sheet["A9"].alignment = alignment_center
        work_sheet["A9"] = f"{document_month_name} {document_year} год"

        set_style_for_area(work_sheet, first_row_table_number, table_data_end_row_number, item_number_col_number, employees_signature_col_number, thin_border, alignment_center_wrap)
        set_style_for_area(work_sheet, third_row_table_number, table_data_end_row_number, fio_col_number, fio_col_number, alignment_style=alignment_left_wrap)

        for day_number in range(1, document_last_day_month):
            date = datetime.date(document_year, document_month, day_number)
            if date.weekday() in (5, 6):
                day_col = day_number + (start_date_col_number - 1)
                set_style_for_area(work_sheet, second_row_table_number, table_data_end_row_number, day_col, day_col, fill_style=weekend_fill)

        merge_cells_by_row(work_sheet, first_row_table_number, second_row_table_number, item_number_col_number, working_shift_col_number)
        work_sheet.merge_cells(start_row=first_row_table_number, start_column=start_date_col_number, end_row=first_row_table_number, end_column=end_date_col_number)
        merge_cells_by_row(work_sheet, first_row_table_number, second_row_table_number, amount_hours_col_number, employees_signature_col_number)

        work_sheet.cell(row=first_row_table_number, column=item_number_col_number, value="№ п/п")
        work_sheet.cell(row=first_row_table_number, column=fio_col_number, value="Фамилия имя отчество")
        work_sheet.cell(row=first_row_table_number, column=position_col_number, value="Должность (профессия)")
        work_sheet.cell(row=first_row_table_number, column=type_employment_col_number, value="Вид занятости (осн, внутр, внеш)")
        work_sheet.cell(row=first_row_table_number, column=occupied_volume_col_number, value="Занимаемый объем (согл ТД), шт ед")
        work_sheet.cell(row=first_row_table_number, column=norm_hours_col_number, value="Норма часов на занимаемый объем")
        work_sheet.cell(row=first_row_table_number, column=working_shift_col_number, value="Рабочая смена")
        work_sheet.cell(row=first_row_table_number, column=start_date_col_number, value="Числа месяца")
        work_sheet.cell(row=first_row_table_number, column=amount_hours_col_number, value="Количество часов согласно графика")
        work_sheet.cell(row=first_row_table_number, column=employees_signature_col_number, value="Подпись работника")
        for date_number in range(document_last_day_month):
            work_sheet.cell(row=second_row_table_number, column=date_number + start_date_col_number, value=date_number + 1)

        for index_row, work_time in enumerate(chunk_data):
            item_number = index_row + 1
            fio = work_time.get("fio")
            position = work_time.get("position")
            type_employment = work_time.get("bidType")
            occupied_volume = ""
            norm_hours = ""
            working_shift = ""
            amount_hours = work_time.get("totalHoursDecimal")
            date_keys = []
            for key in work_time.keys():
                try:
                    date_key = datetime.datetime.strptime(key, "%Y-%m-%d")
                    date_keys.append(date_key)
                except ValueError:
                    pass
            date_keys_sorted = sorted(date_keys)
            date_values = [_parse_cell_data(work_day_statuses, work_time.get(date_key.strftime("%Y-%m-%d"))) for date_key in date_keys_sorted]
            work_sheet.cell(row=third_row_table_number + index_row, column=item_number_col_number, value=item_number)
            work_sheet.cell(row=third_row_table_number + index_row, column=fio_col_number, value=fio)
            work_sheet.cell(row=third_row_table_number + index_row, column=position_col_number, value=position)
            work_sheet.cell(row=third_row_table_number + index_row, column=type_employment_col_number, value=type_employment)
            work_sheet.cell(row=third_row_table_number + index_row, column=occupied_volume_col_number, value=occupied_volume)
            work_sheet.cell(row=third_row_table_number + index_row, column=norm_hours_col_number, value=norm_hours)
            work_sheet.cell(row=third_row_table_number + index_row, column=working_shift_col_number, value=working_shift)
            for index_date, date in enumerate(date_values):
                work_sheet.cell(row=third_row_table_number + index_row, column=start_date_col_number + index_date, value=date)
            work_sheet.cell(row=third_row_table_number + index_row, column=amount_hours_col_number, value=amount_hours)

        work_sheet.cell(row=table_data_end_row_number + 2, column=1, value="Заведующий отделением")
        work_sheet.cell(row=table_data_end_row_number + 2, column=3, value="_____________________")
        work_sheet.cell(row=table_data_end_row_number + 2, column=11, value="Старшая медицинская сестра")
        work_sheet.cell(row=table_data_end_row_number + 2, column=16, value="_____________________")

    return work_book


def form_02(request_data):
    year = request_data.get("year")
    month = request_data.get("month")
    department_id = request_data.get("departmentId")
    fact_time_work_data = FactTimeWork.get_fact_time_work_employee(int(year), int(month), int(department_id))
    tabel_document = fact_time_work_data.get("document")
    fact_time_work = fact_time_work_data.get("result")
    work_day_statuses = WorkDayStatus.get_statuses_dict()
    tabel_document_number = tabel_document.pk
    current_date = datetime.datetime.now()
    current_month_name = pytils.dt.ru_strftime(u"%B", inflected=True, date=current_date)
    current_day = current_date.day
    current_year = current_date.year
    document_date = tabel_document.month_tabel
    document_month_name = pytils.dt.ru_strftime(u"%B", date=document_date)
    document_year = document_date.year
    document_month = document_date.month
    document_first_day_month = 1
    document_last_day_month = calendar.monthrange(document_year, document_month)[1]
    department_title = tabel_document.department.name
    organization: Hospitals = request_data.get("hospital")
    organization_title = organization.safe_short_title

    thin_line = Side(style='thin')
    thin_bottom_border = Border(bottom=thin_line)
    thin_border = Border(left=thin_line, top=thin_line, right=thin_line, bottom=thin_line)
    thick_line = Side(style="thick")
    thick_bottom_border = Border(bottom=thick_line)
    alignment_center = Alignment(horizontal='center', vertical='center')
    alignment_center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignment_right = Alignment(horizontal='right', vertical='center')
    alignment_left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    font_bold = Font(bold=True)
    weekend_fill = PatternFill(start_color="ffe699", end_color="ffe699", fill_type="solid")


    work_book: Workbook = openpyxl.Workbook()
    work_book.remove(work_book.get_sheet_by_name("Sheet"))

    max_row_count_in_list = 15
    for index, chunk_data in enumerate(chunked(fact_time_work, max_row_count_in_list)):
        department_words = department_title.replace("-", " ").split(" ")
        list_name = ''.join([word[0] for word in department_words])
        work_sheet = work_book.create_sheet(f"{list_name}_{index}")

        first_row_table_number = 12
        second_row_table_number = 13
        third_row_table_number = 14
        four_row_table_number = 15
        table_data_end_row_number = four_row_table_number + len(chunk_data) - 1

        fio_col_number = 1
        tabel_number_col_number = 2
        position_col_number = 3
        type_employment_col_number = 4
        occupied_volume_col_number = 5
        norm_hours_col_number = 6
        start_date_col_number = 7
        end_date_col_number = start_date_col_number + document_last_day_month - 1
        sum_common_hours_col_number = end_date_col_number + 1
        sum_night_hours_col_number = end_date_col_number + 2
        sum_holidays_hours_col_number = end_date_col_number + 3

        fio_col_width = 17.83
        tabel_number_col_width = 7
        position_col_width = 16.17
        type_employment_col_width = 9.33
        occupied_volume_col_width = 6.17
        norm_hours_col_width = 6.33
        date_col_width = 3.5
        sum_common_hours_col_width = 6.83
        sum_night_hours_col_width = 6.33
        sum_holidays_hours_col_width = 5.67

        work_sheet.column_dimensions[get_column_letter(fio_col_number)].width = fio_col_width
        work_sheet.column_dimensions[get_column_letter(tabel_number_col_number)].width = tabel_number_col_width
        work_sheet.column_dimensions[get_column_letter(position_col_number)].width = position_col_width
        work_sheet.column_dimensions[get_column_letter(type_employment_col_number)].width = type_employment_col_width
        work_sheet.column_dimensions[get_column_letter(occupied_volume_col_number)].width = occupied_volume_col_width
        work_sheet.column_dimensions[get_column_letter(norm_hours_col_number)].width = norm_hours_col_width
        for col_number in range(start_date_col_number, end_date_col_number + 1):
            work_sheet.column_dimensions[get_column_letter(col_number)].width = date_col_width
        work_sheet.column_dimensions[get_column_letter(sum_common_hours_col_number)].width = sum_common_hours_col_width
        work_sheet.column_dimensions[get_column_letter(sum_night_hours_col_number)].width = sum_night_hours_col_width
        work_sheet.column_dimensions[get_column_letter(sum_holidays_hours_col_number)].width = sum_holidays_hours_col_width

        work_sheet.row_dimensions[first_row_table_number].height = 15
        work_sheet.row_dimensions[third_row_table_number].height = 15
        for row_number in range(four_row_table_number, table_data_end_row_number + 1):
            work_sheet.row_dimensions[row_number].height = 30

        work_sheet["AK1"] = "Приложение № 1 к"
        work_sheet["AK2"] = "приказу от '20' февраля 2025 г № 37"
        work_sheet.merge_cells('A3:AK3')
        work_sheet["A3"].alignment = alignment_center
        work_sheet["A3"].font = font_bold
        work_sheet["A3"] = f"Табель № {tabel_document_number}"

        work_sheet.merge_cells('A4:AK4')
        work_sheet["A4"].alignment = alignment_center
        work_sheet["A4"].font = font_bold
        work_sheet["A4"] = "учета использования рабочего времени"
        work_sheet["AL4"] = "Коды"
        work_sheet["AJ5"] = "Форма ОКУД"
        work_sheet["AJ5"].alignment = alignment_right
        work_sheet["AL5"] = "0504421"
        work_sheet["K6"] = "за период"
        work_sheet["M6"] = "c"
        work_sheet["N6"] = document_first_day_month
        work_sheet.merge_cells('P6:Q6')
        work_sheet["P6"].alignment = alignment_center
        work_sheet["P6"] = "по"
        work_sheet["R6"] = document_last_day_month
        work_sheet.merge_cells('S6:U6')
        work_sheet["S6"] = document_month_name
        work_sheet["V6"] = document_year
        work_sheet["W6"] = "г."
        work_sheet["AJ6"] = "Дата"
        work_sheet["AJ6"].alignment = alignment_right
        work_sheet["A7"] = "Учреждение"
        work_sheet.merge_cells('C7:AD7')
        work_sheet["C7"].alignment = alignment_center
        work_sheet["C7"].font = font_bold
        work_sheet["C7"] = organization_title
        work_sheet["AJ7"] = "по ОКПО"
        work_sheet["AJ7"].alignment = alignment_right
        work_sheet["A8"] = "Структурное подразделение"
        work_sheet.merge_cells('C8:AD8')
        work_sheet["C8"].alignment = alignment_center
        work_sheet["C8"].font = font_bold
        work_sheet["C8"] = department_title
        work_sheet["A9"] = "Вид табеля"
        work_sheet.merge_cells('C9:AD9')
        work_sheet["C9"].alignment = alignment_center
        work_sheet["C9"].font = font_bold
        work_sheet["C9"] = ""
        work_sheet["AJ9"] = "Номер корректировки"
        work_sheet["AJ9"].alignment = alignment_right
        work_sheet.merge_cells('G10:AA10')
        work_sheet["G10"].alignment = alignment_center
        work_sheet["G10"] = "(первичный - 0; корректирующий 1,2 и т.д.)"
        work_sheet["AJ10"] = "Дата формирования документа"
        work_sheet["AJ10"].alignment = alignment_right
        work_sheet["AL10"] = current_date.strftime("%d.%m.%Y")

        cell_c_number = 3
        cell_ag_number = 30
        set_style_for_area(work_sheet, 7, 9, cell_c_number, cell_ag_number, thin_bottom_border)
        cell_al_number = 38
        set_style_for_area(work_sheet, 4, 10, cell_al_number, cell_al_number, thin_border)

        set_style_for_area(work_sheet, first_row_table_number, table_data_end_row_number, fio_col_number, sum_holidays_hours_col_number, thin_border, alignment_center_wrap)
        set_style_for_area(work_sheet, four_row_table_number, table_data_end_row_number, fio_col_number, fio_col_number, alignment_style=alignment_left_wrap)

        for day_number in range(1, document_last_day_month):
            date = datetime.date(document_year, document_month, day_number)
            if date.weekday() in (5, 6):
                day_col = day_number + (start_date_col_number - 1)
                set_style_for_area(work_sheet, second_row_table_number, table_data_end_row_number, day_col, day_col, fill_style=weekend_fill)

        merge_cells_by_row(work_sheet, first_row_table_number, second_row_table_number, fio_col_number, norm_hours_col_number)
        work_sheet.merge_cells(start_row=first_row_table_number, start_column=start_date_col_number, end_row=first_row_table_number, end_column=sum_common_hours_col_number)
        # merge_cells_by_row(work_sheet, first_row_table_number, second_row_table_number, sum_common_hours_col_number, sum_holidays_hours_col_number)

        work_sheet.cell(row=first_row_table_number, column=fio_col_number, value="Фамилия имя отчество")
        work_sheet.cell(row=first_row_table_number, column=tabel_number_col_number, value="Учетный номер")
        work_sheet.cell(row=first_row_table_number, column=position_col_number, value="Должность (профессия)")
        work_sheet.cell(row=first_row_table_number, column=type_employment_col_number, value="Вид занятости (осн, внутр, внеш)")
        work_sheet.cell(row=first_row_table_number, column=occupied_volume_col_number, value="Занимаемый объем (согл ТД), шт ед")
        work_sheet.cell(row=first_row_table_number, column=norm_hours_col_number, value="Норма часов на занимаемый объем")
        work_sheet.cell(row=first_row_table_number, column=start_date_col_number, value="Числа месяца")

        for date_number in range(document_last_day_month):
            work_sheet.cell(row=second_row_table_number, column=date_number + start_date_col_number, value=date_number + 1)

        work_sheet.cell(row=second_row_table_number, column=sum_common_hours_col_number, value="Всего отработано часов")
        work_sheet.cell(row=second_row_table_number, column=sum_night_hours_col_number, value="Ночные")
        work_sheet.cell(row=second_row_table_number, column=sum_holidays_hours_col_number, value="Праздничные, выходные")

        for day_number in range(1, sum_holidays_hours_col_number + 1):
            work_sheet.cell(row=third_row_table_number, column=day_number, value=day_number)


        for index_row, fact_time in enumerate(chunk_data):
            fio = fact_time.get("fio")
            tabel_number = fact_time.get("tabel_number")
            position = fact_time.get("position")
            type_employment = fact_time.get("bid_type")
            occupied_volume = ""
            norm_hours = ""

            sum_common_hours = fact_time.get("sum_common_hours")
            sum_night_hours = fact_time.get("sum_night_hours")
            sum_holidays_hours = fact_time.get("sum_holidays_hours")
            date_keys = []
            for key in fact_time.keys():
                try:
                    date_key = datetime.datetime.strptime(key, "%Y-%m-%d")
                    date_keys.append(date_key)
                except ValueError:
                    pass
            date_keys_sorted = sorted(date_keys)
            date_values = [_parse_tabel_cell_data(work_day_statuses, fact_time.get(date_key.strftime("%Y-%m-%d"))) for date_key in date_keys_sorted]
            work_sheet.cell(row=four_row_table_number + index_row, column=fio_col_number, value=fio)
            work_sheet.cell(row=four_row_table_number + index_row, column=tabel_number_col_number, value=tabel_number)
            work_sheet.cell(row=four_row_table_number + index_row, column=position_col_number, value=position)
            work_sheet.cell(row=four_row_table_number + index_row, column=type_employment_col_number, value=type_employment)
            work_sheet.cell(row=four_row_table_number + index_row, column=occupied_volume_col_number, value=occupied_volume)
            work_sheet.cell(row=four_row_table_number + index_row, column=norm_hours_col_number, value=norm_hours)
            for index_date, date in enumerate(date_values):
                work_sheet.cell(row=four_row_table_number + index_row, column=start_date_col_number + index_date, value=date)
            work_sheet.cell(row=four_row_table_number + index_row, column=sum_common_hours_col_number, value=sum_common_hours)
            work_sheet.cell(row=four_row_table_number + index_row, column=sum_night_hours_col_number, value=sum_night_hours)
            work_sheet.cell(row=four_row_table_number + index_row, column=sum_holidays_hours_col_number, value=sum_holidays_hours)

        organization_head_cell = work_sheet.cell(row=table_data_end_row_number + 2, column=1, value="Главный врач")
        organization_head_cell.border = thin_bottom_border
        organization_head_fio_cell = work_sheet.cell(row=table_data_end_row_number + 2, column=3, value="")
        organization_head_fio_cell.border = thin_bottom_border
        work_sheet.cell(row=table_data_end_row_number + 3, column=1, value="(должность)")
        work_sheet.cell(row=table_data_end_row_number + 3, column=8, value="(расшифровка подписи)")
        work_sheet.merge_cells(start_row=table_data_end_row_number + 3, start_column=8, end_row=table_data_end_row_number + 3, end_column=12)
        work_sheet.cell(row=table_data_end_row_number + 3, column=20, value="Отметка бухгалтерии о принятии настоящего табеля")
        department_head_cell = work_sheet.cell(row=table_data_end_row_number + 5, column=1, value="Зав. отделением")
        department_head_cell.border = thin_bottom_border
        department_head_fio_cell = work_sheet.cell(row=table_data_end_row_number + 5, column=3, value="")
        department_head_fio_cell.border = thin_bottom_border
        work_sheet.cell(row=table_data_end_row_number + 5, column=20, value="Исполнитель")
        work_sheet.merge_cells(start_row=table_data_end_row_number + 5, start_column=22, end_row=table_data_end_row_number + 5, end_column=35)
        work_sheet.cell(row=table_data_end_row_number + 6, column=1, value="(должность)")
        work_sheet.cell(row=table_data_end_row_number + 6, column=8, value="(расшифровка подписи)")
        work_sheet.merge_cells(start_row=table_data_end_row_number + 6, start_column=8, end_row=table_data_end_row_number + 6, end_column=12)
        buh_signature_sub_title = work_sheet.cell(row=table_data_end_row_number + 6, column=22, value="(подпись)")
        buh_signature_sub_title.alignment = alignment_center
        work_sheet.merge_cells(start_row=table_data_end_row_number + 6, start_column=22, end_row=table_data_end_row_number + 6, end_column=35)
        older_nurse_cell = work_sheet.cell(row=table_data_end_row_number + 8, column=1, value="Старшая мед. сестра")
        older_nurse_cell.border = thin_bottom_border
        older_nurse_fio_cell = work_sheet.cell(row=table_data_end_row_number + 8, column=3, value="")
        older_nurse_fio_cell.border = thin_bottom_border
        work_sheet.cell(row=table_data_end_row_number + 8, column=8, value="(расшифровка подписи)")
        work_sheet.merge_cells(start_row=table_data_end_row_number + 8, start_column=8, end_row=table_data_end_row_number + 8, end_column=12)
        work_sheet.cell(row=table_data_end_row_number + 8, column=20, value='"')
        work_sheet.cell(row=table_data_end_row_number + 8, column=21, value='"')
        work_sheet.merge_cells(start_row=table_data_end_row_number + 8, start_column=22, end_row=table_data_end_row_number + 8, end_column=25)
        work_sheet.cell(row=table_data_end_row_number + 8, column=26, value="20")
        number_year_cell = work_sheet.cell(row=table_data_end_row_number + 8, column=27, value="")
        number_year_cell.border = thin_bottom_border
        work_sheet.cell(row=table_data_end_row_number + 8, column=28, value="г.")

        work_sheet.cell(row=table_data_end_row_number + 9, column=1, value="(должность)")
        work_sheet.cell(row=table_data_end_row_number + 9, column=8, value="(расшифровка подписи)")
        work_sheet.merge_cells(start_row=table_data_end_row_number + 9, start_column=8, end_row=table_data_end_row_number + 9, end_column=12)

        hr_specialist_cell = work_sheet.cell(row=table_data_end_row_number + 11, column=1, value="Специалист о.к.")
        hr_specialist_cell.border = thin_bottom_border
        hr_specialist_fio_cell = work_sheet.cell(row=table_data_end_row_number + 11, column=3, value="")
        hr_specialist_fio_cell.border = thin_bottom_border
        work_sheet.cell(row=table_data_end_row_number + 12, column=1, value="(должность)")
        work_sheet.cell(row=table_data_end_row_number + 12, column=8, value="(расшифровка подписи)")
        work_sheet.merge_cells(start_row=table_data_end_row_number + 12, start_column=8, end_row=table_data_end_row_number + 12, end_column=12)

        economist_cell = work_sheet.cell(row=table_data_end_row_number + 13, column=1, value="Экономист")
        economist_cell.border = thin_bottom_border
        economist_fio_cell =  work_sheet.cell(row=table_data_end_row_number + 13, column=3, value="")
        economist_fio_cell.border = thin_bottom_border
        work_sheet.cell(row=table_data_end_row_number + 13, column=1, value="(должность)")
        work_sheet.cell(row=table_data_end_row_number + 13, column=8, value="(расшифровка подписи)")
        work_sheet.merge_cells(start_row=table_data_end_row_number + 13, start_column=8, end_row=table_data_end_row_number + 13, end_column=12)

        work_sheet.cell(row=table_data_end_row_number + 14, column=1, value='"  "     20___г.')



























        # TODO возможно, если надо будет один в один, придется для header, table_col_names и прочего заводить свои стили текста (размер, шрифт и т.д)
        # TODO а тут мы соберем таблицу-табель)
        # TODO и еще ниже блок подписей и прочих)

    return work_book
