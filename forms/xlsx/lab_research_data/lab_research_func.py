from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter


def lab_report_base(ws1):
    style_border = NamedStyle(name="style_border_ca")
    bd = Side(style="thin", color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=12)
    style_border.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

    columns = [
        ("№", 5),
        ("Группа", 30),
        ("Подгруппа", 30),
        ("Наименование услуги", 50),
        ("Внтренний код", 17),
        ("Код НМУ", 17),
        ("Тест", 40),
        ("Код ФСЛИ", 20),
        ("Контейнер тип", 30),
        ("Контейнер ИД", 20),
        ("Статус", 17),
    ]

    row = 1
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border
    return ws1


def fill_lab_report(ws1, data, row=2):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style="thin", color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=12)
    style_border1.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
    r = row
    step = 1
    for val in data:
        ws1.cell(row=r, column=1).value = step
        ws1.cell(row=r, column=2).value = val.group_title
        ws1.cell(row=r, column=3).value = val.subgroup_title
        ws1.cell(row=r, column=4).value = val.research_title
        ws1.cell(row=r, column=5).value = val.internal_code
        ws1.cell(row=r, column=6).value = val.nmu_code
        ws1.cell(row=r, column=7).value = val.fraction_title
        ws1.cell(row=r, column=8).value = val.fraction_fsli
        ws1.cell(row=r, column=9).value = val.tube_title
        ws1.cell(row=r, column=10).value = val.tube_id
        ws1.cell(row=r, column=11).value = "заблокирован" if val.hide_status else "активен"
        for i in range(1, 12):
            ws1.cell(row=r, column=i).style = style_border1
        r += 1
        step += 1
    return ws1
