import os
from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from laboratory.settings import BASE_DIR


def billing_base(ws1):
    style_border = NamedStyle(name="style_border_ca")
    bd = Side(style="thin", color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=8)
    style_border.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

    columns = [
        ("№", 5),
        ("Пациент", 20),
        ("Дата рожд.", 7),
        ("Дата вып.", 7),
        ("Код", 8),
        ("Код НМУ", 11),
        ("Наименование услуги", 19),
        ("Лаб.номер", 13),
        ("Цена", 5),
        ("Кол", 5),
        ("Стоимость, руб", 7),
    ]

    row = 13
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border
    return ws1


def fill_billing(ws1, data, row=14):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style="thin", color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=8)
    style_border1.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

    style_border2 = NamedStyle(name="style_border2")
    bd = Side(style="thin", color="000000")
    style_border2.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border2.font = Font(bold=False, size=10)
    style_border2.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    logo_exists = os.path.exists(f"{BASE_DIR}/forms/xlsx/media/logo.png")
    if logo_exists:
        logo = Image(os.path.join(BASE_DIR, "forms", "xlsx", "media", "logo.png"))
        logo.height = 130
        logo.width = 920
        ws1.add_image(logo, "A1")
    ws1.merge_cells("A9:L11")
    megre_cell = ws1["A9"]
    megre_cell.value = "Реестр № 1YYYY от 00 марта 2024 года \n оказанных медицинских услуг по договору №00 - 00/00/2024 \n с 00.00.00 по 00.00.00"
    megre_cell.style = style_border2
    r = row
    for val in data:
        ws1.cell(row=r, column=1).value = val.get("serialNumber")
        ws1.cell(row=r, column=2).value = val.get("patientFio")
        ws1.cell(row=r, column=3).value = val.get("patientBirthDay")
        ws1.cell(row=r, column=4).value = val.get("executeDate")
        ws1.cell(row=r, column=5).value = val.get("internalId")
        ws1.cell(row=r, column=6).value = val.get("codeNMU")
        ws1.cell(row=r, column=7).value = val.get("researchTitle")
        ws1.cell(row=r, column=8).value = val.get("tubeNumber")
        ws1.cell(row=r, column=9).value = val.get("coast")
        ws1.cell(row=r, column=10).value = 1
        ws1.cell(row=r, column=11).value = val.get("summ") if val.get("summ") else val.get("coast")
        for i in range(1, 12):
            ws1.cell(row=r, column=i).style = style_border1
        r += 1
    return ws1
