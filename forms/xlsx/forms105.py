import json

from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter
from directions.models import Issledovaniya
from podrazdeleniya.models import Podrazdeleniya


def form_01(request_data) -> Workbook:
    direction_list = request_data.get("direction_list")
    department_pk = request_data.get("department_pk")
    podr = Podrazdeleniya.objects.filter(pk=int(department_pk)).first()
    d1 = request_data.get("start_date")
    d2 = request_data.get("end_date")
    if isinstance(direction_list, str):
        direction_list = json.loads(direction_list)
    if not direction_list:
        direction_list = []
    wb = openpyxl.Workbook()
    wb.remove(wb.get_sheet_by_name("Sheet"))
    ws = wb.create_sheet("Выписки")
    iss_objs = Issledovaniya.objects.filter(napravleniye_id__in=direction_list)
    result_detail = [
        {
            "doctor": i.doc_confirmation.get_fio(),
            "patient": i.napravleniye.client.get_fio_w_card(),
            "direction_main": i.napravleniye.parent.napravleniye_id,
            "date_extract": i.medical_examination,
        }
        for i in iss_objs
    ]
    style_border = NamedStyle(name="style_border_ca")
    bd = Side(style="thin", color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=12)
    style_border.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    style_border2 = NamedStyle(name="style_border2")
    style_border2.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border2.font = Font(bold=False, size=12)
    style_border2.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    ws.cell(row=2, column=1).value = "Период:"
    ws.cell(row=3, column=1).value = f"c {d1} по {d2}"
    ws.cell(row=4, column=1).value = f"{podr.title}"

    columns = [
        ('№ п/п.', 5),
        ('Врач', 50),
        ('Пациент', 50),
        ('История', 25),
        ('Дата выписки', 25),
    ]
    row = 6
    for idx, column in enumerate(columns, 1):
        ws.cell(row=row, column=idx).value = column[0]
        ws.column_dimensions[get_column_letter(idx)].width = column[1]
        ws.cell(row=row, column=idx).style = style_border

    step = 0
    for i in result_detail:
        step +=1
        row += 1
        ws.cell(row=row, column=1).value = step
        ws.cell(row=row, column=2).value = i.get("doctor")
        ws.cell(row=row, column=3).value = i.get("patient")
        ws.cell(row=row, column=4).value = i.get("direction_main")
        ws.cell(row=row, column=5).value = i.get("date_extract").strftime("%d.%m.%Y")

        for c in range(5):
            ws.cell(row=row, column=c + 1).style = style_border2

    return wb
