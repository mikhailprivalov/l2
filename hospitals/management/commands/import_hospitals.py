from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from hospitals.models import Hospitals


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл с медорганизациями
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        print(ws)
        starts = False
        code_tfoms, full_title, short_title, address = '', '', '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "код" in cells and "краткое" in cells and "полное" in cells and "адрес" in cells:
                    starts = True
                    code_tfoms = cells.index("код")
                    short_title = cells.index("краткое")
                    full_title = cells.index("полное")
                    address = cells.index("адрес")
            else:
                if Hospitals.objects.filter(code_tfoms=cells[code_tfoms]).exists():
                    continue
                Hospitals.objects.create(full_title=cells[full_title], short_title=cells[short_title], code_tfoms=cells[code_tfoms], address=cells[address])
                print(f'добавлено МО:{cells[code_tfoms]}:{cells[full_title]}:{cells[short_title]}:{cells[address]}:')
