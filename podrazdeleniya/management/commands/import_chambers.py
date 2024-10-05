from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from podrazdeleniya.models import Podrazdeleniya, Chamber, Bed


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        department_title_idx, chamber_title_idx, bed_number_idx = '', '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Название палаты" in cells:
                    department_title_idx = cells.index("Отделение")
                    chamber_title_idx = cells.index("Название палаты")
                    bed_number_idx = cells.index("Номер койки")
                    starts = True
            else:
                current_department_title = cells[department_title_idx].strip()
                current_chamber_title = cells[chamber_title_idx].strip()
                current_bed_number = cells[bed_number_idx].strip()
                if not current_department_title or not current_chamber_title or not current_bed_number:
                    self.stdout.write(f'Пустая строка')
                    continue
                department = Podrazdeleniya.objects.filter(title__iexact=current_department_title).first()
                if not department:
                    self.stdout.write(f' Подразделения {current_department_title} - нет')
                    continue
                chamber = Chamber.objects.filter(title__iexact=current_chamber_title).first()
                if chamber:
                    bed = Bed.objects.filter(chamber_id=chamber.pk, bed_number=current_bed_number).first()
                    if not bed:
                        bed = Bed(chamber_id=chamber.pk, bed_number=current_bed_number)
                        bed.save()
                        self.stdout.write(f'Кровать с номером {current_bed_number} - добавлена')
                else:
                    chamber = Chamber(title=current_chamber_title, podrazdelenie_id=department.pk)
                    chamber.save()
                    self.stdout.write(f'Палата с названием {current_chamber_title} - добавлена')
                    bed = Bed(chamber_id=chamber.pk, bed_number=current_bed_number)
                    bed.save()
                    self.stdout.write(f'Кровать с номером {current_bed_number} - добавлена')
