from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from podrazdeleniya.models import Podrazdeleniya


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл
        """
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        title, type_podr = '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Подразделение" in cells:
                    title = cells.index("Подразделение")
                    type_podr = cells.index("Тип")
                    starts = True
            else:
                podrazdeleniye_obj = Podrazdeleniya.objects.filter(title=cells[title].strip()).first()
                if not podrazdeleniye_obj:
                    Podrazdeleniya(title=cells[title], p_type=int(cells[type_podr].strip())).save()
                    self.stdout.write(f'Подразделение добавлено - {cells[title]}')
