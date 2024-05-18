import csv
import datetime
import io
import logging
import re
import tempfile
from copy import deepcopy
from sys import stdout

from django.http import HttpRequest, JsonResponse
from django.utils.module_loading import import_string

from api.directions.views import eds_documents
from api.models import Application

from api.parse_file.pdf import extract_text_from_pdf
import simplejson as json

from api.patients.sql_func import search_cards_by_numbers
from api.patients.views import patients_search_card
from api.views import endpoint
from openpyxl import load_workbook
from appconf.manager import SettingManager
from contracts.models import PriceCoast, Company, MedicalExamination, CompanyDepartment, PriceName
import directions.models as directions
from directory.models import SetOrderResearch, Researches, ParaclinicInputGroups, ParaclinicInputField
from directory.sql_func import is_paraclinic_filter_research, is_lab_filter_research
from ecp_integration.integration import fill_slot_from_xlsx
from hospitals.models import Hospitals
from laboratory.settings import CONTROL_AGE_MEDEXAM, DAYS_AGO_SEARCH_RESULT
from results.sql_func import check_lab_instrumental_results_by_cards_and_period
from statistic.views import commercial_offer_xls_save_file, data_xls_save_file, data_xls_save_headers_file
from users.models import AssignmentResearches, DoctorProfile
from clients.models import Individual, HarmfulFactor, PatientHarmfullFactor, Card, CardBase, DocumentType, Document
from integration_framework.views import check_enp
from utils.dates import age_for_year, normalize_dots_date
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
from django.db import transaction
from django.utils import timezone
from utils.nsi_directories import NSI
from django.contrib.auth.decorators import login_required

logger = logging.getLogger("IF")


def dnk_covid(request):
    prefixes = []
    key_dnk = SettingManager.get("dnk_kovid", default="false", default_type="s")
    to_return = None
    for x in "ABCDEF":
        prefixes.extend([f"{x}{i}" for i in range(1, 13)])
    file = request.FILES["file"]
    if file.content_type == "application/pdf" and file.size < 100000:
        with tempfile.TemporaryFile() as fp:
            fp.write(file.read())
            text = extract_text_from_pdf(fp)
        if text:
            text = text.replace("\n", "").split("Коронавирусы подобные SARS-CoVВККоронавирус SARS-CoV-2")
        to_return = []
        if text:
            for i in text:
                k = i.split("N")
                if len(k) > 1 and k[1].split(" ")[0].isdigit():
                    result = json.dumps({"pk": k[1].split(" ")[0], "result": [{"dnk_SARS": "Положительно" if "+" in i else "Отрицательно"}]})
                    to_return.append({"pk": k[1].split(" ")[0], "result": "Положительно" if "+" in i else "Отрицательно"})
                    http_func({"key": key_dnk, "result": result}, request.user)

    return to_return


def http_func(data, user):
    http_obj = HttpRequest()
    http_obj.POST.update(data)
    http_obj.user = user
    return endpoint(http_obj)


def create_patient(family_data, name_data, patronymic_data, birthday_data, gender_data):
    patient_indv = Individual(
        family=family_data,
        name=name_data,
        patronymic=patronymic_data,
        birthday=birthday_data,
        sex=gender_data,
    )
    patient_indv.save()
    patient_card = Card.add_l2_card(individual=patient_indv)
    return patient_card


def get_background_token():
    application = Application.objects.filter(active=True, is_background_worker=True).first()
    if application:
        bearer_token = f"Bearer {application.key}"
    else:
        new_application = Application(name="background_worker", is_background_worker=True)
        new_application.save()
        bearer_token = f"Bearer {new_application.key}"
    return bearer_token


def search_patient(snils_data, request_user, family_data, name_data, patronymic_data, birthday_data):
    patient_card = None
    bearer_token = get_background_token()
    params = {"enp": "", "snils": snils_data, "check_mode": "l2-snils"}
    request_obj = HttpRequest()
    request_obj._body = params
    request_obj.user = request_user
    request_obj.method = "POST"
    request_obj.META["HTTP_AUTHORIZATION"] = bearer_token
    current_patient = None
    if snils_data and snils_data != "None":
        current_patient = check_enp(request_obj)
    if not current_patient or current_patient.data.get("message"):
        patient_card = search_by_fio(request_obj, family_data, name_data, patronymic_data, birthday_data)
        if patient_card is None:
            possible_family = find_and_replace(family_data, "е", "ё")
            patient_card = search_by_possible_fio(request_obj, name_data, patronymic_data, birthday_data, possible_family)
            if patient_card is None:
                return patient_card
    elif current_patient.data.get("patient_data") and type(current_patient.data.get("patient_data")) != list:
        patient_card_pk = current_patient.data["patient_data"]["card"]
        patient_card = Card.objects.filter(pk=patient_card_pk).first()
    else:
        patient_card = Individual.import_from_tfoms(current_patient.data["patient_data"], None, None, None, True)

    return patient_card


def find_factors(harmful_factors: list):
    if not harmful_factors:
        return None
    incorrect_factor = []
    harmful_factors_data = []
    for i in harmful_factors:
        current_code = i.replace(" ", "")
        harmful_factor = HarmfulFactor.objects.filter(title=current_code).first()
        if harmful_factor:
            harmful_factors_data.append({"factorId": harmful_factor.pk})
        else:
            incorrect_factor.append(f"{current_code}")

    return harmful_factors_data, incorrect_factor


def add_factors_data(patient_card: Card, position: str, factors_data: list, exam_data: str, company_inn: str, department: str):
    try:
        PatientHarmfullFactor.save_card_harmful_factor(patient_card.pk, factors_data)
        company_obj = Company.objects.filter(inn=company_inn).first()
        department_obj = CompanyDepartment.objects.filter(company_id=company_obj.pk, title=department).first()
        if department_obj:
            patient_card.work_department_db_id = department_obj.pk
        else:
            new_department = CompanyDepartment.save_department(company_obj.pk, department)
            patient_card.work_department_db_id = new_department.pk
        patient_card.work_position = position.strip()
        patient_card.work_place_db = company_obj
        patient_card.save()
        MedicalExamination.save_examination(patient_card, company_obj, exam_data)
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "message": e}


def add_factors_from_file(request):
    incorrect_patients = []
    company_inn = request.POST.get("companyInn", None)
    company_file = request.FILES["file"]
    wb = load_workbook(filename=company_file)
    ws = wb.worksheets[0]
    starts = False
    snils, fio, birthday, gender, inn_company, code_harmful, position, examination_date, department = (
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    for index, row in enumerate(ws.rows, 1):
        cells = [str(x.value) for x in row]
        if not starts:
            if "код вредности" in cells:
                snils = cells.index("снилс")
                fio = cells.index("фио")
                birthday = cells.index("дата рождения")
                gender = cells.index("пол")
                inn_company = cells.index("инн организации")
                code_harmful = cells.index("код вредности")
                position = cells.index("должность")
                examination_date = cells.index("дата мед. осмотра")
                department = cells.index("подразделение")
                starts = True
        else:
            if company_inn != cells[inn_company].strip():
                incorrect_patients.append({"fio": cells[fio], "reason": "ИНН организации не совпадает"})
                continue
            snils_data = cells[snils].replace("-", "").replace(" ", "")
            fio_data, family_data, name_data, patronymic_data = None, None, None, None
            if cells[fio] and cells[fio] != "None":
                fio_data = cells[fio].split(" ")
                family_data = fio_data[0]
                name_data = fio_data[1]
                if len(fio_data) > 2:
                    patronymic_data = fio_data[2]
            birthday_data = cells[birthday].split(" ")[0]
            code_harmful_data = cells[code_harmful].split(",")
            exam_data = cells[examination_date].split(" ")[0]
            try:
                datetime.datetime.strptime(birthday_data, '%Y-%m-%d')
                datetime.datetime.strptime(exam_data, '%Y-%m-%d')
            except ValueError as e:
                incorrect_patients.append({"fio": cells[fio], "reason": f"Неверный формат даты/несуществующая дата в файле: {e}"})
                continue
            gender_data = cells[gender][0]
            department_data = cells[department]
            if fio_data is None and snils_data is None:
                incorrect_patients.append({"fio": f"Строка: {index}", "reason": "Отсутствует данные"})
                continue
            patient_card = search_patient(snils_data, request.user, family_data, name_data, patronymic_data, birthday_data)
            if patient_card is None:
                patient_card = create_patient(family_data, name_data, patronymic_data, birthday_data, gender_data)
            harmful_factors_data, incorrect_factor = find_factors(code_harmful_data)
            if incorrect_factor:
                incorrect_patients.append({"fio": cells[fio], "reason": f"Неверные факторы: {incorrect_factor}"})
            patient_updated = add_factors_data(patient_card, cells[position], harmful_factors_data, exam_data, company_inn, department_data)
            if not patient_updated["ok"]:
                incorrect_patients.append({"fio": cells[fio], "reason": f"Сохранение не удалось, ошибка: {patient_updated['message']}"})

    return incorrect_patients


@csrf_exempt
def load_file(request):
    link = ""
    req_data = dict(request.POST)
    if req_data.get("isGenCommercialOffer")[0] == "true":
        results = gen_commercial_offer(request)
        link = "open-xls"
    elif req_data.get("isWritePatientEcp")[0] == "true":
        results = write_patient_ecp(request)
        link = "open-xls"
    elif req_data.get("researchSet")[0] != "-1":
        research_set = int(req_data.get("researchSet")[0])
        results = data_research_exam_patient(request, research_set)
        link = "open-xls"
    elif len(req_data.get("companyInn")[0]) != 0:
        results = add_factors_from_file(request)
        return JsonResponse({"ok": True, "results": results, "company": True})
    elif req_data.get("isLoadResultService")[0] == "true":
        id_research = int(req_data.get("idService")[0])
        research = Researches.objects.filter(pk=id_research).first()
        id_doc_profile = int(req_data.get("idDoctorProfile")[0])
        doc_profile = DoctorProfile.objects.filter(pk=id_doc_profile).first()
        if not research or not doc_profile:
            return JsonResponse({"ok": True, "results": "", "company": True})
        else:
            results = auto_load_result(request, research, doc_profile)
            return JsonResponse({"ok": True, "results": results, "company": True})
    else:
        results = dnk_covid(request)
    return JsonResponse({"ok": True, "results": results, "link": link})


def gen_commercial_offer(request):
    file_data = request.FILES["file"]
    selected_price = request.POST.get("selectedPrice")

    wb = load_workbook(filename=file_data)
    ws = wb[wb.sheetnames[0]]
    starts = False
    counts_research = {}
    patients = []

    for row in ws.rows:
        cells = [str(x.value) for x in row]
        if not starts:
            if "код вредности" in cells:
                starts = True
                harmful_factor = cells.index("код вредности")
                fio = cells.index("фио")
                born = cells.index("дата рождения")
                position = cells.index("должность")
                sex = cells.index("пол")
        else:
            harmful_factor_data = [i.replace(" ", "") for i in cells[harmful_factor].split(",")]
            born_data = cells[born].split(" ")[0]
            age = -1
            if born_data != "None":
                age = age_for_year(normalize_dots_date(born_data))
                if "м" in cells[sex]:
                    adds_harmfull = CONTROL_AGE_MEDEXAM.get("м")
                else:
                    adds_harmfull = CONTROL_AGE_MEDEXAM.get("ж")
                if adds_harmfull:
                    for i in sorted(adds_harmfull.keys()):
                        if age < i:
                            harmful_factor_data.append(adds_harmfull[i])
                            break
            templates_data = HarmfulFactor.objects.values_list("template_id", flat=True).filter(title__in=harmful_factor_data)
            researches_data = AssignmentResearches.objects.values_list("research_id", flat=True).filter(template_id__in=templates_data)
            researches_data = set(researches_data)
            for r in researches_data:
                if counts_research.get(r):
                    counts_research[r] += 1
                else:
                    counts_research[r] = 1
            patients.append({"fio": cells[fio], "born": born_data, "harmful_factor": cells[harmful_factor], "position": cells[position], "researches": researches_data, "age": age})

    price_data = PriceCoast.objects.filter(price_name__id=selected_price, research_id__in=list(counts_research.keys()))
    data_price = [{"title": k.research.title, "code": k.research.code, "count": counts_research[k.research.pk], "coast": k.coast} for k in price_data]
    research_price = {d.research.pk: f"{d.research.title}@{d.coast}" for d in price_data}
    file_name = commercial_offer_xls_save_file(data_price, patients, research_price)
    return file_name


def auto_load_result_from_xlsx(request, research, doc_profile):
    file_data = request.FILES["file"]
    financing_source_title = request.POST.get("financingSourceTitle")
    title_fields = request.POST.get("titleFields")
    incorrect_patients = auto_create_protocol(title_fields, file_data, financing_source_title, research, doc_profile)
    return incorrect_patients


def auto_load_result(request, research, doc_profile):
    file_data = request.FILES["file"]
    financing_source_title = request.POST.get("financingSourceTitle")
    title_fields = request.POST.get("titleFields")
    title_data = [i.strip() for i in title_fields.split(",")]
    wb = load_workbook(filename=file_data)
    ws = wb[wb.sheetnames[0]]
    starts = False
    snils_type = DocumentType.objects.filter(title__startswith="СНИЛС").first()
    enp_type = DocumentType.objects.filter(title__startswith="Полис ОМС").first()
    date_variant_title_field = ["Дата осмотра"]
    index_cell = {}
    incorrect_patients = []
    for row in ws.rows:
        cells = [str(x.value) for x in row]
        if not starts:
            if "fio" in cells:
                starts = True
                for t in title_data:
                    index_cell[t] = cells.index(t)
        else:
            data_result = {}
            for i_cell in index_cell.keys():
                data_result[i_cell] = cells[index_cell[i_cell]]
            if len(data_result["snils"]) < 11 or len(data_result["enp"]) < 16:
                continue
            snils = data_result["snils"].replace("-", "").replace(" ", "")
            birthday = normalize_dots_date(data_result["birthday"].split(" ")[0])
            sex = data_result["sex"].lower()

            individual, enp_doc, snils_doc = None, None, None
            if data_result["enp"]:
                individuals = Individual.objects.filter(tfoms_enp=data_result["enp"])
                if not individuals.exists():
                    individuals = Individual.objects.filter(document__number=data_result["enp"]).filter(
                        Q(document__document_type__title="Полис ОМС") | Q(document__document_type__title="ЕНП")
                    )
                    individual = individuals.first()

            if not individual:
                individual = Individual(
                    family=data_result["lastname"],
                    name=data_result["firstname"],
                    patronymic=data_result["patronymic"],
                    birthday=birthday,
                    sex=sex,
                )
                individual.save()

            if not Document.objects.filter(individual=individual, document_type=snils_type).exists():
                snils_doc = Document(
                    individual=individual,
                    document_type=snils_type,
                    number=snils,
                )
                snils_doc.save()
            else:
                snils_doc = Document.objects.filter(individual=individual, document_type=snils_type).first()
            if not Document.objects.filter(individual=individual, document_type=enp_type).exists():
                enp_doc = Document(
                    individual=individual,
                    document_type=enp_type,
                    number=data_result["enp"],
                )
                enp_doc.save()
            else:
                enp_doc = Document.objects.filter(individual=individual, document_type=enp_type).first()

            if not Card.objects.filter(base__internal_type=True, individual=individual, is_archive=False).exists():
                card = Card.add_l2_card(individual, polis=enp_doc, snils=snils_doc)
            else:
                card = Card.objects.filter(base__internal_type=True, individual=individual, is_archive=False).first()
            card.main_address = data_result["address"]
            card.save(update_fields=["main_address"])

            financing_source = directions.IstochnikiFinansirovaniya.objects.filter(title__iexact=financing_source_title, base__internal_type=True).first()

            try:
                with transaction.atomic():
                    direction = directions.Napravleniya.objects.create(
                        client=card,
                        istochnik_f=financing_source,
                        polis_who_give=card.polis.who_give if card.polis else None,
                        polis_n=card.polis.number if card.polis else None,
                        hospital=doc_profile.hospital,
                        total_confirmed=True,
                        last_confirmed_at=timezone.now(),
                        eds_required_signature_types=['Врач', 'Медицинская организация'],
                    )

                    iss = directions.Issledovaniya.objects.create(
                        napravleniye=direction,
                        research=research,
                        time_confirmation=timezone.now(),
                        time_save=timezone.now(),
                        doc_confirmation=doc_profile,
                        doc_save=doc_profile,
                        doc_confirmation_string=f"{doc_profile.get_fio_parts()}",
                    )
                    for group in ParaclinicInputGroups.objects.filter(research=research):
                        for f in ParaclinicInputField.objects.filter(group=group):
                            if data_result.get(f.title, None):
                                if f.title.strip() in date_variant_title_field:
                                    tmp_val = data_result[f.title]
                                    data_result[f.title] = normalize_dots_date(tmp_val)
                                if f.title.strip() == "Диагноз":
                                    tmp_val = data_result[f.title].strip()
                                    diag = directions.Diagnoses.objects.filter(d_type="mkb10.4", code=tmp_val.upper()).first()
                                    if not diag:
                                        incorrect_patients.append({"fio": data_result["fio"], "reason": "Неверные данные:"})
                                    res = f'"code": "{tmp_val}", "title": "{diag.title}", "id": "{diag.id}"'
                                    res = "{" + res + "}"
                                    data_result[f.title] = res
                                if f.field_type == 28:
                                    for nsi_key in NSI.values():
                                        if f.title == nsi_key.get("title"):
                                            for key, val in nsi_key.get("values").items():
                                                if val == data_result[f.title].strip():
                                                    res = f'"code": "{key}", "title": "{val}"'
                                                    res = "{" + res + "}"
                                                    data_result[f.title] = res
                                                    continue
                                directions.ParaclinicResult(issledovaniye=iss, field=f, field_type=f.field_type, value=data_result.get(f.title)).save()

                    eds_documents_data = json.dumps({'pk': direction.pk})
                    eds_documents_obj = HttpRequest()
                    eds_documents_obj._body = eds_documents_data
                    eds_documents_obj.user = request.user
                    eds_documents(eds_documents_obj)

            except Exception as e:
                logger.exception(e)
                message = "Серверная ошибка"
                return {"ok": False, "message": message}

    return incorrect_patients


def auto_create_protocol(title_fields, file_data, financing_source_title, research, doc_profile):
    title_data = [i.strip() for i in title_fields.split(",")]
    wb = load_workbook(filename=file_data)
    ws = wb[wb.sheetnames[0]]
    starts = False
    snils_type = DocumentType.objects.filter(title__startswith="СНИЛС").first()
    enp_type = DocumentType.objects.filter(title__startswith="Полис ОМС").first()
    date_variant_title_field = ["Дата осмотра"]
    index_cell = {}
    incorrect_patients = []
    step = 0
    for row in ws.rows:
        cells = [str(x.value) for x in row]
        if not starts:
            if "fio" in cells:
                starts = True
                for t in title_data:
                    index_cell[t] = cells.index(t)
        else:
            data_result = {}
            for i_cell in index_cell.keys():
                data_result[i_cell] = cells[index_cell[i_cell]]
            if len(data_result["snils"]) < 11 or len(data_result["enp"]) < 16:
                continue
            snils = data_result["snils"].replace("-", "").replace(" ", "")
            birthday = normalize_dots_date(data_result["birthday"].split(" ")[0])
            sex = data_result["sex"].lower()

            individual, enp_doc, snils_doc = None, None, None
            if data_result["enp"]:
                individuals = Individual.objects.filter(tfoms_enp=data_result["enp"])
                if not individuals.exists():
                    individuals = Individual.objects.filter(document__number=data_result["enp"]).filter(
                        Q(document__document_type__title="Полис ОМС") | Q(document__document_type__title="ЕНП")
                    )
                    individual = individuals.first()

            if not individual:
                individual = Individual(
                    family=data_result["lastname"],
                    name=data_result["firstname"],
                    patronymic=data_result["patronymic"],
                    birthday=birthday,
                    sex=sex,
                )
                individual.save()

            if not Document.objects.filter(individual=individual, document_type=snils_type).exists():
                snils_doc = Document(
                    individual=individual,
                    document_type=snils_type,
                    number=snils,
                )
                snils_doc.save()
            else:
                snils_doc = Document.objects.filter(individual=individual, document_type=snils_type).first()
            if not Document.objects.filter(individual=individual, document_type=enp_type).exists():
                enp_doc = Document(
                    individual=individual,
                    document_type=enp_type,
                    number=data_result["enp"],
                )
                enp_doc.save()
            else:
                enp_doc = Document.objects.filter(individual=individual, document_type=enp_type).first()

            if not Card.objects.filter(base__internal_type=True, individual=individual, is_archive=False).exists():
                card = Card.add_l2_card(individual, polis=enp_doc, snils=snils_doc)
            else:
                card = Card.objects.filter(base__internal_type=True, individual=individual, is_archive=False).first()
            card.main_address = data_result["address"]
            card.save(update_fields=["main_address"])

            financing_source = directions.IstochnikiFinansirovaniya.objects.filter(title__iexact=financing_source_title, base__internal_type=True).first()

            try:
                with transaction.atomic():
                    direction = directions.Napravleniya.objects.create(
                        client=card,
                        istochnik_f=financing_source,
                        polis_who_give=card.polis.who_give if card.polis else None,
                        polis_n=card.polis.number if card.polis else None,
                        hospital=doc_profile.hospital,
                        total_confirmed=True,
                        last_confirmed_at=timezone.now(),
                        eds_required_signature_types=["Врач", "Медицинская организация"],
                    )

                    iss = directions.Issledovaniya.objects.create(
                        napravleniye=direction,
                        research=research,
                        time_confirmation=timezone.now(),
                        time_save=timezone.now(),
                        doc_confirmation=doc_profile,
                        doc_save=doc_profile,
                        doc_confirmation_string=f"{doc_profile.get_fio_parts()}",
                    )

                    for group in ParaclinicInputGroups.objects.filter(research=research):
                        for f in ParaclinicInputField.objects.filter(group=group):
                            if data_result.get(f.title, None):
                                if f.title.strip() in date_variant_title_field:
                                    tmp_val = data_result[f.title]
                                    data_result[f.title] = normalize_dots_date(tmp_val)
                                if f.title.strip() == "Диагноз":
                                    tmp_val = data_result[f.title].strip()
                                    diag = directions.Diagnoses.objects.filter(d_type="mkb10.4", code=tmp_val.upper(), hide=False).first()
                                    if not diag:
                                        incorrect_patients.append({"fio": data_result["fio"], "reason": "Неверные данные:"})
                                    res = f'"code": "{tmp_val}", "title": "{diag.title}", "id": "{diag.id}"'
                                    res = "{" + res + "}"
                                    data_result[f.title] = res
                                if f.field_type == 28:
                                    for nsi_key in NSI.values():
                                        if f.title == nsi_key.get("title"):
                                            for key, val in nsi_key.get("values").items():
                                                if val == data_result[f.title].strip():
                                                    res = f'"code": "{key}", "title": "{val}"'
                                                    res = "{" + res + "}"
                                                    data_result[f.title] = res
                                                    continue
                                directions.ParaclinicResult(issledovaniye=iss, field=f, field_type=f.field_type, value=data_result.get(f.title)).save()

                    eds_documents_data = json.dumps({"pk": direction.pk})
                    eds_documents_obj = HttpRequest()
                    eds_documents_obj._body = eds_documents_data
                    eds_documents_obj.user = doc_profile.user
                    eds_documents(eds_documents_obj)

            except Exception as e:
                logger.exception(e)
                message = "Серверная ошибка"
                return {"ok": False, "message": message}
            step += 1
            print(f"Карта: {card} Направление: {direction.pk}, Шаг: {step}")  # noqa: T001
    return incorrect_patients


def search_by_fio(request_obj, family, name, patronymic, birthday):
    patient_card = None
    params_tfoms = {
        "enp": "",
        "family": family,
        "name": name,
        "bd": birthday,
        "check_mode": "l2-enp-full",
    }
    params_internal = {
        "type": CardBase.objects.get(internal_type=True).pk,
        "extendedSearch": True,
        "form": {
            "family": family,
            "name": name,
            "patronymic": patronymic,
            "birthday": birthday,
            "archive": False,
        },
        "limit": 1,
    }
    request_obj._body = params_tfoms
    current_patient = check_enp(request_obj)
    if current_patient.data.get("message"):
        request_obj._body = json.dumps(params_internal)
        data = patients_search_card(request_obj)
        results_json = json.loads(data.content.decode("utf-8"))
        if len(results_json["results"]) > 0:
            patient_card_pk = results_json["results"][0]["pk"]
            patient_card = Card.objects.filter(pk=patient_card_pk).first()
    elif len(current_patient.data["list"]) > 1:
        return patient_card
    else:
        patient_card = Individual.import_from_tfoms(current_patient.data["list"], None, None, None, True)
    return patient_card


def find_and_replace(text, symbol1, symbol2):
    result = []
    for i in range(len(text)):
        if text[i] == symbol1:
            current_text = text[0:i] + symbol2 + text[i + 1 :]
            result.append(current_text)
        elif text[i] == symbol2:
            current_text = text[0:i] + symbol1 + text[i + 1 :]
            result.append(current_text)
    return result


def search_by_possible_fio(request_obj, name, patronymic, birthday, possible_family):
    if not possible_family:
        return None
    patient_card = None
    for i in possible_family:
        current_family = i
        patient_card = search_by_fio(request_obj, current_family, name, patronymic, birthday)
        if patient_card is not None:
            break
    return patient_card


def load_price_coasts(price_id: int, file):
    price = PriceName.objects.filter(pk=price_id).first()
    if not price:
        return False
    wb = load_workbook(filename=file)
    ws = wb[wb.sheetnames[0]]
    internal_code_idx, title_idx, code_idx, coast_idx = '', '', '', ''
    starts = False
    for row in ws.rows:
        cells = [str(x.value) for x in row]
        if not starts:
            if "Код по прайсу" in cells:
                internal_code_idx = cells.index("Код по прайсу")
                title_idx = cells.index("Услуга")
                code_idx = cells.index("Код ОКМУ")
                service_cols = [internal_code_idx, title_idx, code_idx]
                coast_idx = None
                for idx, cell in enumerate(cells):
                    if idx in service_cols:
                        continue
                    coast_idx = idx
                price_title = f"{price.title}-{price.symbol_code}"
                if price_title != cells[coast_idx].strip():
                    return False
                starts = True
        else:
            internal_code = cells[internal_code_idx].strip()
            coast = cells[coast_idx].strip()
            if internal_code == "None" or coast == "None" or coast == "0":
                continue
            service = Researches.objects.filter(internal_code=internal_code).first()
            if not service:
                continue
            current_coast = PriceCoast.objects.filter(price_name_id=price.pk, research_id=service.pk).first()
            if current_coast:
                if current_coast.coast != coast:
                    current_coast.coast = coast
                    current_coast.save()
            else:
                new_coast = PriceCoast(price_name_id=price.pk, research_id=service.pk, coast=coast)
                new_coast.save()
    return True


@login_required
def load_csv(request):
    file_data = request.FILES["file"]
    selected_price = request.POST.get('selectedPrice')
    if selected_price:
        result = False
        try:
            result = load_price_coasts(selected_price, file_data)
        except:
            stdout.write('Файл не загружен')
        if result:
            return JsonResponse({"ok": True, "results": [], "methods": []})
        else:
            return JsonResponse({"ok": False, "results": [], "methods": []})
    file_data = file_data.read().decode("utf-8")
    io_string = io.StringIO(file_data)

    data = csv.reader(io_string, delimiter="\t")
    header = next(data)

    application = None

    for app in Application.objects.filter(csv_header__isnull=False).exclude(csv_header__exact=""):
        if app.csv_header in header:
            application = app
            break

    if application is None or application.csv_header not in header:
        return JsonResponse({"ok": False, "message": "Файл не соответствует ни одному приложению"})

    app_key = application.key.hex

    method = re.search(r"^(.+)\s\d+\s.*$", header[1]).group(1)
    results = []
    pattern = re.compile(r"^\d+$")

    for row in data:
        if len(row) > 5 and pattern.match(row[2]):
            r = {
                "pk": row[2],
                "result": row[5],
            }
            result = json.dumps({"pk": r["pk"], "result": {method: r["result"]}})
            resp = http_func({"key": app_key, "result": result, "message_type": "R"}, request.user)
            resp = json.loads(resp.content)
            results.append(
                {
                    "pk": row[2],
                    "result": row[5],
                    "comment": "успешно" if resp["ok"] else "не удалось сохранить результат",
                }
            )

    return JsonResponse({"ok": True, "results": results, "method": method})


@login_required
def load_equipment(request):
    file_data = request.FILES["file"]
    equipment = request.POST.get("equipment")

    application = Application.objects.filter(pk=int(equipment)).first()
    app_key = application.key.hex
    wb = load_workbook(filename=file_data)
    ws = wb[wb.sheetnames[0]]
    test = None
    search_result = "Результат,"
    search_tubes_number = "Схема"
    tube_list = []
    result_list = []
    for row in ws.rows:
        for cell in row:
            data_from_cell = str(cell.value)
            result = re.findall(search_result, data_from_cell)
            if len(result) > 0:
                num_row = cell.row
                result_data = ws[f"A{num_row + 1}:M{num_row+10}"]
                result_list = [[j.value for j in k] for k in result_data]

            result = re.findall(search_tubes_number, data_from_cell)
            if len(result) > 0:
                num_row = cell.row
                tube_data = ws[f"A{num_row + 1}:M{num_row+10}"]
                tube_list = [[j.value for j in k] for k in tube_data]

            if cell.value == "Тест:":
                num_row = cell.row
                test = ws.cell(row=num_row, column=7).value
                test = test.replace(' ', '')
    is_set_analit = False
    for i in result_list[-1]:
        if i:
            is_set_analit = True

    step = 0
    pare_result = []
    for i in tube_list:
        second_step = 0
        for tube in i:
            if is_set_analit:
                pare_result.append({"pk": tube, "result": {"method": f"{test}-{result_list[-1][second_step]}", "value": result_list[step][second_step]}})
            else:
                pare_result.append({"pk": tube, "result": {"method": test, "value": result_list[step][second_step]}})
            second_step += 1
        step += 1

    results = []

    for data in pare_result:
        if data.get("pk"):
            try:
                pk = int(data.get("pk"))
                if pk > 1000:
                    r = {
                        "pk": int(data.get("pk")),
                        "method": data["result"]["method"],
                        "value": data["result"]["value"],
                    }
                    result = json.dumps({"pk": r["pk"], "result": {r["method"]: r["value"]}})
                    resp = http_func({"key": app_key, "result": result, "message_type": "R"}, request.user)
                    resp = json.loads(resp.content)
                    results.append(
                        {
                            "pk": r["pk"],
                            "result": {r["method"]: r["value"]},
                            "comment": "успешно" if resp["ok"] else "не удалось сохранить результат",
                        }
                    )
            except:
                continue

    return JsonResponse({"ok": True, "results": results})


def write_patient_ecp(request):
    file_data = request.FILES["file"]
    wb = load_workbook(filename=file_data)
    ws = wb[wb.sheetnames[0]]
    starts = False
    patients = []
    for row in ws.rows:
        cells = [str(x.value) for x in row]
        if not starts:
            if "врач" in cells:
                starts = True
                family = cells.index("фaмилия")
                name = cells.index("имя")
                patronymic = cells.index("отчество")
                born = cells.index("дата рождения")
                doctor = cells.index("врач")
        else:
            born_data = cells[born].split(" ")[0]
            if "." in born_data:
                born_data = normalize_dots_date(born_data)
            patient = {"family": cells[family], "name": cells[name], "patronymic": cells[patronymic], "birthday": born_data, "snils": ""}
            result = fill_slot_from_xlsx(cells[doctor], patient)
            is_write = "Ошибка"
            message = ""
            if result:
                if result.get("register"):
                    is_write = "записан"
                if result.get("message"):
                    message = result.get("message", "")
            patients.append({**patient, "is_write": is_write, "doctor": cells[doctor], "message": message})
    file_name = data_xls_save_file(patients, "Запись")
    return file_name


def data_research_exam_patient(request, set_research):
    file_data = request.FILES["file"]
    wb = load_workbook(filename=file_data)
    ws = wb[wb.sheetnames[0]]
    starts = False
    set_research_d = SetOrderResearch.objects.filter(set_research_id=set_research).order_by("order")
    head_research_data = {i.research_id: i.research.title for i in set_research_d}
    patients_data = {}
    for row in ws.rows:
        cells = [str(x.value) for x in row]
        if not starts:
            if "снилс" in cells:
                starts = True
                family = cells.index("фaмилия")
                name = cells.index("имя")
                patronymic = cells.index("отчество")
                born = cells.index("дата рождения")
                snils = cells.index("снилс")
                starts = True
        else:
            born_data = cells[born].split(" ")[0]
            if "." in born_data:
                born_data = normalize_dots_date(born_data)
            patient = {"family": cells[family], "name": cells[name], "patronymic": cells[patronymic], "birthday": born_data}
            patients_data[cells[snils]] = patient
    doc_type = DocumentType.objects.filter(title="СНИЛС").first()
    patient_cards = search_cards_by_numbers(tuple(patients_data.keys()), doc_type.id)
    cards_id = [i.card_id for i in patient_cards]
    researches_id = [i.research_id for i in set_research_d]

    purpose_research = {i: 0 for i in head_research_data.keys()}
    meta_patients = {
        pc.card_id: {
            "card_num": pc.card_number,
            "snils": pc.document_number,
            "researches": deepcopy(purpose_research),
            "fio": f"{pc.family} {pc.name} {pc.patronymic}",
            "district": pc.district_title,
        }
        for pc in patient_cards
    }

    is_paraclinic_researches = is_paraclinic_filter_research(tuple(researches_id))
    paraclinic_researches = [i.research_id for i in is_paraclinic_researches]

    is_lab_research = is_lab_filter_research(tuple(researches_id))
    lab_research = [i.research_id for i in is_lab_research]

    lab_days_ago_confirm = DAYS_AGO_SEARCH_RESULT.get("isLab")
    instrumental_days_ago_confirm = DAYS_AGO_SEARCH_RESULT.get("isInstrumental")
    patient_results = check_lab_instrumental_results_by_cards_and_period(
        tuple(cards_id), lab_days_ago_confirm, instrumental_days_ago_confirm, tuple(lab_research), tuple(paraclinic_researches)
    )
    for pr in patient_results:
        if meta_patients.get(pr.client_id):
            meta_patients[pr.client_id]["researches"][pr.research_id] = 1

    head_data = {
        "num_card": "№ карты",
        "district": "участок",
        "family": "Фамилия",
        "name": "имя",
        "patronymic": "отчество",
        "current_age": "возраст текущий",
        "year_age": "возраст в году",
        "snils": "СНИЛС",
        **head_research_data,
    }

    file_name = data_xls_save_headers_file(meta_patients, head_data, "Пройденые услуги", "fill_xls_check_research_exam_data")
    return file_name


def get_parts_fio(fio_data):
    family_data = fio_data[0]
    name_data = fio_data[1]
    patronymic_data = None
    if len(fio_data) > 2:
        patronymic_data = fio_data[2]
    return family_data, name_data, patronymic_data


def upload_file(request):
    # todo - Логирование загрузки файлов
    result = []
    try:
        file = request.FILES["file"]
        request_data = request.POST
        selected_form = request_data.get("selectedForm")
        entity_id = request_data.get("entityId")
        other_need_data = request_data.get("otherNeedData")
        data = {"file": file, "selectedForm": selected_form, "entity_id": entity_id, "other_need_data": other_need_data}
        function = import_string(selected_form)
        result = function(
            request_data={
                **data,
                "user": request.user,
                "hospital": request.user.doctorprofile.get_hospital() if hasattr(request.user, "doctorprofile") else Hospitals.get_default_hospital(),
            }
        )
    except Exception as e:
        # todo - Выводить структуру файла которая нужна, если передано не-то
        stdout.write(f"{e}")
    return JsonResponse({"data": result})
