from typing import Union

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from api.parse_file.normalization import normalize_values
from api.parse_file.utils import check_need_col
from api.parse_file.validation import check_values
from employees.models import EmployeePosition, EmployeeVacation, WorkDayStatus
from hospitals.models import Hospitals


def check_request_data(organization_id, user=None):
    """
    Проверяет есть ли переданная организация и проверяет права пользователя (если он передан)
    """
    result = {"ok": True, "message": ""}
    organization: Union[Hospitals, None] = Hospitals.objects.filter(pk=organization_id).first()
    if not organization:
        result = {"ok": False, "message": "Такой организации нет"}
    elif user:
        user_organization: Union[Hospitals, None] = user.doctorprofile.get_hospital()
        user_is_admin = user.is_superuser
        if user_organization.pk != organization.pk and user_is_admin is False:
            result = {"ok": False, "message": "Запрещено передавать не свою организацию"}
    return result


def normalize_vacation_data(employee_data: dict, normalize_funcs_list: list):
    result = {
        "tabel_number": None,
        "vacation_type": None,
        "vacation_start": None,
        "vacation_end": None,
    }
    for action in normalize_funcs_list:
        fields = action.get("fields", set())
        normalize_funcs = action.get("normalize_funcs", set())
        for field in fields:
            value = employee_data.get(field)
            normalized_value = normalize_values(value, normalize_funcs)
            result[field] = normalized_value
    return result


def validate_vacation_data(normalized_data: dict, russian_keys: dict, values_lens: dict, checks_lists: list):
    result = {"ok": True, "data": {}}
    errors = []
    tabel_number = normalized_data.get("tabel_number")
    if not tabel_number:
        result = {"ok": False, "data": {}, "empty": True}
        return result

    invalid_value_fields = set()

    for check in checks_lists:
        fields = check.get("fields", set())
        check_funcs = check.get("check_funcs", set())
        for field in fields:
            if field not in invalid_value_fields:
                value = normalized_data.get(field)
                value_len = values_lens.get(field)
                ru_key = russian_keys.get(field)
                check_result = check_values(value, check_funcs, value_len, ru_key)
                if not check_result.get("ok"):
                    invalid_value_fields.add(field)
                    errors.append(check_result.get("message"))
    if errors:
        result = {"ok": False, "data": {"tabel_number": tabel_number, "reason": ", ".join(errors)}}
    return result


def parse_work_sheet(ws: Worksheet):

    result = {"ok": True, "message": ""}

    need_col_name = {"Табельный номер", "Вид отпуска (ежегодный, дополнительный)", "Отпуск, с", "Отпуск, по"}
    starts = False
    tabel_number_idx, vacation_type_idx, vacation_start_idx, vacation_end_idx = (
        None,
        None,
        None,
        None,
    )
    employee_vacations = []
    incorrect_employee_vacations = []

    normalize_actions = [
        {"fields": ["tabel_number", "vacation_type", "vacation_start", "vacation_end"], "normalize_funcs": ["strip"]},
        {"fields": ["vacation_start", "vacation_end"], "normalize_funcs": ["normalize_date"]},
    ]

    russian_keys = {
        "tabel_number": "Табельный номер",
        "vacation_type": "Вид отпуска",
        "vacation_start": "Отпуск, с",
        "vacation_end": "Отпуск, по",
    }
    values_lens = {
        "tabel_number": EmployeePosition._meta.get_field("tabel_number").max_length,
    }
    checks_lists = [
        {"fields": ["tabel_number"], "check_funcs": ["check_not_empty", "check_max_len"]},
        {"fields": ["vacation_start", "vacation_end"], "check_funcs": ["check_not_empty", "check_date"]},
    ]
    for row in ws.rows:
        cells = [str(x.value) for x in row]
        if not starts:
            if "Табельный номер" in cells:
                result_check_col = check_need_col(cells, need_col_name)
                if not result_check_col.get("ok"):
                    return result_check_col
                tabel_number_idx = cells.index("Табельный номер")
                vacation_type_idx = cells.index("Вид отпуска (ежегодный, дополнительный)")
                vacation_start_idx = cells.index("Отпуск, с")
                vacation_end_idx = cells.index("Отпуск, по")
                starts = True
        else:
            employee_vacation_data = {
                "tabel_number": cells[tabel_number_idx],
                "vacation_type": cells[vacation_type_idx],
                "vacation_start": cells[vacation_start_idx],
                "vacation_end": cells[vacation_end_idx],
            }
            normalized_employee_data = normalize_vacation_data(employee_vacation_data, normalize_actions)
            validation_result = validate_vacation_data(normalized_employee_data, russian_keys, values_lens, checks_lists)
            if not validation_result.get("ok") and validation_result.get("empty"):
                continue
            if not validation_result.get("ok"):
                incorrect_employee_vacations.append(validation_result["data"])
                continue
            employee_vacations.append(normalized_employee_data)
    if not starts:
        return {"ok": False, "message": "Не найдена колонка 'Табельный номер'", "result": {}}
    result["result"] = {"employee_vacations": employee_vacations, "incorrect_employee_vacations": incorrect_employee_vacations}
    return result


def update_employees_vacation(organization_id, employee_vacations):
    incorrect_employee_vacations = []
    for vacation in employee_vacations:
        tabel_number = vacation.get("tabel_number")
        vacation_type = vacation.get("vacation_type")
        vacation_start = vacation.get("vacation_start")
        vacation_end = vacation.get("vacation_end")
        employee = EmployeePosition.find_by_tabel_number(organization_id, tabel_number)
        if not employee:
            incorrect_employee_vacations.append({"tabel_number": tabel_number, "reason": "Работник не найден"})
            continue
        vacation_type_db = WorkDayStatus.objects.filter(vacation_title=vacation_type).first()
        if not vacation_type_db:
            incorrect_employee_vacations.append({"tabel_number": tabel_number, "reason": "Тип отпуска не найден"})
            continue
        employee_vacation = EmployeeVacation(employee_position_id=employee.pk, work_day_status_id=vacation_type_db.pk, start=vacation_start, end=vacation_end)
        employee_vacation.save()
    return {"ok": True, "incorrect_employee_vacations": incorrect_employee_vacations}


def form_01(request_data):
    """
    Загрузка отпусков работников

    На входе:
    "file" - Файл XLSX со столбцами (указаны в need_col_name)
    "entity_id" - id организации
    "user" (опционально) - user объект из request-а с фронта, может не быть при вызове из django management commands
    """

    file = request_data.get("file")
    organization_id = request_data.get("entity_id")
    user = request_data.get("user")
    result_request_check = check_request_data(organization_id, user)
    if not result_request_check.get("ok"):
        return result_request_check
    wb = load_workbook(filename=file, read_only=True)
    ws = wb[wb.sheetnames[0]]
    result_parse_file = parse_work_sheet(ws)
    if not result_parse_file.get("ok"):
        return result_parse_file
    data_result_parse = result_parse_file.get("result")
    employee_vacations: list = data_result_parse.get("employee_vacations")
    incorrect_employee_vacations: list = data_result_parse.get("incorrect_employee_vacations")
    result_update = update_employees_vacation(organization_id, employee_vacations)
    incorrect_employee_vacations.extend(result_update.get("incorrect_employee_vacations", []))
    columns = [{"field": "tabel_number", "key": "tabel_number", "title": "Табельный номер", "align": "left", "width": 250}, {"field": "reason", "key": "reason", "title": 'Причина ошибки'}]
    result = {
        "colData": columns,
        "data": incorrect_employee_vacations,
    }
    return {"ok": True, "result": result, "message": ""}
