import base64
import json

import requests
from openpyxl.reader.excel import load_workbook
from contracts.models import PriceName, PriceCoast
from directory.models import Researches, CategoryDirectory
from laboratory.settings import RMIS_MIDDLE_SERVER_ADDRESS, RMIS_MIDDLE_SERVER_TOKEN


def _find_column_index(cells, *titles):
    for title in titles:
        if title and title in cells:
            return cells.index(title)
    return None


def _parse_optional_coast(value):
    if value is None or value == "" or value == "None":
        return None
    try:
        coast = float(str(value).strip())
    except (TypeError, ValueError):
        return None
    if coast <= 0:
        return None
    return coast


def form_01(request_data):
    """
    Загрузка цен по прайсу

    На входе:
    Файл XLSX с ценами прайса
    Cтруктура:
    Код по прайсу (internal_code Researches), Услуга (title_researches),
    колонка с названием прайса (priceCoasts.coast),
    опционально колонка "<название прайса> ЦИТО" (priceCoasts.coast_cito)
    """
    price_id = request_data.get("entity_id")
    file = request_data.get("file")
    price = PriceName.objects.filter(pk=price_id).first()
    if not price:
        return {"ok": False, "result": [], "message": "Такого прайса нет"}
    wb = load_workbook(filename=file)
    ws = wb[wb.sheetnames[0]]
    internal_code_idx, coast_idx, coast_cito_idx, category_idx, short_title_research_idx = (
        '',
        '',
        None,
        '',
        '',
    )
    starts = False
    for row in ws.rows:
        cells = [str(x.value) for x in row]
        if not starts:
            if "Код по прайсу" in cells:
                internal_code_idx = cells.index("Код по прайсу")
                category_idx = cells.index("Категория")
                short_title_research_idx = cells.index("Короткое название")
                coast_idx = _find_column_index(cells, price.title, f"{price.title}-{price.symbol_code}")
                if coast_idx is None:
                    return {"ok": False, "result": [], "message": "Название прайса не совпадает"}
                coast_cito_idx = _find_column_index(cells, f"{price.title} ЦИТО", f"{price.title}-{price.symbol_code} ЦИТО")
                starts = True
        else:
            internal_code = cells[internal_code_idx].strip()
            category_title = cells[category_idx].strip()
            short_service_title = cells[short_title_research_idx].strip()
            try:
                coast = float(cells[coast_idx].strip())
            except Exception:
                continue
            coast_cito = _parse_optional_coast(cells[coast_cito_idx]) if coast_cito_idx is not None else None
            if internal_code == "None" or not coast:
                continue
            service = Researches.objects.filter(internal_code=internal_code).first()

            if not service:
                continue
            current_coast = PriceCoast.objects.filter(price_name_id=price.pk, research_id=service.pk).first()
            if current_coast:
                changed = False
                if current_coast.coast != coast:
                    current_coast.coast = coast
                    changed = True
                if coast_cito_idx is not None and current_coast.coast_cito != coast_cito:
                    current_coast.coast_cito = coast_cito
                    changed = True
                if changed:
                    current_coast.save()
            else:
                new_coast = PriceCoast(price_name_id=price.pk, research_id=service.pk, coast=coast, coast_cito=coast_cito)
                new_coast.save()
            category = CategoryDirectory.objects.filter(title=category_title).first()
            if category:
                service.category = category
            if short_service_title == 'None' or not short_service_title:
                short_service_title = ""
            if service.short_title != short_service_title:
                service.short_title = short_service_title
            service.save()

    if not starts:
        return {"ok": False, "result": [], "message": "Не найдены колонка 'Код по прайсу' "}
    return {"ok": True, "result": [], "message": ""}


def form_02(request_data):
    """
    Загрузка посещений из файла

    На входе:
    Файл XLSX с посещениями
    Cтруктура:
    номер карты, Заведующий отделением, Отделение, Услуга, Фамилия, Имя, Отчество, Дата рождения, СНИЛС, Диагноз, Дата услуги, Это травма
    """
    file = request_data.get("file")
    wb = load_workbook(filename=file)
    ws = wb[wb.sheetnames[0]]
    card_number_idx, head_department_idx, department_idx, service_idx, family_idx, name_idx, patronymic_idx, birthday_idx, snils_idx, diagnos_idx, service_date_idx, is_travma_idx = (
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
    )
    starts = False
    file_data = []
    for row in ws.rows:
        cells = [str(x.value) for x in row]
        if not starts:
            if "номер карты" in cells:
                card_number_idx = cells.index("номер карты")
                head_department_idx = cells.index("Заведующий отделением")
                department_idx = cells.index("Отделение")
                service_idx = cells.index("Услуга")
                family_idx = cells.index("Фамилия")
                name_idx = cells.index("Имя")
                patronymic_idx = cells.index("Отчество")
                birthday_idx = cells.index("Дата рождения")
                snils_idx = cells.index("СНИЛС")
                diagnos_idx = cells.index("Диагноз")
                service_date_idx = cells.index("Дата услуги")
                is_travma_idx = cells.index("Это травма")
                starts = True
        else:
            tmp_data = {
                "cardNumber": cells[card_number_idx],
                "headDepartment": cells[head_department_idx],
                "department": cells[department_idx],
                "service": cells[service_idx],
                "family": cells[family_idx],
                "name": cells[name_idx],
                "patronymic": cells[patronymic_idx],
                "birthday": cells[birthday_idx],
                "snils": cells[snils_idx],
                "diagnos": cells[diagnos_idx],
                "serviceDate": cells[service_date_idx],
                "isTravma": cells[is_travma_idx],
            }
            file_data.append(tmp_data)
    if not starts:
        return {"ok": False, "result": [], "message": "Не найдена колонка 'номер карты' "}

    json_str = json.dumps(file_data)
    base64_data = base64.b64encode(json_str.encode())
    json_data = {"data": base64_data}
    headers = {"authorization": f"Bearer {RMIS_MIDDLE_SERVER_TOKEN}"}

    response = requests.post(f"{RMIS_MIDDLE_SERVER_ADDRESS}/send-case-visit", json=json_data, headers=headers)
    result = response.json()

    return {"ok": True, "result": [], "message": f"{result}"}
