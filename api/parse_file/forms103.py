from typing import Union
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from api.parse_file.normalization import normalize_values
from api.parse_file.validation import check_values
from employees.models import Department, Position, Employee, EmployeePosition, TypeWorkTimeEmployee
from hospitals.models import Hospitals

BULK_BATCH_SIZE = 500


def check_request_data(organization_id, user=None):
    """
    Проверяет есть ли переданная организация и проверяет права пользователя (если он передан)
    """
    result = {"ok": True, "message": ""}
    organization: Union[Hospitals, None] = Hospitals.objects.filter(pk=organization_id).first()
    if not organization:
        result = {"ok": False, "message": "Такой организации нет"}
    elif user:
        user_organization: Union[Hospitals, None] = user.doctorprofile.get_hospital()
        user_is_admin = user.is_superuser
        if user_organization.pk != organization.pk and user_is_admin is False:
            result = {"ok": False, "message": "Запрещено передавать не свою организацию"}
    return result


def check_need_col(cols: list, need_cols: set):
    """
    Проверяет что все необходимые колонки есть
    """
    remains_need_cols = set(need_cols - set(cols))
    if remains_need_cols:
        return False
    return True


def parse_work_sheet(ws: Worksheet):
    """
    Разбор xlsx листа из файла
    """
    result = {"ok": True, "message": ""}
    need_col_name = {"Вид занятости", "СНИЛС", "Табельный номер", "Сотрудник", "Подразделение", "Должность", "Количество ставок", "Дата приема", "Дата увольнения", "График работы"}
    starts = False
    employment_form_idx, snils_idx, tabel_number_idx, employee_fio_idx, department_title_idx, position_title_idx, rate_idx, date_employment_idx, date_dismissal_idx, work_schedule_idx = (
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
    )
    employees = []
    incorrect_employees = []
    departments_titles = set()
    positions_titles = set()

    normalize_actions = [
        {
            "fields": ["employment_form", "tabel_number", "fio", "department_title", "position_title", "rate"],
            "normalize_funcs": ["remove_double_spaces"],
        },
        {"fields": ["snils"], "normalize_funcs": ["normalize_snils"]},
        {"fields": ["rate"], "normalize_funcs": ["normalize_rate"]},
        {"fields": ["date_employment", "date_dismissal"], "normalize_funcs": ["normalize_date"]},
        {"fields": ["work_schedule"], "normalize_funcs": ["get_first_number", "convert_hours_to_minutes"]},
    ]

    russian_keys = {
        "employment_form": "Вид занятости",
        "snils": "СНИЛС",
        "tabel_number": "Табельный номер",
        "fio": "ФИО",
        "department_title": "Подразделение",
        "position_title": "Должность",
        "rate": "Количество ставок",
        "date_employment": "Дата приема",
        "date_dismissal": "Дата увольнения",
        "work_schedule": "График работы",
    }
    values_lens = {
        "employment_form": TypeWorkTimeEmployee._meta.get_field("title").max_length,
        "snils": 11,
        "tabel_number": EmployeePosition._meta.get_field("tabel_number").max_length,
        "fio": sum([Employee._meta.get_field("family").max_length, Employee._meta.get_field("name").max_length, Employee._meta.get_field("patronymic").max_length]),
        "department_title": Department._meta.get_field("name").max_length,
        "position_title": Position._meta.get_field("name").max_length,
    }
    checks_lists = [
        {
            "fields": ["employment_form", "snils", "tabel_number", "fio", "department_title", "position_title"],
            "check_funcs": ["check_not_empty", "check_max_len"],
        },
        {"fields": ["snils"], "check_funcs": ["check_snils"]},
        {"fields": ["rate"], "check_funcs": ["check_not_empty", "check_rate"]},
        {"fields": ["date_employment"], "check_funcs": ["check_not_empty"]},
        {"fields": ["date_employment", "date_dismissal"], "check_funcs": ["check_date"]},
    ]
    for row in ws.rows:
        cells = [str(x.value) for x in row]
        if not starts:
            if "Табельный номер" in cells:
                if not check_need_col(cells, need_col_name):
                    return {"ok": False, "message": "Нет обязательных полей", "result": {}}
                employment_form_idx = cells.index("Вид занятости")
                snils_idx = cells.index("СНИЛС")
                tabel_number_idx = cells.index("Табельный номер")
                employee_fio_idx = cells.index("Сотрудник")
                department_title_idx = cells.index("Подразделение")
                position_title_idx = cells.index("Должность")
                rate_idx = cells.index("Количество ставок")
                date_employment_idx = cells.index("Дата приема")
                date_dismissal_idx = cells.index("Дата увольнения")
                work_schedule_idx = cells.index("График работы")
                starts = True
        else:
            employee_data = {
                "employment_form": cells[employment_form_idx],
                "snils": cells[snils_idx],
                "tabel_number": cells[tabel_number_idx],
                "fio": cells[employee_fio_idx],
                "department_title": cells[department_title_idx],
                "position_title": cells[position_title_idx],
                "rate": cells[rate_idx],
                "date_employment": cells[date_employment_idx],
                "date_dismissal": cells[date_dismissal_idx],
                "work_schedule": cells[work_schedule_idx],
            }
            normalized_employee_data = normalize_employee_data(employee_data, normalize_actions)
            validation_result = validate_employee_data(normalized_employee_data, russian_keys, values_lens, checks_lists)
            if not validation_result.get("ok") and validation_result.get("empty"):
                continue
            if not validation_result.get("ok"):
                incorrect_employees.append(validation_result["data"])
                continue
            departments_titles.add(normalized_employee_data["department_title"])
            positions_titles.add(normalized_employee_data["position_title"])
            employees.append(normalized_employee_data)
    if not starts:
        return {"ok": False, "message": "Не найдена колонка 'Табельный номер'", "result": {}}
    result["result"] = {"employees": employees, "incorrect_employees": incorrect_employees, "departments_titles": departments_titles, "positions_titles": positions_titles}
    return result


def normalize_employee_data(employee_data: dict, normalize_funcs_list: list):
    result = {
        "employment_form": None,
        "snils": None,
        "tabel_number": None,
        "fio": None,
        "family": None,
        "name": None,
        "patronymic": None,
        "department_title": None,
        "position_title": None,
        "rate": None,
        "date_employment": None,
        "date_dismissal": None,
        "work_schedule": None,
    }
    for action in normalize_funcs_list:
        fields = action.get("fields", set())
        normalize_funcs = action.get("normalize_funcs", set())
        for field in fields:
            value = employee_data.get(field)
            normalized_value = normalize_values(value, normalize_funcs)
            result[field] = normalized_value

    if result["fio"]:
        fio_data = result["fio"].split(" ")
        result["family"] = fio_data[0]
        result["name"] = fio_data[1]
        if len(fio_data) > 2:
            result["patronymic"] = " ".join(fio_data[2:])
    return result


def validate_employee_data(normalized_data: dict, russian_keys: dict, values_lens: dict, checks_lists: list):
    result = {"ok": True, "data": {}}
    fio = normalized_data.get("fio")
    snils = normalized_data.get("snils")
    name_local = fio if fio else snils
    errors = []
    if not name_local:
        result = {"ok": False, "data": {}, "empty": True}
        return result

    invalid_value_fields = set()

    for check in checks_lists:
        fields = check.get("fields", set())
        check_funcs = check.get("check_funcs", set())
        for field in fields:
            if field not in invalid_value_fields:
                value = normalized_data.get(field)
                value_len = values_lens.get(field)
                ru_key = russian_keys.get(field)
                check_result = check_values(value, check_funcs, value_len, ru_key)
                if not check_result.get("ok"):
                    invalid_value_fields.add(field)
                    errors.append(check_result.get("message"))
    if errors:
        result = {"ok": False, "data": {"fio": name_local, "reason": ", ".join(errors)}}
    return result


def update_organization_departments(organization_id: int, new_departments_titles: Union[list, set]):
    existing_titles = set(Department.objects.filter(is_active=True, hospital_id=organization_id).values_list("name", flat=True))
    missing_titles = [title for title in new_departments_titles if title not in existing_titles]
    if not missing_titles:
        return
    Department.objects.bulk_create(
        [Department(hospital_id=organization_id, name=title, is_active=True) for title in missing_titles],
        ignore_conflicts=True,
        batch_size=BULK_BATCH_SIZE,
    )


def update_organization_positions(organization_id: int, new_positions_titles: Union[list, set]):
    existing_titles = set(Position.objects.filter(is_active=True, hospital_id=organization_id).values_list("name", flat=True))
    missing_titles = [title for title in new_positions_titles if title not in existing_titles]
    if not missing_titles:
        return
    Position.objects.bulk_create(
        [Position(hospital_id=organization_id, name=title, is_active=True) for title in missing_titles],
        ignore_conflicts=True,
        batch_size=BULK_BATCH_SIZE,
    )


def _make_employee_position_key(*, is_active, employee_id, position_id, department_id, tabel_number):
    return (is_active, employee_id, position_id, department_id, tabel_number)


def _employee_position_key(employee, department, position, tabel_number, is_active):
    return _make_employee_position_key(
        is_active=is_active,
        employee_id=employee.pk,
        position_id=position.pk,
        department_id=department.pk,
        tabel_number=tabel_number,
    )


def _find_employee_position(employee_position_index, employee, department, position, tabel_number, is_active):
    return employee_position_index.get(_employee_position_key(employee, department, position, tabel_number, is_active))


def _load_employee_position_index(organization_id):
    return {
        _make_employee_position_key(
            is_active=employee_position.is_active,
            employee_id=employee_position.employee_id,
            position_id=employee_position.position_id,
            department_id=employee_position.department_id,
            tabel_number=employee_position.tabel_number,
        ): employee_position
        for employee_position in EmployeePosition.objects.filter(employee__hospital_id=organization_id)
    }


def _apply_employee_position_fields(employee_position, employee_data):
    if employee_data.get("date_dismissal"):
        employee_position.date_dismissal = employee_data.get("date_dismissal")
        employee_position.is_active = False
    work_schedule_minutes_weekly = employee_data.get("work_schedule")
    employee_position.weekly_hours_norm = work_schedule_minutes_weekly
    if employee_position.work_days_per_week and work_schedule_minutes_weekly:
        work_schedule_minutes_day = work_schedule_minutes_weekly // employee_position.work_days_per_week
        work_schedule_minutes_per_rate = work_schedule_minutes_day * employee_position.rate if employee_position.rate else work_schedule_minutes_day
        employee_position.daily_hours_norm = int(work_schedule_minutes_per_rate)
    return employee_position


def _build_employee_position(employee_data, employee, department, position, employment_form):
    active = False if employee_data.get("date_dismissal") else True
    work_schedule_minutes_weekly = employee_data.get("work_schedule")
    employee_position = EmployeePosition(
        is_active=active,
        employee_id=employee.pk,
        position_id=position.pk,
        department_id=department.pk,
        tabel_number=employee_data.get("tabel_number"),
        rate=employee_data.get("rate"),
        type_work_time_id=employment_form.pk,
        date_employment=employee_data.get("date_employment"),
        date_dismissal=employee_data.get("date_dismissal"),
        weekly_hours_norm=work_schedule_minutes_weekly,
    )
    if employee_position.work_days_per_week and work_schedule_minutes_weekly:
        work_schedule_minutes_day = work_schedule_minutes_weekly // employee_position.work_days_per_week
        work_schedule_minutes_per_rate = work_schedule_minutes_day * employee_position.rate if employee_position.rate else work_schedule_minutes_day
        employee_position.daily_hours_norm = int(work_schedule_minutes_per_rate)
    return employee_position


def _load_employees_by_snils(organization_id):
    return {employee.snils: employee for employee in Employee.objects.filter(hospital_id=organization_id).exclude(snils__isnull=True).exclude(snils="")}


def _load_departments_by_name(organization_id):
    return {department.name: department for department in Department.objects.filter(is_active=True, hospital_id=organization_id)}


def _load_positions_by_name(organization_id):
    return {position.name: position for position in Position.objects.filter(is_active=True, hospital_id=organization_id)}


def update_organization_employee_positions(organization_id: int, employees):
    """
    Обновление и создание трудовых договоров (EmployeePosition)
    Ищет активный трудовой договор - обновляет, если нет и стоит дата увольнения, ищет архивный, если нет - создает архивный
    Если не нашло ничего - создает новый трудовой договор
    """
    departments_by_name = _load_departments_by_name(organization_id)
    positions_by_name = _load_positions_by_name(organization_id)
    employment_forms_by_title = {employment_form.title: employment_form for employment_form in TypeWorkTimeEmployee.objects.all()}
    employees_by_snils = _load_employees_by_snils(organization_id)
    employee_position_index = _load_employee_position_index(organization_id)

    incorrect_employees = []
    employees_to_create = {}

    for employee_data in employees:
        snils = employee_data.get("snils")
        if snils in employees_by_snils or snils in employees_to_create:
            continue
        employees_to_create[snils] = Employee(
            hospital_id=organization_id,
            family=employee_data.get("family"),
            name=employee_data.get("name"),
            patronymic=employee_data.get("patronymic"),
            snils=snils,
        )

    if employees_to_create:
        Employee.objects.bulk_create(list(employees_to_create.values()), batch_size=BULK_BATCH_SIZE)
        for employee in Employee.objects.filter(hospital_id=organization_id, snils__in=employees_to_create.keys()):
            employees_by_snils[employee.snils] = employee

    employee_positions_to_create = []
    employee_positions_to_update = {}
    planned_employee_position_keys = set()

    for employee_data in employees:
        snils = employee_data.get("snils")
        employment_form_title = employee_data.get("employment_form")
        tabel_number = employee_data.get("tabel_number")

        employment_form = employment_forms_by_title.get(employment_form_title)
        if not employment_form:
            incorrect_employees.append({"fio": employee_data["fio"], "reason": f"Нет такого вида занятости в справочнике ({employment_form_title})"})
            continue

        department = departments_by_name.get(employee_data.get("department_title"))
        if not department:
            incorrect_employees.append({"fio": employee_data["fio"], "reason": f"Нет подразделения ({employee_data.get('department_title')})"})
            continue

        position = positions_by_name.get(employee_data.get("position_title"))
        if not position:
            incorrect_employees.append({"fio": employee_data["fio"], "reason": f"Нет должности ({employee_data.get('position_title')})"})
            continue

        employee = employees_by_snils.get(snils)
        if not employee:
            incorrect_employees.append({"fio": employee_data["fio"], "reason": "Не удалось найти или создать работника"})
            continue

        active_employee_position = _find_employee_position(employee_position_index, employee, department, position, tabel_number, is_active=True)
        if active_employee_position:
            _apply_employee_position_fields(active_employee_position, employee_data)
            employee_positions_to_update[active_employee_position.pk] = active_employee_position
            continue

        if employee_data.get("date_dismissal"):
            inactive_employee_position = _find_employee_position(employee_position_index, employee, department, position, tabel_number, is_active=False)
            if inactive_employee_position:
                continue
            create_as_active = False
        else:
            create_as_active = True

        employee_position_key = _employee_position_key(employee, department, position, tabel_number, create_as_active)
        if employee_position_key in employee_position_index or employee_position_key in planned_employee_position_keys:
            continue

        employee_positions_to_create.append(_build_employee_position(employee_data, employee, department, position, employment_form))
        planned_employee_position_keys.add(employee_position_key)

    if employee_positions_to_create:
        EmployeePosition.objects.bulk_create(employee_positions_to_create, batch_size=BULK_BATCH_SIZE)

    if employee_positions_to_update:
        EmployeePosition.objects.bulk_update(
            list(employee_positions_to_update.values()),
            ["date_dismissal", "is_active", "weekly_hours_norm", "daily_hours_norm"],
            batch_size=BULK_BATCH_SIZE,
        )

    return {"ok": True, "message": "", "data": incorrect_employees}


def unmerge_cells(work_sheet: Worksheet):
    merged_cells = list(work_sheet.merged_cells.ranges)
    for merged_range in merged_cells:
        work_sheet.unmerge_cells(str(merged_range))


def delete_cols(work_sheet: Worksheet, nums_col):
    for col in nums_col:
        work_sheet.delete_cols(col, 1)


def prepare_employee_file(work_sheet: Worksheet):
    """
    Предварительная обработка файла с работниками
    """
    a2_cell = work_sheet['A2'].value
    if a2_cell == "Личные данные сотрудников":
        unmerge_cells(work_sheet)
        work_sheet.delete_rows(1, 8)
        nums_columns_to_delete = [17, 16, 14, 13, 12, 11, 10, 9, 7, 6, 5, 3, 2]
        delete_cols(work_sheet, nums_columns_to_delete)


def form_01(request_data):
    """
    Загрузка работников организации

    На входе:
    "file" - Файл XLSX со столбцами (указаны в need_col_name)
    "entity_id" - ID организации куда загружать данные с файла
    "user" (опционально) - user объект из request-а с фронта, может не быть при вызове из django management commands
    """

    file = request_data.get("file")
    organization_id = request_data.get("entity_id")
    user = request_data.get("user")
    result_request_check = check_request_data(organization_id, user)
    if not result_request_check.get("ok"):
        return result_request_check
    wb = load_workbook(filename=file)
    ws = wb[wb.sheetnames[0]]
    prepare_employee_file(ws)
    result_parse_file = parse_work_sheet(ws)
    if not result_parse_file.get("ok"):
        return result_parse_file
    data_result_parse = result_parse_file.get("result")
    employees: list = data_result_parse.get("employees")
    incorrect_employees: list = data_result_parse.get("incorrect_employees")
    departments_titles: set = data_result_parse.get("departments_titles")
    positions_titles: set = data_result_parse.get("positions_titles")
    update_organization_departments(organization_id, departments_titles)
    update_organization_positions(organization_id, positions_titles)
    result_update = update_organization_employee_positions(organization_id, employees)
    incorrect_employees.extend(result_update.get("data", []))
    columns = [{"field": "fio", "key": "fio", "title": "Работник", "align": "left", "width": 250}, {"field": "reason", "key": "reason", "title": 'Причины ошибки'}]
    result = {
        "colData": columns,
        "data": incorrect_employees,
    }
    return {"ok": True, "result": result, "message": ""}
