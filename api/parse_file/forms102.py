from openpyxl.reader.excel import load_workbook

from researches.models import Tubes


def check_need_col(cols: list, need_cols: set):
    other_need_cols = set(set(cols) - need_cols)
    if len(other_need_cols) + len(need_cols) != len(cols):
        return False
    return True


def form_01(request_data):
    """
    Загрузка цен по прайсу

    На входе:
    Файл XLSX с ёмкостями для биоматериала
    Cтруктура:
    Наименование (Tube.title), Краткое наименование (Tube.short_title), Цвет (красный, зеленый)
    """

    colors = {
        "красн": "#FF0000",
        "оранж": "#FF8C00",
        "желт": "#FFFF00",
        "жёлт": "#FFFF00",
        "зелен": "#008000",
        "голуб": "#00FFFF",
        "синий": "#0000FF",
        "синего": "#0000FF",
        "фиолет": "#9400D3",
        "сирен": "#c8a2c8",
        "розов": "#FF1493",
        "белый": "#FFFFFF",
        "белая": "#FFFFFF",
        "белого": "#FFFFFF",
        "белой": "#FFFFFF",
        "серый": "#808080",
        "серая": "#808080",
        "серого": "#808080",
        "серой": "#808080",
        "корич": "#8B4513",
    }

    file = request_data.get("file")
    wb = load_workbook(filename=file)
    ws = wb[wb.sheetnames[0]]
    title_idx, short_title_idx, color_idx = (
        '',
        '',
        '',
    )
    need_col_name = {"Наименование", "код", "цвет (rgb, hex)"}
    starts = False
    for row in ws.rows:
        cells = [str(x.value) for x in row]
        if not starts:
            if "цвет (rgb, hex)" in cells:
                if not check_need_col(cells, need_col_name):
                    return {"ok": False, "result": {}, "message": "Нет обязательных полей"}
                title_idx = cells.index("Наименование")
                short_title_idx = cells.index("код")
                color_idx = cells.index("цвет")
                starts = True
        else:
            title = cells[title_idx].strip()
            short_title = cells[short_title_idx].strip()
            color = cells[color_idx].strip()
            valid = Tubes.check_tube(title, short_title, color)
            if not valid:
                return {"ok": False, "result": {}, "message": "Валидация не пройдена"}
            new_tube = Tubes(title=title, short_title=short_title, color=color)
            new_tube.save()

    if not starts:
        return {"ok": False, "result": [], "message": "Не найдены колонка 'цвет (rgb, hex)'"}
    return {"ok": True, "result": [], "message": ""}
