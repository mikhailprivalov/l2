from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils.cell import get_column_letter
import json


def monitoring_xlsx(ws1, monitoring_title, table_data, date):
    """
    :param ws1:
    :return:
    """
    style_border = NamedStyle(name="style_border")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=False, size=12)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    header_size = 8
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=header_size)
    ws1.cell(row=1, column=1).value = f'Мониторинг - {monitoring_title}, от {date}'
    for i in range(header_size):
        ws1.cell(row=1, column=1 + i).style = style_border

    ws1.column_dimensions[get_column_letter(1)].width = 22

    current_row = 3
    ws1.cell(row=current_row, column=1).value = "МО"
    ws1.cell(row=current_row, column=1).style = style_border

    current_column = 2
    # строим заголовки
    for column_group in table_data['titles']:
        end_column = current_column + len(column_group['fields']) - 1
        # Заголовок группы
        ws1.merge_cells(start_row=current_row, start_column=current_column, end_row=current_row, end_column=end_column)
        ws1.cell(row=current_row, column=current_column).value = f'{column_group["groupTitle"]}'
        for i in range(end_column - current_column + 1):
            ws1.cell(row=current_row, column=current_column + i).style = style_border

        # Заголовок поля
        current_row += 1
        current_column = end_column - len(column_group['fields'])
        for field_colimn in column_group['fields']:
            current_column += 1
            ws1.column_dimensions[get_column_letter(current_column)].width = 15
            ws1.cell(row=current_row, column=current_column).value = f'{field_colimn}'
            ws1.cell(row=current_row, column=current_column).style = style_border
        current_row -= 1
        current_column += 1

    current_column = 1
    current_row += 1
    is_show_tables_title = False
    json_data = {}
    for row in table_data['rows']:
        current_row += 1
        ws1.cell(row=current_row, column=current_column).value = row['hospTitle']
        ws1.cell(row=current_row, column=current_column).style = style_border
        for value in row['values']:
            for v in value:
                is_dict = False
                if type(v) is str and "columns" in v:
                    try:
                        json_data = json.loads(v)
                        is_dict = True
                    except:
                        is_dict = False
                if not is_dict:
                    current_column += 1
                    ws1.cell(row=current_row, column=current_column).value = v
                    ws1.cell(row=current_row, column=current_column).style = style_border
                if is_dict:
                    col_title_data = json_data.get("columns")
                    col_title = col_title_data.get("titles")
                    start_col = current_column
                    if not is_show_tables_title:
                        for c_title in col_title:
                            start_col += 1
                            ws1.cell(row=current_row - 1, column=start_col).value = c_title
                            ws1.cell(row=current_row - 1, column=start_col).style = style_border
                            is_show_tables_title = True
                    rows_data = json_data.get("rows")
                    for r_data in rows_data:
                        start_col = current_column
                        for rd in r_data:
                            start_col += 1
                            ws1.cell(row=current_row, column=start_col).value = rd
                            ws1.cell(row=current_row, column=start_col).style = style_border
                        current_row += 1
        current_column = 1

    current_row += 1
    current_column = 1
    ws1.cell(row=current_row, column=current_column).value = "Итого"
    ws1.cell(row=current_row, column=current_column).style = style_border

    for total in table_data['total']:
        for v in total:
            current_column += 1
            ws1.cell(row=current_row, column=current_column).value = v
            ws1.cell(row=current_row, column=current_column).style = style_border

    current_row += 2
    current_column = 1
    for empty_hospital in table_data['empty_hospital']:
        current_row += 1
        ws1.cell(row=current_row, column=current_column).value = empty_hospital

    return ws1
