from openpyxl.reader.excel import load_workbook

from contracts.models import PriceName, PriceCoast
from directory.models import Researches


def form_01(request_data):
    """
    На входе:
    Файл XLSX с ценами прайса
    Cтруктура:
    Код по прайсу (internal_code Researches), Услуга (title_researches), следующее поле - любое текстовое название прайса (priceCoasts.coast)
    """
    price_id = request_data.get("entity_id")
    file = request_data.get("file")
    price = PriceName.objects.filter(pk=price_id).first()
    if not price:
        return {"ok": False, "result": [], "message": "Такого прайса нет"}
    wb = load_workbook(filename=file)
    ws = wb[wb.sheetnames[0]]
    internal_code_idx, coast_idx = (
        '',
        '',
    )
    starts = False
    for row in ws.rows:
        cells = [str(x.value) for x in row]
        if not starts:
            if "Код по прайсу" in cells:
                internal_code_idx = cells.index("Код по прайсу")
                try:
                    coast_idx = cells.index(price.title)
                except ValueError:
                    return {"ok": False, "result": [], "message": "Название прайса не совпадает"}
                starts = True
        else:
            internal_code = cells[internal_code_idx].strip()
            try:
                coast = float(cells[coast_idx].strip())
            except Exception:
                continue
            if internal_code == "None" or not coast:
                continue
            service = Researches.objects.filter(internal_code=internal_code).first()
            if not service:
                continue
            current_coast = PriceCoast.objects.filter(price_name_id=price.pk, research_id=service.pk).first()
            if current_coast:
                if current_coast.coast != coast:
                    current_coast.coast = coast
                    current_coast.save()
            else:
                new_coast = PriceCoast(price_name_id=price.pk, research_id=service.pk, coast=coast)
                new_coast.save()
    if not starts:
        return {"ok": False, "result": [], "message": "Не найдены колонка 'Код по прайсу' "}
    return {"ok": True, "result": [], "message": ""}
