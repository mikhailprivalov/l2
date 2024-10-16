from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils.cell import get_column_letter
import json


def patologistology_buh_base(ws1):
    style_border = NamedStyle(name="style_border")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=2, column=1).value = 'Период:'

    columns = [
        ('Медицинская организация ', 36),
        ('ФИО пациента ', 25),
        ('Дата рождения ', 17),
        ('Номер полиса ОМС ', 30),
        ('СНИЛС ', 15),
        ('Дата регистрации', 17),
        ('Код МКБ 10 заключение ', 12),
        ('Источник ', 12),
        ('Код оплаты (категория)', 30),
        ('Цель ', 26),
        ('Код мед услуги и категория сложности', 35),
        ('Направление', 30),
        ('Врач', 30),
        ('Подтверждено', 30),
    ]
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=4, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=4, column=idx).style = style_border
    return ws1


def patologistology_buh_data(ws1, data):
    """
    res - результат выборки SQL
    порядок возврата:
    napr, date_confirm, time_confirm, create_date_napr, create_time_napr,
    doc_fio, coast, discount, how_many, ((coast + (coast/100 * discount)) * how_many)::NUMERIC(10,2) AS sum_money,
    ist_f, time_confirmation, num_card, ind_family, ind_name,
    patronymic, birthday, date_born, to_char(EXTRACT(YEAR from age(time_confirmation, date_born)), '999') as ind_age
    :return:
    """
    style_border_res = NamedStyle(name="style_border_res")
    bd = Side(style='thin', color="000000")
    style_border_res.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border_res.font = Font(bold=False, size=11)
    style_border_res.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    r = 4
    for res in data:
        r += 1
        ws1.cell(row=r, column=1).value = res["hospital"]
        ws1.cell(row=r, column=2).value = res["fio_patient"]
        ws1.cell(row=r, column=3).value = res["born_patient"]
        ws1.cell(row=r, column=4).value = res["polis"]
        ws1.cell(row=r, column=5).value = res["snils"]
        ws1.cell(row=r, column=6).value = res["visit_date"]
        try:
            mcb10 = json.loads(res["mcb10_code"])
            result_mcb10 = mcb10.get("code", "")
        except:
            result_mcb10 = ""
        ws1.cell(row=r, column=7).value = result_mcb10
        ws1.cell(row=r, column=8).value = res["fin_source"]
        ws1.cell(row=r, column=9).value = res["price_category"]
        ws1.cell(row=r, column=10).value = res["purpose"]
        try:
            service = json.loads(res["service_code"])
            service_code = service.get('code', '')
            service_title = service.get('title', '')
        except:
            service_code = ""
            service_title = ""
        ws1.cell(row=r, column=11).value = f"{service_code} - {service_title}"
        ws1.cell(row=r, column=12).value = res["direction"]
        ws1.cell(row=r, column=13).value = res["doctor_fio"]
        ws1.cell(row=r, column=14).value = res["date_confirm"]

        rows = ws1[f'A{r}:L{r}']
        for row in rows:
            for cell in row:
                cell.style = style_border_res

    return ws1
