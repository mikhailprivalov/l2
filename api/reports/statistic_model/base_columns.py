from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils.cell import get_column_letter
import json


def custom_research_data(query_sql):
    result = []
    prev_direction = None
    step = 0
    tmp_result = {
        "permanent_field": {
            "Случай": "",
            "Протокол": "",
            "Пациент": "",
            "Пол": "",
            "Дата рождения": "",
            "Возраст": "",
            "Адрес": "",
        },
        "custom_field": {}
    }
    custom_fields = []
    for i in query_sql:
        if prev_direction != i.direction_id and step != 0:
            result.append(tmp_result.copy())
        if prev_direction != i.direction_number:
            tmp_result = {
                "permanent_field": {
                    "Случай": f"{i.hosp_direction}",
                    "Протокол": f"{i.direction_id}",
                    "Пациент": f"{i.family} {i.name} {i.patronymic}",
                    "Пол": i.patient_sex,
                    "Дата рождения": i.born,
                    "Возраст": i.ind_age,
                    "Адрес": i.patient_fact_address if i.patient_fact_address else i.patient_main_address,
                }
            }
        tmp_result["custom_field"][i.field_title] = i.field_value
        if i.field_title not in custom_fields:
            custom_fields.append(i.field_title)
        step += 1
        prev_direction = i.direction_id
    result.append(tmp_result.copy())
    fields = ["Случай", "Протокол", "Пациент", "Пол", "Дата рождения", "Возраст", "Адрес"]
    fields.extend(custom_fields)
    return {"result": result, "custom_fields": custom_fields, "fields": fields}


def custom_research_base(ws1, d1, d2, result_query, research_title):
    style_border = NamedStyle(name="style_border_ca")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=1, column=1).value = 'Услуга:'
    ws1.cell(row=1, column=2).value = research_title
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'

    columns = [
        ('Направление', 15),
        ('Источник', 15),
        ('Пациент', 45),
        ('Пол', 10),
        ('Дата рождения', 26),
        ('Возраст', 10),
        ('Адрес', 40),
        ('Исполнитель', 35),
        ('Код врача', 15),
    ]

    columns2 = [(i, 25) for i in result_query["custom_fields"]]
    columns.extend(columns2)
    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border

    return ws1