from sys import stdout

from django.core.management.base import BaseCommand
from django.db.models import Q
from openpyxl import load_workbook
import logging

import clients.models as clients
import directions.models as directions
from directory.models import ParaclinicInputGroups, Researches, ParaclinicInputField
from users.models import DoctorProfile
from utils.dates import normalize_dots_date
from django.db import transaction
from django.utils import timezone

from utils.nsi_directories import NSI
from api.directions.views import process_to_sign_direction

logger = logging.getLogger("IF")


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл с пациентами
               path_distr - файл с участками
        """
        parser.add_argument("path", type=str)
        parser.add_argument("id_research", type=str)
        parser.add_argument("id_doc_profile", type=str)

    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        stdout.write(fp)
        wb = load_workbook(filename=fp)
        title_fields = "fio,lastname,firstname,patronymic,sex,birthday,address,snils,enp,Диагноз,Дата осмотра,Группа здоровья,Вид места жительства"
        id_doc_profile = int(kwargs["id_doc_profile"])
        stdout.write(f"{id_doc_profile}")
        id_research = int(kwargs["id_research"])
        stdout.write(f"{id_research}")
        stdout.write(title_fields)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        doc_profile = DoctorProfile.objects.filter(pk=id_doc_profile).first()
        research = Researches.objects.filter(pk=id_research).first()
        starts = False
        snils_type = clients.DocumentType.objects.filter(title__startswith="СНИЛС").first()
        enp_type = clients.DocumentType.objects.filter(title__startswith="Полис ОМС").first()
        date_variant_title_field = ["Дата осмотра"]
        index_cell = {}
        incorrect_patients = []
        title_data = [i.strip() for i in title_fields.split(",")]
        financing_source_title = "омс"
        step = 0
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "fio" in cells:
                    starts = True
                    for t in title_data:
                        index_cell[t] = cells.index(t)
            else:
                data_result = {}
                for i_cell in index_cell.keys():
                    data_result[i_cell] = cells[index_cell[i_cell]]
                if len(data_result["snils"]) < 11 or len(data_result["enp"]) < 16:
                    continue
                snils = data_result["snils"].replace("-", "").replace(" ", "")
                birthday = normalize_dots_date(data_result["birthday"].split(" ")[0])
                sex = data_result["sex"].lower()

                individual, enp_doc, snils_doc = None, None, None
                if data_result["enp"]:
                    individuals = clients.Individual.objects.filter(tfoms_enp=data_result["enp"])
                    if not individuals.exists():
                        individuals = clients.Individual.objects.filter(document__number=data_result["enp"]).filter(
                            Q(document__document_type__title="Полис ОМС") | Q(document__document_type__title="ЕНП")
                        )
                        individual = individuals.first()

                if not individual:
                    individual = clients.Individual(
                        family=data_result["lastname"],
                        name=data_result["firstname"],
                        patronymic=data_result["patronymic"],
                        birthday=birthday,
                        sex=sex,
                    )
                    individual.save()

                if not clients.Document.objects.filter(individual=individual, document_type=snils_type).exists():
                    snils_doc = clients.Document(
                        individual=individual,
                        document_type=snils_type,
                        number=snils,
                    )
                    snils_doc.save()
                else:
                    snils_doc = clients.Document.objects.filter(individual=individual, document_type=snils_type).first()
                if not clients.Document.objects.filter(individual=individual, document_type=enp_type).exists():
                    enp_doc = clients.Document(
                        individual=individual,
                        document_type=enp_type,
                        number=data_result["enp"],
                    )
                    enp_doc.save()
                else:
                    enp_doc = clients.Document.objects.filter(individual=individual, document_type=enp_type).first()

                if not clients.Card.objects.filter(base__internal_type=True, individual=individual, is_archive=False).exists():
                    card = clients.Card.add_l2_card(individual, polis=enp_doc, snils=snils_doc)
                else:
                    card = clients.Card.objects.filter(base__internal_type=True, individual=individual, is_archive=False).first()
                card.main_address = data_result["address"]
                card.save(update_fields=["main_address"])

                financing_source = directions.IstochnikiFinansirovaniya.objects.filter(title__iexact=financing_source_title, base__internal_type=True).first()

                try:
                    with transaction.atomic():
                        direction = directions.Napravleniya.objects.create(
                            client=card,
                            istochnik_f=financing_source,
                            polis_who_give=card.polis.who_give if card.polis else None,
                            polis_n=card.polis.number if card.polis else None,
                            hospital=doc_profile.hospital,
                            total_confirmed=True,
                            last_confirmed_at=timezone.now(),
                            eds_required_signature_types=["Врач", "Медицинская организация"],
                        )

                        iss = directions.Issledovaniya.objects.create(
                            napravleniye=direction,
                            research=research,
                            time_confirmation=timezone.now(),
                            time_save=timezone.now(),
                            doc_confirmation=doc_profile,
                            doc_save=doc_profile,
                            doc_confirmation_string=f"{doc_profile.get_fio_parts()}",
                        )

                        for group in ParaclinicInputGroups.objects.filter(research=research):
                            for f in ParaclinicInputField.objects.filter(group=group):
                                if data_result.get(f.title, None):
                                    if f.title.strip() in date_variant_title_field:
                                        tmp_val = data_result[f.title]
                                        data_result[f.title] = normalize_dots_date(tmp_val)
                                    if f.title.strip() == "Диагноз":
                                        tmp_val = data_result[f.title].strip()
                                        diag = directions.Diagnoses.objects.filter(d_type="mkb10.4", code__exact=tmp_val.upper()).first()
                                        if not diag:
                                            incorrect_patients.append({"fio": data_result["fio"], "reason": "Неверные данные:"})
                                        res = f'"code": "{tmp_val}", "title": "{diag.title}", "id": "{diag.id}"'
                                        res = "{" + res + "}"
                                        data_result[f.title] = res
                                    if f.field_type == 28:
                                        for nsi_key in NSI.values():
                                            if f.title == nsi_key.get("title"):
                                                for key, val in nsi_key.get("values").items():
                                                    if val == data_result[f.title].strip():
                                                        res = f'"code": "{key}", "title": "{val}"'
                                                        res = "{" + res + "}"
                                                        data_result[f.title] = res
                                                        continue
                                    directions.ParaclinicResult(issledovaniye=iss, field=f, field_type=f.field_type, value=data_result.get(f.title)).save()
                        stdout.write(f"Добавлена карта: \n, {direction.pk}, {card}")
                        result = process_to_sign_direction(direction, direction.pk, doc_profile.user, iss)
                        stdout.write(f"sign_data: {result}")
                except Exception as e:
                    logger.exception(e)
                    message = "Серверная ошибка"
                    return {"ok": False, "message": message}
            step += 1
            stdout.write(f"{step}")
            stdout.write(f"Добавлена карта: \n, {direction.pk}, {card}")
