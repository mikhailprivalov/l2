from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from contracts.models import PriceName, PriceCoast
from directory.models import Researches


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл  со столбцами:
        Код прайса, Услуга, Код ОКМУ, далее - динамические колонки (коды прайсов)

        """
        fp = kwargs["path"]
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        internal_code_idx, service_name_idx, nmy_code_idx = '', '', ''
        not_price_col = []
        columns = {}
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Код прайса" in cells:
                    internal_code_idx = cells.index("Код прайса")
                    service_name_idx = cells.index("Услуга")
                    nmy_code_idx = cells.index("Код ОКМУ")
                    not_price_col = [internal_code_idx, service_name_idx, nmy_code_idx]
                    for idx, column in enumerate(cells):
                        columns[idx] = column.strip()
                    starts = True
            else:
                internal_code = cells[internal_code_idx].strip()
                if internal_code == "None":
                    self.stdout.write('Нет внутреннего кода')
                    continue
                service: Researches = Researches.objects.filter(internal_code=internal_code).first()
                if not service:
                    self.stdout.write('Услуга не найдена')
                    continue
                for idx, cell in enumerate(cells):
                    if idx in not_price_col:
                        continue
                    service_coast = cell.strip()
                    if service_coast == "None":
                        continue
                    price_code = columns[idx]
                    price: PriceName = PriceName.objects.filter(symbol_code=price_code).first()
                    if not price:
                        self.stdout.write('Нет такого прайса')
                        continue
                    current_price_coasts: PriceCoast = PriceCoast.objects.filter(price_name_id=price.pk, research_id=service.pk).first()
                    if current_price_coasts:
                        if current_price_coasts.coast != service_coast:
                            current_price_coasts.coast = service_coast
                            current_price_coasts.save()
                            self.stdout.write(f'Цена услуги {service.title} в прайсе {price.title} обновлена')
                        else:
                            self.stdout.write(f'Цена услуги {service.title} в прайсе {price.title} совпадает')
                    else:
                        new_price_coasts: PriceCoast = PriceCoast(price_name_id=price.pk, research_id=service.pk, coast=service_coast, number_services_by_contract=1)
                        self.stdout.write(f'Добавлена услуга {service.title} в прайс {price.title}')
                        new_price_coasts.save()
