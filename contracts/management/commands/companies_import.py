from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from contracts.models import Company
from django.db.models import Q


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл со столбцами:
        Название, Короткое название, ИНН (опционально)
        """
        fp = kwargs["path"]
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        title, short_title, inn = '', '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Название" in cells:
                    title = cells.index("Название")
                    short_title = cells.index("Короткое название")
                    starts = True
                if "ИНН" in cells:
                    inn = cells.index("ИНН")
            else:
                title_string = cells[title].strip()
                short_title_string = cells[short_title].strip()
                inn_string = None
                if inn:
                    inn_string = cells[inn].strip()
                    company = Company.objects.filter(inn=inn_string)
                else:
                    company = Company.objects.filter(Q(title__iexact=title_string) | Q(short_title__iexact=short_title_string))
                if not company.exists():
                    new_company = Company(title=title_string, short_title=short_title_string)
                    if inn:
                        new_company.inn = inn_string
                    new_company.save()
                    self.stdout.write('компания создана ' + title_string)
