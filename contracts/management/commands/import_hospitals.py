from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from hospitals.models import Hospitals


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл со столбцами:
        Имя, Полное название учреждения, ИНН
        """
        fp = kwargs["path"]
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        short_title_idx, title_idx, inn_idx = '', '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Полное название учреждения" in cells:
                    short_title_idx = cells.index("Имя")
                    title_idx = cells.index("Полное название учреждения")
                    starts = True
            else:
                short_title = cells[short_title_idx].strip()
                title = cells[title_idx].strip()
                if title == "None":
                    continue
                current_hospital = Hospitals.objects.filter(title__iexact=title).first()
                if current_hospital:
                    self.stdout.write('Такая больница уже есть')
                else:
                    new_hospital = Hospitals(title=title, short_title=short_title)
                    new_hospital.save()
                    self.stdout.write('компания создана ')
