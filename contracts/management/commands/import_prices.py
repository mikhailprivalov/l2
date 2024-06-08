from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from contracts.models import PriceName
from django.db.models import Q

from hospitals.models import Hospitals


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл  со столбцами:
        Плательщик/Контрагент, Тариф, Действует С, Код
        """
        fp = kwargs["path"]
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        hospital_idx, price_name_idx, price_start_idx, code_idx = '', '', '', ''
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "Тарифный план" in cells:
                    hospital_idx = cells.index("Плательщик/Контрагент")
                    price_name_idx = cells.index("Тариф")
                    price_start_idx = cells.index("Действует С")
                    code_idx = cells.index("Код")
                    starts = True
            else:
                price_start = cells[price_start_idx].split(" ")[0]
                price_code = cells[code_idx].strip()
                if price_code != "None":
                    hospital_string = cells[hospital_idx].strip()
                    hospital = Hospitals.objects.filter(Q(title__iexact=hospital_string) | Q(short_title__iexact=hospital_string)).first()
                    if hospital:
                        price_title = f"{hospital.short_title}-{cells[price_name_idx].strip()}-{price_code}"
                    else:
                        self.stdout.write(f" Больницы {hospital_string} нет")
                        continue
                    current_price = PriceName.objects.filter(symbol_code=price_code)
                    if not current_price.exists():
                        new_price = PriceName(title=price_title, date_start=price_start, hospital_id=hospital.pk, subcontract=True, symbol_code=price_code)
                        new_price.save()
                        self.stdout.write(f"Прайс {price_title} создан")
                    else:
                        self.stdout.write(f"Такой прайс уже есть - {price_title}")
                else:
                    self.stdout.write("нет кода прайса")
