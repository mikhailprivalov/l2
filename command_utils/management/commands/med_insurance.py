import os

from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from laboratory.settings import BASE_DIR


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл с микроорганизмами со столбцами:
        """
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        smo_project_file_path = os.path.join('utils', 'nsi_medinsurance.py')
        smo_file_path = os.path.join(BASE_DIR, smo_project_file_path)
        smo_data = {}
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "ID" in cells:
                    smo_id = cells.index("ID")
                    smocod = cells.index("SMOCOD")
                    starts = True
            else:
                smo_data[cells[smocod]] = cells[smo_id]
        import json

        with open(smo_file_path, 'w') as file:
            file.write(json.dumps(smo_data, ensure_ascii=False))
