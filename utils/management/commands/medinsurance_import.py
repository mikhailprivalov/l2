import os

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from directory.models import Antibiotic, GroupAntibiotic
from laboratory.settings import BASE_DIR


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('path', type=str)

    def handle(self, *args, **kwargs):
        """
        :param path - xlsx файл с микроорганизмами со столбцами:
        Название, Группа, LIS(код)
        """

        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        title, group, lis = '', '', ''
        smo_project_file_path = os.path.join('utils', 'nsi_medinsurance.py')
        smo_file_path = os.path.join(BASE_DIR, smo_project_file_path)
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "ID" in cells:
                    smo_id = cells.index("ID")
                    smocod = cells.index("SMOCOD")
                    ogrn = cells.index("OGRN")
                    kpp = cells.index("KPP")
                    nam_smop = cells.index("NAM_SMOP")
                    nam_smok = cells.index("NAM_SMOK")
                    addr_f = cells.index("ADDR_F")
                    n_doc = cells.index("N_DOC")
                    print(cells[smocod]) # noqa: T001
                    starts = True
